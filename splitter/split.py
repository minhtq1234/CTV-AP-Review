#!/usr/bin/env python3
"""Deterministic slice of the CTV folder splitter.

Turns ONE eForm submission — the schedule Excel + the CCCD photo Excel + the
scanned PDF — into one folder per collaborator (CTV), each holding a manifest
with the *claimed* Excel values plus evidence slots.

This slice does only what needs NO OCR:
  - parse the schedule Excel        (real, authoritative — drives the folder list)
  - count the embedded CCCD images  (real)
  - read the PDF page count         (real)
  - segment the PDF into per-CTV blocks   (STUB: order-based, flagged needsOcr)
  - emit the folder tree + manifests + a report

Everything that needs OCR — reading values off the documents, and linking each
scan to the right person by CCCD — is left as an explicit stub (`needsOcr: true`).
That is exactly the gap the real extraction pass would fill; the manifest shape
already matches what the review app consumes.

Usage:
  python3 split.py [--out DIR] [--extract-media] [--render-pdf]
"""
import argparse, json, re, unicodedata, zipfile
from pathlib import Path

DOWNLOADS = Path("/Users/lap16603/Downloads/Chi phí Cộng tác viên")
SCHEDULE = DOWNLOADS / "BẢNG KÊ THANH TOÁN CTV -THÁNG 2.2026.xlsx"
PHOTOS = DOWNLOADS / "CCCD_T2.xlsx"
PDF = DOWNLOADS / "FA-PM260226080.pdf"
FRONT_MATTER = 5  # pp1-2 signed schedule, pp3-5 tax form (Mẫu 01/TNDN)

# (field key, label, which claimed column, which docs the real OCR pass would cross-check)
FIELDS = [
    ("name", "Họ và tên", "name", ["id_front", "contract", "pit", "bbnt"]),
    ("cccd", "Số CCCD", "cccd", ["id_front", "contract", "pit", "bbnt"]),
    ("dob", "Ngày sinh", "dob", ["id_front"]),
    ("bank_acct", "Số tài khoản", "bank_account", ["contract"]),
    ("bank_name", "Ngân hàng", "bank_name", ["contract"]),
    ("gross", "Phí dịch vụ (Gross)", "gross", ["contract", "pit"]),
    ("pit", "Thuế PIT", "pit", ["pit"]),
    ("net", "Thực nhận", "net", []),
]


def slug(s: str) -> str:
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c)).replace("đ", "d").replace("Đ", "D")
    return re.sub(r"[^a-z0-9]+", "-", s.lower()).strip("-")


def parse_schedule(path: Path):
    import openpyxl
    ws = openpyxl.load_workbook(path, data_only=True).active
    rows = []
    for r in range(6, ws.max_row + 1):
        stt = ws.cell(r, 1).value
        if not isinstance(stt, int):  # stop at the totals row / blanks
            continue

        def g(c):
            v = ws.cell(r, c).value
            return "" if v is None else (v if isinstance(v, (int, float)) else str(v).strip())

        rows.append({
            "stt": stt, "name": g(2), "cccd": g(3), "mst": g(4), "dob": g(5), "sex": g(6),
            "bank_account": g(7), "bank_name": g(8), "work_period": g(9),
            "gross": g(11), "commit": g(12), "pit": g(13), "net": g(14), "product": g(15),
        })
    return rows


def cccd_image_names(path: Path):
    with zipfile.ZipFile(path) as z:
        return [n for n in z.namelist() if n.startswith("xl/media/image")]


def pdf_page_count(path: Path) -> int:
    import fitz
    d = fitz.open(path)
    try:
        return d.page_count
    finally:
        d.close()


def stub_pdf_blocks(page_count: int, n_rows: int):
    """Order-based placeholder: evenly split the post-front-matter pages across
    the CTV rows. The real splitter would cut on document-title / page-number
    boundaries and confirm by OCR'd CCCD — hence `method: stub-order`."""
    total = page_count - FRONT_MATTER
    base, extra = divmod(max(total, 0), n_rows)
    blocks, p = [], FRONT_MATTER + 1
    for i in range(n_rows):
        size = base + (1 if i < extra else 0)
        blocks.append((p, p + size - 1))
        p += size
    return blocks


def extract_media(out: Path, render_pdf: bool):
    media = out / "_media"
    (media / "cccd").mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(PHOTOS) as z:
        for n in cccd_image_names(PHOTOS):
            (media / "cccd" / Path(n).name).write_bytes(z.read(n))
    if render_pdf:
        import fitz
        (media / "pdf").mkdir(parents=True, exist_ok=True)
        d = fitz.open(PDF)
        for i in range(d.page_count):
            pix = d[i].get_pixmap(matrix=fitz.Matrix(100 / 72, 100 / 72))
            pix.save(str(media / "pdf" / f"p{i + 1:03d}.png"))
        d.close()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(Path(__file__).parent / "output"))
    ap.add_argument("--extract-media", action="store_true", help="also write CCCD images (PII)")
    ap.add_argument("--render-pdf", action="store_true", help="also rasterize PDF pages (implies media)")
    args = ap.parse_args()
    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)

    rows = parse_schedule(SCHEDULE)
    images = cccd_image_names(PHOTOS)
    pages = pdf_page_count(PDF)
    blocks = stub_pdf_blocks(pages, len(rows))

    for i, row in enumerate(rows):
        cid = f"ctv-{row['stt']:02d}-{slug(row['name'])}"
        folder = out / cid
        folder.mkdir(parents=True, exist_ok=True)
        exempt = "có" in str(row["commit"]).lower()
        lo, hi = blocks[i]
        flags = ["evidence-linking-unconfirmed: cần OCR số CCCD để gán chứng từ"]
        if exempt:
            flags.append("PIT-exempt: cần xác định bản cam kết trong khối PDF")
        manifest = {
            "id": cid, "stt": row["stt"], "name": row["name"], "product": row["product"],
            "status": "pending", "exempt": exempt,
            "claimed": {k: row[k] for k in
                        ("name", "cccd", "mst", "dob", "sex", "bank_account", "bank_name",
                         "work_period", "gross", "pit", "net")},
            "evidence": {
                "id_card": {"source": PHOTOS.name, "status": "in-pool", "needsOcr": True},
                "pdf_block": {"source": PDF.name, "pages": f"{lo}-{hi}",
                              "method": "stub-order", "needsOcr": True},
            },
            "fields": [
                {"key": k, "label": lbl, "expected": row[ck], "crossCheckDocs": docs,
                 "sources": [], "needsOcr": True}
                for (k, lbl, ck, docs) in FIELDS
            ],
            "flags": flags,
        }
        (folder / "manifest.json").write_text(json.dumps(manifest, ensure_ascii=False, indent=2))

    report = {
        "inputs": {"schedule": SCHEDULE.name, "photos": PHOTOS.name, "pdf": PDF.name},
        "rows": len(rows),
        "stt_range": [rows[0]["stt"], rows[-1]["stt"]] if rows else [],
        "missing_stt": [n for n in range(rows[0]["stt"], rows[-1]["stt"] + 1)
                        if n not in {r["stt"] for r in rows}] if rows else [],
        "cccd_images": len(images),
        "cccd_pairs_est": len(images) // 2,
        "pdf_pages": pages,
        "front_matter_pages": f"1-{FRONT_MATTER}",
        "folders_created": len(rows),
        "pdf_block_size_est": (pages - FRONT_MATTER) // len(rows) if rows else 0,
        "stubbed": ["cccd-image linking (needs OCR)", "pdf block boundaries (order-based)",
                    "field extraction (needs OCR)"],
    }
    (out / "report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2))

    if args.extract_media or args.render_pdf:
        extract_media(out, args.render_pdf)

    print(f"✓ {len(rows)} folders → {out}")
    print(f"  schedule rows: {report['rows']}  (STT {report['stt_range']}, missing {report['missing_stt']})")
    print(f"  CCCD images: {report['cccd_images']}  (~{report['cccd_pairs_est']} pairs)")
    print(f"  PDF: {pages} pages  (front-matter 1-{FRONT_MATTER}, ~{report['pdf_block_size_est']} pp/CTV)")
    print(f"  media written: {bool(args.extract_media or args.render_pdf)}")


if __name__ == "__main__":
    main()
