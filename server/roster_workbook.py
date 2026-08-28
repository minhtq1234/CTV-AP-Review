"""Bounded validation and loading for roster XLSX workbooks."""
from __future__ import annotations

import os
import re
from pathlib import PurePosixPath
from xml.etree import ElementTree as ET
import zipfile

import openpyxl
from openpyxl.xml.constants import CONTYPES_NS, XLSM, XLSX, XLTM, XLTX

from ooxml import OoxmlRelationshipError, resolve_internal_relationship_target
import workbook_layout


MAX_WORKBOOK_BYTES = 25 * 1024 * 1024
MAX_ARCHIVE_MEMBERS = 2_000
MAX_ARCHIVE_MEMBER_BYTES = 25 * 1024 * 1024
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 100 * 1024 * 1024
MAX_XML_BYTES = 16 * 1024 * 1024
MAX_WORKSHEET_ROWS = 10_000
MAX_WORKSHEET_COLUMNS = 128
MAX_WORKSHEET_CELLS = 500_000
MAX_STRING_CHARACTERS = 32_767

_SHEET_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_DOC_REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_CONTENT_TYPES_NS = f"{{{CONTYPES_NS}}}"
_WORKSHEET_REL_TYPE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet"
)
_CONTENT_TYPES_PART = "[Content_Types].xml"
_DEFAULT_WORKBOOK_PART = "xl/workbook.xml"
_WORKBOOK_CONTENT_TYPES = frozenset({XLTM, XLTX, XLSM, XLSX})
_CELL_REF_RE = re.compile(r"^\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")


class RosterWorkbookError(ValueError):
    """Raised when a roster workbook is malformed or exceeds a safety limit."""


class _ByteBudget:
    def __init__(self, limit: int):
        self._remaining = limit

    @property
    def remaining(self) -> int:
        return self._remaining

    def consume(self, size: int) -> None:
        if size > self._remaining:
            raise RosterWorkbookError("archive-uncompressed-too-large")
        self._remaining -= size


class _BoundedReader:
    def __init__(
        self,
        stream,
        member_limit: int,
        byte_budget: _ByteBudget,
    ):
        self._stream = stream
        self._member_limit = member_limit
        self._byte_budget = byte_budget
        self._read = 0

    def read(self, size: int = -1) -> bytes:
        member_remaining = self._member_limit - self._read
        requested = (
            member_remaining + 1
            if size is None or size < 0
            else min(size, member_remaining + 1)
        )
        requested = min(requested, self._byte_budget.remaining + 1)
        content = self._stream.read(requested)
        if len(content) > member_remaining:
            raise RosterWorkbookError("xml-too-large")
        self._byte_budget.consume(len(content))
        self._read += len(content)
        return content

    def close(self) -> None:
        self._stream.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        self.close()


def load_roster_rows(xlsx_source) -> list[list]:
    """Preflight an XLSX, then load the sheet that looks like the bảng kê.

    Not `workbook.active`: that is whichever tab the submitter happened to have
    selected when they saved, and on the combined template it resolves to the
    CCCD sheet -- which parses cleanly with only name, CCCD and STT, leaving
    every money and identity criterion with nothing to compare against, and no
    error anywhere. The sheet is chosen by which one maps the most known
    columns (`workbook_layout.select_roster_sheet`).
    """
    preflight_roster_workbook(xlsx_source)
    _seek_start(xlsx_source)
    workbook = openpyxl.load_workbook(
        xlsx_source,
        read_only=True,
        data_only=True,
    )
    try:
        sheets = {
            name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
            for name in workbook.sheetnames
        }
        chosen = workbook_layout.select_roster_sheet(sheets)
        if chosen is None:
            # Preflight should already have refused this; treat it as a bug
            # rather than silently reading an arbitrary sheet.
            raise RosterWorkbookError("no-roster-sheet")
        return sheets[chosen]
    finally:
        workbook.close()
        _seek_start(xlsx_source)


def preflight_roster_workbook(xlsx_source) -> None:
    """Stream-validate the workbook container and XML before OpenPyXL runs."""
    original_position = _position(xlsx_source)
    try:
        if _source_size(xlsx_source) > MAX_WORKBOOK_BYTES:
            raise RosterWorkbookError("workbook-too-large")
        _seek_start(xlsx_source)
        with zipfile.ZipFile(xlsx_source) as archive:
            members = _validate_archive(archive)
            if _CONTENT_TYPES_PART not in members:
                raise RosterWorkbookError("missing-required-part")
            budget = _ByteBudget(MAX_ARCHIVE_UNCOMPRESSED_BYTES)
            workbook_part, workbook_rels_part = _workbook_parts(
                archive,
                members,
                budget,
            )
            worksheets = _worksheet_targets(
                archive,
                members,
                budget,
                workbook_part,
                workbook_rels_part,
            )
            for name in sorted(members):
                if name in {
                    _CONTENT_TYPES_PART,
                    workbook_part,
                    workbook_rels_part,
                }:
                    continue
                if name.endswith((".xml", ".rels")):
                    _stream_xml(
                        archive,
                        name,
                        budget,
                        worksheet=name in worksheets,
                    )
    except RosterWorkbookError:
        raise
    except (
        ET.ParseError,
        KeyError,
        OSError,
        ValueError,
        zipfile.BadZipFile,
        zipfile.LargeZipFile,
    ) as error:
        raise RosterWorkbookError("invalid-workbook") from error
    finally:
        _restore_position(xlsx_source, original_position)

    # The container is sound. Now: is any sheet usable as a bảng kê? A workbook
    # that parses but has no name/CCCD/money column produces a case where every
    # comparison silently has nothing to compare against, and the reviewer only
    # finds out after a full run.
    _seek_start(xlsx_source)
    workbook = openpyxl.load_workbook(xlsx_source, read_only=True, data_only=True)
    try:
        sheets = {
            name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
            for name in workbook.sheetnames
        }
    finally:
        workbook.close()
        _seek_start(xlsx_source)

    # A sheet can be selected and still be unusable: `select_roster_sheet` only
    # requires name and CCCD, while a bảng kê with no money column has nothing
    # to pay against. Refuse on the same set the diagnostic reports, so the gate
    # and the message can never disagree.
    chosen = workbook_layout.select_roster_sheet(sheets)
    best = chosen or max(
        sheets,
        key=lambda n: workbook_layout.score_roster_sheet(sheets[n]),
        default=None,
    )
    missing = (
        workbook_layout.missing_required_columns(sheets[best])
        if best else ["name", "cccd", "money"]
    )
    if missing:
        raise RosterWorkbookError(
            f"no-roster-sheet: read '{best}', missing {', '.join(missing)}"
        )


def _validate_archive(archive: zipfile.ZipFile) -> set[str]:
    infos = archive.infolist()
    if len(infos) > MAX_ARCHIVE_MEMBERS:
        raise RosterWorkbookError("archive-member-limit")
    names: set[str] = set()
    total_uncompressed = 0
    for info in infos:
        if info.filename in names:
            raise RosterWorkbookError("duplicate-archive-member")
        names.add(info.filename)
        path = PurePosixPath(info.filename)
        if path.is_absolute() or ".." in path.parts or "\\" in info.filename:
            raise RosterWorkbookError("invalid-archive-member")
        if info.flag_bits & 0x1:
            raise RosterWorkbookError("encrypted-entry")
        if info.file_size > MAX_ARCHIVE_MEMBER_BYTES:
            raise RosterWorkbookError("archive-member-too-large")
        total_uncompressed += info.file_size
        if total_uncompressed > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
            raise RosterWorkbookError("archive-uncompressed-too-large")
    return names


def _workbook_parts(
    archive: zipfile.ZipFile,
    members: set[str],
    byte_budget: _ByteBudget,
) -> tuple[str, str]:
    content_types = _parse_xml_root(
        archive,
        _CONTENT_TYPES_PART,
        byte_budget,
    )
    if content_types.tag != f"{_CONTENT_TYPES_NS}Types":
        raise RosterWorkbookError("invalid-content-types")

    overrides: list[str] = []
    default_declarations = 0
    seen_parts: set[str] = set()
    seen_extensions: set[str] = set()
    for element in content_types:
        if element.tag == f"{_CONTENT_TYPES_NS}Override":
            part_name = element.attrib.get("PartName")
            content_type = element.attrib.get("ContentType")
            if not part_name or not content_type:
                raise RosterWorkbookError("invalid-content-types")
            part = _normalize_package_part_name(part_name)
            if part in seen_parts:
                raise RosterWorkbookError("duplicate-content-type-part")
            seen_parts.add(part)
            if content_type in _WORKBOOK_CONTENT_TYPES:
                overrides.append(part)
        elif element.tag == f"{_CONTENT_TYPES_NS}Default":
            extension = element.attrib.get("Extension")
            content_type = element.attrib.get("ContentType")
            if not extension or not content_type:
                raise RosterWorkbookError("invalid-content-types")
            normalized_extension = extension.casefold()
            if normalized_extension in seen_extensions:
                raise RosterWorkbookError("duplicate-content-type-default")
            seen_extensions.add(normalized_extension)
            if content_type in _WORKBOOK_CONTENT_TYPES:
                default_declarations += 1

    declaration_count = len(overrides) + default_declarations
    if declaration_count == 0:
        raise RosterWorkbookError("missing-workbook-part")
    if declaration_count > 1:
        raise RosterWorkbookError("multiple-workbook-parts")
    workbook_part = overrides[0] if overrides else _DEFAULT_WORKBOOK_PART
    workbook_rels_part = _relationship_part_name(workbook_part)
    if workbook_part not in members or workbook_rels_part not in members:
        raise RosterWorkbookError("missing-required-part")
    return workbook_part, workbook_rels_part


def _normalize_package_part_name(part_name: str) -> str:
    if (
        not part_name.startswith("/")
        or part_name.startswith("//")
        or "\\" in part_name
        or "\x00" in part_name
        or "?" in part_name
        or "#" in part_name
    ):
        raise RosterWorkbookError("invalid-workbook-part")
    raw_part = part_name[1:]
    part = PurePosixPath(raw_part)
    if (
        raw_part in {"", "."}
        or part.is_absolute()
        or ".." in part.parts
        or str(part) != raw_part
    ):
        raise RosterWorkbookError("invalid-workbook-part")
    return str(part)


def _relationship_part_name(source_part: str) -> str:
    source = PurePosixPath(source_part)
    return str(source.parent / "_rels" / f"{source.name}.rels")


def _worksheet_targets(
    archive: zipfile.ZipFile,
    members: set[str],
    byte_budget: _ByteBudget,
    workbook_part: str,
    workbook_rels_part: str,
) -> set[str]:
    workbook = _parse_xml_root(archive, workbook_part, byte_budget)
    rels_root = _parse_xml_root(
        archive,
        workbook_rels_part,
        byte_budget,
    )
    relationships: dict[str, dict[str, object]] = {}
    for element in rels_root.findall(f"{_REL_NS}Relationship"):
        rel_id = element.attrib.get("Id")
        rel_type = element.attrib.get("Type")
        target = element.attrib.get("Target")
        if not rel_id or not rel_type or not target or rel_id in relationships:
            raise RosterWorkbookError("invalid-workbook-relationships")
        target_mode = element.attrib.get("TargetMode")
        relationships[rel_id] = {
            "type": rel_type,
            "target": target,
            "external": target_mode is not None and target_mode != "Internal",
        }

    sheet_nodes = workbook.findall(f".//{_SHEET_NS}sheet")
    if not sheet_nodes:
        raise RosterWorkbookError("missing-worksheet")
    targets: set[str] = set()
    used_relationships: set[str] = set()
    for sheet in sheet_nodes:
        rel_id = sheet.attrib.get(f"{_DOC_REL_NS}id")
        if not rel_id or rel_id in used_relationships:
            raise RosterWorkbookError("invalid-workbook-relationships")
        used_relationships.add(rel_id)
        relationship = relationships.get(rel_id)
        if relationship is None or relationship["type"] != _WORKSHEET_REL_TYPE:
            raise RosterWorkbookError("invalid-workbook-relationships")
        try:
            target = resolve_internal_relationship_target(
                workbook_part,
                relationship["target"],
                external=relationship["external"],
            )
        except OoxmlRelationshipError as error:
            raise RosterWorkbookError(str(error)) from error
        if (
            not target.endswith(".xml")
            or target not in members
            or target in targets
        ):
            raise RosterWorkbookError("invalid-workbook-relationships")
        targets.add(target)
    return targets


def _parse_xml_root(
    archive: zipfile.ZipFile,
    member: str,
    byte_budget: _ByteBudget,
) -> ET.Element:
    info = archive.getinfo(member)
    if info.file_size > MAX_XML_BYTES:
        raise RosterWorkbookError("xml-too-large")
    with _BoundedReader(
        archive.open(info),
        MAX_XML_BYTES,
        byte_budget,
    ) as stream:
        root = ET.parse(stream).getroot()
    for element in root.iter():
        if len(element.text or "") > MAX_STRING_CHARACTERS:
            raise RosterWorkbookError("string-too-large")
    return root


def _stream_xml(
    archive: zipfile.ZipFile,
    member: str,
    byte_budget: _ByteBudget,
    *,
    worksheet: bool,
) -> None:
    info = archive.getinfo(member)
    if info.file_size > MAX_XML_BYTES:
        raise RosterWorkbookError("xml-too-large")
    row_count = 0
    cell_count = 0
    string_characters = 0
    string_depth = 0
    with _BoundedReader(
        archive.open(info),
        MAX_XML_BYTES,
        byte_budget,
    ) as stream:
        for event, element in ET.iterparse(stream, events=("start", "end")):
            if event == "start":
                if element.tag in {f"{_SHEET_NS}si", f"{_SHEET_NS}is"}:
                    string_depth += 1
                    string_characters = 0
                continue
            text = element.text or ""
            if len(text) > MAX_STRING_CHARACTERS:
                raise RosterWorkbookError("string-too-large")
            if string_depth and element.tag == f"{_SHEET_NS}t":
                string_characters += len(text)
                if string_characters > MAX_STRING_CHARACTERS:
                    raise RosterWorkbookError("string-too-large")
            if element.tag in {f"{_SHEET_NS}si", f"{_SHEET_NS}is"}:
                string_depth -= 1
                string_characters = 0
            if worksheet:
                if element.tag == f"{_SHEET_NS}dimension":
                    _validate_dimension(element.attrib.get("ref", ""))
                elif element.tag == f"{_SHEET_NS}row":
                    row_count += 1
                    if row_count > MAX_WORKSHEET_ROWS:
                        raise RosterWorkbookError("worksheet-row-limit")
                    row_reference = element.attrib.get("r")
                    if row_reference:
                        _validate_row_number(row_reference)
                elif element.tag == f"{_SHEET_NS}c":
                    cell_count += 1
                    if cell_count > MAX_WORKSHEET_CELLS:
                        raise RosterWorkbookError("worksheet-cell-limit")
                    cell_reference = element.attrib.get("r")
                    if cell_reference:
                        _validate_cell_reference(cell_reference)
            element.clear()


def _validate_dimension(reference: str) -> None:
    if not reference:
        raise RosterWorkbookError("invalid-worksheet-dimension")
    endpoints = reference.split(":")
    if len(endpoints) not in {1, 2}:
        raise RosterWorkbookError("invalid-worksheet-dimension")
    start_column, start_row = _parse_cell_reference(endpoints[0])
    end_column, end_row = _parse_cell_reference(endpoints[-1])
    if end_column < start_column or end_row < start_row:
        raise RosterWorkbookError("invalid-worksheet-dimension")
    _validate_grid_bounds(end_row, end_column)
    if (
        (end_row - start_row + 1) * (end_column - start_column + 1)
        > MAX_WORKSHEET_CELLS
    ):
        raise RosterWorkbookError("worksheet-cell-limit")


def _validate_row_number(reference: str) -> None:
    try:
        row = int(reference)
    except ValueError as error:
        raise RosterWorkbookError("invalid-row-reference") from error
    if row < 1 or row > MAX_WORKSHEET_ROWS:
        raise RosterWorkbookError("worksheet-row-limit")


def _validate_cell_reference(reference: str) -> None:
    column, row = _parse_cell_reference(reference)
    _validate_grid_bounds(row, column)


def _parse_cell_reference(reference: str) -> tuple[int, int]:
    match = _CELL_REF_RE.fullmatch(reference)
    if match is None:
        raise RosterWorkbookError("invalid-cell-reference")
    column = 0
    for character in match.group(1).upper():
        column = column * 26 + ord(character) - ord("A") + 1
    return column, int(match.group(2))


def _validate_grid_bounds(row: int, column: int) -> None:
    if row > MAX_WORKSHEET_ROWS:
        raise RosterWorkbookError("worksheet-row-limit")
    if column > MAX_WORKSHEET_COLUMNS:
        raise RosterWorkbookError("worksheet-column-limit")


def _position(source) -> int | None:
    if isinstance(source, (str, bytes, os.PathLike)):
        return None
    return source.tell()


def _source_size(source) -> int:
    if isinstance(source, (str, bytes, os.PathLike)):
        return os.path.getsize(source)
    position = source.tell()
    source.seek(0, os.SEEK_END)
    size = source.tell()
    source.seek(position)
    return size


def _seek_start(source) -> None:
    if not isinstance(source, (str, bytes, os.PathLike)):
        source.seek(0)


def _restore_position(source, position: int | None) -> None:
    if position is not None:
        source.seek(position)
