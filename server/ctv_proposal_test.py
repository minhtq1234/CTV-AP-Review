from io import BytesIO
import json

from openpyxl import Workbook
from PIL import Image
import pytest

import ctv_proposal
from ctv_grouping_evidence import GroupingEvidence
from ctv_proposal import ProposalState
from ctv_inspection import inspect_observation
from ctv_inspection_model import (
    InspectionResult,
    InspectionSource,
    InspectionTotals,
    InspectionUnit,
)
from ctv_inventory import InventoryError, InventoryObservation, open_inventory_observation


def _roster_bytes(rows=(("Alice", "CTV-001"), ("Bao", "CTV-002"))):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Payment roster"
    sheet.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    for row in rows:
        if len(row) == 2:
            sheet.append((*row, "FA-SYNTHETIC-001", 100))
        elif len(row) == 3 and any(value is not None for value in row):
            sheet.append((row[0], row[1], "FA-SYNTHETIC-001", row[2]))
        else:
            sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _source(tmp_path, *, roster_rows=(("Alice", "CTV-001"), ("Bao", "CTV-002"))):
    source = tmp_path / "source"
    source.mkdir()
    (source / "roster.xlsx").write_bytes(_roster_bytes(roster_rows))
    image = BytesIO()
    with Image.new("RGB", (10, 10)) as value:
        value.save(image, format="PNG")
    (source / "identity.png").write_bytes(image.getvalue())
    return source


def _source_with_two_rosters(tmp_path):
    source = tmp_path / "source"
    source.mkdir()
    workbook = Workbook()
    first = workbook.active
    first.title = "First roster"
    first.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    first.append(("Alice", "CTV-001", "FA-SYNTHETIC-001", 100))
    first.append(("Bao", "CTV-002", "FA-SYNTHETIC-001", 100))
    second = workbook.create_sheet("Second roster")
    second.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    second.append(("Carol", "CTV-101", "FA-SYNTHETIC-002", 100))
    second.append(("Duy", "CTV-102", "FA-SYNTHETIC-002", 100))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    (source / "rosters.xlsx").write_bytes(output.getvalue())
    return source


def _source_with_valid_and_invalid_rosters(tmp_path):
    source = tmp_path / "source"
    source.mkdir()
    workbook = Workbook()
    valid = workbook.active
    valid.title = "Valid roster"
    valid.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    valid.append(("Alice", "CTV-001", "FA-SYNTHETIC-001", 100))
    invalid = workbook.create_sheet("Invalid roster")
    invalid.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    invalid.append(("Bao", None, "FA-SYNTHETIC-002", 100))
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    (source / "rosters.xlsx").write_bytes(output.getvalue())
    return source


def _state(tmp_path):
    source = _source(tmp_path)
    observation_context = open_inventory_observation(source)
    observation = observation_context.__enter__()
    inspection = inspect_observation(observation)
    return observation_context, observation, ProposalState.from_inspection(observation, inspection)


def _grouping_state(
    tmp_path,
    source_roles,
    *,
    source_only=(),
    leading_source_only=(),
    text_by_unit_number=None,
    unit_issue_codes_by_unit_number=None,
    duplicate_group_by_evidence_number=None,
):
    source = tmp_path / "generated-source"
    source.mkdir()
    (source / "roster.xlsx").write_bytes(_roster_bytes())
    context = open_inventory_observation(source)
    observation = context.__enter__()
    roster_inspection = inspect_observation(observation)
    roster_unit = roster_inspection.units[0]
    units = [roster_unit]
    sources = [roster_inspection.sources[0]]
    facts = GroupingEvidence()
    facts.capture(
        roster_unit.evidence_id,
        roster_unit.unit_kind,
        roster_unit.unit_index,
        "payment roster",
    )
    for status, detected_type, issue_codes in leading_source_only:
        evidence_number = len(sources) + 1
        sources.append(
            InspectionSource(
                evidence_id=f"evidence-{evidence_number:04d}",
                detected_type=detected_type,
                inspection_status=status,
                unit_count=None
                if status in {"unreadable", "encrypted", "over-limit"}
                else 0,
                issue_codes=issue_codes,
            )
        )
    unit_number = 2
    for evidence_number, roles in enumerate(source_roles, start=len(sources) + 1):
        evidence_id = f"evidence-{evidence_number:04d}"
        sources.append(
            InspectionSource(
                evidence_id=evidence_id,
                detected_type="pdf",
                inspection_status="inspected",
                unit_count=len(roles),
                issue_codes=(),
            )
        )
        for unit_index, role in enumerate(roles, start=1):
            issue_codes = (unit_issue_codes_by_unit_number or {}).get(
                unit_number, ()
            )
            unit = InspectionUnit(
                unit_id=f"unit-{unit_number:04d}",
                evidence_id=evidence_id,
                unit_kind="pdf-page",
                unit_index=unit_index,
                suggested_role=role,
                confidence_band="none" if role == "unknown" else "high",
                needs_user_review=role == "unknown" or bool(issue_codes),
                inspection_method="embedded-text",
                signal_codes=(),
                issue_codes=issue_codes,
            )
            units.append(unit)
            facts.capture(
                evidence_id,
                "pdf-page",
                unit_index,
                (text_by_unit_number or {}).get(
                    unit_number,
                    "unclassified continuation"
                    if role == "unknown"
                    else f"whole case {role} continuation",
                ),
            )
            unit_number += 1
    for status, detected_type, issue_codes in source_only:
        evidence_number = len(sources) + 1
        sources.append(
            InspectionSource(
                evidence_id=f"evidence-{evidence_number:04d}",
                detected_type=detected_type,
                inspection_status=status,
                unit_count=None if status in {"unreadable", "encrypted", "over-limit"} else 0,
                issue_codes=issue_codes,
            )
        )
    for evidence_number, duplicate_group_id in (
        duplicate_group_by_evidence_number or {}
    ).items():
        facts.capture_source_duplicate(
            f"evidence-{evidence_number:04d}", duplicate_group_id
        )
    issue_count = sum(len(source.issue_codes) for source in sources) + sum(
        len(unit.issue_codes) for unit in units
    )
    inspection = InspectionResult(
        inspection_version="1.0",
        inspection_status="complete-with-issues" if issue_count else "complete",
        observation_id=observation.observation_id,
        totals=InspectionTotals(
            sources=len(sources),
            units=len(units),
            classified=sum(unit.suggested_role != "unknown" for unit in units),
            unknown=sum(unit.suggested_role == "unknown" for unit in units),
            needs_user_review=sum(unit.needs_user_review for unit in units),
            issues=issue_count,
        ),
        sources=tuple(sources),
        units=tuple(units),
    )
    state = ProposalState.from_inspection(
        observation,
        inspection,
        _grouping_evidence=facts,
    )
    return context, state


def _missing_roster_grouping_state(tmp_path, *, unit_count=536):
    source = tmp_path / "missing-roster-source"
    source.mkdir()
    context = open_inventory_observation(source)
    observation = context.__enter__()
    remaining = unit_count - 301
    worksheet_splits = tuple(
        min(100, remaining - offset)
        for offset in range(0, remaining, 100)
    )
    source_specs = (
        ("pdf", "pdf-page", "embedded-text", 300),
        ("image", "image", "image-structure", 1),
        *(
            ("xlsx", "worksheet", "worksheet-structure", count)
            for count in worksheet_splits
        ),
    )
    sources = []
    units = []
    unit_number = 1
    for evidence_number, (
        detected_type,
        unit_kind,
        inspection_method,
        count,
    ) in enumerate(source_specs, start=1):
        evidence_id = f"evidence-{evidence_number:04d}"
        sources.append(
            InspectionSource(
                evidence_id=evidence_id,
                detected_type=detected_type,
                inspection_status="inspected",
                unit_count=count,
                issue_codes=(),
            )
        )
        for unit_index in range(1, count + 1):
            units.append(
                InspectionUnit(
                    unit_id=f"unit-{unit_number:04d}",
                    evidence_id=evidence_id,
                    unit_kind=unit_kind,
                    unit_index=unit_index,
                    suggested_role="unknown",
                    confidence_band="none",
                    needs_user_review=True,
                    inspection_method=inspection_method,
                    signal_codes=(),
                    issue_codes=(),
                )
            )
            unit_number += 1
    inspection = InspectionResult(
        inspection_version="1.0",
        inspection_status="complete",
        observation_id=observation.observation_id,
        totals=InspectionTotals(
            sources=len(sources),
            units=len(units),
            classified=0,
            unknown=len(units),
            needs_user_review=len(units),
            issues=0,
        ),
        sources=tuple(sources),
        units=tuple(units),
    )
    state = ProposalState.from_inspection(
        observation,
        inspection,
        _grouping_evidence=GroupingEvidence(),
        _roster_candidates=(),
    )
    return context, state


def test_missing_roster_fallback_accounts_for_every_large_mixed_unit_without_guessing(
    tmp_path,
):
    context, state = _missing_roster_grouping_state(tmp_path)
    try:
        local = state.local_review_snapshot()
        exceptions = local["review"]["exceptions"]
        groups = local["review"]["groups"]

        assert local["roster"]["status"] == "missing"
        assert exceptions == [
            {
                "exceptionId": "exception-0001",
                "kind": "roster",
                "issueCode": "roster-missing",
                "allowedActions": [],
                "similarityKey": exceptions[0]["similarityKey"],
            }
        ]
        assert local["review"]["coverage"] == {
            "groups": 5,
            "automaticallyOrganizedUnits": 0,
            "exceptionClusters": 1,
            "exceptionUnits": 536,
            "unaccountedUnits": 0,
        }
        assert [len(group["memberUnitIds"]) for group in groups] == [
            300,
            1,
            100,
            100,
            35,
        ]
        assert all(
            group["state"] == "exception"
            and group["role"] == "unknown"
            and group["target"] == {
                "scope": "case",
                "participantHandles": [],
            }
            and group["issueCodes"] == ["roster-missing"]
            for group in groups
        )
        assignments, dispositions = state._public_assignments()
        assert len(assignments) == 536
        assert all(item == {"unitId": item["unitId"], "decision": "unresolved"} for item in assignments)
        assert dispositions == []
        summary = state.approval_summary()
        assert summary["readyToPrepare"] is False
        assert summary["counts"]["unresolved"] == 536
        with pytest.raises(ValueError, match="not ready"):
            state.approve(summary["proposalDigest"])
        with pytest.raises(ValueError, match="snapshot is unavailable"):
            state.consume_approved_package_snapshot(summary["proposalDigest"])
        with pytest.raises(ValueError, match="eligible roster candidate"):
            state.resolve_exception(
                {
                    "exceptionId": "exception-0001",
                    "action": "choose-roster",
                    "rosterUnitId": "unit-0001",
                    "applyToSimilar": False,
                }
            )
    finally:
        context.__exit__(None, None, None)


def test_ambiguous_roster_fallback_is_atomic_and_retires_ids_after_choice(
    tmp_path,
    monkeypatch,
):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        inspection = inspect_observation(observation)
        state = ProposalState.from_inspection(
            observation,
            inspection,
            _grouping_evidence=GroupingEvidence(),
        )
        before = _review_bytes(state)
        local = state.local_review_snapshot()
        old_group_ids = {
            group["groupId"] for group in local["review"]["groups"]
        }
        roster_exception = local["review"]["exceptions"][0]

        assert local["roster"]["status"] == "ambiguous"
        assert len(local["review"]["exceptions"]) == 1
        assert roster_exception["kind"] == "roster"
        assert roster_exception["issueCode"] == "roster-ambiguous"
        assert local["review"]["coverage"] == {
            "groups": 1,
            "automaticallyOrganizedUnits": 0,
            "exceptionClusters": 1,
            "exceptionUnits": 2,
            "unaccountedUnits": 0,
        }

        with pytest.raises(ValueError, match="eligible roster candidate"):
            state.resolve_exception(
                {
                    "exceptionId": roster_exception["exceptionId"],
                    "action": "choose-roster",
                    "rosterUnitId": "unit-9999",
                    "applyToSimilar": False,
                }
            )
        assert _review_bytes(state) == before

        original_builder = ctv_proposal.build_grouping_plan
        monkeypatch.setattr(
            ctv_proposal,
            "build_grouping_plan",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(
                ValueError("synthetic grouping failure")
            ),
        )
        with pytest.raises(ValueError, match="synthetic grouping failure"):
            state.resolve_exception(
                {
                    "exceptionId": roster_exception["exceptionId"],
                    "action": "choose-roster",
                    "rosterUnitId": local["roster"]["candidateUnitIds"][0],
                    "applyToSimilar": False,
                }
            )
        assert _review_bytes(state) == before
        monkeypatch.setattr(ctv_proposal, "build_grouping_plan", original_builder)

        state.resolve_exception(
            {
                "exceptionId": roster_exception["exceptionId"],
                "action": "choose-roster",
                "rosterUnitId": local["roster"]["candidateUnitIds"][1],
                "applyToSimilar": False,
            }
        )
        selected = state.local_review_snapshot()
        assert selected["roster"]["status"] == "selected"
        assert selected["review"]["coverage"]["unaccountedUnits"] == 0
        assert old_group_ids.isdisjoint(
            group["groupId"] for group in selected["review"]["groups"]
        )
        assert all(
            item["exceptionId"] != roster_exception["exceptionId"]
            for item in selected["review"]["exceptions"]
        )


def test_automatic_groups_leave_only_exact_exception_clusters_unresolved(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (
            ("service-contract",) * 10,
            ("unknown", "unknown"),
            ("unknown",),
        ),
    )
    try:
        local = state.local_review_snapshot()

        assert set(local) == {"roster", "review", "summary"}
        assert local["roster"]["status"] == "selected"
        assert local["review"]["coverage"] == {
            "groups": 4,
            "automaticallyOrganizedUnits": 11,
            "exceptionClusters": 2,
            "exceptionUnits": 3,
            "unaccountedUnits": 0,
        }
        assert len(local["review"]["exceptions"]) == 2
        assert state.approval_summary()["readyToPrepare"] is False
    finally:
        context.__exit__(None, None, None)


def test_local_review_snapshot_is_ready_after_automatic_groups_need_no_exceptions(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("service-contract", "service-contract"),),
    )
    try:
        local = state.local_review_snapshot()
        expected_count_keys = {
            "sources",
            "units",
            "participants",
            "accepted",
            "reassigned",
            "excluded",
            "unresolved",
        }

        assert local["review"]["exceptions"] == []
        assert local["summary"]["readyToPrepare"] is True
        assert set(local["summary"]["counts"]) == expected_count_keys
        digest = state.approval_summary()["proposalDigest"]
        assert set(state.approve(digest)["counts"]) == expected_count_keys
    finally:
        context.__exit__(None, None, None)


def _review_bytes(state):
    return json.dumps(
        state.local_review_snapshot(),
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")


def test_resolve_exception_assigns_exact_cluster_and_expands_every_member(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("unknown", "unknown"), ("unknown",)),
    )
    try:
        before = state.approval_summary()["proposalDigest"]
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "assign",
                "role": "acceptance-record",
                "target": {
                    "scope": "individual",
                    "participantHandles": ["participant-0001"],
                },
                "applyToSimilar": False,
            }
        )

        local = state.local_review_snapshot()
        assert state.approval_summary()["proposalDigest"] != before
        assert local["review"]["coverage"] == {
            "groups": 3,
            "automaticallyOrganizedUnits": 1,
            "exceptionClusters": 1,
            "exceptionUnits": 1,
            "unaccountedUnits": 0,
        }
        assert state.approval_summary()["counts"] == {
            "sources": 3,
            "units": 4,
            "participants": 2,
            "accepted": 1,
            "reassigned": 2,
            "excluded": 0,
            "unresolved": 1,
        }
    finally:
        context.__exit__(None, None, None)


def test_apply_to_similar_touches_only_unresolved_matching_clusters(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("unknown",), ("unknown",), ("identity-front",)),
    )
    try:
        state.resolve_exception(
            {
                "exceptionId": "exception-0002",
                "action": "exclude",
                "reason": "irrelevant",
                "applyToSimilar": False,
            }
        )
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "assign",
                "role": "acceptance-record",
                "target": {
                    "scope": "individual",
                    "participantHandles": ["participant-0001"],
                },
                "applyToSimilar": True,
            }
        )

        local = state.local_review_snapshot()
        assert [
            item["exceptionId"] for item in local["review"]["exceptions"]
        ] == ["exception-0003"]
        assert state.approval_summary()["counts"] == {
            "sources": 4,
            "units": 4,
            "participants": 2,
            "accepted": 1,
            "reassigned": 1,
            "excluded": 1,
            "unresolved": 1,
        }
    finally:
        context.__exit__(None, None, None)


def test_undo_exception_and_reopen_group_restore_the_exact_prior_digest(tmp_path):
    context, state = _grouping_state(tmp_path, (("unknown",),))
    try:
        before = state.approval_summary()["proposalDigest"]
        before_review = _review_bytes(state)
        request = {
            "exceptionId": "exception-0001",
            "action": "exclude",
            "reason": "irrelevant",
            "applyToSimilar": False,
        }

        state.resolve_exception(request)
        state.undo_exception({"exceptionId": "exception-0001"})
        assert state.approval_summary()["proposalDigest"] == before
        assert _review_bytes(state) == before_review

        state.resolve_exception(request)
        state.reopen_group({"groupId": "group-0002"})
        assert state.approval_summary()["proposalDigest"] == before
        assert _review_bytes(state) == before_review
    finally:
        context.__exit__(None, None, None)


@pytest.mark.parametrize(
    "mapping",
    (
        {
            "exceptionId": "exception-9999",
            "action": "exclude",
            "reason": "irrelevant",
            "applyToSimilar": False,
        },
        {
            "exceptionId": "exception-0001",
            "action": "split",
            "splitBeforeUnitId": "unit-0004",
            "applyToSimilar": False,
        },
        {
            "exceptionId": "exception-0001",
            "action": "merge-next",
            "applyToSimilar": False,
        },
        {
            "exceptionId": "exception-0001",
            "action": "exclude",
            "reason": "irrelevant",
            "applyToSimilar": False,
            "extra": False,
        },
    ),
    ids=("stale", "invalid-member", "cross-source-merge", "extra-key"),
)
def test_resolve_exception_invalid_requests_leave_digest_and_state_byte_identical(
    tmp_path,
    mapping,
):
    context, state = _grouping_state(
        tmp_path,
        (("unknown", "unknown"), ("unknown",)),
    )
    try:
        before_digest = state.approval_summary()["proposalDigest"]
        before_review = _review_bytes(state)

        with pytest.raises(ValueError):
            state.resolve_exception(mapping)

        assert state.approval_summary()["proposalDigest"] == before_digest
        assert _review_bytes(state) == before_review
    finally:
        context.__exit__(None, None, None)


def test_resolve_exception_split_changes_only_contiguous_group_review_and_digest(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("unknown", "unknown", "unknown"),),
    )
    try:
        before = state.approval_summary()["proposalDigest"]
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "split",
                "splitBeforeUnitId": "unit-0003",
                "applyToSimilar": False,
            }
        )
        local = state.local_review_snapshot()

        assert state.approval_summary()["proposalDigest"] != before
        assert [
            group["memberUnitIds"] for group in local["review"]["groups"]
        ] == [["unit-0001"], ["unit-0002"], ["unit-0003", "unit-0004"]]
        assert local["review"]["coverage"] == {
            "groups": 3,
            "automaticallyOrganizedUnits": 1,
            "exceptionClusters": 2,
            "exceptionUnits": 3,
            "unaccountedUnits": 0,
        }
        assert len(
            {item["similarityKey"] for item in local["review"]["exceptions"]}
        ) == 2
        assert "merge-next" in local["review"]["exceptions"][0][
            "allowedActions"
        ]
        assert "merge-next" not in local["review"]["exceptions"][1][
            "allowedActions"
        ]
        assert state.approval_summary()["counts"]["unresolved"] == 3
    finally:
        context.__exit__(None, None, None)


def test_exact_coverage_holds_after_every_exception_cluster_is_resolved(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("unknown",), ("unknown",)),
    )
    try:
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "assign",
                "role": "acceptance-record",
                "target": {
                    "scope": "individual",
                    "participantHandles": ["participant-0001"],
                },
                "applyToSimilar": True,
            }
        )
        summary = state.approval_summary()
        approved = state.approve(summary["proposalDigest"])

        assert summary["readyToPrepare"] is True
        assert [
            assignment["unitId"] for assignment in approved["unitAssignments"]
        ] == ["unit-0001", "unit-0002", "unit-0003"]
        assert len({
            assignment["unitId"] for assignment in approved["unitAssignments"]
        }) == 3
    finally:
        context.__exit__(None, None, None)


def test_resolve_exception_accepts_fixed_source_recommendation_and_disposition(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("service-contract",),),
        source_only=(("unreadable", "pdf", ("document-unreadable",)),),
    )
    try:
        assert state.local_review_snapshot()["review"]["exceptions"] == [
            {
                "exceptionId": "exception-0001",
                "kind": "source",
                "issueCode": "source-unreadable",
                "recommendedAction": "exclude",
                "allowedActions": ["exclude"],
                "similarityKey": "similarity-8af016f0fcf7d9a6",
                "evidenceId": "evidence-0003",
            }
        ]
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "accept-recommendation",
                "applyToSimilar": False,
            }
        )
        summary = state.approval_summary()
        approved = state.approve(summary["proposalDigest"])

        assert approved["sourceDispositions"] == [
            {
                "evidenceId": "evidence-0003",
                "decision": "excluded",
                "reason": "unreadable-replacement-available",
            }
        ]
        assert summary["readyToPrepare"] is True
    finally:
        context.__exit__(None, None, None)


def test_split_then_merge_next_is_atomic_and_undo_restores_split_state(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("unknown", "unknown", "unknown"),),
    )
    try:
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "split",
                "splitBeforeUnitId": "unit-0003",
                "applyToSimilar": False,
            }
        )
        split_digest = state.approval_summary()["proposalDigest"]
        split_review = _review_bytes(state)

        merge_exception_id = state.local_review_snapshot()["review"][
            "exceptions"
        ][0]["exceptionId"]
        state.resolve_exception(
            {
                "exceptionId": merge_exception_id,
                "action": "merge-next",
                "applyToSimilar": False,
            }
        )
        assert [
            group["memberUnitIds"]
            for group in state.local_review_snapshot()["review"]["groups"]
        ] == [["unit-0001"], ["unit-0002", "unit-0003", "unit-0004"]]

        state.undo_exception({"exceptionId": merge_exception_id})
        assert state.approval_summary()["proposalDigest"] == split_digest
        assert _review_bytes(state) == split_review
    finally:
        context.__exit__(None, None, None)


def test_choose_roster_exception_recomputes_groups_from_preloaded_candidates(
    tmp_path,
):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        facts = GroupingEvidence()
        inspection = inspect_observation(
            observation,
            _private_text_sink=facts.capture,
        )
        state = ProposalState.from_inspection(
            observation,
            inspection,
            _grouping_evidence=facts,
        )
        roster_units = [
            unit["unitId"]
            for unit in state.units
            if unit["suggestedRole"] == "payment-roster"
        ]

        assert state.local_review_snapshot()["roster"]["status"] == "ambiguous"
        assert state.local_review_snapshot()["review"]["exceptions"][0][
            "kind"
        ] == "roster"
        state.resolve_exception(
            {
                "exceptionId": "exception-0001",
                "action": "choose-roster",
                "rosterUnitId": roster_units[1],
                "applyToSimilar": False,
            }
        )

        local = state.local_review_snapshot()
        assert local["roster"]["status"] == "selected"
        assert local["roster"]["rosterUnitId"] == roster_units[1]
        assert local["review"]["coverage"]["unaccountedUnits"] == 0
        assert state.approval_summary()["counts"]["unresolved"] == 1
        for exception in list(local["review"]["exceptions"]):
            state.resolve_exception(
                {
                    "exceptionId": exception["exceptionId"],
                    "action": "assign",
                    "role": "payment-roster",
                    "target": {"scope": "case", "participantHandles": []},
                    "applyToSimilar": False,
                }
            )
        assert state.approval_summary()["readyToPrepare"] is True
        assert all(
            private not in repr(local)
            for private in ("Alice", "CTV-001", "Carol", "CTV-101")
        )


def test_accept_recommendation_applies_each_similar_groups_own_participant_target(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("acceptance-record",), ("acceptance-record",)),
        text_by_unit_number={
            2: "Alice CTV-001 acceptance",
            3: "Bao CTV-002 acceptance",
        },
        unit_issue_codes_by_unit_number={
            2: ("classification-conflict",),
            3: ("classification-conflict",),
        },
    )
    try:
        exceptions = state.local_review_snapshot()["review"]["exceptions"]
        assert len(exceptions) == 2
        assert exceptions[0]["similarityKey"] == exceptions[1]["similarityKey"]

        state.resolve_exception(
            {
                "exceptionId": exceptions[0]["exceptionId"],
                "action": "accept-recommendation",
                "applyToSimilar": True,
            }
        )
        assignments = {
            item["unitId"]: item
            for item in state.approve(
                state.approval_summary()["proposalDigest"]
            )["unitAssignments"]
        }

        assert assignments["unit-0002"]["target"]["participantHandles"] == [
            "participant-0001"
        ]
        assert assignments["unit-0003"]["target"]["participantHandles"] == [
            "participant-0002"
        ]
    finally:
        context.__exit__(None, None, None)


def test_split_and_merge_retire_ids_without_recycling_stale_targets(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("unknown", "unknown", "unknown"),),
    )
    try:
        initial = state.local_review_snapshot()["review"]
        stale_exception_id = initial["exceptions"][0]["exceptionId"]
        stale_group_id = initial["groups"][1]["groupId"]
        state.resolve_exception(
            {
                "exceptionId": stale_exception_id,
                "action": "split",
                "splitBeforeUnitId": "unit-0003",
                "applyToSimilar": False,
            }
        )

        split = state.local_review_snapshot()["review"]
        assert stale_exception_id not in {
            item["exceptionId"] for item in split["exceptions"]
        }
        assert stale_group_id not in {item["groupId"] for item in split["groups"]}
        split_digest = state.approval_summary()["proposalDigest"]
        split_bytes = _review_bytes(state)
        with pytest.raises(ValueError):
            state.resolve_exception(
                {
                    "exceptionId": stale_exception_id,
                    "action": "exclude",
                    "reason": "irrelevant",
                    "applyToSimilar": False,
                }
            )
        with pytest.raises(ValueError):
            state.reopen_group({"groupId": stale_group_id})
        assert state.approval_summary()["proposalDigest"] == split_digest
        assert _review_bytes(state) == split_bytes

        merge_exception_id = split["exceptions"][0]["exceptionId"]
        merge_group_id = split["groups"][1]["groupId"]
        state.resolve_exception(
            {
                "exceptionId": merge_exception_id,
                "action": "merge-next",
                "applyToSimilar": False,
            }
        )
        merged = state.local_review_snapshot()["review"]
        assert merge_exception_id not in {
            item["exceptionId"] for item in merged["exceptions"]
        }
        assert merge_group_id not in {
            item["groupId"] for item in merged["groups"]
        }
        merged_digest = state.approval_summary()["proposalDigest"]
        merged_bytes = _review_bytes(state)
        with pytest.raises(ValueError):
            state.resolve_exception(
                {
                    "exceptionId": merge_exception_id,
                    "action": "exclude",
                    "reason": "irrelevant",
                    "applyToSimilar": False,
                }
            )
        with pytest.raises(ValueError):
            state.reopen_group({"groupId": merge_group_id})
        assert state.approval_summary()["proposalDigest"] == merged_digest
        assert _review_bytes(state) == merged_bytes
    finally:
        context.__exit__(None, None, None)


def test_source_and_unit_exceptions_publish_in_interleaved_canonical_id_order(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("unknown",),),
        leading_source_only=(("unreadable", "pdf", ("document-unreadable",)),),
    )
    try:
        exceptions = state.local_review_snapshot()["review"]["exceptions"]

        assert [item["exceptionId"] for item in exceptions] == [
            "exception-0001",
            "exception-0002",
        ]
        assert [item["kind"] for item in exceptions] == ["source", "unit-cluster"]
    finally:
        context.__exit__(None, None, None)


def test_reopen_group_converts_an_automatic_assignment_to_unresolved(tmp_path):
    context, state = _grouping_state(tmp_path, (("service-contract",),))
    try:
        review = state.local_review_snapshot()["review"]
        automatic = next(
            group for group in review["groups"] if group["memberUnitIds"] == ["unit-0002"]
        )

        state.reopen_group({"groupId": automatic["groupId"]})

        reopened = state.local_review_snapshot()["review"]
        assert reopened["exceptions"][0]["kind"] == "unit-cluster"
        assert state.approval_summary()["counts"]["unresolved"] == 1
    finally:
        context.__exit__(None, None, None)


def test_reopen_group_materializes_automatic_duplicate_exclusion(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("service-contract",), ("service-contract",)),
        duplicate_group_by_evidence_number={
            2: "duplicate-0007",
            3: "duplicate-0007",
        },
    )
    try:
        review = state.local_review_snapshot()["review"]
        duplicate = next(
            group for group in review["groups"] if group["memberUnitIds"] == ["unit-0003"]
        )
        assert state.approval_summary()["counts"]["excluded"] == 1

        state.reopen_group({"groupId": duplicate["groupId"]})

        reopened = state.local_review_snapshot()["review"]
        assert any(
            item.get("groupIds") == [duplicate["groupId"]]
            for item in reopened["exceptions"]
        )
        assert reopened["coverage"]["exceptionUnits"] == 1
        assert state.approval_summary()["counts"]["excluded"] == 0
        assert state.approval_summary()["counts"]["unresolved"] == 1
    finally:
        context.__exit__(None, None, None)


def test_original_adjacent_compatible_exceptions_offer_merge_next(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("acceptance-record", "acceptance-record"),),
        text_by_unit_number={
            2: "Alice CTV-001 acceptance",
            3: "Bao CTV-002 acceptance",
        },
        unit_issue_codes_by_unit_number={
            2: ("classification-conflict",),
            3: ("classification-conflict",),
        },
    )
    try:
        exceptions = state.local_review_snapshot()["review"]["exceptions"]
        assert len(exceptions) == 2
        assert "merge-next" in exceptions[0]["allowedActions"]

        state.resolve_exception(
            {
                "exceptionId": exceptions[0]["exceptionId"],
                "action": "merge-next",
                "applyToSimilar": False,
            }
        )

        assert state.local_review_snapshot()["review"]["coverage"] == {
            "groups": 2,
            "automaticallyOrganizedUnits": 1,
            "exceptionClusters": 1,
            "exceptionUnits": 2,
            "unaccountedUnits": 0,
        }
    finally:
        context.__exit__(None, None, None)


def test_grouped_select_roster_recomputes_the_complete_review(tmp_path):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        facts = GroupingEvidence()
        inspection = inspect_observation(observation, _private_text_sink=facts.capture)
        state = ProposalState.from_inspection(
            observation, inspection, _grouping_evidence=facts
        )
        roster_units = [
            unit["unitId"]
            for unit in state.units
            if unit["suggestedRole"] == "payment-roster"
        ]

        state.select_roster({"rosterUnitId": roster_units[1]})

        local = state.local_review_snapshot()
        assert local["roster"] == {
            "status": "selected",
            "rosterUnitId": roster_units[1],
            "candidateUnitIds": roster_units,
            "candidateSummaries": [
                {
                    "rosterUnitId": roster_unit_id,
                    "participantCount": 2,
                    "eligible": True,
                    "issueCodes": [],
                }
                for roster_unit_id in roster_units
            ],
            "participantHandles": ["participant-0001", "participant-0002"],
            "issueCodes": [],
        }
        assert local["review"]["coverage"]["unaccountedUnits"] == 0


def test_local_roster_projection_exposes_bounded_candidate_eligibility_facts(
    tmp_path,
):
    source = _source_with_valid_and_invalid_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        facts = GroupingEvidence()
        inspection = inspect_observation(
            observation, _private_text_sink=facts.capture
        )
        state = ProposalState.from_inspection(
            observation, inspection, _grouping_evidence=facts
        )

        roster = state.local_review_snapshot()["roster"]

        assert roster["candidateSummaries"] == [
            {
                "rosterUnitId": "unit-0001",
                "participantCount": 1,
                "eligible": True,
                "issueCodes": [],
            },
            {
                "rosterUnitId": "unit-0002",
                "participantCount": 0,
                "eligible": False,
                "issueCodes": [
                    "roster-row-invalid",
                    "roster-fa-code-blank",
                ],
            },
        ]
        assert "Alice" not in repr(roster["candidateSummaries"])
        assert "CTV-001" not in repr(roster["candidateSummaries"])


def test_grouped_select_roster_failure_rolls_back_every_visible_state(tmp_path):
    source = _source_with_valid_and_invalid_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        facts = GroupingEvidence()
        inspection = inspect_observation(observation, _private_text_sink=facts.capture)
        state = ProposalState.from_inspection(
            observation, inspection, _grouping_evidence=facts
        )
        invalid_roster_id = next(
            unit["unitId"]
            for unit in state.units
            if unit["suggestedRole"] == "payment-roster"
            and unit["unitId"] != state.approval_summary()["rosterUnitId"]
        )
        before_digest = state.approval_summary()["proposalDigest"]
        before_review = _review_bytes(state)
        before_participants = state.participants_for_local_review()

        with pytest.raises(ValueError):
            state.select_roster({"rosterUnitId": invalid_roster_id})

        assert state.approval_summary()["proposalDigest"] == before_digest
        assert _review_bytes(state) == before_review
        assert state.participants_for_local_review() == before_participants


def test_grouped_select_roster_build_failure_is_atomic(tmp_path, monkeypatch):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        facts = GroupingEvidence()
        inspection = inspect_observation(observation, _private_text_sink=facts.capture)
        state = ProposalState.from_inspection(
            observation, inspection, _grouping_evidence=facts
        )
        roster_units = [
            unit["unitId"]
            for unit in state.units
            if unit["suggestedRole"] == "payment-roster"
        ]
        before_digest = state.approval_summary()["proposalDigest"]
        before_review = _review_bytes(state)
        before_participants = state.participants_for_local_review()
        real_builder = ctv_proposal.build_grouping_plan

        def fail_for_selected_candidate(inspection_value, candidate, evidence):
            if candidate.unit_id == roster_units[1]:
                raise ValueError("synthetic-group-build-failure")
            return real_builder(inspection_value, candidate, evidence)

        monkeypatch.setattr(
            ctv_proposal, "build_grouping_plan", fail_for_selected_candidate
        )

        with pytest.raises(ValueError, match="synthetic-group-build-failure"):
            state.select_roster({"rosterUnitId": roster_units[1]})

        assert state.approval_summary()["proposalDigest"] == before_digest
        assert _review_bytes(state) == before_review
        assert state.participants_for_local_review() == before_participants


def test_source_not_applicable_can_be_approved_and_consumed_coherently(tmp_path):
    context, state = _grouping_state(
        tmp_path,
        (("service-contract",),),
        source_only=(("inspected", "pdf", ()),),
    )
    try:
        exception = state.local_review_snapshot()["review"]["exceptions"][0]
        assert exception["issueCode"] == "source-not-applicable"
        state.resolve_exception(
            {
                "exceptionId": exception["exceptionId"],
                "action": "accept-recommendation",
                "applyToSimilar": False,
            }
        )
        digest = state.approval_summary()["proposalDigest"]

        state.approve(digest)
        snapshot = state.consume_approved_package_snapshot(digest)

        assert snapshot.source_dispositions[0].acquisition_status == "opaque"
        assert snapshot.source_dispositions[0].reason == "intentionally-omitted"
    finally:
        context.__exit__(None, None, None)


@pytest.mark.parametrize(
    "entrypoint",
    (
        "resolve-accept",
        "resolve-split",
        "resolve-merge",
        "undo-exception",
        "reopen-group",
        "select-roster",
        "set-unit-decision",
        "set-source-disposition",
    ),
)
def test_exact_mapping_rejects_hostile_keys_without_hash_or_equality_calls(
    tmp_path, entrypoint
):
    class HostileKey(str):
        equality_calls = 0
        hash_calls = 0

        def __hash__(self):
            type(self).hash_calls += 1
            return super().__hash__()

        def __eq__(self, other):
            type(self).equality_calls += 1
            return super().__eq__(other)

    context, state = _grouping_state(
        tmp_path, (("unknown", "unknown", "unknown"),)
    )
    try:
        exception_id = state.local_review_snapshot()["review"]["exceptions"][0][
            "exceptionId"
        ]
        group_id = state.local_review_snapshot()["review"]["groups"][1][
            "groupId"
        ]
        if entrypoint == "resolve-accept":
            mapping = {
                HostileKey("action"): "accept-recommendation",
                "exceptionId": exception_id,
                "applyToSimilar": False,
            }
            invoke = state.resolve_exception
        elif entrypoint == "resolve-split":
            mapping = {
                HostileKey("action"): "split",
                "exceptionId": exception_id,
                "splitBeforeUnitId": "unit-0003",
                "applyToSimilar": False,
            }
            invoke = state.resolve_exception
        elif entrypoint == "resolve-merge":
            mapping = {
                HostileKey("action"): "merge-next",
                "exceptionId": exception_id,
                "applyToSimilar": False,
            }
            invoke = state.resolve_exception
        elif entrypoint == "undo-exception":
            mapping = {HostileKey("exceptionId"): exception_id}
            invoke = state.undo_exception
        elif entrypoint == "reopen-group":
            mapping = {HostileKey("groupId"): group_id}
            invoke = state.reopen_group
        elif entrypoint == "select-roster":
            mapping = {
                HostileKey("rosterUnitId"): state.approval_summary()[
                    "rosterUnitId"
                ]
            }
            invoke = state.select_roster
        elif entrypoint == "set-unit-decision":
            mapping = {
                HostileKey("decision"): "excluded",
                "unitId": "unit-0002",
                "reason": "irrelevant",
            }
            invoke = state.set_unit_decision
        else:
            mapping = {
                HostileKey("decision"): "excluded",
                "evidenceId": "evidence-9999",
                "reason": "irrelevant",
            }
            invoke = state.set_source_disposition
        HostileKey.equality_calls = 0
        HostileKey.hash_calls = 0

        with pytest.raises(
            ValueError, match="proposal request must use its exact object shape"
        ):
            invoke(mapping)

        assert HostileKey.equality_calls == 0
        assert HostileKey.hash_calls == 0
    finally:
        context.__exit__(None, None, None)


def test_unknown_role_exception_does_not_advertise_an_unexecutable_recommendation(
    tmp_path,
):
    context, state = _grouping_state(tmp_path, (("unknown",),))
    try:
        local = state.local_review_snapshot()
        exception = local["review"]["exceptions"][0]
        before_digest = state.approval_summary()["proposalDigest"]
        before_review = _review_bytes(state)

        assert exception["issueCode"] == "role-uncertain"
        assert "recommendedAction" not in exception
        assert exception["allowedActions"] == ["assign", "exclude", "split"]
        with pytest.raises(ValueError):
            state.resolve_exception(
                {
                    "exceptionId": exception["exceptionId"],
                    "action": "accept-recommendation",
                    "applyToSimilar": False,
                }
            )
        assert state.approval_summary()["proposalDigest"] == before_digest
        assert _review_bytes(state) == before_review

        state.resolve_exception(
            {
                "exceptionId": exception["exceptionId"],
                "action": "assign",
                "role": "acceptance-record",
                "target": {"scope": "case", "participantHandles": []},
                "applyToSimilar": False,
            }
        )
        approved = state.approve(state.approval_summary()["proposalDigest"])

        assert approved["unitAssignments"][-1] == {
            "unitId": "unit-0002",
            "decision": "reassigned",
            "role": "acceptance-record",
            "target": {"scope": "case", "participantHandles": []},
        }
    finally:
        context.__exit__(None, None, None)


def test_local_review_effective_facts_equal_the_approved_package_snapshot(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("unknown",),),
        source_only=(("unreadable", "pdf", ("document-unreadable",)),),
    )
    try:
        initial = state.local_review_snapshot()["review"]
        unit_exception = next(
            item for item in initial["exceptions"] if item["kind"] == "unit-cluster"
        )
        source_exception = next(
            item for item in initial["exceptions"] if item["kind"] == "source"
        )
        state.resolve_exception(
            {
                "exceptionId": unit_exception["exceptionId"],
                "action": "assign",
                "role": "acceptance-record",
                "target": {
                    "scope": "individual",
                    "participantHandles": ["participant-0002"],
                },
                "applyToSimilar": False,
            }
        )
        state.resolve_exception(
            {
                "exceptionId": source_exception["exceptionId"],
                "action": "exclude",
                "reason": "irrelevant",
                "applyToSimilar": False,
            }
        )

        local = state.local_review_snapshot()
        resolved_group = next(
            group
            for group in local["review"]["groups"]
            if group["memberUnitIds"] == ["unit-0002"]
        )
        assert resolved_group["state"] == "user-resolved"
        assert resolved_group["effectiveResolution"] == {
            "action": "assign",
            "role": "acceptance-record",
            "target": {
                "scope": "individual",
                "participantHandles": ["participant-0002"],
            },
        }
        assert local["review"]["resolvedExclusions"] == [
            {
                "exceptionId": source_exception["exceptionId"],
                "kind": "source",
                "evidenceId": "evidence-0003",
                "issueCode": "source-unreadable",
                "reason": "irrelevant",
            }
        ]
        assert local["review"]["coverage"] == {
            "groups": 2,
            "automaticallyOrganizedUnits": 1,
            "exceptionClusters": 0,
            "exceptionUnits": 0,
            "unaccountedUnits": 0,
        }
        assert local["summary"]["counts"] == {
            "sources": 3,
            "units": 2,
            "participants": 2,
            "accepted": 1,
            "reassigned": 1,
            "excluded": 1,
            "unresolved": 0,
        }
        assert local["summary"]["readyToPrepare"] is True
        assert "Alice" not in repr(local["review"])
        assert "CTV-001" not in repr(local["review"])
        digest_group = next(
            group
            for group in state._digest_input()["groupReview"]["groups"]
            if group["memberUnitIds"] == ["unit-0002"]
        )
        assert "effectiveResolution" not in digest_group

        digest = local["summary"]["proposalDigest"]
        state.approve(digest)
        approved = state.consume_approved_package_snapshot(digest)
        unit = next(
            item for item in approved.unit_decisions if item.unit_id == "unit-0002"
        )
        source = next(
            item
            for item in approved.source_dispositions
            if item.evidence_id == "evidence-0003"
        )
        assert (unit.role, unit.scope, unit.participant_handles) == (
            "acceptance-record",
            "individual",
            ("participant-0002",),
        )
        assert (source.decision, source.reason) == ("excluded", "irrelevant")
    finally:
        context.__exit__(None, None, None)


def test_local_review_projects_the_exact_effective_group_exclusion(tmp_path):
    context, state = _grouping_state(tmp_path, (("unknown",),))
    try:
        exception = next(
            item
            for item in state.local_review_snapshot()["review"]["exceptions"]
            if item["kind"] == "unit-cluster"
        )
        state.resolve_exception(
            {
                "exceptionId": exception["exceptionId"],
                "action": "exclude",
                "reason": "irrelevant",
                "applyToSimilar": False,
            }
        )

        resolved = next(
            group
            for group in state.local_review_snapshot()["review"]["groups"]
            if group["memberUnitIds"] == ["unit-0002"]
        )
        assert resolved["effectiveResolution"] == {
            "action": "exclude",
            "reason": "irrelevant",
        }
    finally:
        context.__exit__(None, None, None)


def test_recommendation_requires_role_target_scope_compatibility_and_valid_still_executes(
    tmp_path,
):
    context, state = _grouping_state(
        tmp_path,
        (("payment-tax-form",), ("service-contract",)),
        text_by_unit_number={
            2: "unmatched payment tax form",
            3: "whole case service contract",
        },
        unit_issue_codes_by_unit_number={
            3: ("classification-conflict",),
        },
    )
    try:
        local = state.local_review_snapshot()
        tax_exception = next(
            item
            for item in local["review"]["exceptions"]
            if item["memberUnitIds"] == ["unit-0002"]
        )
        valid_exception = next(
            item
            for item in local["review"]["exceptions"]
            if item["memberUnitIds"] == ["unit-0003"]
        )
        tax_group = next(
            item
            for item in local["review"]["groups"]
            if item["memberUnitIds"] == ["unit-0002"]
        )
        before_digest = state.approval_summary()["proposalDigest"]
        before_review = _review_bytes(state)

        assert tax_exception["issueCode"] == "participant-no-match"
        assert tax_group["role"] == "payment-tax-form"
        assert tax_group["target"] == {
            "scope": "case",
            "participantHandles": [],
        }
        assert "recommendedAction" not in tax_exception
        with pytest.raises(ValueError, match="no executable recommendation"):
            state.resolve_exception(
                {
                    "exceptionId": tax_exception["exceptionId"],
                    "action": "accept-recommendation",
                    "applyToSimilar": False,
                }
            )
        assert state.approval_summary()["proposalDigest"] == before_digest
        assert _review_bytes(state) == before_review

        assert valid_exception["recommendedAction"] == "assign"
        state.resolve_exception(
            {
                "exceptionId": valid_exception["exceptionId"],
                "action": "accept-recommendation",
                "applyToSimilar": False,
            }
        )
        state.resolve_exception(
            {
                "exceptionId": tax_exception["exceptionId"],
                "action": "assign",
                "role": "payment-tax-form",
                "target": {
                    "scope": "individual",
                    "participantHandles": ["participant-0001"],
                },
                "applyToSimilar": False,
            }
        )
        approved = state.approve(state.approval_summary()["proposalDigest"])
        assignments = {
            item["unitId"]: item for item in approved["unitAssignments"]
        }

        assert assignments["unit-0002"]["target"] == {
            "scope": "individual",
            "participantHandles": ["participant-0001"],
        }
        assert assignments["unit-0003"]["target"] == {
            "scope": "case",
            "participantHandles": [],
        }
    finally:
        context.__exit__(None, None, None)


def test_caller_owned_observation_is_inspected_without_being_closed(tmp_path):
    source = _source(tmp_path)
    with open_inventory_observation(source) as observation:
        inspection = inspect_observation(observation)
        assert inspection.observation_id == observation.observation_id
        assert observation.snapshot(inspection.sources[0].evidence_id, max_bytes=25 * 1024 * 1024)


def test_selected_roster_maps_usable_rows_to_opaque_handles_in_row_order(tmp_path):
    context, observation, state = _state(tmp_path)
    try:
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        summary = state.approval_summary()
        assert summary["rosterUnitId"] == roster["unitId"]
        assert summary["participantHandles"] == [
            "participant-0001", "participant-0002"
        ]
        assert "Alice" not in repr(summary)
    finally:
        context.__exit__(None, None, None)


def test_roster_selection_uses_explicit_owned_snapshot_capability(
    tmp_path, monkeypatch
):
    source = _source(tmp_path)
    with open_inventory_observation(source) as observation:
        inspection = inspect_observation(observation)
        owned_snapshot = observation.snapshot
        calls = []

        def snapshot_source(evidence_id, *, max_bytes):
            calls.append(evidence_id)
            return owned_snapshot(evidence_id, max_bytes=max_bytes)

        state = ProposalState.from_inspection(
            observation, inspection, _snapshot_source=snapshot_source
        )
        monkeypatch.setattr(
            InventoryObservation,
            "snapshot",
            lambda *_args, **_kwargs: pytest.fail(
                "default observation acquisition bypassed injected capability"
            ),
        )
        roster = next(
            unit for unit in state.units if unit["suggestedRole"] == "payment-roster"
        )
        state.select_roster({"rosterUnitId": roster["unitId"]})

        assert state.approval_summary()["participantHandles"] == [
            "participant-0001", "participant-0002"
        ]
        assert calls == [roster["evidenceId"]]


def test_local_participant_display_keeps_name_and_masked_identity_out_of_public_results(tmp_path):
    context, observation, state = _state(tmp_path)
    try:
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        state.select_roster({"rosterUnitId": roster["unitId"]})
        assert state.participants_for_local_review() == [
            {"participantHandle": "participant-0001", "name": "Alice", "identityHint": "***-001"},
            {"participantHandle": "participant-0002", "name": "Bao", "identityHint": "***-002"},
        ]
        public = state.draft_result()
        assert "Alice" not in repr(public)
        assert "CTV-001" not in repr(public)
        assert "***-001" not in repr(public)
    finally:
        context.__exit__(None, None, None)


def test_roster_blank_separator_is_ignored_while_duplicate_identity_still_blocks(
    tmp_path,
):
    source = _source(
        tmp_path,
        roster_rows=(
            ("Alice", "CTV-001"),
            (None, None, None),
            ("Bao", "CTV-001"),
        ),
    )
    with open_inventory_observation(source) as observation:
        state = ProposalState.from_inspection(observation, inspect_observation(observation))
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        state.select_roster({"rosterUnitId": roster["unitId"]})
        summary = state.approval_summary()
        assert summary["readyToPrepare"] is False
        assert summary["participantHandles"] == [
            "participant-0001",
            "participant-0002",
        ]
        assert "roster-row-invalid" not in summary["issueCodes"]
        assert "roster-identity-duplicate" in summary["issueCodes"]


@pytest.mark.parametrize(
    "invalid_row",
    [
        ("Alice", "", 100),
        ("", "CTV-001", 100),
        ("", "", 100),
    ],
    ids=("missing-identity", "missing-name", "trailing-data"),
)
def test_roster_nonblank_incomplete_rows_remain_invalid(tmp_path, invalid_row):
    source = _source(
        tmp_path,
        roster_rows=(("Alice", "CTV-001"), invalid_row),
    )
    with open_inventory_observation(source) as observation:
        state = ProposalState.from_inspection(
            observation, inspect_observation(observation)
        )
        roster = next(
            unit
            for unit in state.units
            if unit["suggestedRole"] == "payment-roster"
        )
        state.select_roster({"rosterUnitId": roster["unitId"]})

        summary = state.approval_summary()
        assert summary["participantHandles"] == ["participant-0001"]
        assert "roster-row-invalid" in summary["issueCodes"]
        assert summary["readyToPrepare"] is False


def test_header_only_roster_is_invalid_and_has_no_usable_participants(tmp_path):
    source = _source(tmp_path, roster_rows=())
    with open_inventory_observation(source) as observation:
        state = ProposalState.from_inspection(observation, inspect_observation(observation))
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        state.select_roster({"rosterUnitId": roster["unitId"]})
        assert state.approval_summary()["participantHandles"] == []
        assert "roster-row-invalid" in state.approval_summary()["issueCodes"]
        assert state.approval_summary()["readyToPrepare"] is False


def test_assignments_require_exact_mappings_and_resolve_every_unit_and_source(tmp_path):
    context, observation, state = _state(tmp_path)
    try:
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        state.select_roster({"rosterUnitId": roster["unitId"]})
        with pytest.raises(ValueError):
            state.set_unit_decision({"unitId": roster["unitId"], "decision": "accepted"})
        for unit in state.units:
            if unit["unitId"] == roster["unitId"]:
                state.set_unit_decision({
                    "unitId": unit["unitId"], "decision": "excluded", "reason": "irrelevant"
                })
            else:
                state.set_unit_decision({
                    "unitId": unit["unitId"], "decision": "reassigned", "role": "identity-front",
                    "target": {"scope": "individual", "participantHandles": ["participant-0001"]},
                })
        for source in state.sources:
            if source["evidenceId"] not in {unit["evidenceId"] for unit in state.units}:
                state.set_source_disposition({
                    "evidenceId": source["evidenceId"], "decision": "excluded", "reason": "irrelevant"
                })
        assert state.approval_summary()["readyToPrepare"] is True
    finally:
        context.__exit__(None, None, None)


def test_roster_change_invalidates_existing_participant_targeted_assignments(tmp_path):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        state = ProposalState.from_inspection(observation, inspect_observation(observation))
        rosters = [unit for unit in state.units if unit["suggestedRole"] == "payment-roster"]
        state.select_roster({"rosterUnitId": rosters[0]["unitId"]})
        state.set_unit_decision({
            "unitId": rosters[0]["unitId"], "decision": "accepted", "role": "payment-roster",
            "target": {"scope": "individual", "participantHandles": ["participant-0001"]},
        })
        state.select_roster({"rosterUnitId": rosters[1]["unitId"]})
        assert state.approval_summary()["counts"]["unresolved"] == len(state.units)


def test_equal_strong_rosters_remain_explicitly_resolvable_from_preloaded_candidates(
    tmp_path,
):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        calls = []
        owned_snapshot = observation.snapshot

        def snapshot_source(evidence_id, *, max_bytes):
            calls.append((evidence_id, max_bytes))
            return owned_snapshot(evidence_id, max_bytes=max_bytes)

        state = ProposalState.from_inspection(
            observation,
            inspect_observation(observation),
            _snapshot_source=snapshot_source,
        )
        rosters = [
            unit
            for unit in state.units
            if unit["suggestedRole"] == "payment-roster"
        ]

        assert state.approval_summary()["rosterUnitId"] is None
        assert state.approval_summary()["issueCodes"] == ["roster-ambiguous"]
        assert len(calls) == 1

        state.select_roster({"rosterUnitId": rosters[1]["unitId"]})

        assert state.approval_summary()["rosterUnitId"] == rosters[1]["unitId"]
        assert "roster-ambiguous" not in state.approval_summary()["issueCodes"]
        assert len(calls) == 1


def test_accepted_and_reassigned_enforce_their_distinct_suggested_role_meanings(tmp_path):
    context, observation, state = _state(tmp_path)
    try:
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        unknown = next(unit for unit in state.units if unit["suggestedRole"] == "unknown")
        state.select_roster({"rosterUnitId": roster["unitId"]})
        with pytest.raises(ValueError):
            state.set_unit_decision({
                "unitId": roster["unitId"], "decision": "reassigned", "role": "payment-roster",
                "target": {"scope": "individual", "participantHandles": ["participant-0001"]},
            })
        with pytest.raises(ValueError):
            state.set_unit_decision({
                "unitId": unknown["unitId"], "decision": "accepted", "role": "identity-front",
                "target": {"scope": "individual", "participantHandles": ["participant-0001"]},
            })
        state.set_unit_decision({
            "unitId": unknown["unitId"], "decision": "reassigned", "role": "identity-front",
            "target": {"scope": "individual", "participantHandles": ["participant-0001"]},
        })
    finally:
        context.__exit__(None, None, None)


def test_digest_excludes_private_roster_values_and_changes_for_each_decision(tmp_path):
    context, observation, state = _state(tmp_path)
    try:
        roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
        state.select_roster({"rosterUnitId": roster["unitId"]})
        before = state.approval_summary()["proposalDigest"]
        state.set_unit_decision({"unitId": roster["unitId"], "decision": "excluded", "reason": "irrelevant"})
        after = state.approval_summary()["proposalDigest"]
        assert before != after
        assert "Alice" not in repr(state.approval_summary())
        assert "CTV-001" not in repr(state.approval_summary())
    finally:
        context.__exit__(None, None, None)


def test_approved_proposal_stays_private_and_observation_close_rejects_source_mutation(tmp_path):
    source = _source(tmp_path)
    with pytest.raises(InventoryError, match="inventory-tree-changed"):
        with open_inventory_observation(source) as observation:
            state = ProposalState.from_inspection(observation, inspect_observation(observation))
            roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
            state.select_roster({"rosterUnitId": roster["unitId"]})
            for unit in state.units:
                state.set_unit_decision({"unitId": unit["unitId"], "decision": "excluded", "reason": "irrelevant"})
            digest = state.approval_summary()["proposalDigest"]
            approved = state.approve(digest)
            assert approved["approval"] == {"status": "user-approved", "approvedProposalDigest": digest}
            assert "Alice" not in repr(approved)
            (source / "identity.png").write_bytes(b"changed after approval")


def test_draft_and_cancelled_results_have_their_fixed_public_shapes(tmp_path):
    context, observation, state = _state(tmp_path)
    try:
        assert set(state.draft_result()) == {"version", "outcome", "observationId", "readyToPrepare", "counts", "issueCodes"}
        assert state.cancelled_result() == {"version": "1.0", "outcome": "cancelled", "readyToPrepare": False}
    finally:
        context.__exit__(None, None, None)


def test_roster_switch_keeps_public_digest_and_review_values_private(tmp_path):
    source = _source_with_two_rosters(tmp_path)
    with open_inventory_observation(source) as observation:
        state = ProposalState.from_inspection(observation, inspect_observation(observation))
        rosters = [unit for unit in state.units if unit["suggestedRole"] == "payment-roster"]
        state.select_roster({"rosterUnitId": rosters[0]["unitId"]})
        first = state.approval_summary()
        state.select_roster({"rosterUnitId": rosters[1]["unitId"]})
        second = state.approval_summary()
        assert first["proposalDigest"] != second["proposalDigest"]
        assert all(value not in repr(second) for value in ("Alice", "CTV-001", "Carol", "CTV-101"))
