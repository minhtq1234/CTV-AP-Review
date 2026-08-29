"""A synthetic workbook shaped like the PUBGm submission template.

Real submissions are identity documents and are never committed. This builds a
file with the same *structure* -- three sheets, a merged image header spanning
two columns, three populations of image in known columns -- carrying invented
names and numbers, so the layout logic can be developed and tested without any
real data.
"""
from __future__ import annotations

import io

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

#: Invented people, as (name, CCCD, MST). Sequential numbers, obviously not
#: real. This is the repository's ONE fabricated identity series -- the demo
#: case builds on it rather than inventing a second set, so a number seen in a
#: fixture always means the same person.
PEOPLE = [
    ("NGUYEN VAN MOT", "001100000001", "0011000001"),
    ("TRAN THI HAI", "001100000002", "0011000002"),
    ("LE VAN BA", "001100000003", "0011000003"),
]


def _png(color: tuple[int, int, int]) -> io.BytesIO:
    buf = io.BytesIO()
    Image.new("RGB", (12, 8), color).save(buf, format="PNG")
    buf.seek(0)
    return buf


def build(path: str, *, active_sheet: str = "CCCD") -> str:
    """Write the workbook to `path` and return it.

    `active_sheet` defaults to `CCCD` on purpose: that is what the real file
    does, and it is the condition that makes `workbook.active` pick the wrong
    sheet. Tests rely on this default to reproduce the bug.
    """
    wb = openpyxl.Workbook()
    ctv = wb.active
    ctv.title = "CTV"

    # Header block above the table, as the real template has.
    ctv["A1"] = "THANH TOÁN DỊCH VỤ"
    ctv["A2"] = "Mã eform plan:"
    ctv["A3"] = "Mã eform thanh toán:"
    headers = ["STT", "Họ và tên", "CCCD/ PP", "MST", "Ngày/ tháng/ năm sinh",
               "Giới tính", "Số tài khoản", "Ngân hàng", "Gross", "Thuế PIT", "Thực Nhận"]
    for col, text in enumerate(headers, start=1):
        ctv.cell(row=5, column=col, value=text)
    for i, (name, cccd, mst) in enumerate(PEOPLE):
        r = 7 + i
        ctv.cell(r, 1, i + 1)
        ctv.cell(r, 2, name)
        ctv.cell(r, 3, cccd)
        ctv.cell(r, 4, mst)
        ctv.cell(r, 5, "01/01/1990")
        ctv.cell(r, 6, "NAM")
        ctv.cell(r, 7, "0123456789")
        ctv.cell(r, 8, "Ngân hàng Thử Nghiệm")
        ctv.cell(r, 9, 8000000)
        ctv.cell(r, 10, 0)
        ctv.cell(r, 11, 8000000)

    cccd_sheet = wb.create_sheet("CCCD")
    for col, text in enumerate(["STT", "Họ tên", "Số CCCD", "Hình CCCD", None, "STK", "Hình Ảnh"],
                               start=1):
        if text is not None:
            cccd_sheet.cell(row=1, column=col, value=text)
    # The header that makes the layout self-describing: one label over two columns.
    cccd_sheet.merge_cells("D1:E1")
    for i, (name, cccd, _) in enumerate(PEOPLE):
        r = 2 + i
        cccd_sheet.cell(r, 1, i + 1)
        cccd_sheet.cell(r, 2, name)
        cccd_sheet.cell(r, 3, cccd)
        cccd_sheet.cell(r, 6, "0123456789 - Ngân hàng Thử Nghiệm")
        cccd_sheet.add_image(XLImage(_png((200, 30, 30))), f"D{r}")   # front
        cccd_sheet.add_image(XLImage(_png((30, 30, 200))), f"E{r}")   # back
        cccd_sheet.add_image(XLImage(_png((30, 200, 30))), f"G{r}")   # bank screenshot

    mst_sheet = wb.create_sheet("MST")
    for col, text in enumerate(["STT", "Họ tên", "MST", "Hình Ảnh"], start=1):
        mst_sheet.cell(row=1, column=col, value=text)
    for i, (name, _, mst) in enumerate(PEOPLE):
        r = 2 + i
        mst_sheet.cell(r, 1, i + 1)
        mst_sheet.cell(r, 2, name)
        mst_sheet.cell(r, 3, mst)
        mst_sheet.add_image(XLImage(_png((200, 200, 30))), f"D{r}")

    wb.active = wb.sheetnames.index(active_sheet)
    wb.save(path)
    return path


def build_july(path: str) -> str:
    """The other template: one sheet, no images, header on row 1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Thông tin CK"
    for col, text in enumerate(["STT", "Họ và tên", "CCCD", "MST", "Số tài khoản",
                                "Gross", "Thuế PIT", "Thực Nhận"], start=1):
        ws.cell(row=1, column=col, value=text)
    for i, (name, cccd, mst) in enumerate(PEOPLE):
        r = 2 + i
        ws.cell(r, 1, i + 1); ws.cell(r, 2, name); ws.cell(r, 3, cccd); ws.cell(r, 4, mst)
        ws.cell(r, 5, "0123456789"); ws.cell(r, 6, 8000000); ws.cell(r, 7, 0); ws.cell(r, 8, 8000000)
    wb.save(path)
    return path
