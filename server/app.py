"""FastAPI backend: accepts an uploaded scanned PDF (+ optional roster), runs
split -> OCR/extract on a background thread with progress, and serves the
per-packet CtvFolder manifests + rendered page PNGs the frontend consumes.

Binds 127.0.0.1 only (run via `uvicorn app:app --host 127.0.0.1`). Each
upload becomes a durable **case** under `server/data/cases/<id>/` (see
`cases.py`) — unlike the old temp-dir job store, this survives a backend
restart: `case.json` holds the metadata + per-packet reviews, and
`packets/<i>/` holds that packet's manifest + rendered pages. Never commit
`server/data` — see server/README.md for the full PII note.
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
from copy import deepcopy
from datetime import datetime, timezone

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel, Field, field_validator
from starlette.concurrency import run_in_threadpool
from typing import Literal
import threading
import tempfile

import openpyxl

import cccd_workbook
import roster_checks
import workbook_layout

from cases import CaseStore, compact_cccd_summary, progress_of
from cccd_manual import (
    CccdManualError,
    assign_card,
    card_side_path,
    list_cards,
)
from cccd_workbook import MAX_WORKBOOK_BYTES as MAX_CCCD_WORKBOOK_BYTES
from pipeline import run_pipeline  # noqa: F401 - referenced as `run_pipeline` at call
                                    # time below so tests can monkeypatch this name.
from report import build_report
from roster_workbook import (
    RosterWorkbookError,
    load_roster_rows,
    preflight_roster_workbook,
)
import criteria as cr
from cases import effective_overrides
from evaluate import as_payload as criteria_payload
from summary_criteria import as_payload as summary_payload

app = FastAPI()

# Dev origins only: the Vite dev server's default port plus its usual
# fallbacks when 5173 is taken, on both localhost and 127.0.0.1.
_DEV_ORIGINS = [
    f"http://{host}:{port}"
    for host in ("localhost", "127.0.0.1")
    for port in (5173, 5174, 5175)
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=_DEV_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

store = CaseStore(os.path.join(os.path.dirname(__file__), "data", "cases"))

# Live progress while a case is `processing` (not persisted — it's transient
# and only meaningful for the in-flight run); keyed by case id.
_progress: dict[str, dict] = {}

_PAGE_NAME_RE = re.compile(r"^[A-Za-z0-9_.-]+\.(?:png|jpe?g)$")


def rewrite_manifest_urls(manifest: dict, base: str) -> dict:
    """Point every page's `src` at `{base}/page/{basename}`.

    `ocr_extract.render_pages` writes `src` as an absolute on-disk path (fine
    for the offline/no-server use case); the server instead exposes pages
    under the case/packet-scoped page endpoint, keyed by basename only.
    """
    out = deepcopy(manifest)
    for doc in out.get("docs", []):
        for page in doc.get("pages", []):
            page["src"] = f"{base}/page/{os.path.basename(page['src'])}"
    return out


def _review_field_count(cid: str, index: int) -> int:
    path = os.path.join(
        store.case_dir(cid), "packets", str(index), "manifest.json",
    )
    try:
        with open(path, "r", encoding="utf-8") as f:
            manifest = json.load(f)
        fields = manifest.get("fields") if isinstance(manifest, dict) else None
        return len(fields) if isinstance(fields, list) else 0
    except (OSError, ValueError, TypeError):
        return 0


def _packet_for_response(cid: str, packet: dict, rollup: dict | None = None) -> dict:
    rollup = rollup or {}
    return {
        **packet,
        "reviewFieldCount": _review_field_count(cid, packet["index"]),
        # What the engine found and nobody has decided on yet. A candidate for
        # resubmission, not a resubmission -- see cases.needs_resubmit.
        "findingCount": rollup.get("findings", 0),
        # The engine's own rollup for this packet, so a list column and the
        # 25-criterion matrix cannot disagree. Absent when there is no roster to
        # compare against -- the list must not imply a verdict it cannot support.
        "aiStatus": rollup.get("aiStatus"),
        "documents": rollup.get("documents"),
        "hasCommitment": rollup.get("hasCommitment"),
    }


#: Statuses the engine produces that a reviewer would want to look at. `rv` is
#: not among them: six criteria open `rv` on every packet by construction, so
#: counting those would make every packet a candidate and say nothing.
_FINDING_STATUSES = ("no", "missing")


#: The bảng kê column. It is the reference being compared against, not a
#: document the CTV submits, so it never counts toward document completeness.
_ROSTER_COLUMN = "Excel"


def _packet_rollups(cid: str, case: dict) -> dict[int, dict]:
    """`{packet_index: {findings, aiStatus, documents, hasCommitment}}`, fresh.

    One loop, because reading each manifest and evaluating every criterion was
    already being done here for the finding count alone; the rollup, the document
    span and the commitment check are all derived from results already in hand,
    and measured at 0ms on top.

    Measured on the 41-packet July case: 13ms to load the manifests, 12ms to
    evaluate all 41 packets. The whole list request is ~0.5s warm, dominated by
    the roster .xlsx read rather than by anything here -- worth knowing before
    optimising the wrong half. Stays per-request rather than persisted: one less
    thing that can go stale against a re-ingest.

    `aiStatus` is the engine's own worst-wins rollup (`criteria.roll_up`) over
    the packet's criteria, so the list column and the 25-criterion matrix cannot
    disagree about the same packet.

    Document completeness comes from the engine too, rather than from a
    hand-written "required documents" list: a criterion whose document is absent
    already yields `missing` cells naming that document, so the span is the
    documents the criteria actually reach for this packet and `missing` is the
    subset that is not there.
    """
    rows = _roster_rows(cid)
    if not rows:
        return {}
    import criteria as cr
    import evaluate as ev

    out: dict[int, dict] = {}
    for packet in case["packets"]:
        path = os.path.join(store.case_dir(cid), "packets",
                            str(packet["index"]), "manifest.json")
        if not os.path.isfile(path):
            continue
        try:
            with open(path, "r", encoding="utf-8") as f:
                manifest = json.load(f)
        except (OSError, ValueError):
            continue
        decided = effective_overrides(packet.get("review"))
        results = ev.evaluate_packet(
            manifest, _roster_row_for(cid, packet), decided,
        )
        touched = {int(key.split(":", 1)[0]) for key in decided}
        # A finding a person has decided on is no longer a candidate -- it has
        # been looked at. Counting it in both places would make the two numbers
        # move together, which defeats separating them.
        cells = [c for r in results for c in r.cells if c.document != _ROSTER_COLUMN]
        span = {c.document for c in cells if c.status is not cr.Status.NOT_APPLICABLE}
        out[packet["index"]] = {
            "findings": sum(
                1 for r in results
                if r.status.value in _FINDING_STATUSES and r.stt not in touched),
            "aiStatus": cr.roll_up([r.status for r in results]).value,
            "documents": {
                "span": len(span),
                "missing": sorted(
                    {c.document for c in cells if c.status is cr.Status.MISSING}),
            },
            "hasCommitment": any(
                (d or {}).get("kind") == "commitment"
                for d in (manifest.get("docs") or [])),
        }
    return out


def _upload_size(upload: UploadFile) -> int:
    upload.file.seek(0, os.SEEK_END)
    size = upload.file.tell()
    upload.file.seek(0)
    return size


def _same_bytes(path: str, upload_file) -> bool:
    """Whether the stored file and the uploaded stream hold identical bytes.

    Compared by digest in bounded chunks rather than read whole: these are up to
    25 MB workbooks and the check runs on the request thread.
    """
    stored = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            stored.update(chunk)
    incoming = hashlib.sha256()
    upload_file.seek(0)
    for chunk in iter(lambda: upload_file.read(1024 * 1024), b""):
        incoming.update(chunk)
    upload_file.seek(0)
    return stored.digest() == incoming.digest()


def _roster_rejection(upload: UploadFile) -> str | None:
    """None if the workbook is usable as a bảng kê, else why it is not.

    Returns the reason rather than a bool so the reviewer learns which sheet was
    read and which columns were missing. A wrong-sheet read and a legitimately
    empty column otherwise report the same thing in every Excel cell of the
    matrix, and only after a full processing run.
    """
    if not (upload.filename or "").casefold().endswith(".xlsx"):
        return "not-xlsx"
    try:
        upload.file.seek(0)
        preflight_roster_workbook(upload.file)
        return None
    except RosterWorkbookError as error:
        return str(error)
    except Exception:
        return "invalid-workbook"
    finally:
        upload.file.seek(0)


async def _validate_uploads(
    roster: UploadFile | None,
    cccd: UploadFile | None,
) -> None:
    if cccd is not None and roster is None:
        raise HTTPException(
            status_code=422,
            detail={"code": "cccd-requires-roster"},
        )
    if roster is not None:
        reason = await run_in_threadpool(_roster_rejection, roster)
        if reason is not None:
            raise HTTPException(
                status_code=422,
                detail={"code": "invalid-roster-workbook", "reason": reason},
            )
    if cccd is None:
        return
    if not (cccd.filename or "").casefold().endswith(".xlsx"):
        raise HTTPException(
            status_code=422,
            detail={"code": "invalid-cccd-workbook"},
        )
    if _upload_size(cccd) > MAX_CCCD_WORKBOOK_BYTES:
        raise HTTPException(
            status_code=413,
            detail={"code": "cccd-workbook-too-large"},
        )


def _run_case(
    cid: str,
    pdf_path: str,
    roster_path: str | None,
    cccd_path: str | None = None,
) -> None:
    case_dir = store.case_dir(cid)

    def cb(stage: str, done: int, total: int, detail: str) -> None:
        _progress[cid] = {"stage": stage, "done": done, "total": total, "detail": detail}

    try:
        result = run_pipeline(
            pdf_path,
            roster_path,
            case_dir,
            cb,
            cccd_xlsx_path=cccd_path,
        )
        store.set_result(
            cid,
            summary=result.get("summary"),
            packets=result.get("packets", []),
            cccd_workbook=result.get("cccdWorkbook"),
            purchase_total=result.get("purchaseTotal"),
        )
    except Exception as e:  # noqa: BLE001 - surfaced to the caller via case["error"]
        store.set_error(cid, str(e))
    finally:
        _progress.pop(cid, None)


@app.post("/api/cases")
async def post_case(
    pdf: UploadFile = File(...),
    roster: UploadFile | None = File(None),
    cccd: UploadFile | None = File(None),
):
    await _validate_uploads(roster, cccd)
    now = datetime.now(timezone.utc).isoformat()
    cid = store.create(name=pdf.filename or "case", pdf_name=pdf.filename or "input.pdf",
                        roster_name=roster.filename if roster is not None else None, now=now,
                        cccd_name=cccd.filename if cccd is not None else None)
    case_dir = store.case_dir(cid)

    pdf_path = os.path.join(case_dir, "input.pdf")
    with open(pdf_path, "wb") as f:
        shutil.copyfileobj(pdf.file, f)

    roster_path = None
    if roster is not None:
        roster_path = os.path.join(case_dir, "roster.xlsx")
        with open(roster_path, "wb") as f:
            shutil.copyfileobj(roster.file, f)

    cccd_path = None
    if cccd is not None:
        # The combined template is one file serving as both the bảng kê and the
        # card source, so a reviewer selects it in both fields. Store it once and
        # let both roles point at it, rather than keeping two copies of an 18 MB
        # workbook and OCRing a second set of identical drawings.
        cccd.file.seek(0)
        if roster_path is not None and _same_bytes(roster_path, cccd.file):
            cccd_path = roster_path
        else:
            cccd_path = os.path.join(case_dir, "cccd.xlsx")
            cccd.file.seek(0)
            with open(cccd_path, "wb") as f:
                shutil.copyfileobj(cccd.file, f)
    elif roster_path is not None and _holds_cards(roster_path):
        # One file, one upload. The combined template is both the bảng kê and
        # the card source; a reviewer who selects it once should not have to
        # know to select it again in the CCCD field. Selecting it once used to
        # ingest no cards at all, and every packet then reported
        # CCCD/Passport missing with nothing on screen to say why.
        cccd_path = roster_path

    t = threading.Thread(
        target=_run_case,
        args=(cid, pdf_path, roster_path, cccd_path),
        daemon=True,
    )
    t.start()
    return {"case_id": cid}


@app.get("/api/cases")
async def list_cases():
    return store.list()


@app.get("/api/cases/{cid}")
async def get_case(cid: str):
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    out = dict(case)
    out.pop("cccdWorkbook", None)
    out["cccdSummary"] = compact_cccd_summary(
        case.get("cccdWorkbook"),
    )
    rollups = await run_in_threadpool(_packet_rollups, cid, case)
    out["packets"] = [
        _packet_for_response(cid, packet, rollups.get(packet["index"]))
        for packet in case["packets"]
    ]
    out["progress"] = {
        **progress_of(case["packets"]),
        "candidates": sum(1 for r in rollups.values() if r.get("findings")),
    }
    if case["status"] == "processing" and cid in _progress:
        out["liveProgress"] = _progress[cid]
    return out


def _roster_rows(cid: str) -> list[list]:
    """The bảng kê rows as uploaded, or [] if this case has no roster."""
    path = os.path.join(store.case_dir(cid), "roster.xlsx")
    if not os.path.isfile(path):
        return []
    try:
        with open(path, "rb") as f:
            return load_roster_rows(f)
    except (OSError, ValueError):
        return []


@app.get("/api/cases/{cid}/summary")
async def get_summary(cid: str):
    """The five roster-level criteria — Acc's Tổng hợp tab.

    These span the whole bảng kê rather than one CTV, so they are computed here
    from the roster as uploaded plus the packets' duplicate-identity flags.
    """
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    rows = await run_in_threadpool(_roster_rows, cid)
    payload = await run_in_threadpool(
        summary_payload, rows, case["packets"], case.get("purchaseTotal"),
    )
    payload["rosterName"] = case.get("rosterName")
    return payload


def _roster_row_for(cid: str, packet: dict) -> dict | None:
    """The bảng kê row this packet was matched to, by CCCD then name."""
    identity = packet.get("rosterIdentity") or {}
    if not identity:
        return None
    import roster_checks

    rows = _roster_rows(cid)
    if not rows:
        return None
    columns, first_data = roster_checks.locate_columns(rows)
    people, _ = roster_checks.read_people(rows, columns, first_data)
    wanted = roster_checks.digits(identity.get("cccd", ""))
    if wanted:
        for person in people:
            if roster_checks.digits(person.get("cccd", "")) == wanted:
                return person
    name = _norm(identity.get("name", ""))
    if name:
        for person in people:
            if _norm(person.get("name", "")) == name:
                return person
    return None


def _norm(value: str) -> str:
    from ocr_extract import norm

    return norm(value or "").strip()


class CriterionDecisionBody(BaseModel):
    """A reviewer's decision on one criteria cell.

    No reason and no second confirmation, per Acc: a decision is one click.
    `reason` stays available for a reviewer who wants to record why.
    """

    toStatus: str
    reason: str = ""


def _parse_override_key(key: str) -> tuple[int, str]:
    """`"21:Hợp đồng"` -> `(21, "Hợp đồng")`, or 422."""
    stt, _, document = key.partition(":")
    if not stt.isdigit() or not document:
        raise HTTPException(status_code=422,
                            detail={"code": "invalid-criterion-key"})
    return int(stt), document


@app.put("/api/cases/{cid}/packets/{i}/criteria/{key}")
async def put_criterion_decision(
    cid: str, i: int, key: str, body: CriterionDecisionBody,
):
    """Record a reviewer's decision on one cell, keeping what the engine thought.

    The engine is re-run first so `fromStatus` is the status actually being
    replaced rather than whatever the client believed -- a stale matrix must not
    be able to write a wrong provenance into the audit trail.
    """
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    packet = next((p for p in case["packets"] if p["index"] == i), None)
    if packet is None:
        raise HTTPException(status_code=404, detail="packet not found")
    stt, document = _parse_override_key(key)

    path = os.path.join(store.case_dir(cid), "packets", str(i), "manifest.json")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="manifest not found")
    with open(path, "r", encoding="utf-8") as f:
        manifest = json.load(f)
    row = await run_in_threadpool(_roster_row_for, cid, packet)
    current = await run_in_threadpool(
        criteria_payload, manifest, row,
        effective_overrides(packet.get("review")),
    )
    row_now = next((c for c in current["criteria"] if c["stt"] == stt), None)
    cell_now = next((c for c in (row_now or {}).get("cells", [])
                     if c["document"] == document), None)
    if cell_now is None:
        raise HTTPException(
            status_code=422,
            detail={"code": "cell-not-in-criterion", "key": key})

    try:
        override = cr.Override(
            stt=stt, document=document,
            from_status=cr.Status(cell_now["status"]),
            to_status=cr.Status(body.toStatus),
            reason=body.reason,
            at=datetime.now(timezone.utc).isoformat(),
            by="",   # no auth yet; the spec's §6 allows an empty author
        )
    except ValueError as e:
        raise HTTPException(status_code=422,
                            detail={"code": "invalid-decision",
                                    "message": str(e)}) from e

    updated = store.add_override(cid, i, override)
    if updated is None:
        raise HTTPException(status_code=404, detail="case or packet not found")
    saved = next(p for p in updated["packets"] if p["index"] == i)
    return {
        "override": override.as_dict(),
        "history": saved["review"]["overrides"][override.key],
        "packet": _packet_for_response(cid, saved),
        "status": updated["status"],
    }


@app.get("/api/cases/{cid}/packets/{i}/criteria")
async def get_criteria(cid: str, i: int):
    """Acc's 25-criterion matrix for one packet, computed from its manifest.

    The matrix used to be hand-typed; this is the engine's output. Every cell
    carries the value read, why it got its status, and the evidence behind it.
    """
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    packet = next((p for p in case["packets"] if p["index"] == i), None)
    if packet is None:
        raise HTTPException(status_code=404, detail="packet not found")
    path = os.path.join(store.case_dir(cid), "packets", str(i), "manifest.json")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="manifest not found")
    with open(path, "r", encoding="utf-8") as f:
        manifest = json.load(f)
    row = await run_in_threadpool(_roster_row_for, cid, packet)
    payload = await run_in_threadpool(
        criteria_payload, manifest, row,
        effective_overrides(packet.get("review")),
    )
    payload["packet"] = i
    payload["name"] = packet.get("name") or ""
    return payload


PacketRejectionReason = Literal[
    "missing_documents",
    "wrong_template",
    "missing_signature",
]


class PacketRejectionBody(BaseModel):
    reasons: list[PacketRejectionReason] = Field(min_length=1)
    note: str = ""

    @field_validator("reasons")
    @classmethod
    def reasons_must_be_unique(cls, reasons):
        if len(reasons) != len(set(reasons)):
            raise ValueError("rejection reasons must be unique")
        return reasons


class ReviewBody(BaseModel):
    done: bool = False
    fields: dict = Field(default_factory=dict)
    rejection: PacketRejectionBody | None = None


@app.put("/api/cases/{cid}/packets/{i}/review")
async def put_review(cid: str, i: int, body: ReviewBody):
    updated = store.set_review(cid, i, body.model_dump())
    if updated is None:
        raise HTTPException(status_code=404, detail="case or packet not found")
    packet = next((p for p in updated["packets"] if p["index"] == i), None)
    return {"packet": _packet_for_response(cid, packet),
            "progress": progress_of(updated["packets"]),
            "status": updated["status"]}


def _load_manifests(cid: str, packets: list[dict]) -> dict:
    out = {}
    for p in packets:
        path = os.path.join(store.case_dir(cid), "packets", str(p["index"]), "manifest.json")
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                out[p["index"]] = json.load(f)
    return out


@app.post("/api/cases/{cid}/report")
async def post_report(cid: str):
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    manifests = _load_manifests(cid, case["packets"])
    now = datetime.now(timezone.utc).isoformat()
    # With the roster, the report carries the criteria engine's own findings and
    # the roster-level section -- not only what a reviewer flagged by hand.
    rows = await run_in_threadpool(_roster_rows, cid)
    decisions = {
        p["index"]: effective_overrides(p.get("review"))
        for p in case["packets"]
    }
    report = await run_in_threadpool(
        build_report, case, manifests, now, rows or None,
        case.get("purchaseTotal"), decisions,
    )
    case_dir = store.case_dir(cid)
    with open(os.path.join(case_dir, "report.md"), "w", encoding="utf-8") as f:
        f.write(report["markdown"])
    with open(os.path.join(case_dir, "report.csv"), "w", encoding="utf-8") as f:
        f.write(report["csv"])
    return report


@app.get("/api/cases/{cid}/report.md")
async def get_report_md(cid: str):
    path = os.path.join(store.case_dir(cid), "report.md")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="report not generated")
    return FileResponse(path, media_type="text/markdown", filename="bao-cao.md")


@app.get("/api/cases/{cid}/report.csv")
async def get_report_csv(cid: str):
    path = os.path.join(store.case_dir(cid), "report.csv")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="report not generated")
    return FileResponse(path, media_type="text/csv", filename="bao-cao.csv")


@app.delete("/api/cases/{cid}")
async def delete_case(cid: str):
    store.delete(cid)
    return {"ok": True}


@app.get("/api/cases/{cid}/packets/{i}/manifest.json")
async def get_manifest(cid: str, i: int):
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    path = os.path.join(store.case_dir(cid), "packets", str(i), "manifest.json")
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="manifest not found")
    with open(path, "r", encoding="utf-8") as f:
        manifest = json.load(f)
    base = f"/api/cases/{cid}/packets/{i}"
    return JSONResponse(rewrite_manifest_urls(manifest, base))


@app.get("/api/cases/{cid}/packets/{i}/page/{name}")
async def get_page(cid: str, i: int, name: str):
    if not _PAGE_NAME_RE.match(name):
        raise HTTPException(status_code=400, detail="invalid page name")
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    path = os.path.join(store.case_dir(cid), "packets", str(i), name)
    if not os.path.isfile(path):
        raise HTTPException(status_code=404, detail="page not found")
    return FileResponse(path)


# --- manual CCCD card assignment -------------------------------------------
# Roughly half the cards in a real workbook never yield a readable number, so
# the matcher can never place them. These three routes let the reviewer do it
# by eye. Deliberately narrow: card ids and images for UNATTACHED cards only —
# never file paths, never roster values (see cases.compact_cccd_summary).


class CccdAssignBody(BaseModel):
    packetIndex: int | None = None


def _manifest_paths(cid: str, packets: list[dict]) -> dict[int, str]:
    return {
        p["index"]: os.path.join(
            store.case_dir(cid), "packets", str(p["index"]), "manifest.json"
        )
        for p in packets
    }


def _case_or_404(cid: str) -> dict:
    case = store.get(cid)
    if case is None:
        raise HTTPException(status_code=404, detail="case not found")
    return case


@app.get("/api/cases/{cid}/cccd-cards")
async def get_cccd_cards(cid: str):
    case = _case_or_404(cid)
    try:
        return {"cards": list_cards(case)}
    except CccdManualError as error:
        raise HTTPException(status_code=409, detail={"code": error.code})


@app.get("/api/cases/{cid}/cccd-cards/{card_id}/image/{side}")
async def get_cccd_card_image(cid: str, card_id: str, side: str):
    case = _case_or_404(cid)
    try:
        path = card_side_path(case, store.case_dir(cid), card_id, side)
    except CccdManualError as error:
        raise HTTPException(status_code=404, detail={"code": error.code})
    return FileResponse(path)


@app.put("/api/cases/{cid}/cccd-cards/{card_id}")
async def put_cccd_card(cid: str, card_id: str, body: CccdAssignBody):
    case = _case_or_404(cid)
    try:
        updated = assign_card(
            case,
            card_id,
            body.packetIndex,
            store.case_dir(cid),
            _manifest_paths(cid, case["packets"]),
        )
    except CccdManualError as error:
        raise HTTPException(status_code=409, detail={"code": error.code})
    store.set_cccd_workbook(cid, updated["cccdWorkbook"])
    return {
        "cards": list_cards(updated),
        "cccdSummary": compact_cccd_summary(updated["cccdWorkbook"]),
    }
@app.post("/api/uploads/inspect")
async def post_inspect(
    roster: UploadFile = File(...),
    cccd: UploadFile | None = File(None),
):
    """What the tool inferred from these workbooks, before anything is committed.

    Inference can be confidently wrong and silent -- reading the CCCD sheet as
    the bảng kê used to look exactly like a roster with empty columns. So it is
    declared: which sheet was read, how many people it holds, and which column
    each population of images was found in. A reviewer confirms that in seconds
    rather than discovering it after a ~50-minute run.
    """
    reason = await run_in_threadpool(_roster_rejection, roster)
    if reason is not None:
        raise HTTPException(
            status_code=422,
            detail={"code": "invalid-roster-workbook", "reason": reason},
        )

    def describe() -> dict:
        roster.file.seek(0)
        rows = load_roster_rows(roster.file)
        columns, first_data = roster_checks.locate_columns(rows)
        people, _ = roster_checks.read_people(rows, columns or {}, first_data)
        roster.file.seek(0)
        sheet = _selected_sheet_name(roster.file)

        images: list[dict] = []
        for upload in _distinct_workbooks(roster, cccd):
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as scratch:
                upload.file.seek(0)
                shutil.copyfileobj(upload.file, scratch)
                scratch_path = scratch.name
            upload.file.seek(0)
            try:
                images.extend(cccd_workbook.describe_image_columns(scratch_path))
            except Exception:
                # A workbook whose drawings cannot be walked still has a usable
                # roster half; report what is known rather than failing the
                # whole declaration.
                pass
            finally:
                os.unlink(scratch_path)
        return {
            "rosterSheet": sheet,
            "people": len(people),
            "columns": sorted(columns or {}),
            "images": images,
        }

    return await run_in_threadpool(describe)


def _holds_cards(xlsx_path: str) -> bool:
    """Whether this workbook carries a column of ID-card images.

    Deliberately narrower than "has any images": the combined template also
    holds bank and tax-lookup screenshots, and the older single-sheet bảng kê
    holds none at all. Only a column the header identifies as cards makes a
    workbook worth walking for cards.
    """
    try:
        columns = cccd_workbook.describe_image_columns(xlsx_path)
    except Exception:
        # A workbook whose drawings cannot be walked still has a usable roster
        # half. Treat it as card-less rather than failing the upload.
        return False
    return any(column.get("kind") == "card" for column in columns)


def _distinct_workbooks(roster: UploadFile, cccd: UploadFile | None):
    """The workbooks to walk for images, without walking the same one twice --
    the combined template arrives in both fields."""
    if cccd is None:
        return [roster]
    cccd.file.seek(0)
    roster.file.seek(0)
    same = hashlib.sha256(roster.file.read()).digest() == hashlib.sha256(
        cccd.file.read()
    ).digest()
    roster.file.seek(0)
    cccd.file.seek(0)
    return [roster] if same else [roster, cccd]


def _selected_sheet_name(xlsx_source) -> str | None:
    """Which sheet `load_roster_rows` chose, for the declaration to name."""
    xlsx_source.seek(0)
    workbook = openpyxl.load_workbook(xlsx_source, read_only=True, data_only=True)
    try:
        sheets = {
            name: [list(row) for row in workbook[name].iter_rows(values_only=True)]
            for name in workbook.sheetnames
        }
        return workbook_layout.select_roster_sheet(sheets)
    finally:
        workbook.close()
        xlsx_source.seek(0)
