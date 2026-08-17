"""Capability-aware semantic validation for prepared intake v2 packages."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
import hmac
from io import BytesIO
import json
import threading

import fitz
import openpyxl
from PIL import Image
from pydantic import ValidationError

from ctv_inspection import inspect_observation
from ctv_inventory import InventoryObservation
from ctv_package_assignment import _source_id
from ctv_package_builder import (
    ArtifactReceipt,
    PackageIdentity,
    build_manifest_bytes,
    create_build_plan,
    iter_rendered_artifacts,
)
from ctv_proposal import (
    ApprovedProposalSnapshot,
    ProposalState,
    RosterRowSnapshot,
    SourceDispositionSnapshot,
    UnitDecisionSnapshot,
)
from intake_contract_v2 import (
    AssignmentsDocumentV2,
    CanonicalRosterDocumentV2,
    CanonicalRosterRowV2,
    CanonicalRosterValuesV2,
    ExceptionsDocumentV2,
    MAX_INPUT_PDF_BYTES,
    MAX_JSON_BYTES,
    MAX_PACKAGE_BYTES,
    MAX_ROSTER_OR_EVIDENCE_BYTES,
    PackageManifestV2,
    UnacquiredSourceV2,
    ValidationCheckV2,
    ValidationReportV2,
    VerifiedSourceV2,
)
from intake_package_validator import MAX_MANIFEST_BYTES, _PackageReader
from roster_workbook import preflight_roster_workbook


VALIDATOR_VERSION_V2 = "ctv-intake-v2-validator/1.0"
_MANIFEST_NAME = "case-manifest.json"
_RECEIPT_NAME = "validation-report.json"
_ROSTER_HEADERS = (
    "Roster Row ID", "Name", "Identity", "FA Code", "Tax ID",
    "Birth Date", "Bank Account", "Service Fee", "Product",
)
_ROSTER_FIELDS = (
    "name", "identity", "fa_code", "tax_id", "birth_date",
    "bank_account", "service_fee", "product",
)
_ROLES_BY_KIND = {
    "pdf-page": {
        "service-contract", "acceptance-record", "payment-tax-form",
        "shared-supporting-evidence", "other-supporting-evidence",
    },
    "worksheet": {"payment-roster", "other-supporting-evidence"},
    "image": {
        "identity-front", "identity-back", "shared-supporting-evidence",
        "other-supporting-evidence",
    },
}


@dataclass(frozen=True)
class V2ValidationExpectation:
    observation_id: str
    proposal_digest: str
    expected_manifest_sha256: str | None = None

    def __post_init__(self) -> None:
        if (
            type(self.observation_id) is not str
            or not self.observation_id.startswith("observation-")
            or len(self.observation_id) != 76
            or type(self.proposal_digest) is not str
            or len(self.proposal_digest) != 64
            or any(character not in "0123456789abcdef" for character in self.proposal_digest)
            or (
                self.expected_manifest_sha256 is not None
                and (
                    type(self.expected_manifest_sha256) is not str
                    or len(self.expected_manifest_sha256) != 64
                    or any(
                        character not in "0123456789abcdef"
                        for character in self.expected_manifest_sha256
                    )
                )
            )
        ):
            raise ValueError("v2-validation-expectation-invalid")


@dataclass(frozen=True)
class ContentValidationV2:
    report: ValidationReportV2
    manifest_sha256: str
    declared_artifact_set_sha256: str
    tree_sha256: str
    artifact_sha256: tuple[tuple[str, str], ...]


class _SourceSnapshotCache:
    """Reuse one bounded acquisition through strict production consumers."""

    def __init__(self, observation: InventoryObservation) -> None:
        self._observation = observation
        self._sources = {
            item.evidence_id: item for item in observation.sources
        }
        self._content: dict[str, bytes] = {}
        self._reserved_bytes = 0
        self._lock = threading.RLock()

    @property
    def charged_bytes(self) -> int:
        with self._lock:
            return self._reserved_bytes

    def snapshot(self, evidence_id: str, *, max_bytes: int) -> bytes:
        with self._lock:
            cached = self._content.get(evidence_id)
            if cached is not None:
                if len(cached) > max_bytes:
                    raise ValueError("source exceeds max_bytes")
                return cached
            source = self._sources.get(evidence_id)
            if source is None or type(source.size) is not int:
                raise ValueError("source is not bound to observation")
            if type(max_bytes) is not int or max_bytes < 0 or source.size > max_bytes:
                raise ValueError("source exceeds max_bytes")
            next_total = self._reserved_bytes + source.size
            if next_total > MAX_PACKAGE_BYTES:
                raise ValueError("source aggregate exceeds validator bound")
            self._reserved_bytes = next_total
            content = self._observation.snapshot(
                evidence_id, max_bytes=max_bytes
            )
            self._content[evidence_id] = content
            return content


class _DuplicateKey(ValueError):
    pass


def _unique_object(pairs):
    result = {}
    for key, value in pairs:
        if key in result:
            raise _DuplicateKey()
        result[key] = value
    return result


def _reject_constant(_value):
    raise ValueError("non-finite JSON number")


def _strict_json(content: bytes) -> object:
    return json.loads(
        content.decode("utf-8"),
        object_pairs_hook=_unique_object,
        parse_constant=_reject_constant,
    )


def _canonical_json(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8") + b"\n"


class _Checks:
    def __init__(self) -> None:
        self._checks: list[ValidationCheckV2] = []
        self._codes: set[str] = set()
        self.errors: list[str] = []

    def add(self, code: str, passed: bool, *evidence_refs: str) -> None:
        if code in self._codes:
            return
        self._codes.add(code)
        self._checks.append(
            ValidationCheckV2(
                code=code,
                passed=passed,
                evidenceRefs=sorted(set(evidence_refs)),
            )
        )
        if not passed:
            self.errors.append(code)

    @property
    def checks(self) -> list[ValidationCheckV2]:
        return list(self._checks)


def _report(
    checks: _Checks,
    *,
    package_id: str,
    observation_id: str,
    proposal_digest: str,
    manifest_sha256: str,
    artifact_set_sha256: str,
) -> ValidationReportV2:
    return ValidationReportV2(
        schemaVersion="2.0",
        outcome="invalid" if checks.errors else "valid",
        packageStatus="prepared",
        checks=checks.checks,
        errors=list(checks.errors),
        warnings=[],
        validatedAt=datetime.now(timezone.utc),
        validatorVersion=VALIDATOR_VERSION_V2,
        packageId=package_id,
        sourceObservationId=observation_id,
        proposalDigest=proposal_digest,
        manifestSha256=manifest_sha256,
        declaredArtifactSetSha256=artifact_set_sha256,
    )


def _fallback_result(
    checks: _Checks,
    expectation: V2ValidationExpectation,
    *,
    manifest_sha256: str = "0" * 64,
) -> ContentValidationV2:
    report = _report(
        checks,
        package_id="package-" + "0" * 64,
        observation_id=expectation.observation_id,
        proposal_digest=expectation.proposal_digest,
        manifest_sha256=manifest_sha256,
        artifact_set_sha256="0" * 64,
    )
    return ContentValidationV2(report, manifest_sha256, "0" * 64, "0" * 64, ())


def _artifact_limit(kind: str) -> int:
    if kind == "input-pdf":
        return MAX_INPUT_PDF_BYTES
    if kind in {"roster", "evidence"}:
        return MAX_ROSTER_OR_EVIDENCE_BYTES
    return MAX_JSON_BYTES


def _declared_artifact_set_sha256(manifest: PackageManifestV2) -> str:
    value = [
        {
            "artifactId": item.artifact_id,
            "path": item.path,
            "size": item.size,
            "sha256": item.sha256,
        }
        for item in manifest.artifacts
    ]
    return sha256(_canonical_json(value)).hexdigest()


def _load_roster(content: bytes, artifact_id: str) -> CanonicalRosterDocumentV2:
    preflight_roster_workbook(BytesIO(content))
    workbook = openpyxl.load_workbook(
        BytesIO(content), read_only=True, data_only=True, keep_links=False
    )
    try:
        if workbook.sheetnames != ["Roster"]:
            raise ValueError("roster sheet mismatch")
        rows = list(workbook.active.iter_rows(values_only=True))
        if not rows or tuple(rows[0]) != _ROSTER_HEADERS:
            raise ValueError("roster header mismatch")
        canonical_rows = []
        for row in rows[1:]:
            if len(row) != len(_ROSTER_HEADERS) or not row[0]:
                raise ValueError("roster row mismatch")
            values = dict(zip(_ROSTER_FIELDS, row[1:]))
            canonical_rows.append(
                CanonicalRosterRowV2(
                    rosterRowId=row[0],
                    values=CanonicalRosterValuesV2(
                        name=values["name"],
                        identity=values["identity"],
                        faCode=values["fa_code"],
                        taxId=values["tax_id"],
                        birthDate=values["birth_date"],
                        bankAccount=values["bank_account"],
                        serviceFee=values["service_fee"],
                        product=values["product"],
                    ),
                )
            )
        return CanonicalRosterDocumentV2(
            schemaVersion="2.0", artifactId=artifact_id, rows=canonical_rows
        )
    finally:
        workbook.close()


def _source_facts_are_consistent(
    manifest: PackageManifestV2,
    observation: InventoryObservation,
    inspection,
) -> bool:
    observed = {item.evidence_id: item for item in observation.sources}
    inventory = {item.evidence_id: item for item in observation.result.items}
    inspected = {item.evidence_id: item for item in inspection.sources}
    source_by_id = {
        _source_id(manifest.proposal_digest, evidence_id): evidence_id
        for evidence_id in observed
    }
    if set(source_by_id) != {item.source_id for item in manifest.sources}:
        return False
    for source in manifest.sources:
        evidence_id = source_by_id.get(source.source_id)
        if evidence_id is None:
            return False
        observed_source = observed[evidence_id]
        inventory_item = inventory[evidence_id]
        inspection_source = inspected[evidence_id]
        expected_path = (
            f"sources/{evidence_id}{observed_source.extension}"
            if observed_source.extension.startswith(".")
            else f"sources/{evidence_id}"
        )
        if source.path != expected_path:
            return False
        if isinstance(source, VerifiedSourceV2):
            expected_media = (
                "application/pdf"
                if inspection_source.detected_type == "pdf"
                else "application/octet-stream"
            )
            if (
                inspection_source.inspection_status != "inspected"
                or source.size != inventory_item.size
                or source.sha256 != inventory_item.sha256
                or inventory_item.hash_status != "computed"
                or source.media_type != expected_media
                or source.page_count
                != (
                    inspection_source.unit_count
                    if inspection_source.detected_type == "pdf"
                    else None
                )
            ):
                return False
        elif isinstance(source, UnacquiredSourceV2):
            if (
                source.acquisition_status != inspection_source.inspection_status
                or source.issue_codes != list(inspection_source.issue_codes)
            ):
                return False
        else:
            return False
    return True


def _assignment_semantics_are_consistent(
    assignments: AssignmentsDocumentV2,
    manifest: PackageManifestV2,
    roster: CanonicalRosterDocumentV2,
    inspection,
    artifacts: dict[str, bytes],
) -> bool:
    assignments.validate_against_manifest(manifest)
    assignments.validate_against_roster(roster)
    inspected_units = {item.unit_id: item for item in inspection.units}
    assigned_ids = {item.unit_id for item in assignments.units}
    excluded_unit_ids = {
        item.record_id for item in assignments.exclusions if item.record_type == "unit"
    }
    if assigned_ids | excluded_unit_ids != set(inspected_units):
        return False
    participant_order = {
        item.participant_handle: index
        for index, item in enumerate(assignments.participants)
    }
    for unit in assignments.units:
        inspected = inspected_units.get(unit.unit_id)
        if inspected is None:
            return False
        if (
            unit.source_id != _source_id(manifest.proposal_digest, inspected.evidence_id)
            or unit.source_unit_index != inspected.unit_index
            or unit.unit_kind != inspected.unit_kind
            or unit.role not in _ROLES_BY_KIND[unit.unit_kind]
        ):
            return False
        if unit.decision == "accepted":
            if inspected.suggested_role == "unknown" or unit.role != inspected.suggested_role:
                return False
        elif inspected.suggested_role != "unknown" and unit.role == inspected.suggested_role:
            return False
        handles = list(unit.target.participant_handles)
        if handles != sorted(handles, key=participant_order.__getitem__):
            return False
        if unit.role == "payment-roster" and (
            unit.unit_kind != "worksheet" or unit.target.scope != "case"
        ):
            return False

    workbook_units: dict[str, list] = defaultdict(list)
    for unit in assignments.units:
        if unit.output_locator.kind == "worksheet":
            workbook_units[unit.output_locator.artifact_id].append(unit)
    manifest_artifacts = {item.artifact_id: item for item in manifest.artifacts}
    for artifact_id, units in workbook_units.items():
        units.sort(key=lambda item: item.source_unit_index)
        if [item.output_locator.worksheet_index for item in units] != list(
            range(1, len(units) + 1)
        ):
            return False
        artifact = manifest_artifacts[artifact_id]
        content = artifacts[artifact_id]
        preflight_roster_workbook(BytesIO(content))
        workbook = openpyxl.load_workbook(
            BytesIO(content), read_only=True, data_only=True, keep_links=False
        )
        try:
            if len(workbook.worksheets) != len(units):
                return False
            if workbook.sheetnames != [
                f"Worksheet {index:04d}" for index in range(1, len(units) + 1)
            ]:
                return False
        finally:
            workbook.close()
    for artifact in manifest.artifacts:
        if artifact.kind != "evidence" or not artifact.path.endswith(".png"):
            continue
        with Image.open(BytesIO(artifacts[artifact.artifact_id])) as image:
            if image.format != "PNG" or image.n_frames != 1:
                return False
    return True


def _normalized_projection(
    manifest: PackageManifestV2,
    assignments: AssignmentsDocumentV2,
    observation: InventoryObservation,
    inspection,
    snapshot_source,
) -> ApprovedProposalSnapshot:
    """Rebuild only the normalized persisted projection, not user intent."""
    inspected_units = {item.unit_id: item for item in inspection.units}
    roster_units = [
        item
        for item in assignments.units
        if item.role == "payment-roster" and item.output_locator.kind == "roster"
    ]
    if len(roster_units) != 1:
        raise ValueError("roster projection mismatch")
    assigned_roster = roster_units[0]
    inspected_roster = inspected_units.get(assigned_roster.unit_id)
    if (
        inspected_roster is None
        or inspected_roster.unit_kind != "worksheet"
        or inspected_roster.suggested_role != "payment-roster"
    ):
        raise ValueError("roster projection mismatch")

    proposal = ProposalState.from_inspection(
        observation, inspection, _snapshot_source=snapshot_source
    )
    rows, row_issues, package_issues, columns = proposal._roster_rows(
        inspected_roster
    )
    if not rows or row_issues or package_issues:
        raise ValueError("roster projection mismatch")
    roster_rows = tuple(
        RosterRowSnapshot(
            participant_handle=f"participant-{index:04d}",
            row_index=row_index,
            name=row["name"],
            identity=row["identity"],
            fa_code=row.get("faCode", ""),
            tax_id=row.get("taxId", ""),
            birth_date=row.get("birthDate", ""),
            bank_account=row.get("bankAccount", ""),
            service_fee=row.get("serviceFee", ""),
            product=row.get("product", ""),
        )
        for index, (row, row_index) in enumerate(rows, start=1)
    )

    included = {item.unit_id: item for item in assignments.units}
    excluded = {
        item.record_id: item
        for item in assignments.exclusions
        if item.record_type == "unit"
    }
    if set(included) & set(excluded) or set(included) | set(excluded) != set(
        inspected_units
    ):
        raise ValueError("unit projection mismatch")
    pdf_pages = {
        (item.source_id, item.source_page): item for item in manifest.pdf_pages
    }
    source_models = {item.source_id: item for item in manifest.sources}
    unit_decisions = []
    for unit in sorted(
        inspection.units, key=lambda item: int(item.unit_id.rsplit("-", 1)[1])
    ):
        assigned = included.get(unit.unit_id)
        if assigned is not None:
            unit_decisions.append(
                UnitDecisionSnapshot(
                    unit_id=unit.unit_id,
                    evidence_id=unit.evidence_id,
                    unit_kind=unit.unit_kind,
                    unit_index=unit.unit_index,
                    decision=assigned.decision,
                    role=assigned.role,
                    scope=assigned.target.scope,
                    participant_handles=tuple(
                        assigned.target.participant_handles
                    ),
                )
            )
            continue
        source_id = _source_id(manifest.proposal_digest, unit.evidence_id)
        source_model = source_models.get(source_id)
        if source_model is None:
            raise ValueError("source projection mismatch")
        exclusion = excluded[unit.unit_id]
        if unit.unit_kind == "pdf-page":
            page = pdf_pages.get(
                (source_id, unit.unit_index)
            )
            if page is None:
                raise ValueError("PDF projection mismatch")
            reason = page.coverage_state
        elif source_model.coverage_state == "duplicate":
            reason = "duplicate"
        elif exclusion.reason == "excluded-by-user":
            reason = exclusion.reason
        else:
            raise ValueError("unit projection mismatch")
        if reason not in {"duplicate", "excluded-by-user"}:
            raise ValueError("unit projection mismatch")
        unit_decisions.append(
            UnitDecisionSnapshot(
                unit_id=unit.unit_id,
                evidence_id=unit.evidence_id,
                unit_kind=unit.unit_kind,
                unit_index=unit.unit_index,
                decision="excluded",
                reason=reason,
            )
        )

    unit_evidence_ids = {item.evidence_id for item in inspection.units}
    source_dispositions = []
    for source in inspection.sources:
        if source.evidence_id in unit_evidence_ids:
            continue
        source_model = source_models.get(
            _source_id(manifest.proposal_digest, source.evidence_id)
        )
        if not isinstance(source_model, UnacquiredSourceV2):
            raise ValueError("source projection mismatch")
        source_dispositions.append(
            SourceDispositionSnapshot(
                evidence_id=source.evidence_id,
                decision="excluded",
                reason=source_model.coverage_state,
                acquisition_status=source_model.acquisition_status,
                coverage_state=source_model.coverage_state,
                issue_codes=tuple(source_model.issue_codes),
            )
        )
    return ApprovedProposalSnapshot(
        observation_id=observation.observation_id,
        proposal_digest=manifest.proposal_digest,
        roster_unit_id=inspected_roster.unit_id,
        roster_evidence_id=inspected_roster.evidence_id,
        roster_worksheet_index=inspected_roster.unit_index,
        roster_rows=roster_rows,
        unit_decisions=tuple(unit_decisions),
        source_dispositions=tuple(source_dispositions),
        fa_code=roster_rows[0].fa_code,
        canonical_to_source_columns=columns,
    )


def _production_projection_is_consistent(
    manifest_content: bytes,
    manifest: PackageManifestV2,
    approved: ApprovedProposalSnapshot,
    observation: InventoryObservation,
    inspection,
    artifacts: dict[str, bytes],
    snapshot_source,
) -> bool:
    plan = create_build_plan(observation, inspection, approved)
    rendered = tuple(iter_rendered_artifacts(
        plan, observation, _snapshot_source=snapshot_source
    ))
    if {item.artifact_id for item in rendered} != set(artifacts):
        return False
    if any(artifacts[item.artifact_id] != item.content for item in rendered):
        return False
    receipts = tuple(ArtifactReceipt.from_rendered(item) for item in rendered)
    return manifest_content == build_manifest_bytes(plan, receipts)


def _validate_content(
    reader: _PackageReader,
    observation: InventoryObservation,
    expectation: V2ValidationExpectation,
    *,
    receipt_allowed: bool,
) -> ContentValidationV2:
    checks = _Checks()
    if (
        type(reader) is not _PackageReader
        or not reader.has_secure_open_provenance()
        or type(observation) is not InventoryObservation
    ):
        checks.add("secure-open-unavailable", False, "package-tree")
        return _fallback_result(checks, expectation)

    manifest_content, failure = reader.read_manifest()
    if manifest_content is None or failure is not None:
        checks.add("manifest-invalid", False, "manifest")
        return _fallback_result(checks, expectation)
    manifest_digest = sha256(manifest_content).hexdigest()
    if (
        expectation.expected_manifest_sha256 is not None
        and not hmac.compare_digest(
            manifest_digest, expectation.expected_manifest_sha256
        )
    ):
        checks.add(
            "writer-manifest-binding-mismatch", False, "manifest"
        )
    try:
        manifest = PackageManifestV2.model_validate(_strict_json(manifest_content))
    except (UnicodeError, ValueError, ValidationError):
        checks.add("manifest-invalid", False, "manifest")
        return _fallback_result(
            checks, expectation, manifest_sha256=manifest_digest
        )
    checks.add("manifest-valid", True, "manifest")
    artifact_set_digest = _declared_artifact_set_sha256(manifest)

    required = {_MANIFEST_NAME, *(item.path for item in manifest.artifacts)}
    optional = {_RECEIPT_NAME} if receipt_allowed else set()
    limits = {_MANIFEST_NAME: MAX_MANIFEST_BYTES}
    limits.update({item.path: _artifact_limit(item.kind) for item in manifest.artifacts})
    if receipt_allowed:
        limits[_RECEIPT_NAME] = MAX_JSON_BYTES
    tree, tree_failure = reader.snapshot_tree(
        set(required),
        max_bytes_by_path=limits,
        max_total_bytes=MAX_PACKAGE_BYTES,
        optional_paths=optional,
        exclude_from_digest={_RECEIPT_NAME} if receipt_allowed else set(),
    )
    if tree is None or tree_failure is not None:
        code = (
            "validation-report-present"
            if not receipt_allowed and tree_failure == "extra"
            else "package-tree-invalid"
        )
        checks.add(code, False, "package-tree")
        return ContentValidationV2(
            _report(
                checks,
                package_id=manifest.package_id,
                observation_id=manifest.source_observation_id,
                proposal_digest=manifest.proposal_digest,
                manifest_sha256=manifest_digest,
                artifact_set_sha256=artifact_set_digest,
            ),
            manifest_digest,
            artifact_set_digest,
            "0" * 64,
            (),
        )
    checks.add("package-tree-valid", True, "package-tree")

    artifacts: dict[str, bytes] = {}
    artifacts_valid = True
    artifact_digests = []
    for artifact in manifest.artifacts:
        content, read_failure = reader.read_cached(
            artifact.path, max_bytes=_artifact_limit(artifact.kind)
        )
        if (
            content is None
            or read_failure is not None
            or len(content) != artifact.size
            or sha256(content).hexdigest() != artifact.sha256
        ):
            artifacts_valid = False
            continue
        artifacts[artifact.artifact_id] = content
        artifact_digests.append((artifact.artifact_id, artifact.sha256))
    checks.add(
        "artifacts-valid" if artifacts_valid else "artifact-digest-mismatch",
        artifacts_valid,
        "artifacts",
    )

    expectation_valid = (
        observation.observation_id == expectation.observation_id
        and manifest.source_observation_id == expectation.observation_id
        and manifest.proposal_digest == expectation.proposal_digest
    )
    checks.add(
        "source-binding-valid" if expectation_valid else "source-binding-mismatch",
        expectation_valid,
        "source-observation",
    )
    identity = PackageIdentity.derive(
        observation_id=manifest.source_observation_id,
        proposal_digest=manifest.proposal_digest,
        writer_version=manifest.package_version,
        schema_version=manifest.schema_version,
        compatibility_target=manifest.compatibility_target,
    )
    identity_valid = (
        identity.package_id == manifest.package_id
        and identity.batch_id == manifest.batch_id
        and identity.case_id == manifest.case_id
        and manifest.validator_version == VALIDATOR_VERSION_V2
    )
    checks.add(
        "package-identity-valid" if identity_valid else "package-identity-mismatch",
        identity_valid,
        "package-identity",
    )

    source_cache = _SourceSnapshotCache(observation)
    try:
        inspection = inspect_observation(
            observation, _snapshot_source=source_cache.snapshot
        )
        sources_valid = _source_facts_are_consistent(
            manifest, observation, inspection
        )
    except Exception:
        inspection = None
        sources_valid = False
    checks.add(
        "sources-valid" if sources_valid else "source-verification-mismatch",
        sources_valid,
        "sources",
    )

    pdf_valid = False
    input_artifact = next(
        (item for item in manifest.artifacts if item.kind == "input-pdf"), None
    )
    if input_artifact is not None and input_artifact.artifact_id in artifacts:
        try:
            with fitz.open(
                stream=artifacts[input_artifact.artifact_id], filetype="pdf"
            ) as document:
                included = [
                    item
                    for item in manifest.pdf_pages
                    if item.coverage_state in {"assigned", "shared"}
                ]
                pdf_valid = (
                    document.page_count == len(included)
                    and sorted(item.target_page for item in included)
                    == list(range(1, document.page_count + 1))
                )
        except Exception:
            pdf_valid = False
    checks.add(
        "pdf-coverage-valid" if pdf_valid else "pdf-coverage-invalid",
        pdf_valid,
        "artifact-input-pdf",
    )

    assignments = None
    exceptions = None
    roster = None
    assignments_artifact = next(
        (item for item in manifest.artifacts if item.kind == "assignments"), None
    )
    exceptions_artifact = next(
        (item for item in manifest.artifacts if item.kind == "exceptions"), None
    )
    roster_artifact = next(
        (item for item in manifest.artifacts if item.kind == "roster"), None
    )
    try:
        if assignments_artifact is not None:
            assignments = AssignmentsDocumentV2.model_validate(
                _strict_json(artifacts[assignments_artifact.artifact_id])
            )
        if exceptions_artifact is not None:
            exceptions = ExceptionsDocumentV2.model_validate(
                _strict_json(artifacts[exceptions_artifact.artifact_id])
            )
        if roster_artifact is not None:
            roster = _load_roster(
                artifacts[roster_artifact.artifact_id], roster_artifact.artifact_id
            )
    except (KeyError, UnicodeError, ValueError, ValidationError, OSError, RuntimeError):
        pass

    roster_valid = roster is not None and all(
        row.values.fa_code == manifest.fa_code for row in roster.rows
    )
    checks.add(
        "roster-valid" if roster_valid else "roster-invalid",
        roster_valid,
        "roster",
    )
    exceptions_valid = (
        exceptions is not None
        and not exceptions.items
        and not manifest.exception_ids
    )
    checks.add(
        "exceptions-valid" if exceptions_valid else "exceptions-not-empty",
        exceptions_valid,
        "artifact-exceptions",
    )
    assignments_valid = False
    if (
        assignments is not None
        and roster is not None
        and inspection is not None
        and artifacts_valid
    ):
        try:
            assignments_valid = _assignment_semantics_are_consistent(
                assignments, manifest, roster, inspection, artifacts
            )
        except Exception:
            assignments_valid = False
    checks.add(
        "assignments-valid" if assignments_valid else "assignment-invalid",
        assignments_valid,
        "artifact-assignments",
    )
    projection_valid = False
    if (
        assignments is not None
        and inspection is not None
        and artifacts_valid
    ):
        try:
            approved = _normalized_projection(
                manifest,
                assignments,
                observation,
                inspection,
                source_cache.snapshot,
            )
        except Exception:
            approved = None
        if approved is not None:
            try:
                projection_valid = _production_projection_is_consistent(
                    manifest_content,
                    manifest,
                    approved,
                    observation,
                    inspection,
                    artifacts,
                    source_cache.snapshot,
                )
            except Exception:
                projection_valid = False
    checks.add(
        (
            "production-projection-valid"
            if projection_valid
            else "production-projection-mismatch"
        ),
        projection_valid,
        "package-projection",
    )

    report = _report(
        checks,
        package_id=manifest.package_id,
        observation_id=manifest.source_observation_id,
        proposal_digest=manifest.proposal_digest,
        manifest_sha256=manifest_digest,
        artifact_set_sha256=artifact_set_digest,
    )
    return ContentValidationV2(
        report=report,
        manifest_sha256=manifest_digest,
        declared_artifact_set_sha256=artifact_set_digest,
        tree_sha256=tree.tree_sha256,
        artifact_sha256=tuple(artifact_digests),
    )


def validate_v2_content_reader(
    reader: _PackageReader,
    observation: InventoryObservation,
    expectation: V2ValidationExpectation,
) -> ContentValidationV2:
    """Validate content through caller-owned package and source capabilities."""
    if type(expectation) is not V2ValidationExpectation:
        raise TypeError("v2 validation expectation is required")
    return _validate_content(
        reader, observation, expectation, receipt_allowed=False
    )


def canonical_v2_receipt_bytes(content: ContentValidationV2) -> bytes:
    """Serialize exactly the report produced by one validation result."""
    if type(content) is not ContentValidationV2:
        raise TypeError("v2 content validation is required")
    return _canonical_json(content.report.model_dump(by_alias=True, mode="json"))


def validate_v2_publication_reader(
    reader: _PackageReader,
    observation: InventoryObservation,
    expectation: V2ValidationExpectation,
) -> ContentValidationV2:
    """Recompute content, then prove the existing receipt describes it exactly."""
    if type(expectation) is not V2ValidationExpectation:
        raise TypeError("v2 validation expectation is required")
    content = _validate_content(
        reader, observation, expectation, receipt_allowed=True
    )
    consistent = not content.report.errors
    receipt = None
    receipt_content, receipt_failure = reader.read_cached(
        _RECEIPT_NAME, max_bytes=MAX_JSON_BYTES
    )
    if receipt_content is None or receipt_failure is not None:
        consistent = False
    else:
        try:
            receipt = ValidationReportV2.model_validate(
                _strict_json(receipt_content)
            )
            if receipt_content != _canonical_json(
                receipt.model_dump(by_alias=True, mode="json")
            ):
                consistent = False
        except (UnicodeError, ValueError, ValidationError):
            consistent = False
    if receipt is not None:
        expected_checks = [
            check.model_dump(by_alias=True) for check in content.report.checks
        ]
        actual_checks = [
            check.model_dump(by_alias=True) for check in receipt.checks
        ]
        consistent = consistent and (
            receipt.source_observation_id == content.report.source_observation_id
            and receipt.proposal_digest == content.report.proposal_digest
            and receipt.package_id == content.report.package_id
            and receipt.manifest_sha256 == content.manifest_sha256
            and receipt.declared_artifact_set_sha256
            == content.declared_artifact_set_sha256
            and receipt.outcome == content.report.outcome
            and receipt.package_status == content.report.package_status
            and receipt.validator_version == content.report.validator_version
            and receipt.errors == content.report.errors
            and receipt.warnings == content.report.warnings
            and actual_checks == expected_checks
            and all(
                check.code != "validation-report-consistent"
                for check in receipt.checks
            )
        )

    checks = _Checks()
    for check in content.report.checks:
        checks.add(check.code, check.passed, *check.evidence_refs)
    checks.add("validation-report-consistent", consistent, "receipt")
    report = _report(
        checks,
        package_id=content.report.package_id,
        observation_id=content.report.source_observation_id,
        proposal_digest=content.report.proposal_digest,
        manifest_sha256=content.manifest_sha256,
        artifact_set_sha256=content.declared_artifact_set_sha256,
    )
    return ContentValidationV2(
        report=report,
        manifest_sha256=content.manifest_sha256,
        declared_artifact_set_sha256=content.declared_artifact_set_sha256,
        tree_sha256=content.tree_sha256,
        artifact_sha256=content.artifact_sha256,
    )
