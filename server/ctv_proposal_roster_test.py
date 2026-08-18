"""TDD coverage for immutable private roster candidates and selection."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
import pytest

from ctv_inspection_model import (
    InspectionResult,
    InspectionSource,
    InspectionTotals,
    InspectionUnit,
)
from ctv_proposal_roster import (
    RosterCandidate,
    RosterCandidateRow,
    RosterSelection,
    choose_automatic_roster,
    load_roster_candidates,
)


_PRIVATE = ("Synthetic Person", "079123456781", "FA-SYNTHETIC-001")


class _HostileValue:
    def __eq__(self, _other):
        raise AssertionError("hostile equality must not run")

    def __int__(self):
        raise AssertionError("hostile coercion must not run")


def _candidate_row(**overrides):
    values = {
        "row_index": 2,
        "name": _PRIVATE[0],
        "identity": _PRIVATE[1],
        "values": (
            ("faCode", _PRIVATE[2]),
            ("identity", _PRIVATE[1]),
            ("name", _PRIVATE[0]),
        ),
    }
    values.update(overrides)
    return RosterCandidateRow(**values)


def _candidate(**overrides):
    values = {
        "unit_id": "unit-0001",
        "evidence_id": "evidence-0001",
        "worksheet_index": 1,
        "rows": (_candidate_row(),),
        "blocking_issue_codes": (),
        "package_issue_codes": (),
        "canonical_to_source_columns": (
            ("faCode", "faCode"),
            ("identity", "identity"),
            ("name", "name"),
        ),
        "score": (1, 1, 1),
    }
    values.update(overrides)
    return RosterCandidate(**values)


def _workbook_bytes(*, headers=None, rows=(), formula_only_row=False):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Payment roster"
    worksheet.append(headers or ("name", "identity", "faCode"))
    for row in rows:
        worksheet.append(row)
    if formula_only_row:
        worksheet.cell(row=worksheet.max_row + 1, column=3, value="=1+1")
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _inspection_for_snapshots(snapshots, *, unit_numbers=None):
    if unit_numbers is None:
        unit_numbers = tuple(range(1, len(snapshots) + 1))
    sources = tuple(
        InspectionSource(
            evidence_id=f"evidence-{index:04d}",
            detected_type="xlsx",
            inspection_status="inspected",
            unit_count=1,
            issue_codes=(),
        )
        for index in range(1, len(snapshots) + 1)
    )
    units = tuple(
        InspectionUnit(
            unit_id=f"unit-{unit_number:04d}",
            evidence_id=source.evidence_id,
            unit_kind="worksheet",
            unit_index=1,
            suggested_role="payment-roster",
            confidence_band="high",
            needs_user_review=False,
            inspection_method="worksheet-structure",
            signal_codes=("roster-column-pattern", "roster-row-pattern"),
            issue_codes=(),
        )
        for source, unit_number in zip(sources, unit_numbers)
    )
    inspection = InspectionResult(
        inspection_version="1.0",
        inspection_status="complete",
        observation_id="observation-" + "0" * 64,
        totals=InspectionTotals(
            sources=len(sources),
            units=len(units),
            classified=len(units),
            unknown=0,
            needs_user_review=0,
            issues=0,
        ),
        sources=sources,
        units=units,
    )
    by_evidence_id = {
        source.evidence_id: snapshot
        for source, snapshot in zip(sources, snapshots)
    }

    def snapshot_source(evidence_id, *, max_bytes):
        assert max_bytes == 25 * 1024 * 1024
        return by_evidence_id[evidence_id]

    return inspection, snapshot_source


def generated_inspection_with_rosters(tmp_path, *, valid):
    del tmp_path
    snapshots = tuple(
        _workbook_bytes(
            rows=((f"Synthetic Person {index}", f"07912345678{index}", "FA-SYNTHETIC-001"),)
        )
        for index in range(1, valid + 1)
    )
    return _inspection_for_snapshots(snapshots)


@pytest.mark.parametrize(
    "overrides",
    [
        {"row_index": True},
        {"row_index": _HostileValue()},
        {"name": ["private-name"]},
        {"identity": _HostileValue()},
        {"values": [("identity", _PRIVATE[1]), ("name", _PRIVATE[0])]},
        {"values": (("identity", _PRIVATE[1]), ["name", _PRIVATE[0]])},
    ],
    ids=(
        "bool-row-index",
        "hostile-row-index",
        "list-name",
        "hostile-identity",
        "list-values",
        "list-nested-pair",
    ),
)
def test_candidate_rows_require_exact_built_in_types_and_nested_tuples(overrides):
    with pytest.raises((TypeError, ValueError)):
        _candidate_row(**overrides)


@pytest.mark.parametrize(
    "overrides",
    [
        {
            "values": (
                ("faCode", _PRIVATE[2]),
                ("identity", _PRIVATE[1]),
                ("name", "Different Private Name"),
            )
        },
        {
            "values": (
                ("faCode", _PRIVATE[2]),
                ("identity", "Different Private Identity"),
                ("name", _PRIVATE[0]),
            )
        },
        {
            "values": (
                ("identity", _PRIVATE[1]),
                ("name", _PRIVATE[0]),
                ("name", _PRIVATE[0]),
            )
        },
    ],
    ids=("name-mismatch", "identity-mismatch", "duplicate-private-field"),
)
def test_candidate_rows_reject_private_field_mismatches_without_disclosure(
    overrides,
):
    with pytest.raises(ValueError) as captured:
        _candidate_row(**overrides)

    assert all(value not in str(captured.value) for value in _PRIVATE)
    assert "Different Private" not in str(captured.value)


@pytest.mark.parametrize(
    "overrides",
    [
        {"unit_id": _HostileValue()},
        {"evidence_id": ["evidence-0001"]},
        {"worksheet_index": True},
        {"rows": [_candidate_row()]},
        {"rows": (_HostileValue(),)},
        {"blocking_issue_codes": ["roster-row-invalid"]},
        {"package_issue_codes": ["roster-fa-code-blank"]},
        {
            "canonical_to_source_columns": [
                ("faCode", "faCode"),
                ("identity", "identity"),
                ("name", "name"),
            ]
        },
        {
            "canonical_to_source_columns": (
                ("faCode", "faCode"),
                ["identity", "identity"],
                ("name", "name"),
            )
        },
        {"score": [1, 1, 1]},
        {"score": (True, 1, 1)},
    ],
    ids=(
        "hostile-unit-id",
        "list-evidence-id",
        "bool-worksheet-index",
        "list-rows",
        "hostile-row",
        "list-blocking-issues",
        "list-package-issues",
        "list-columns",
        "list-nested-column",
        "list-score",
        "bool-score-bit",
    ),
)
def test_candidates_require_exact_built_in_types_and_nested_tuples(overrides):
    with pytest.raises((TypeError, ValueError)):
        _candidate(**overrides)


@pytest.mark.parametrize(
    "overrides",
    [
        {
            "blocking_issue_codes": ("roster-row-invalid",),
            "score": (1, 1, 1),
        },
        {
            "package_issue_codes": ("roster-fa-code-blank",),
            "score": (1, 1, 1),
        },
        {"score": (1, 1, 0)},
    ],
    ids=("blocking-issue", "package-issue", "usable-row-count"),
)
def test_candidates_reject_scores_that_contradict_rows_or_issues(overrides):
    with pytest.raises(ValueError):
        _candidate(**overrides)


@pytest.mark.parametrize(
    "overrides",
    [
        {"status": _HostileValue()},
        {"roster_unit_id": ["unit-0001"]},
        {"candidate_unit_ids": ["unit-0001"]},
        {"candidate_unit_ids": (_HostileValue(),)},
        {"issue_codes": []},
        {"roster_unit_id": "unit-0002"},
        {"issue_codes": ("roster-ambiguous",)},
    ],
    ids=(
        "hostile-status",
        "list-roster-id",
        "list-candidate-ids",
        "hostile-candidate-id",
        "list-issues",
        "unlisted-selected-id",
        "selected-with-issue",
    ),
)
def test_selections_require_exact_types_and_status_coherence(overrides):
    values = {
        "status": "selected",
        "roster_unit_id": "unit-0001",
        "candidate_unit_ids": ("unit-0001",),
        "issue_codes": (),
    }
    values.update(overrides)

    with pytest.raises((TypeError, ValueError)):
        RosterSelection(**values)


@pytest.mark.parametrize(
    "values",
    [
        {
            "status": "missing",
            "roster_unit_id": None,
            "candidate_unit_ids": ("unit-0001",),
            "issue_codes": ("roster-missing",),
        },
        {
            "status": "ambiguous",
            "roster_unit_id": None,
            "candidate_unit_ids": ("unit-0001",),
            "issue_codes": ("roster-ambiguous",),
        },
        {
            "status": "invalid",
            "roster_unit_id": "unit-0001",
            "candidate_unit_ids": ("unit-0001",),
            "issue_codes": ("roster-invalid",),
        },
    ],
    ids=("missing-with-candidate", "ambiguous-with-one", "invalid-with-selection"),
)
def test_nonselected_statuses_reject_contradictory_selection_fields(values):
    with pytest.raises(ValueError):
        RosterSelection(**values)


def test_one_valid_roster_is_selected_without_user_action(tmp_path):
    inspection, snapshots = generated_inspection_with_rosters(tmp_path, valid=1)
    candidates = load_roster_candidates(inspection, snapshots)
    selection = choose_automatic_roster(candidates)

    assert selection.status == "selected"
    assert selection.roster_unit_id == candidates[0].unit_id
    assert selection.issue_codes == ()


def test_equal_valid_rosters_create_one_roster_exception(tmp_path):
    inspection, snapshots = generated_inspection_with_rosters(tmp_path, valid=2)
    selection = choose_automatic_roster(
        load_roster_candidates(inspection, snapshots)
    )

    assert selection.status == "ambiguous"
    assert selection.roster_unit_id is None
    assert selection.issue_codes == ("roster-ambiguous",)


def test_no_payment_roster_candidate_is_missing():
    inspection, snapshots = _inspection_for_snapshots(())

    selection = choose_automatic_roster(
        load_roster_candidates(inspection, snapshots)
    )

    assert selection.status == "missing"
    assert selection.roster_unit_id is None
    assert selection.candidate_unit_ids == ()
    assert selection.issue_codes == ("roster-missing",)


@pytest.mark.parametrize(
    ("snapshot", "expected_codes"),
    [
        (
            _workbook_bytes(headers=("name", "other"), rows=((_PRIVATE[0], "x"),)),
            ("roster-header-missing", "roster-row-invalid"),
        ),
        (_workbook_bytes(rows=()), ("roster-row-invalid",)),
        (
            _workbook_bytes(
                rows=(
                    (_PRIVATE[0], _PRIVATE[1], _PRIVATE[2]),
                    ("Other Person", _PRIVATE[1], _PRIVATE[2]),
                )
            ),
            ("roster-identity-duplicate",),
        ),
        (
            _workbook_bytes(
                rows=((_PRIVATE[0], _PRIVATE[1], _PRIVATE[2]),),
                formula_only_row=True,
            ),
            ("roster-row-invalid",),
        ),
    ],
    ids=("malformed-header", "zero-rows", "duplicate-identity", "formula-only-row"),
)
def test_blocking_roster_defects_are_ineligible(snapshot, expected_codes):
    inspection, snapshots = _inspection_for_snapshots((snapshot,))

    candidates = load_roster_candidates(inspection, snapshots)
    selection = choose_automatic_roster(candidates)

    assert candidates[0].blocking_issue_codes == expected_codes
    assert selection.status == "invalid"
    assert selection.issue_codes == ("roster-invalid",)


@pytest.mark.parametrize(
    ("headers", "rows", "expected_issue"),
    [
        (
            ("name", "identity"),
            ((_PRIVATE[0], _PRIVATE[1]),),
            "roster-fa-code-missing",
        ),
        (
            ("name", "identity", "faCode"),
            ((_PRIVATE[0], _PRIVATE[1], ""),),
            "roster-fa-code-blank",
        ),
        (
            ("name", "identity", "faCode"),
            (
                (_PRIVATE[0], _PRIVATE[1], _PRIVATE[2]),
                ("Other Person", "079123456782", "FA-SYNTHETIC-002"),
            ),
            "roster-fa-code-conflict",
        ),
    ],
    ids=("missing-column", "blank-value", "conflicting-values"),
)
def test_fa_code_incompleteness_prevents_automatic_selection(
    headers, rows, expected_issue
):
    inspection, snapshots = _inspection_for_snapshots(
        (_workbook_bytes(headers=headers, rows=rows),)
    )

    candidates = load_roster_candidates(inspection, snapshots)
    selection = choose_automatic_roster(candidates)

    assert expected_issue in candidates[0].package_issue_codes
    assert candidates[0].score[1] == 0
    assert selection.status == "invalid"
    assert selection.issue_codes == ("roster-invalid",)


def test_unique_eligible_roster_with_most_rows_wins_in_stable_numeric_order():
    inspection, snapshots = _inspection_for_snapshots(
        (
            _workbook_bytes(rows=(("Person 10", "ID-10", "FA-1"),)),
            _workbook_bytes(
                rows=(("Person 2A", "ID-2A", "FA-1"), ("Person 2B", "ID-2B", "FA-1"))
            ),
        ),
        unit_numbers=(10, 2),
    )

    candidates = load_roster_candidates(inspection, snapshots)
    selection = choose_automatic_roster(candidates)

    assert tuple(candidate.unit_id for candidate in candidates) == (
        "unit-0002",
        "unit-0010",
    )
    assert selection.status == "selected"
    assert selection.roster_unit_id == "unit-0002"
    assert selection.candidate_unit_ids == ("unit-0002", "unit-0010")


def test_row_cap_is_blocking():
    rows = tuple(
        (f"Person {index}", f"ID-{index}", "FA-1")
        for index in range(1, 10_001)
    )
    inspection, snapshots = _inspection_for_snapshots(
        (_workbook_bytes(rows=rows),)
    )

    candidate = load_roster_candidates(inspection, snapshots)[0]

    assert candidate.blocking_issue_codes == ("roster-over-limit",)


def test_cell_cap_is_blocking():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.cell(row=1, column=1, value="x")
    worksheet.cell(row=7, column=15_000, value="x")
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    inspection, snapshots = _inspection_for_snapshots((output.getvalue(),))

    candidate = load_roster_candidates(inspection, snapshots)[0]

    assert candidate.blocking_issue_codes == ("roster-over-limit",)


def test_workbook_cap_and_snapshot_failure_use_one_fixed_private_safe_result():
    inspection, _snapshots = _inspection_for_snapshots((_workbook_bytes(),))
    calls = []

    def failing_snapshot_source(evidence_id, *, max_bytes):
        calls.append((evidence_id, max_bytes))
        raise RuntimeError("private-path-and-cell-value")

    candidate = load_roster_candidates(inspection, failing_snapshot_source)[0]

    assert calls == [("evidence-0001", 25 * 1024 * 1024)]
    assert candidate.blocking_issue_codes == ("roster-unreadable",)
    assert "private-path-and-cell-value" not in repr(candidate)


def test_snapshot_callback_return_over_workbook_cap_is_rejected_before_parsing(
    monkeypatch,
):
    inspection, _snapshots = _inspection_for_snapshots((_workbook_bytes(),))

    monkeypatch.setattr(
        "ctv_proposal_roster.worksheet_nonblank_row_indexes",
        lambda *_args, **_kwargs: pytest.fail(
            "oversized snapshot reached workbook parsing"
        ),
    )

    def oversized_snapshot_source(_evidence_id, *, max_bytes):
        return b"private-cell-value" + b"x" * max_bytes

    candidate = load_roster_candidates(inspection, oversized_snapshot_source)[0]

    assert candidate.blocking_issue_codes == ("roster-over-limit",)
    assert "private-cell-value" not in repr(candidate)


def test_private_roster_values_are_absent_from_candidate_row_and_selection_repr():
    inspection, snapshots = _inspection_for_snapshots(
        (_workbook_bytes(rows=((_PRIVATE[0], _PRIVATE[1], _PRIVATE[2]),)),)
    )

    candidates = load_roster_candidates(inspection, snapshots)
    selection = choose_automatic_roster(candidates)

    assert all(value not in repr(candidates[0]) for value in _PRIVATE)
    assert all(value not in repr(candidates[0].rows[0]) for value in _PRIVATE)
    assert all(value not in repr(selection) for value in _PRIVATE)
