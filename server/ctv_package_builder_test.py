"""TDD coverage for deterministic, bounded prepared-package artifacts."""

from __future__ import annotations

import dataclasses
from datetime import date
import hashlib
from io import BytesIO
import json
import zipfile

import fitz
from openpyxl import Workbook, load_workbook
from PIL import Image
from pydantic import BaseModel, Field
import pytest

from ctv_inspection import inspect_observation
from ctv_inventory import InventoryObservation, open_inventory_observation
from ctv_package_builder import (
    ArtifactReceipt,
    PackageIdentity,
    PackageBuildError,
    RenderedArtifact,
    build_manifest_bytes,
    create_build_plan,
    iter_rendered_artifacts,
    writer_version_string,
)
from ctv_proposal import ProposalState


def _save_workbook(path, sheets):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for title, rows in sheets:
        worksheet = workbook.create_sheet(title)
        for row in rows:
            worksheet.append(row)
    workbook.save(path)
    workbook.close()


def _mark_as_xlsm(path):
    source = BytesIO(path.read_bytes())
    target = BytesIO()
    with zipfile.ZipFile(source, "r") as source_zip:
        with zipfile.ZipFile(target, "w") as target_zip:
            for info in source_zip.infolist():
                content = source_zip.read(info.filename)
                if info.filename == "[Content_Types].xml":
                    content = content.replace(
                        b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
                        b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
                    )
                target_zip.writestr(info, content)
    path.write_bytes(target.getvalue())


def _approved_inputs(
    tmp_path,
    *,
    fa_code="FA-SYNTHETIC-001",
    image_format="PNG",
    image_suffix=".png",
    workbook_suffix=".xlsx",
):
    source = tmp_path / "synthetic-source"
    source.mkdir()
    document = fitz.open()
    for page_number in range(1, 5):
        page = document.new_page()
        page.insert_text((72, 72), f"HOP DONG DICH VU SYNTHETIC PAGE {page_number}\nBEN A\nBEN B\nCHU KY")
    document.save(source / "a-pages.pdf")
    document.close()

    image_bytes = BytesIO()
    with Image.new("RGB", (3, 2), (10, 20, 30)) as image:
        image.save(image_bytes, format=image_format)
    (source / f"b-image{image_suffix}").write_bytes(image_bytes.getvalue())

    evidence_path = source / f"c-evidence{workbook_suffix}"
    _save_workbook(
        evidence_path,
        (
            ("Synthetic support one", (("Reference", "Amount"), ("SYN-1", 10))),
            ("Synthetic support two", (("Reference", "Amount"), ("SYN-2", 20))),
        ),
    )
    evidence_book = load_workbook(evidence_path)
    evidence_book.worksheets[0]["A2"].comment = __import__("openpyxl").comments.Comment(
        "PRIVATE EVIDENCE COMMENT", "Synthetic"
    )
    evidence_book.worksheets[0]["B2"].hyperlink = "https://private.invalid/evidence"
    evidence_book.worksheets[1].sheet_state = "hidden"
    evidence_book.save(evidence_path)
    evidence_book.close()
    if workbook_suffix == ".xlsm":
        _mark_as_xlsm(evidence_path)
    (source / "d-excluded.bin").write_bytes(b"PRIVATE-EXCLUDED-BYTES-079123456789")
    _save_workbook(
        source / "z-roster.xlsx",
        ((
            "Payment roster",
            (
                ("name", "identity", "faCode", "taxId", "birthDate", "bankAccount", "serviceFee", "product", "So tien"),
                ("Synthetic Person 1", "079123456781", fa_code, "TAX-1", date(1990, 1, 1), "BANK-1", "100", "Product 1", "100"),
                ("Synthetic Person 2", "079123456782", fa_code, "TAX-2", "1990-01-02", "BANK-2", "200", "Product 2", "200"),
            ),
        ),),
    )

    context = open_inventory_observation(source)
    observation = context.__enter__()
    inspection = inspect_observation(observation)
    state = ProposalState.from_inspection(observation, inspection)
    roster = next(unit for unit in state.units if unit["suggestedRole"] == "payment-roster")
    state.select_roster({"rosterUnitId": roster["unitId"]})
    pdf_targets = {
        "unit-0001": ("individual", ["participant-0001"]),
        "unit-0002": ("shared", ["participant-0001", "participant-0002"]),
        "unit-0003": ("individual", ["participant-0002"]),
        "unit-0004": ("case", []),
    }
    for unit in state.units:
        if unit["unitId"] == roster["unitId"]:
            role, target = "payment-roster", ("case", [])
        elif unit["unitKind"] == "pdf-page":
            role, target = "service-contract", pdf_targets[unit["unitId"]]
        elif unit["unitKind"] == "image":
            role, target = "identity-front", ("individual", ["participant-0001"])
        else:
            role, target = "other-supporting-evidence", ("case", [])
        decision = "accepted" if unit["suggestedRole"] == role else "reassigned"
        state.set_unit_decision({
            "unitId": unit["unitId"],
            "decision": decision,
            "role": role,
            "target": {"scope": target[0], "participantHandles": target[1]},
        })
    unit_evidence_ids = {unit["evidenceId"] for unit in state.units}
    for source_record in state.sources:
        if source_record["evidenceId"] not in unit_evidence_ids:
            state.set_source_disposition({
                "evidenceId": source_record["evidenceId"],
                "decision": "excluded",
                "reason": "irrelevant",
            })
    digest = state.approval_summary()["proposalDigest"]
    state.approve(digest)
    approved = state.consume_approved_package_snapshot(digest)
    return context, observation, inspection, approved


def _assert_private_free_boundary(error, private_marker, expected_code):
    assert str(error) == expected_code
    assert private_marker not in str(error)
    assert private_marker not in repr(error)
    assert error.__cause__ is None
    assert error.__context__ is None


def test_private_rows_and_rendered_bytes_are_excluded_from_repr():
    from ctv_inspection_workbook import WorksheetValues

    private_marker = "PRIVATE-REPR-079123456789"
    worksheet = WorksheetValues(1, ((private_marker,),))
    artifact = RenderedArtifact(
        artifact_id="artifact-test",
        kind="evidence",
        path="evidence/test.png",
        source_ids=("source-test",),
        content=private_marker.encode(),
    )

    assert private_marker not in repr(worksheet)
    assert private_marker not in repr(artifact)


def test_public_builder_boundaries_translate_contract_diagnostics_without_context(
    tmp_path, monkeypatch
):
    import ctv_package_builder as builder

    private_marker = "PRIVATE-PYDANTIC-079123456789"

    class PrivateFailure(BaseModel):
        value: str = Field(max_length=1)

    class RejectingColumns:
        def __new__(cls, *_args, **_kwargs):
            PrivateFailure(value=private_marker)

        @classmethod
        def model_validate(cls, _value):
            PrivateFailure(value=private_marker)

    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        monkeypatch.setattr(builder, "CanonicalSourceColumnsV2", RejectingColumns)
        with pytest.raises(PackageBuildError) as raised:
            create_build_plan(observation, inspection, approved)
        _assert_private_free_boundary(
            raised.value, private_marker, "package-plan-invalid"
        )
    finally:
        context.__exit__(None, None, None)
    monkeypatch.undo()

    (tmp_path / "render-errors").mkdir()
    context, observation, inspection, approved = _approved_inputs(
        tmp_path / "render-errors"
    )
    try:
        plan = create_build_plan(observation, inspection, approved)
        monkeypatch.setattr(builder, "CanonicalRosterRowV2", RejectingColumns)
        iterator = iter_rendered_artifacts(plan, observation)
        assert next(iterator).path == "input.pdf"
        with pytest.raises(PackageBuildError) as raised:
            next(iterator)
        _assert_private_free_boundary(
            raised.value, private_marker, "package-roster-unavailable"
        )

        monkeypatch.setattr(builder, "PackageManifestV2", RejectingColumns)
        receipts = tuple(
            ArtifactReceipt(
                artifact_id=recipe.artifact_id,
                kind=recipe.kind,
                path=recipe.path,
                source_ids=recipe.source_ids,
                size=0,
                sha256="0" * 64,
            )
            for recipe in plan.recipes
        )
        with pytest.raises(PackageBuildError) as raised:
            build_manifest_bytes(plan, receipts)
        _assert_private_free_boundary(
            raised.value, private_marker, "package-receipt-invalid"
        )
    finally:
        context.__exit__(None, None, None)

    (tmp_path / "long-fa").mkdir()
    context, observation, inspection, approved = _approved_inputs(
        tmp_path / "long-fa", fa_code="FA-" + private_marker * 5
    )
    try:
        with pytest.raises(PackageBuildError) as raised:
            create_build_plan(observation, inspection, approved)
        _assert_private_free_boundary(
            raised.value, private_marker, "package-plan-invalid"
        )
    finally:
        context.__exit__(None, None, None)


def test_package_identity_is_domain_separated_full_length_and_path_independent():
    identity = PackageIdentity.derive(
        observation_id="observation-" + "b" * 64,
        proposal_digest="a" * 64,
        writer_version="ctv-package-writer-test/1",
        schema_version="2.0",
        compatibility_target="ctv-intake-v2",
    )

    assert identity.digest == "c04febde9345540afe67bc389c4fb41f22172f4977b2e413e7ebcb83a4bd6251"
    assert identity.package_id == "package-" + identity.digest
    assert identity.batch_id == "batch-8a67959ab55b86429fa101ff83fd67b1345d604fbc907615d2088a20b049d062"
    assert identity.case_id == "case-51a4a86f3f60b88f59c37f0bde3388b7900fb4b0ac8a97bfffc6a86ce7ea87a6"
    assert identity.final_directory == "ctv-package-c04febde9345540afe67bc38"
    assert len(identity.final_directory.removeprefix("ctv-package-")) == 24
    assert len({identity.package_id, identity.batch_id, identity.case_id}) == 3

    changed_writer = PackageIdentity.derive(
        observation_id="observation-" + "b" * 64,
        proposal_digest="a" * 64,
        writer_version="ctv-package-writer-test/2",
        schema_version="2.0",
        compatibility_target="ctv-intake-v2",
    )
    assert changed_writer != identity

    with pytest.raises(TypeError):
        PackageIdentity.derive(
            observation_id="observation-" + "b" * 64,
            proposal_digest="a" * 64,
            writer_version="ctv-package-writer-test/1",
            schema_version="2.0",
            compatibility_target="ctv-intake-v2",
            source_path="private/source.pdf",
        )


def test_effective_serializer_fingerprints_each_change_writer_and_package_identity(
    monkeypatch
):
    import ctv_package_builder as builder

    private_marker = "PRIVATE-DEPENDENCY-/tmp/079123456789"
    baseline_writer = writer_version_string()
    baseline_identity = PackageIdentity.derive(
        observation_id="observation-" + "b" * 64,
        proposal_digest="a" * 64,
        writer_version=baseline_writer,
        schema_version="2.0",
        compatibility_target="ctv-intake-v2",
    )
    for probe_name in (
        "_mupdf_fingerprint",
        "_lxml_fingerprint",
        "_zlib_fingerprint",
        "_pillow_fingerprint",
    ):
        with monkeypatch.context() as local_patch:
            local_patch.setattr(builder, probe_name, lambda: private_marker)
            changed_writer = writer_version_string()
            changed_identity = PackageIdentity.derive(
                observation_id="observation-" + "b" * 64,
                proposal_digest="a" * 64,
                writer_version=changed_writer,
                schema_version="2.0",
                compatibility_target="ctv-intake-v2",
            )
        assert changed_writer != baseline_writer
        assert changed_identity != baseline_identity
        assert private_marker not in changed_writer
        assert "/tmp/" not in changed_writer


def test_build_plan_fixes_pdf_order_evidence_grouping_and_all_locators_before_render(
    tmp_path, monkeypatch
):
    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        expected_identity = PackageIdentity.derive(
            observation_id=observation.observation_id,
            proposal_digest=approved.proposal_digest,
            writer_version=writer_version_string(),
            schema_version="2.0",
            compatibility_target="ctv-intake-v2",
        )
        monkeypatch.setattr(
            InventoryObservation,
            "snapshot",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("planning acquired bytes")),
        )
        plan = create_build_plan(observation, inspection, approved)

        assert plan.identity == expected_identity
        assert [page.unit_id for page in plan.pdf_pages] == [
            "unit-0004", "unit-0002", "unit-0001", "unit-0003"
        ]
        assert [page.target_page for page in plan.pdf_pages] == [1, 2, 3, 4]
        assert len({page.source_page_key for page in plan.pdf_pages}) == len(plan.pdf_pages)
        assert [recipe.path for recipe in plan.recipes] == [
            "input.pdf",
            "roster.xlsx",
            "evidence/evidence-0001.png",
            "evidence/evidence-0002.xlsx",
            "assignments.json",
            "exceptions.json",
        ]
        workbook_recipe = plan.recipes[3]
        assert workbook_recipe.source_unit_indexes == (1, 2)
        workbook_units = [
            unit for unit in plan.assignments.document.units
            if unit.output_locator.artifact_id == workbook_recipe.artifact_id
        ]
        assert [unit.output_locator.worksheet_index for unit in workbook_units] == [1, 2]
        assert set(plan.locators) == {
            item.unit_id for item in approved.unit_decisions
            if item.decision in {"accepted", "reassigned"}
        }
    finally:
        context.__exit__(None, None, None)


@pytest.mark.parametrize(
    ("image_format", "image_suffix", "expected_image_format"),
    (("JPEG", ".jpg", "JPEG"), ("TIFF", ".tiff", "TIFF")),
)
def test_original_jpeg_and_tiff_sources_keep_conservative_mime_size_hash_and_path(
    tmp_path, image_format, image_suffix, expected_image_format
):
    context, observation, inspection, approved = _approved_inputs(
        tmp_path,
        image_format=image_format,
        image_suffix=image_suffix,
    )
    try:
        plan = create_build_plan(observation, inspection, approved)
        source = next(item for item in plan.sources if item.path.endswith(image_suffix))
        observed = next(
            item for item in observation.sources if item.extension == image_suffix
        )
        original = observation.snapshot(
            observed.evidence_id,
            max_bytes=observed.size,
        )
        with Image.open(BytesIO(original)) as image:
            assert image.format == expected_image_format
        assert source.media_type == "application/octet-stream"
        assert source.path.endswith(image_suffix)
        assert source.size == len(original)
        assert source.sha256 == hashlib.sha256(original).hexdigest()
        rendered = next(
            item for item in iter_rendered_artifacts(plan, observation)
            if item.path.endswith(".png")
        )
        assert rendered.content.startswith(b"\x89PNG\r\n\x1a\n")
    finally:
        context.__exit__(None, None, None)


def test_original_xlsm_source_keeps_macro_mime_size_hash_and_path(tmp_path):
    context, observation, inspection, approved = _approved_inputs(
        tmp_path,
        workbook_suffix=".xlsm",
    )
    try:
        plan = create_build_plan(observation, inspection, approved)
        source = next(item for item in plan.sources if item.path.endswith(".xlsm"))
        observed = next(item for item in observation.sources if item.extension == ".xlsm")
        original = observation.snapshot(observed.evidence_id, max_bytes=observed.size)
        with zipfile.ZipFile(BytesIO(original)) as archive:
            assert b"macroEnabled.main+xml" in archive.read("[Content_Types].xml")
        assert source.media_type == "application/octet-stream"
        assert source.path.endswith(".xlsm")
        assert source.size == len(original)
        assert source.sha256 == hashlib.sha256(original).hexdigest()
        rendered = next(
            item for item in iter_rendered_artifacts(plan, observation)
            if item.path == "evidence/evidence-0002.xlsx"
        )
        assert rendered.content.startswith(b"PK")
    finally:
        context.__exit__(None, None, None)


def _zip_contract(snapshot):
    with zipfile.ZipFile(BytesIO(snapshot)) as archive:
        infos = archive.infolist()
        return [info.filename for info in infos], [
            (info.date_time, info.create_system, info.external_attr >> 16)
            for info in infos
        ]


def test_rendering_produces_exact_canonical_artifacts_and_manifest_twice(tmp_path):
    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        plan = create_build_plan(observation, inspection, approved)
        first = tuple(iter_rendered_artifacts(plan, observation))
        second = tuple(iter_rendered_artifacts(plan, observation))
        assert [(item.path, item.content) for item in first] == [
            (item.path, item.content) for item in second
        ]
        assert [item.path for item in first] == [recipe.path for recipe in plan.recipes]

        by_path = {item.path: item.content for item in first}
        with fitz.open(stream=by_path["input.pdf"], filetype="pdf") as document:
            assert document.page_count == 4
            assert [document[index].get_text().split("PAGE ")[1][0] for index in range(4)] == [
                "4", "2", "1", "3"
            ]
            assert all(not document.metadata.get(key) for key in (
                "author", "title", "subject", "keywords", "creator", "producer",
                "creationDate", "modDate",
            ))
            trailer_id = document.xref_get_key(-1, "ID")[1]
            assert plan.pdf_trailer_id in trailer_id.lower()

        roster = load_workbook(BytesIO(by_path["roster.xlsx"]), data_only=False, keep_links=False)
        try:
            assert roster.sheetnames == ["Roster"]
            rows = list(roster.active.iter_rows(values_only=True))
            assert rows[0] == (
                "Roster Row ID", "Name", "Identity", "FA Code", "Tax ID",
                "Birth Date", "Bank Account", "Service Fee", "Product",
            )
            assert [row[0] for row in rows[1:]] == [
                participant.roster_row_id for participant in plan.assignments.document.participants
            ]
            assert rows[1][5] == "1990-01-01"
            assert not any(
                cell.data_type == "f" for row in roster.active.iter_rows() for cell in row
            )
            assert not roster.defined_names
        finally:
            roster.close()

        evidence = load_workbook(
            BytesIO(by_path["evidence/evidence-0002.xlsx"]),
            data_only=False,
            keep_links=False,
        )
        try:
            assert evidence.sheetnames == ["Worksheet 0001", "Worksheet 0002"]
            assert all(sheet.sheet_state == "visible" for sheet in evidence.worksheets)
            assert list(evidence.worksheets[0].values) == [
                ("Reference", "Amount"), ("SYN-1", 10)
            ]
            assert not any(
                cell.data_type == "f"
                for sheet in evidence.worksheets
                for row in sheet.iter_rows()
                for cell in row
            )
            assert not any(
                cell.comment or cell.hyperlink
                for sheet in evidence.worksheets
                for row in sheet.iter_rows()
                for cell in row
            )
        finally:
            evidence.close()
        for path in ("roster.xlsx", "evidence/evidence-0002.xlsx"):
            names, metadata = _zip_contract(by_path[path])
            assert names == sorted(names)
            assert all(value == ((1980, 1, 1, 0, 0, 0), 3, 0o100644) for value in metadata)
            with zipfile.ZipFile(BytesIO(by_path[path])) as archive:
                core = archive.read("docProps/core.xml")
            assert core.count(b"1980-01-01T00:00:00Z") == 2
            assert b"PRIVATE FIRST" not in by_path[path]
            assert b"PRIVATE SECOND" not in by_path[path]

        with Image.open(BytesIO(by_path["evidence/evidence-0001.png"])) as image:
            assert (image.format, image.mode, image.n_frames) == ("PNG", "RGBA", 1)
        assignments = json.loads(by_path["assignments.json"])
        exceptions = json.loads(by_path["exceptions.json"])
        assert exceptions == {"items": [], "schemaVersion": "2.0"}
        assert assignments["packageId"] == plan.identity.package_id
        for path in ("assignments.json", "exceptions.json"):
            value = by_path[path]
            assert value.endswith(b"\n") and value.count(b"\n") == 1
            assert value == json.dumps(
                json.loads(value), ensure_ascii=False, sort_keys=True, separators=(",", ":")
            ).encode() + b"\n"
        assert all(b"PRIVATE-EXCLUDED-BYTES" not in item.content for item in first)

        receipts = tuple(ArtifactReceipt.from_rendered(item) for item in first)
        manifest_first = build_manifest_bytes(plan, receipts)
        manifest_second = build_manifest_bytes(plan, tuple(ArtifactReceipt.from_rendered(item) for item in second))
        assert manifest_first == manifest_second
        manifest = json.loads(manifest_first)
        assert manifest_first == json.dumps(
            manifest, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        ).encode() + b"\n"
        assert manifest["packageId"] == plan.identity.package_id
        assert [artifact["path"] for artifact in manifest["artifacts"]] == [
            recipe.path for recipe in plan.recipes
        ]
        assert [page["targetPage"] for page in manifest["pdfPages"] if page["targetPage"]] == [
            3, 2, 4, 1
        ]
        assert any(source["bindingStatus"] == "unacquired-exclusion" for source in manifest["sources"])

        tampered = dataclasses.replace(receipts[0], path="different.pdf")
        with pytest.raises(PackageBuildError, match="package-receipt-invalid"):
            build_manifest_bytes(plan, (tampered, *receipts[1:]))
        malformed_hash = dataclasses.replace(receipts[0], sha256="z" * 64)
        with pytest.raises(PackageBuildError, match="package-receipt-invalid"):
            build_manifest_bytes(plan, (malformed_hash, *receipts[1:]))
        oversized = dataclasses.replace(receipts[0], size=256 * 1024 * 1024 + 1)
        with pytest.raises(PackageBuildError, match="package-receipt-invalid"):
            build_manifest_bytes(plan, (oversized, *receipts[1:]))
    finally:
        context.__exit__(None, None, None)


def test_rendering_never_acquires_excluded_source_and_stops_at_resource_ceiling(
    tmp_path, monkeypatch
):
    import ctv_package_builder as builder

    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        plan = create_build_plan(observation, inspection, approved)
        excluded_id = approved.source_dispositions[0].evidence_id
        acquired = []
        original_snapshot = InventoryObservation.snapshot

        def recording_snapshot(self, evidence_id, *, max_bytes):
            acquired.append(evidence_id)
            return original_snapshot(self, evidence_id, max_bytes=max_bytes)

        monkeypatch.setattr(InventoryObservation, "snapshot", recording_snapshot)
        rendered = tuple(iter_rendered_artifacts(plan, observation))
        assert excluded_id not in acquired
        assert set(acquired) <= {recipe.evidence_id for recipe in plan.recipes if recipe.evidence_id} | {
            page.evidence_id for page in plan.pdf_pages
        }

        first_size = rendered[0].size
        monkeypatch.setattr(builder, "MAX_PACKAGE_BYTES", first_size + 1)
        acquired.clear()
        iterator = iter_rendered_artifacts(plan, observation)
        assert next(iterator).path == "input.pdf"
        with pytest.raises(PackageBuildError, match="package-aggregate-over-limit"):
            next(iterator)
        assert all(
            evidence_id not in acquired
            for evidence_id in {
                recipe.evidence_id for recipe in plan.recipes[2:] if recipe.evidence_id
            }
        )

        monkeypatch.setattr(builder, "MAX_PACKAGE_BYTES", 1024 * 1024 * 1024)
        monkeypatch.setattr(builder, "MAX_ROSTER_OR_EVIDENCE_BYTES", 1)
        acquired.clear()
        iterator = iter_rendered_artifacts(plan, observation)
        assert next(iterator).path == "input.pdf"
        with pytest.raises(PackageBuildError, match="package-artifact-over-limit"):
            next(iterator)
    finally:
        context.__exit__(None, None, None)


def test_pdf_and_manifest_abort_on_the_crossing_write_before_more_acquisition(
    tmp_path, monkeypatch
):
    import ctv_package_builder as builder

    sink = builder._CappedBytesIO(4)
    assert sink.write(b"1234") == 4
    with pytest.raises(builder._OutputLimitExceeded):
        sink.write(b"5")
    assert sink.getvalue() == b"1234"

    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        plan = create_build_plan(observation, inspection, approved)
        acquired = []
        original_snapshot = InventoryObservation.snapshot

        def recording_snapshot(self, evidence_id, *, max_bytes):
            acquired.append(evidence_id)
            return original_snapshot(self, evidence_id, max_bytes=max_bytes)

        monkeypatch.setattr(InventoryObservation, "snapshot", recording_snapshot)
        monkeypatch.setattr(builder, "MAX_INPUT_PDF_BYTES", 64)
        with pytest.raises(PackageBuildError, match="package-artifact-over-limit"):
            next(iter_rendered_artifacts(plan, observation))
        assert acquired
        assert set(acquired) == {page.evidence_id for page in plan.pdf_pages}
        assert all(
            recipe.evidence_id not in acquired
            for recipe in plan.recipes[1:]
            if recipe.evidence_id is not None
        )

        receipts = tuple(
            ArtifactReceipt(
                artifact_id=recipe.artifact_id,
                kind=recipe.kind,
                path=recipe.path,
                source_ids=recipe.source_ids,
                size=0,
                sha256="0" * 64,
            )
            for recipe in plan.recipes
        )
        monkeypatch.setattr(builder, "MAX_JSON_BYTES", 32)
        with pytest.raises(PackageBuildError, match="package-artifact-over-limit"):
            build_manifest_bytes(plan, receipts)
    finally:
        context.__exit__(None, None, None)


def test_planning_rejects_zero_pdf_and_page_ceiling_before_acquisition(
    tmp_path, monkeypatch
):
    import ctv_package_builder as builder

    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        decisions_without_pdf = tuple(
            dataclasses.replace(item, decision="excluded", reason="irrelevant", role="", scope="", participant_handles=())
            if item.unit_kind == "pdf-page" else item
            for item in approved.unit_decisions
        )
        no_pdf = dataclasses.replace(approved, unit_decisions=decisions_without_pdf)
        with pytest.raises(ValueError, match="included PDF"):
            create_build_plan(observation, inspection, no_pdf)

        monkeypatch.setattr(builder, "MAX_PACKAGE_PDF_PAGES", 3)
        monkeypatch.setattr(
            InventoryObservation,
            "snapshot",
            lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("ceiling acquired bytes")),
        )
        with pytest.raises(ValueError, match="PDF page ceiling"):
            create_build_plan(observation, inspection, approved)
    finally:
        context.__exit__(None, None, None)


def test_three_complete_builds_in_fresh_directories_are_byte_identical(tmp_path):
    context, observation, inspection, approved = _approved_inputs(tmp_path)
    try:
        plan = create_build_plan(observation, inspection, approved)
        complete_builds = []
        for build_number in range(3):
            output = tmp_path / f"fresh-output-{build_number}"
            output.mkdir()
            rendered = tuple(iter_rendered_artifacts(plan, observation))
            receipts = tuple(ArtifactReceipt.from_rendered(item) for item in rendered)
            manifest = build_manifest_bytes(plan, receipts)
            declared = {"case-manifest.json": manifest}
            for artifact in rendered:
                target = output / artifact.path
                target.parent.mkdir(parents=True, exist_ok=True)
                target.write_bytes(artifact.content)
                declared[artifact.path] = target.read_bytes()
            (output / "case-manifest.json").write_bytes(manifest)
            complete_builds.append(declared)
        assert complete_builds[0] == complete_builds[1] == complete_builds[2]
    finally:
        context.__exit__(None, None, None)
