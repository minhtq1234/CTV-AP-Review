import os
import tempfile

import openpyxl

from test_fixtures.combined_workbook import build, build_july


def test_the_fixture_reproduces_the_active_sheet_trap():
    """The real PUBGm file is saved with CCCD selected. If the fixture did not
    do the same, every test below would pass for the wrong reason."""
    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["CTV", "CCCD", "MST"]
    assert wb.active.title == "CCCD"          # NOT the bảng kê


def test_the_fixture_has_a_merged_image_header():
    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    ws = openpyxl.load_workbook(path)["CCCD"]
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert "D1:E1" in merged
    assert ws["D1"].value == "Hình CCCD"


def test_the_july_fixture_is_single_sheet():
    path = build_july(os.path.join(tempfile.mkdtemp(), "roster.xlsx"))
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Thông tin CK"]


import openpyxl
from workbook_layout import score_roster_sheet, select_roster_sheet


def _rows(path, sheet):
    ws = openpyxl.load_workbook(path, data_only=True)[sheet]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def test_the_ctv_sheet_outscores_its_neighbours():
    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    scores = {name: score_roster_sheet(_rows(path, name)) for name in ("CTV", "CCCD", "MST")}
    assert scores["CTV"] > scores["CCCD"]
    assert scores["CTV"] > scores["MST"]


def test_select_picks_ctv_even_though_cccd_is_the_active_sheet():
    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    wb = openpyxl.load_workbook(path, data_only=True)
    sheets = {name: _rows(path, name) for name in wb.sheetnames}
    assert wb.active.title == "CCCD"            # the trap
    assert select_roster_sheet(sheets) == "CTV"  # what we want instead


def test_select_on_a_single_sheet_workbook_picks_that_sheet():
    path = build_july(os.path.join(tempfile.mkdtemp(), "roster.xlsx"))
    sheets = {"Thông tin CK": _rows(path, "Thông tin CK")}
    assert select_roster_sheet(sheets) == "Thông tin CK"


def test_select_returns_none_when_no_sheet_looks_like_a_roster():
    assert select_roster_sheet({"Sheet1": [["hello"], ["world"]]}) is None


from roster_workbook import load_roster_rows


def test_load_roster_rows_reads_the_ctv_sheet_not_the_active_one():
    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    with open(path, "rb") as handle:
        rows = load_roster_rows(handle)
    flat = [str(c) for row in rows for c in row if c is not None]
    # The bảng kê carries MST and a bank name; the CCCD sheet carries neither.
    assert any("MST" in s for s in flat), "read a sheet with no MST column — probably CCCD"
    assert any("Ngân hàng" in s for s in flat)


def test_load_roster_rows_still_reads_a_single_sheet_workbook():
    path = build_july(os.path.join(tempfile.mkdtemp(), "roster.xlsx"))
    with open(path, "rb") as handle:
        rows = load_roster_rows(handle)
    assert any("Họ và tên" in str(c) for row in rows for c in row if c is not None)


import openpyxl as _openpyxl
import pytest

from roster_workbook import RosterWorkbookError, preflight_roster_workbook
from workbook_layout import missing_required_columns


def _workbook_with_only(path, headers):
    wb = _openpyxl.Workbook()
    ws = wb.active
    for col, text in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=text)
    ws.cell(row=2, column=1, value=1)
    wb.save(path)
    return path


def test_missing_required_columns_names_what_is_absent():
    rows = [["STT", "Họ tên", "Số CCCD"], [1, "NGUYEN VAN MOT", "001100000001"]]
    assert missing_required_columns(rows) == ["money"]
    rows_no_name = [["STT", "Số CCCD"], [1, "001100000001"]]
    assert "name" in missing_required_columns(rows_no_name)


def test_preflight_refuses_a_workbook_with_no_usable_roster_sheet():
    path = _workbook_with_only(os.path.join(tempfile.mkdtemp(), "bad.xlsx"),
                               ["STT", "Họ tên", "Số CCCD"])
    with open(path, "rb") as handle:
        with pytest.raises(RosterWorkbookError) as caught:
            preflight_roster_workbook(handle)
    assert "roster" in str(caught.value)


def test_preflight_accepts_both_real_templates():
    for builder in (build, build_july):
        path = builder(os.path.join(tempfile.mkdtemp(), "ok.xlsx"))
        with open(path, "rb") as handle:
            preflight_roster_workbook(handle)     # must not raise


def test_the_upload_response_names_the_sheet_read_and_what_was_missing(tmp_path, monkeypatch):
    """Task 4's whole point: the reviewer must be able to tell a wrong-sheet
    read from a legitimately empty column, and before a full processing run."""
    import app as appmod
    from fastapi.testclient import TestClient

    monkeypatch.setattr(appmod, "store", appmod.CaseStore(str(tmp_path)))
    path = _workbook_with_only(os.path.join(tempfile.mkdtemp(), "bad.xlsx"),
                               ["STT", "Họ tên", "Số CCCD"])
    with open(path, "rb") as handle:
        payload = handle.read()

    response = TestClient(appmod.app).post(
        "/api/cases",
        files={
            "pdf": ("packet.pdf", b"%PDF-1.4 synthetic", "application/pdf"),
            "roster": (
                "roster.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )

    assert response.status_code == 422
    detail = response.json()["detail"]
    assert detail["code"] == "invalid-roster-workbook"
    assert "money" in detail["reason"]
    assert "Sheet" in detail["reason"]        # names the sheet it actually read


from workbook_layout import classify_image_columns


def test_a_merged_header_covers_both_card_columns():
    # "Hình CCCD" merged across D:E -- one label, two columns, front and back.
    header = {3: "Hình CCCD", 4: "Hình CCCD", 5: "STK", 6: "Hình Ảnh"}
    kinds = classify_image_columns(header, sheet_name="CCCD")
    assert kinds[3] == "card"
    assert kinds[4] == "card"


def test_an_image_column_beside_stk_is_a_bank_screenshot():
    header = {3: "Hình CCCD", 4: "Hình CCCD", 5: "STK", 6: "Hình Ảnh"}
    kinds = classify_image_columns(header, sheet_name="CCCD")
    assert kinds[6] == "bank"


def test_an_image_column_on_an_mst_sheet_is_a_tax_screenshot():
    header = {0: "STT", 1: "Họ tên", 2: "MST", 3: "Hình Ảnh"}
    kinds = classify_image_columns(header, sheet_name="MST")
    assert kinds[3] == "tax"


def test_a_sheet_with_no_image_headers_classifies_nothing():
    assert classify_image_columns({0: "STT", 1: "Họ và tên"}, sheet_name="Thông tin CK") == {}


def test_extraction_labels_the_fixture_images_by_column():
    """End-to-end over a real .xlsx: the header reader, the merge expansion and
    the classifier together must label D/E as card sides, G as bank and the MST
    sheet's D as tax. The unit tests above hand-build kinds; this one earns them."""
    from cccd_workbook import extract_drawings

    directory = tempfile.mkdtemp()
    path = build(os.path.join(directory, "combined.xlsx"))
    result = extract_drawings(path, os.path.join(directory, "out"))

    by_sheet_col = {
        (d.anchor.sheet, d.anchor.from_col): d.kind for d in result.drawings
    }
    assert by_sheet_col[("CCCD", 3)] == "card"      # D, merged header
    assert by_sheet_col[("CCCD", 4)] == "card"      # E, same merged header
    assert by_sheet_col[("CCCD", 6)] == "bank"      # G, beside STK
    assert by_sheet_col[("MST", 3)] == "tax"        # D on the MST sheet


def test_a_sheet_without_image_headers_leaves_every_kind_none():
    """The July template's path: no image headers, so nothing is labelled and
    pairing keeps its original proximity-only behaviour."""
    from cccd_workbook import extract_drawings

    directory = tempfile.mkdtemp()
    path = build_july(os.path.join(directory, "roster.xlsx"))
    result = extract_drawings(path, os.path.join(directory, "out"))

    assert all(d.kind is None for d in result.drawings)


def test_one_combined_workbook_serves_as_both_roster_and_cards(tmp_path, monkeypatch):
    """Task 7: the combined template is a single file that is both the bảng kê
    and the card source. Uploaded in both fields it must be stored once, and the
    pipeline must receive the same path for each."""
    import app as appmod
    from fastapi.testclient import TestClient

    seen = {}

    def fake_pipeline(pdf, roster, out_dir, cb, cccd_xlsx_path=None):
        seen["roster"] = roster
        seen["cccd"] = cccd_xlsx_path
        cb("done", 1, 1, "")
        return {"summary": {"found": 0, "rosterN": 0, "autoMerged": 0}, "packets": []}

    monkeypatch.setattr(appmod, "run_pipeline", fake_pipeline)
    monkeypatch.setattr(appmod, "store", appmod.CaseStore(str(tmp_path)))

    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    with open(path, "rb") as handle:
        payload = handle.read()
    sheet_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    client = TestClient(appmod.app)
    response = client.post("/api/cases", files={
        "pdf": ("packet.pdf", b"%PDF-1.4 synthetic", "application/pdf"),
        "roster": ("combined.xlsx", payload, sheet_type),
        "cccd": ("combined.xlsx", payload, sheet_type),
    })
    assert response.status_code == 200

    cid = response.json()["case_id"]
    for _ in range(200):
        if client.get(f"/api/cases/{cid}").json()["status"] != "processing":
            break

    # Stored once, and both roles point at that one file.
    assert seen["roster"] == seen["cccd"]
    assert os.path.isfile(seen["roster"])
    case_dir = appmod.store.case_dir(cid)
    assert not os.path.exists(os.path.join(case_dir, "cccd.xlsx"))


def test_two_different_workbooks_are_still_stored_separately(tmp_path, monkeypatch):
    """The July pair must keep its two files; dedupe applies only when the same
    bytes arrive in both fields."""
    import app as appmod
    from fastapi.testclient import TestClient

    seen = {}

    def fake_pipeline(pdf, roster, out_dir, cb, cccd_xlsx_path=None):
        seen["roster"] = roster
        seen["cccd"] = cccd_xlsx_path
        cb("done", 1, 1, "")
        return {"summary": {"found": 0, "rosterN": 0, "autoMerged": 0}, "packets": []}

    monkeypatch.setattr(appmod, "run_pipeline", fake_pipeline)
    monkeypatch.setattr(appmod, "store", appmod.CaseStore(str(tmp_path)))

    directory = tempfile.mkdtemp()
    with open(build_july(os.path.join(directory, "roster.xlsx")), "rb") as handle:
        roster_payload = handle.read()
    with open(build(os.path.join(directory, "cards.xlsx")), "rb") as handle:
        cards_payload = handle.read()
    sheet_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    client = TestClient(appmod.app)
    response = client.post("/api/cases", files={
        "pdf": ("packet.pdf", b"%PDF-1.4 synthetic", "application/pdf"),
        "roster": ("roster.xlsx", roster_payload, sheet_type),
        "cccd": ("cards.xlsx", cards_payload, sheet_type),
    })
    assert response.status_code == 200

    cid = response.json()["case_id"]
    for _ in range(200):
        if client.get(f"/api/cases/{cid}").json()["status"] != "processing":
            break

    assert seen["roster"] != seen["cccd"]
    assert os.path.isfile(seen["cccd"])


from workbook_layout import column_letter


def test_column_letters_match_what_excel_shows():
    assert column_letter(0) == "A"
    assert column_letter(3) == "D"
    assert column_letter(6) == "G"
    assert column_letter(25) == "Z"
    assert column_letter(26) == "AA"
    assert column_letter(27) == "AB"


def test_describing_a_workbook_names_every_image_column_and_its_kind():
    """The declaration the reviewer confirms before a full run. Counts come from
    anchors alone -- no image is decoded or written."""
    from cccd_workbook import describe_image_columns

    directory = tempfile.mkdtemp()
    path = build(os.path.join(directory, "combined.xlsx"))
    described = describe_image_columns(path)

    by_column = {(d["sheet"], d["column"]): d for d in described}
    assert by_column[("CCCD", "D")] == {
        "sheet": "CCCD", "column": "D", "kind": "card", "count": 3,
    }
    assert by_column[("CCCD", "E")]["kind"] == "card"
    assert by_column[("CCCD", "G")]["kind"] == "bank"
    assert by_column[("MST", "D")]["kind"] == "tax"
    # Nothing was written: describing is a preview, not an extraction.
    assert os.listdir(directory) == ["combined.xlsx"]


def test_describing_reports_an_unrecognised_image_column_rather_than_hiding_it():
    """A column the header did not explain is what the reviewer most needs to
    see -- silence there is the failure mode this whole declaration exists for."""
    from cccd_workbook import describe_image_columns

    directory = tempfile.mkdtemp()
    path = build(os.path.join(directory, "combined.xlsx"))
    # Strip the headers so nothing can be classified.
    workbook = openpyxl.load_workbook(path)
    sheet = workbook["CCCD"]
    # Passing value=None to openpyxl cell() is a no-op -- it assigns only when
    # the value is not None -- and a MergedCell refuses assignment outright, so
    # the range has to come apart first.
    sheet.unmerge_cells("D1:E1")
    for column in range(1, 8):
        sheet.cell(row=1, column=column).value = None
    stripped = os.path.join(directory, "stripped.xlsx")
    workbook.save(stripped)

    described = describe_image_columns(stripped)
    cccd_columns = [d for d in described if d["sheet"] == "CCCD"]
    assert cccd_columns, "the images are still there and must still be reported"
    assert all(d["kind"] is None for d in cccd_columns)


def test_the_inspect_endpoint_declares_what_was_inferred(tmp_path, monkeypatch):
    """ver3-scope §1: inference has to be shown, because being confidently wrong
    and silent is the failure this whole feature exists to prevent."""
    import app as appmod
    from fastapi.testclient import TestClient

    monkeypatch.setattr(appmod, "store", appmod.CaseStore(str(tmp_path)))
    path = build(os.path.join(tempfile.mkdtemp(), "combined.xlsx"))
    with open(path, "rb") as handle:
        payload = handle.read()
    sheet_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    response = TestClient(appmod.app).post("/api/uploads/inspect", files={
        "roster": ("combined.xlsx", payload, sheet_type),
        "cccd": ("combined.xlsx", payload, sheet_type),
    })

    assert response.status_code == 200
    body = response.json()
    # It names the sheet it chose -- the whole point, since `active` is CCCD.
    assert body["rosterSheet"] == "CTV"
    assert body["people"] == 3
    assert "gross" in body["columns"]
    # The same file in both fields is walked once, not twice.
    cards = [i for i in body["images"] if i["kind"] == "card"]
    assert sorted(c["column"] for c in cards) == ["D", "E"]
    assert all(c["count"] == 3 for c in cards)
    assert [i["column"] for i in body["images"] if i["kind"] == "bank"] == ["G"]
    assert [i["sheet"] for i in body["images"] if i["kind"] == "tax"] == ["MST"]


def test_the_inspect_endpoint_refuses_the_same_workbooks_the_upload_does(tmp_path, monkeypatch):
    import app as appmod
    from fastapi.testclient import TestClient

    monkeypatch.setattr(appmod, "store", appmod.CaseStore(str(tmp_path)))
    path = _workbook_with_only(os.path.join(tempfile.mkdtemp(), "bad.xlsx"),
                               ["STT", "Họ tên", "Số CCCD"])
    with open(path, "rb") as handle:
        payload = handle.read()

    response = TestClient(appmod.app).post("/api/uploads/inspect", files={
        "roster": ("roster.xlsx", payload,
                   "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    })

    assert response.status_code == 422
    assert "money" in response.json()["detail"]["reason"]
