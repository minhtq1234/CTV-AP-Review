"""Generated end-to-end acceptance for the local CTV v2 package workflow."""

from __future__ import annotations

from datetime import datetime
from hashlib import sha256
import http.client
from io import BytesIO
import json
import os
from pathlib import Path
import stat
from urllib.parse import urlsplit

import fitz
from openpyxl import Workbook, load_workbook
from PIL import Image

from ctv_inspection_workbook import _canonical_package_workbook_bytes
from ctv_intake_cli import main as intake_main
from ctv_proposal_review import run_local_review
from intake_contract_v2 import MAX_PACKAGE_BYTES
from intake_package_validator import _PackageReader
from validate_intake_package import main as validate_main


_SYNTHETIC_NAME_1 = "Synthetic Acceptance Person 0001"
_SYNTHETIC_NAME_2 = "Synthetic Acceptance Person 0002"
_SYNTHETIC_ID_1 = "SYNTHETIC-ID-0001"
_SYNTHETIC_ID_2 = "SYNTHETIC-ID-0002"
_SYNTHETIC_FA = "FA-SYNTHETIC-ACCEPTANCE"
_INCLUDED_WORKSHEET_MARKER = "SYNTHETIC-INCLUDED-WORKSHEET"
_EXCLUDED_WORKSHEET_MARKER = "SYNTHETIC-EXCLUDED-WORKSHEET"
_UNSUPPORTED_MARKER = b"SYNTHETIC-UNSUPPORTED-SOURCE-ONLY"
_EXPECTED_PACKAGE_FILES = (
    "assignments.json",
    "case-manifest.json",
    "evidence/evidence-0001.png",
    "evidence/evidence-0002.xlsx",
    "exceptions.json",
    "input.pdf",
    "roster.xlsx",
    "validation-report.json",
)


def _workbook_bytes(sheets) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    fixed = datetime(1980, 1, 1)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    for title, rows in sheets:
        worksheet = workbook.create_sheet(title)
        for row in rows:
            worksheet.append(row)
    return _canonical_package_workbook_bytes(
        workbook, max_bytes=25 * 1024 * 1024
    )


def _write_generated_source(source: Path) -> None:
    source.mkdir(mode=0o700)

    document = fitz.open()
    stream = BytesIO()
    try:
        for page_number in (1, 2):
            page = document.new_page()
            page.insert_text(
                (72, 72),
                "HOP DONG DICH VU SYNTHETIC ACCEPTANCE "
                f"PAGE {page_number}\nBEN A\nBEN B\nCHU KY",
            )
        document.set_metadata({})
        fixed_id = "abcdef0123456789abcdef0123456789"
        document.xref_set_key(-1, "ID", f"[<{fixed_id}><{fixed_id}>]")
        document.save(stream, no_new_id=1)
    finally:
        document.close()
    (source / "a-pages.pdf").write_bytes(stream.getvalue())

    image_stream = BytesIO()
    with Image.new("RGB", (5, 3), (17, 34, 51)) as image:
        image.save(image_stream, format="PNG", compress_level=9, optimize=False)
    (source / "b-image.png").write_bytes(image_stream.getvalue())

    (source / "c-evidence.xlsx").write_bytes(
        _workbook_bytes(
            (
                (
                    "Synthetic included",
                    (("Reference", "Amount"), (_INCLUDED_WORKSHEET_MARKER, 10)),
                ),
                (
                    "Synthetic excluded",
                    (("Reference", "Amount"), (_EXCLUDED_WORKSHEET_MARKER, 20)),
                ),
            )
        )
    )
    (source / "d-unsupported.bin").write_bytes(_UNSUPPORTED_MARKER)
    (source / "z-roster.xlsx").write_bytes(
        _workbook_bytes(
            ((
                "Synthetic payment roster",
                (
                    (
                        "name",
                        "identity",
                        "faCode",
                        "taxId",
                        "birthDate",
                        "bankAccount",
                        "serviceFee",
                        "product",
                        "So tien",
                    ),
                    (
                        _SYNTHETIC_NAME_1,
                        _SYNTHETIC_ID_1,
                        _SYNTHETIC_FA,
                        "SYNTHETIC-TAX-0001",
                        "1990-01-01",
                        "SYNTHETIC-BANK-0001",
                        "100",
                        "Synthetic Product 0001",
                        "100",
                    ),
                    (
                        _SYNTHETIC_NAME_2,
                        _SYNTHETIC_ID_2,
                        _SYNTHETIC_FA,
                        "SYNTHETIC-TAX-0002",
                        "1990-01-02",
                        "SYNTHETIC-BANK-0002",
                        "200",
                        "Synthetic Product 0002",
                        "200",
                    ),
                ),
            ),)
        )
    )


def _request(parsed, method, target, *, cookie=None, csrf=None, body=None):
    connection = http.client.HTTPConnection(parsed.hostname, parsed.port, timeout=5)
    headers = {}
    if cookie is not None:
        headers["Cookie"] = cookie
    if csrf is not None:
        headers["X-CSRF-Token"] = csrf
    encoded = body
    if type(body) is dict:
        encoded = json.dumps(body, separators=(",", ":")).encode("utf-8")
        headers["Content-Type"] = "application/json"
        headers["Origin"] = f"http://127.0.0.1:{parsed.port}"
    connection.request(method, target, body=encoded, headers=headers)
    response = connection.getresponse()
    payload = response.read()
    result = response.status, response.getheaders(), payload
    connection.close()
    return result


def _one_header(headers, name: str) -> str:
    values = [value for key, value in headers if key.lower() == name.lower()]
    assert len(values) == 1
    return values[0]


def _json(payload: bytes) -> dict:
    value = json.loads(payload.decode("utf-8"))
    assert type(value) is dict
    return value


def _post(parsed, route: str, cookie: str, csrf: str, body: dict) -> dict:
    status, _headers, payload = _request(
        parsed, "POST", route, cookie=cookie, csrf=csrf, body=body
    )
    assert status == 200, (route, status, payload)
    return _json(payload)


def _decision(unit: dict, role: str, scope: str, handles: list[str]) -> dict:
    return {
        "unitId": unit["unitId"],
        "decision": (
            "accepted" if unit["suggestedRole"] == role else "reassigned"
        ),
        "role": role,
        "target": {"scope": scope, "participantHandles": handles},
    }


def _http_review_driver(state):
    def drive(browser_url: str) -> bool:
        parsed = urlsplit(browser_url)
        assert parsed.scheme == "http"
        assert parsed.hostname == "127.0.0.1"
        assert parsed.path == "/bootstrap"
        status, headers, payload = _request(
            parsed, "GET", f"{parsed.path}?{parsed.query}"
        )
        assert status == 303
        assert payload == b""
        cookie = _one_header(headers, "Set-Cookie").split(";", 1)[0]

        for route in ("/", "/review.css", "/review.js"):
            status, _headers, payload = _request(
                parsed, "GET", route, cookie=cookie
            )
            assert status == 200
            assert payload

        status, _headers, payload = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        current = _json(payload)
        csrf = current["csrfToken"]
        roster = next(
            unit
            for unit in current["units"]
            if unit["suggestedRole"] == "payment-roster"
        )
        current = _post(
            parsed,
            "/api/roster",
            cookie,
            csrf,
            {"rosterUnitId": roster["unitId"]},
        )
        handles = [
            participant["participantHandle"]
            for participant in current["participants"]
        ]
        assert handles == ["participant-0001", "participant-0002"]

        pdf_units = sorted(
            (unit for unit in current["units"] if unit["unitKind"] == "pdf-page"),
            key=lambda unit: unit["unitId"],
        )
        image = next(
            unit for unit in current["units"] if unit["unitKind"] == "image"
        )
        evidence_worksheets = sorted(
            (
                unit
                for unit in current["units"]
                if unit["unitKind"] == "worksheet"
                and unit["unitId"] != roster["unitId"]
            ),
            key=lambda unit: unit["unitId"],
        )
        assert len(pdf_units) == 2
        assert len(evidence_worksheets) == 2

        decisions = (
            _decision(roster, "payment-roster", "case", []),
            _decision(
                pdf_units[0],
                "service-contract",
                "individual",
                [handles[0]],
            ),
            _decision(
                pdf_units[1],
                "service-contract",
                "shared",
                handles,
            ),
            _decision(image, "identity-front", "individual", [handles[0]]),
            _decision(
                evidence_worksheets[0],
                "other-supporting-evidence",
                "case",
                [],
            ),
            {
                "unitId": evidence_worksheets[1]["unitId"],
                "decision": "excluded",
                "reason": "irrelevant",
            },
        )
        for decision in decisions:
            current = _post(
                parsed, "/api/unit", cookie, csrf, decision
            )

        unit_evidence_ids = {
            unit["evidenceId"] for unit in current["units"]
        }
        source_only = [
            source
            for source in current["sources"]
            if source["evidenceId"] not in unit_evidence_ids
        ]
        assert len(source_only) == 1
        _post(
            parsed,
            "/api/source",
            cookie,
            csrf,
            {
                "evidenceId": source_only[0]["evidenceId"],
                "decision": "excluded",
                "reason": "irrelevant",
            },
        )
        summary = _post(parsed, "/api/summary", cookie, csrf, {})
        assert summary["readyToPrepare"] is True
        approved = _post(
            parsed,
            "/api/approve",
            cookie,
            csrf,
            {"expectedProposalDigest": summary["proposalDigest"]},
        )
        assert approved["outcome"] == "approved"
        return True

    return run_local_review(state, browser_open=drive)


def _filesystem_state(root: Path) -> tuple[tuple[object, ...], ...]:
    paths = (root, *sorted(root.rglob("*"), key=lambda item: item.as_posix()))
    entries = []
    for path in paths:
        metadata = path.lstat()
        relative = "." if path == root else path.relative_to(root).as_posix()
        content_digest = (
            sha256(path.read_bytes()).hexdigest()
            if stat.S_ISREG(metadata.st_mode)
            else None
        )
        entries.append(
            (
                relative,
                stat.S_IFMT(metadata.st_mode),
                stat.S_IMODE(metadata.st_mode),
                metadata.st_size,
                metadata.st_mtime_ns,
                content_digest,
            )
        )
    return tuple(entries)


def _canonical_cli_bytes(value: dict) -> bytes:
    return (
        json.dumps(value, ensure_ascii=False, sort_keys=True, indent=2).encode(
            "utf-8"
        )
        + b"\n"
    )


def _published_tree_sha256(package: Path) -> str:
    reader, failure = _PackageReader.open(package)
    assert reader is not None and failure is None
    try:
        limits = {path: 256 * 1024 * 1024 for path in _EXPECTED_PACKAGE_FILES}
        tree, failure = reader.snapshot_tree(
            set(_EXPECTED_PACKAGE_FILES),
            max_bytes_by_path=limits,
            max_total_bytes=MAX_PACKAGE_BYTES,
        )
        assert tree is not None and failure is None
        return tree.tree_sha256
    finally:
        reader.close()


def test_generated_cli_review_writer_validator_and_collision_acceptance(
    tmp_path, capsys
):
    source = tmp_path / "source"
    output = tmp_path / "output"
    _write_generated_source(source)
    output.mkdir(mode=0o700)
    source_before = _filesystem_state(source)

    argv = [
        "package",
        "prepare",
        "--source-root",
        str(source),
        "--output-root",
        str(output),
        "--json",
    ]
    assert intake_main(argv, package_review_driver=_http_review_driver) == 0
    captured = capsys.readouterr()
    assert captured.err == ""
    envelope = json.loads(captured.out)
    assert captured.out.encode("utf-8") == _canonical_cli_bytes(envelope)
    for private_value in (
        _SYNTHETIC_NAME_1,
        _SYNTHETIC_NAME_2,
        _SYNTHETIC_ID_1,
        _SYNTHETIC_ID_2,
        _SYNTHETIC_FA,
        "SYNTHETIC-BANK-0001",
        "SYNTHETIC-BANK-0002",
        str(source),
        str(output),
        "a-pages.pdf",
        "b-image.png",
        "c-evidence.xlsx",
        "d-unsupported.bin",
        "z-roster.xlsx",
    ):
        assert private_value not in captured.out
    assert set(envelope) == {
        "schemaVersion",
        "operation",
        "status",
        "summary",
        "result",
        "errors",
        "retryable",
    }
    assert envelope["schemaVersion"] == "1.0"
    assert envelope["operation"] == "package.prepare"
    assert envelope["status"] == "succeeded"
    assert envelope["errors"] == []
    assert envelope["retryable"] is False

    result = envelope["result"]
    assert set(result) == {
        "version",
        "outcome",
        "packageId",
        "packageDirectoryName",
        "manifestSha256",
        "declaredArtifactSetSha256",
        "publishedTreeSha256",
        "contractVersion",
        "counts",
        "validation",
        "readyForCtvReview",
    }
    assert result["version"] == "1.0"
    assert result["outcome"] == "prepared"
    assert result["contractVersion"] == "2.0"
    assert result["readyForCtvReview"] is True
    assert result["counts"] == {
        "sources": 5,
        "participants": 2,
        "pdfPages": 2,
        "evidenceArtifacts": 2,
        "assignments": 5,
        "exclusions": 2,
    }
    assert result["validation"] == {
        "outcome": "valid",
        "checkCodes": [
            "manifest-valid",
            "assignments-valid",
            "source-verification-complete",
            "validation-report-consistent",
        ],
        "warningCodes": [],
    }
    assert len(result["packageId"].removeprefix("package-")) == 64
    assert result["packageDirectoryName"] == (
        "ctv-package-" + result["packageId"].removeprefix("package-")[:24]
    )
    assert all(
        type(result[name]) is str and len(result[name]) == 64
        for name in (
            "manifestSha256",
            "declaredArtifactSetSha256",
            "publishedTreeSha256",
        )
    )

    package = output / result["packageDirectoryName"]
    assert [path.name for path in output.iterdir()] == [package.name]
    assert not list(output.glob(".ctv-staging-*"))
    assert stat.S_IMODE(package.stat().st_mode) == 0o700
    assert stat.S_IMODE((package / "evidence").stat().st_mode) == 0o700
    assert tuple(
        sorted(
            path.relative_to(package).as_posix()
            for path in package.rglob("*")
            if path.is_file()
        )
    ) == _EXPECTED_PACKAGE_FILES
    assert all(
        stat.S_IMODE(path.stat().st_mode) == 0o600
        and path.stat().st_nlink == 1
        for path in package.rglob("*")
        if path.is_file()
    )

    with fitz.open(package / "input.pdf") as document:
        assert document.page_count == 2
        page_text = [document[index].get_text() for index in range(2)]
    assert "PAGE 2" in page_text[0]
    assert "PAGE 1" in page_text[1]

    roster = load_workbook(package / "roster.xlsx", data_only=False, keep_links=False)
    try:
        assert roster.sheetnames == ["Roster"]
        rows = list(roster.active.iter_rows(values_only=True))
        assert rows[0] == (
            "Roster Row ID",
            "Name",
            "Identity",
            "FA Code",
            "Tax ID",
            "Birth Date",
            "Bank Account",
            "Service Fee",
            "Product",
        )
        roster_row_ids = [row[0] for row in rows[1:]]
        assert len(roster_row_ids) == 2
        assert len(set(roster_row_ids)) == 2
        assert all(
            type(row_id) is str
            and row_id.startswith("roster-row-")
            and len(row_id.removeprefix("roster-row-")) == 32
            and all(
                character in "0123456789abcdef"
                for character in row_id.removeprefix("roster-row-")
            )
            for row_id in roster_row_ids
        )
        assert [row[1] for row in rows[1:]] == [
            _SYNTHETIC_NAME_1,
            _SYNTHETIC_NAME_2,
        ]
        assert [row[2] for row in rows[1:]] == [
            _SYNTHETIC_ID_1,
            _SYNTHETIC_ID_2,
        ]
        assert [row[3] for row in rows[1:]] == [
            _SYNTHETIC_FA,
            _SYNTHETIC_FA,
        ]
        assert not any(
            cell.data_type == "f" or cell.comment or cell.hyperlink
            for row in roster.active.iter_rows()
            for cell in row
        )
        assert not roster.defined_names
        assert all(sheet.sheet_state == "visible" for sheet in roster.worksheets)
        assert not getattr(roster, "_external_links", [])
        assert not getattr(roster.active, "_images", [])
        assert not any(
            dimension.hidden for dimension in roster.active.row_dimensions.values()
        )
        assert not any(
            dimension.hidden
            for dimension in roster.active.column_dimensions.values()
        )
    finally:
        roster.close()

    evidence = load_workbook(
        package / "evidence/evidence-0002.xlsx",
        data_only=False,
        keep_links=False,
    )
    try:
        assert evidence.sheetnames == ["Worksheet 0001"]
        values = list(evidence.active.values)
        assert values == [
            ("Reference", "Amount"),
            (_INCLUDED_WORKSHEET_MARKER, 10),
        ]
        assert _EXCLUDED_WORKSHEET_MARKER not in repr(values)
        assert not any(
            cell.data_type == "f" or cell.comment or cell.hyperlink
            for row in evidence.active.iter_rows()
            for cell in row
        )
        assert not evidence.defined_names
        assert not getattr(evidence, "_external_links", [])
        assert not getattr(evidence.active, "_images", [])
        assert not any(
            dimension.hidden for dimension in evidence.active.row_dimensions.values()
        )
        assert not any(
            dimension.hidden
            for dimension in evidence.active.column_dimensions.values()
        )
    finally:
        evidence.close()

    with Image.open(package / "evidence/evidence-0001.png") as image:
        assert (image.format, image.mode, image.n_frames, image.size) == (
            "PNG",
            "RGBA",
            1,
            (5, 3),
        )
        assert image.info == {}

    manifest = json.loads((package / "case-manifest.json").read_bytes())
    assignments_bytes = (package / "assignments.json").read_bytes()
    assignments = json.loads(assignments_bytes)
    exceptions = json.loads((package / "exceptions.json").read_bytes())
    assert exceptions == {"items": [], "schemaVersion": "2.0"}
    assert assignments["packageId"] == manifest["packageId"] == result["packageId"]
    assert assignments["sourceObservationId"] == manifest["sourceObservationId"]
    assert assignments["proposalDigest"] == manifest["proposalDigest"]
    assert len(assignments["units"]) == 5
    assert len(assignments["exclusions"]) == 2
    assert [
        participant["rosterRowId"] for participant in assignments["participants"]
    ] == roster_row_ids
    decision_ids = {decision["decisionId"] for decision in manifest["decisions"]}
    source_ids = {source_record["sourceId"] for source_record in manifest["sources"]}
    artifact_ids = {artifact["artifactId"] for artifact in manifest["artifacts"]}
    participant_handles = {
        participant["participantHandle"] for participant in assignments["participants"]
    }
    for unit in assignments["units"]:
        assert unit["decisionId"] in decision_ids
        assert unit["sourceId"] in source_ids
        assert unit["outputLocator"]["artifactId"] in artifact_ids
        assert set(unit["target"]["participantHandles"]) <= participant_handles
    for exclusion in assignments["exclusions"]:
        assert exclusion["decisionId"] in decision_ids
    for private_value in (
        _SYNTHETIC_NAME_1,
        _SYNTHETIC_NAME_2,
        _SYNTHETIC_ID_1,
        _SYNTHETIC_ID_2,
        _SYNTHETIC_FA,
        "SYNTHETIC-BANK-0001",
        "SYNTHETIC-BANK-0002",
    ):
        assert private_value.encode("utf-8") not in assignments_bytes
    assert not any(
        _UNSUPPORTED_MARKER in path.read_bytes()
        for path in package.rglob("*")
        if path.is_file()
    )

    assert sha256((package / "case-manifest.json").read_bytes()).hexdigest() == result[
        "manifestSha256"
    ]
    assert _published_tree_sha256(package) == result["publishedTreeSha256"]
    assert _filesystem_state(source) == source_before

    assert validate_main([str(package), "--source-root", str(source)]) == 0
    validation_capture = capsys.readouterr()
    assert validation_capture.err == ""
    validation_report = json.loads(validation_capture.out)
    for private_value in (
        _SYNTHETIC_NAME_1,
        _SYNTHETIC_NAME_2,
        _SYNTHETIC_ID_1,
        _SYNTHETIC_ID_2,
        _SYNTHETIC_FA,
        str(source),
        str(output),
    ):
        assert private_value not in validation_capture.out
    assert validation_report["schemaVersion"] == "2.0"
    assert validation_report["outcome"] == "valid"
    assert validation_report["packageStatus"] == "prepared"
    assert validation_report["errors"] == []
    assert validation_report["warnings"] == []

    output_before_collision = _filesystem_state(output)
    assert intake_main(argv, package_review_driver=_http_review_driver) == 2
    collision_capture = capsys.readouterr()
    assert collision_capture.err == ""
    collision = json.loads(collision_capture.out)
    assert collision_capture.out.encode("utf-8") == _canonical_cli_bytes(collision)
    assert collision["operation"] == "package.prepare"
    assert collision["status"] == "failed"
    assert collision["result"] == {}
    assert collision["retryable"] is False
    assert [error["code"] for error in collision["errors"]] == [
        "package-output-collision"
    ]
    assert _filesystem_state(output) == output_before_collision
    assert _filesystem_state(source) == source_before
    assert not list(output.glob(".ctv-staging-*"))
