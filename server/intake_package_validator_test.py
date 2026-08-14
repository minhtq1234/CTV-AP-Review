import gc
import hashlib
import json
import os
from pathlib import Path
import weakref

import fitz
import openpyxl
import pytest


SYNTHETIC_TIMESTAMP = "2026-08-11T00:00:00Z"


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_pdf(path: Path, page_count: int = 2) -> None:
    document = fitz.open()
    try:
        for page_number in range(1, page_count + 1):
            page = document.new_page()
            page.insert_text((72, 72), f"Synthetic page {page_number}")
        document.save(path)
    finally:
        document.close()


def _write_roster(
    path: Path,
    *,
    title: str = "Roster",
    headers: tuple[str, ...] = ("Display label", "Synthetic identity"),
    rows: tuple[tuple[str, ...], ...] = (
        ("SUBJECT-ALPHA", "SYNTHETIC-IDENTITY-A"),
        ("SUBJECT-BETA", "SYNTHETIC-IDENTITY-B"),
    ),
) -> None:
    workbook = openpyxl.Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = title
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        workbook.save(path)
    finally:
        workbook.close()


def _artifact(artifact_id: str, kind: str, path: Path, source_ids: list[str]) -> dict:
    return {
        "artifactId": artifact_id,
        "kind": kind,
        "path": path.name,
        "size": path.stat().st_size,
        "sha256": _sha256(path),
        "sourceIds": source_ids,
    }


def _write_package(
    package_dir: Path,
    *,
    status: str = "prepared",
    exception_items: list[dict] | None = None,
) -> dict:
    package_dir.mkdir()
    pdf_path = package_dir / "input.pdf"
    roster_path = package_dir / "roster.xlsx"
    exceptions_path = package_dir / "exceptions.json"
    _write_pdf(pdf_path)
    _write_roster(roster_path)
    source_root = _source_root(package_dir)
    (source_root / "workspace").mkdir(parents=True, exist_ok=True)
    source_pdf = source_root / "workspace" / "source.pdf"
    source_pdf.write_bytes(pdf_path.read_bytes() + b"\n% immutable source snapshot\n")
    (source_root / "workspace" / "source.xlsx").write_bytes(roster_path.read_bytes())
    exceptions_path.write_text(
        json.dumps({"schemaVersion": "1.0", "items": exception_items or []}),
        encoding="utf-8",
    )
    manifest = {
        "schemaVersion": "1.0",
        "batchId": "batch-synthetic",
        "caseId": "case-synthetic",
        "faCode": None,
        "packageVersion": "1.0",
        "status": status,
        "compatibilityTarget": "ctv-intake-v1",
        "sources": [
            {
                "sourceId": "source-pdf",
                "path": "workspace/source.pdf",
                "mediaType": "application/pdf",
                "pageCount": 2,
                "size": source_pdf.stat().st_size,
                "sha256": _sha256(source_pdf),
                "coverageState": "assigned",
            },
            {
                "sourceId": "source-roster",
                "path": "workspace/source.xlsx",
                "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "size": roster_path.stat().st_size,
                "sha256": _sha256(roster_path),
                "coverageState": "assigned",
            },
        ],
        "pdfPages": [
            {
                "sourceId": "source-pdf",
                "sourcePage": 1,
                "coverageState": "assigned",
                "targetPage": 1,
            },
            {
                "sourceId": "source-pdf",
                "sourcePage": 2,
                "coverageState": "assigned",
                "targetPage": 2,
            },
        ],
        "artifacts": [
            _artifact("artifact-pdf", "input-pdf", pdf_path, ["source-pdf"]),
            _artifact("artifact-roster", "roster", roster_path, ["source-roster"]),
            _artifact("artifact-exceptions", "exceptions", exceptions_path, []),
        ],
        "rosterMapping": {
            "sourceId": "source-roster",
            "sheetName": "Roster",
            "canonicalToSourceColumns": {
                "name": "Display label",
                "identity": "Synthetic identity",
            },
        },
        "decisions": [
            {
                "decisionId": "decision-preview-current",
                "proposalVersion": "1.0",
                "type": "approve-preview",
                "actor": "user",
                "timestamp": SYNTHETIC_TIMESTAMP,
                "evidenceRefs": ["source-pdf", "source-roster"],
            }
        ],
        "exceptionIds": [item["exceptionId"] for item in exception_items or []],
        "validatedAt": SYNTHETIC_TIMESTAMP,
        "validatorVersion": "1.0.0",
    }
    _save_manifest(package_dir, manifest)
    return manifest


def _save_manifest(package_dir: Path, manifest: dict) -> None:
    (package_dir / "case-manifest.json").write_text(
        json.dumps(manifest), encoding="utf-8"
    )


def _refresh_artifact(package_dir: Path, manifest: dict, kind: str) -> None:
    artifact = next(item for item in manifest["artifacts"] if item["kind"] == kind)
    path = package_dir / artifact["path"]
    artifact["size"] = path.stat().st_size
    artifact["sha256"] = _sha256(path)
    _save_manifest(package_dir, manifest)


def _add_artifact(
    package_dir: Path,
    manifest: dict,
    *,
    artifact_id: str,
    kind: str,
    filename: str,
    content: bytes,
) -> None:
    path = package_dir / filename
    path.write_bytes(content)
    manifest["artifacts"].append(_artifact(artifact_id, kind, path, []))
    _save_manifest(package_dir, manifest)


def _historical_validation_report() -> dict:
    return {
        "schemaVersion": "1.0",
        "outcome": "invalid",
        "packageStatus": "partially_prepared",
        "checks": [
            {
                "code": "historical-check",
                "passed": False,
                "evidenceRefs": ["historical-synthetic-evidence"],
            }
        ],
        "errors": ["historical-check"],
        "warnings": [],
        "validatedAt": SYNTHETIC_TIMESTAMP,
        "validatorVersion": "0.9.0",
    }


def _validate(package_dir: Path):
    from intake_package_validator import validate_package

    return validate_package(package_dir, source_root=_source_root(package_dir))


def _source_root(package_dir: Path) -> Path:
    return package_dir.parent / f"{package_dir.name}-sources"


def _validate_with_source(package_dir: Path, source_root: Path | None = None):
    from intake_package_validator import validate_package

    root = _source_root(package_dir) if source_root is None else source_root
    return validate_package(package_dir, source_root=root)


def _check(report, code: str):
    return next(check for check in report.checks if check.code == code)


def test_valid_prepared_package_passes_semantic_validation(tmp_path):
    package_dir = tmp_path / "package"
    _write_package(package_dir)

    report = _validate(package_dir)

    assert report.outcome == "valid"
    assert report.package_status == "prepared"
    assert report.errors == []
    assert report.warnings == []
    assert report.validator_version == "1.0.0"


def test_missing_required_artifact_is_rejected(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["artifacts"] = [
        artifact for artifact in manifest["artifacts"] if artifact["kind"] != "roster"
    ]
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "missing-required-artifact" in report.errors
    assert _check(report, "missing-required-artifact").evidence_refs == ["roster"]


def test_duplicate_kind_is_not_snapshotted_while_unaffected_sibling_is_validated(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    input_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    duplicate_ids = [
        "artifact-pdf",
        "artifact-pdf-01",
        "artifact-pdf-02",
        "artifact-pdf-03",
        "artifact-pdf-04",
        "artifact-pdf-05",
        "artifact-pdf-06",
        "artifact-pdf-07",
        "artifact-pdf-08",
    ]
    manifest["artifacts"].extend(
        {**input_artifact, "artifactId": artifact_id}
        for artifact_id in duplicate_ids[1:]
    )
    (package_dir / "roster.xlsx").write_bytes(b"not a workbook")
    _refresh_artifact(package_dir, manifest, "roster")
    real_os_open = validator.os.open

    def reject_duplicate_snapshot(path, flags, mode=0o777, *, dir_fd=None):
        if path == "input.pdf":
            pytest.fail("a declaration from a duplicate artifact kind was opened")
        return real_os_open(path, flags, mode, dir_fd=dir_fd)

    def reject_duplicate_parser(*_args, **_kwargs):
        pytest.fail("a declaration from a duplicate artifact kind was parsed")

    monkeypatch.setattr(validator.os, "open", reject_duplicate_snapshot)
    derived_content = (package_dir / "input.pdf").read_bytes()
    real_fitz_open = validator.fitz.open

    def reject_only_duplicate_artifact_parse(*args, **kwargs):
        if kwargs.get("stream") == derived_content:
            return reject_duplicate_parser(*args, **kwargs)
        return real_fitz_open(*args, **kwargs)

    monkeypatch.setattr(validator.fitz, "open", reject_only_duplicate_artifact_parse)

    first = _validate(package_dir)
    second = _validate(package_dir)

    assert first.errors == ["duplicate-artifact-kind", "roster-unreadable"]
    assert _check(first, "duplicate-artifact-kind").evidence_refs == duplicate_ids
    assert first.model_dump(exclude={"validated_at"}) == second.model_dump(
        exclude={"validated_at"}
    )


def test_byte_size_mismatch_is_rejected_before_digesting_or_reading(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    pdf_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    pdf_path = package_dir / pdf_artifact["path"]
    pdf_inode = pdf_path.stat().st_ino
    with pdf_path.open("r+b") as stream:
        stream.truncate(2 * 1024 * 1024)
    real_fdopen = validator.os.fdopen

    def reject_artifact_stream_read(file_descriptor, *args, **kwargs):
        if validator.os.fstat(file_descriptor).st_ino == pdf_inode:
            pytest.fail("size-mismatched artifact content was opened for reading")
        return real_fdopen(file_descriptor, *args, **kwargs)

    monkeypatch.setattr(validator.os, "fdopen", reject_artifact_stream_read)
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "artifact-size-mismatch" in report.errors
    assert "artifact-digest-mismatch" not in report.errors
    assert _check(report, "artifact-size-mismatch").evidence_refs == ["artifact-pdf"]


def test_digest_mismatch_is_rejected_when_declared_size_matches(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    pdf_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    pdf_artifact["sha256"] = "f" * 64
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "artifact-size-mismatch" not in report.errors
    assert "artifact-digest-mismatch" in report.errors


def test_artifact_over_the_absolute_kind_limit_is_rejected_before_reading(tmp_path):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    pdf_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    pdf_path = package_dir / pdf_artifact["path"]
    hard_limit = validator.MAX_ARTIFACT_BYTES_BY_KIND["input-pdf"]
    with pdf_path.open("r+b") as stream:
        stream.truncate(hard_limit + 1)
    pdf_artifact["size"] = hard_limit + 1
    pdf_artifact["sha256"] = "f" * 64
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "artifact-too-large" in report.errors
    assert "artifact-size-mismatch" not in report.errors
    assert "artifact-digest-mismatch" not in report.errors
    assert "pdf-unreadable" not in report.errors


def test_artifact_appended_after_fstat_is_bounded_and_rejected(tmp_path, monkeypatch):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    pdf_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    pdf_path = package_dir / pdf_artifact["path"]
    pdf_inode = pdf_path.stat().st_ino
    accepted_limit = pdf_path.stat().st_size + 16
    monkeypatch.setitem(
        validator.MAX_ARTIFACT_BYTES_BY_KIND, "input-pdf", accepted_limit
    )
    real_fdopen = validator.os.fdopen
    appended = False

    def append_before_read(file_descriptor, *args, **kwargs):
        nonlocal appended
        if not appended and validator.os.fstat(file_descriptor).st_ino == pdf_inode:
            with pdf_path.open("ab") as stream:
                stream.write(b"X" * 32)
            appended = True
        return real_fdopen(file_descriptor, *args, **kwargs)

    monkeypatch.setattr(validator.os, "fdopen", append_before_read)

    report = _validate(package_dir)

    assert appended
    assert "artifact-too-large" in report.errors
    assert "artifact-size-mismatch" not in report.errors
    assert "artifact-digest-mismatch" not in report.errors
    assert "pdf-unreadable" not in report.errors


def test_missing_and_extra_pdf_page_coverage_are_reported(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["pdfPages"] = [
        manifest["pdfPages"][0],
        {
            "sourceId": "source-pdf",
            "sourcePage": 3,
            "coverageState": "assigned",
            "targetPage": 2,
        },
    ]
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "page-coverage-missing" in report.errors
    assert "page-coverage-extra" in report.errors
    assert _check(report, "page-coverage-missing").evidence_refs == [
        "source-pdf#page=2"
    ]
    assert _check(report, "page-coverage-extra").evidence_refs == [
        "source-pdf#page=3"
    ]


def test_pdf_source_requires_a_declared_page_count(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0].pop("pageCount")
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "page-coverage-missing" in report.errors
    assert _check(report, "page-coverage-missing").evidence_refs == [
        "source-pdf#page-count"
    ]


def test_each_pdf_source_uses_its_own_page_count_in_a_merged_input_pdf(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    _write_pdf(package_dir / "input.pdf", page_count=3)
    _refresh_artifact(package_dir, manifest, "input-pdf")
    manifest["sources"].append(
        {
            "sourceId": "source-pdf-secondary",
            "path": "workspace/source-secondary.pdf",
            "mediaType": "application/pdf",
            "pageCount": 1,
            "size": 1,
            "sha256": "e" * 64,
            "coverageState": "assigned",
        }
    )
    secondary_source = _source_root(package_dir) / "workspace" / "source-secondary.pdf"
    _write_pdf(secondary_source, page_count=1)
    manifest["sources"][-1]["size"] = secondary_source.stat().st_size
    manifest["sources"][-1]["sha256"] = _sha256(secondary_source)
    input_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    input_artifact["sourceIds"].append("source-pdf-secondary")
    manifest["pdfPages"].append(
        {
            "sourceId": "source-pdf-secondary",
            "sourcePage": 1,
            "coverageState": "assigned",
            "targetPage": 3,
        }
    )
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert report.outcome == "valid"
    assert report.errors == []


@pytest.mark.parametrize(
    ("target_pages", "missing_evidence", "extra_evidence"),
    [
        ((1, 1), "artifact-pdf#target-page=2", "artifact-pdf#target-page=1"),
        ((1, 3), "artifact-pdf#target-page=2", "artifact-pdf#target-page=3"),
    ],
)
def test_derived_pdf_target_pages_must_be_covered_exactly_once(
    tmp_path, target_pages, missing_evidence, extra_evidence
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    for page, target_page in zip(manifest["pdfPages"], target_pages, strict=True):
        page["targetPage"] = target_page
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert missing_evidence in _check(report, "page-coverage-missing").evidence_refs
    assert extra_evidence in _check(report, "page-coverage-extra").evidence_refs


@pytest.mark.parametrize(
    ("relationship", "expected_evidence"),
    [
        ("omitted", "artifact-pdf#omitted-source=1"),
        ("extra-pdf", "artifact-pdf#extra-source=1"),
        ("non-pdf", "artifact-pdf#non-pdf-source=1"),
    ],
)
def test_input_pdf_provenance_matches_sources_represented_by_target_pages(
    tmp_path, relationship, expected_evidence
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    input_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "input-pdf"
    )
    if relationship == "omitted":
        input_artifact["sourceIds"] = []
    elif relationship == "extra-pdf":
        manifest["sources"].append(
            {
                "sourceId": "source-pdf-unmapped",
                "path": "workspace/source-unmapped.pdf",
                "mediaType": "application/pdf",
                "pageCount": 1,
                "size": 1,
                "sha256": "e" * 64,
                "coverageState": "assigned",
            }
        )
        manifest["pdfPages"].append(
            {
                "sourceId": "source-pdf-unmapped",
                "sourcePage": 1,
                "coverageState": "assigned",
                "targetPage": None,
            }
        )
        input_artifact["sourceIds"].append("source-pdf-unmapped")
    else:
        input_artifact["sourceIds"].append("source-roster")
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "input-pdf-provenance-mismatch" in report.errors
    assert expected_evidence in _check(
        report, "input-pdf-provenance-mismatch"
    ).evidence_refs


def test_unknown_source_decision_exception_and_evidence_references_are_rejected(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0]["duplicateSourceId"] = "source-unknown"
    manifest["sources"][0]["decisionId"] = "decision-unknown"
    manifest["artifacts"][0]["sourceIds"].append("source-artifact-unknown")
    manifest["decisions"] = [
        {
            "decisionId": "decision-present",
            "proposalVersion": "1.0",
            "type": "approve-preview",
            "actor": "user",
            "timestamp": SYNTHETIC_TIMESTAMP,
            "evidenceRefs": ["artifact-unknown", "source-pdf#page=99"],
        }
    ]
    manifest["exceptionIds"] = ["exception-unknown"]
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "source-reference-unknown" in report.errors
    assert "decision-reference-unknown" in report.errors
    assert "exception-reference-unknown" in report.errors
    assert _check(report, "source-reference-unknown").evidence_refs == [
        "source-artifact-unknown",
        "source-unknown",
    ]
    assert "evidence-reference-unknown" in report.errors
    assert _check(report, "evidence-reference-unknown").evidence_refs == [
        "decision-present#evidence-ref=0",
        "decision-present#evidence-ref=1",
    ]


def test_unknown_evidence_reports_owner_and_index_without_echoing_raw_values(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(
        package_dir,
        exception_items=[
            {
                "exceptionId": "exception-evidence-owner",
                "code": "artifact-outside-package",
                "severity": "warning",
                "evidenceRefs": ["../../private/location/synthetic-subject"],
                "explanation": "Synthetic unknown exception evidence.",
                "requiredAction": "Review the synthetic evidence reference.",
                "resolution": "open",
            }
        ],
    )
    manifest["decisions"] = [
        {
            "decisionId": "decision-evidence-owner",
            "proposalVersion": "1.0",
            "type": "approve-preview",
            "actor": "user",
            "timestamp": SYNTHETIC_TIMESTAMP,
            "evidenceRefs": ["synthetic.person@example.invalid"],
        }
    ]
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)
    serialized_report = report.model_dump_json(by_alias=True)

    assert _check(report, "evidence-reference-unknown").evidence_refs == [
        "decision-evidence-owner#evidence-ref=0",
        "exception-evidence-owner#evidence-ref=0",
    ]
    assert "../../private/location/synthetic-subject" not in serialized_report
    assert "synthetic.person@example.invalid" not in serialized_report


def test_evidence_can_reference_an_authoritative_source_page_missing_coverage(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["pdfPages"] = [manifest["pdfPages"][0]]
    manifest["decisions"] = [
        {
            "decisionId": "decision-missing-coverage",
            "proposalVersion": "1.0",
            "type": "assign-page",
            "actor": "user",
            "timestamp": SYNTHETIC_TIMESTAMP,
            "evidenceRefs": ["source-pdf#page=2"],
        }
    ]
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "page-coverage-missing" in report.errors
    assert "evidence-reference-unknown" not in report.errors


def test_unreadable_pdf_and_roster_are_reported_as_independent_sibling_failures(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    (package_dir / "input.pdf").write_bytes(b"not a pdf")
    (package_dir / "roster.xlsx").write_bytes(b"not a workbook")
    _refresh_artifact(package_dir, manifest, "input-pdf")
    _refresh_artifact(package_dir, manifest, "roster")

    report = _validate(package_dir)

    assert "pdf-unreadable" in report.errors
    assert "roster-unreadable" in report.errors


@pytest.mark.parametrize(
    "mapping",
    [
        None,
        {
            "sourceId": "source-roster",
            "sheetName": "Roster",
            "canonicalToSourceColumns": {"identity": "Synthetic identity"},
        },
        {
            "sourceId": "source-roster",
            "sheetName": "Roster",
            "canonicalToSourceColumns": {
                "name": "Display label",
                "identity": "Display label",
            },
        },
    ],
)
def test_missing_or_ambiguous_canonical_name_and_identity_mapping_is_rejected(
    tmp_path, mapping
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["rosterMapping"] = mapping
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "roster-mapping-missing" in report.errors


def test_selected_roster_sheet_and_mapped_columns_must_exist(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["rosterMapping"]["sheetName"] = "Missing sheet"
    manifest["rosterMapping"]["canonicalToSourceColumns"]["identity"] = (
        "Missing identity column"
    )
    _save_manifest(package_dir, manifest)

    missing_sheet = _validate(package_dir)
    assert "roster-sheet-missing" in missing_sheet.errors

    manifest["rosterMapping"]["sheetName"] = "Roster"
    _save_manifest(package_dir, manifest)
    missing_column = _validate(package_dir)
    assert "roster-column-missing" in missing_column.errors


def test_nonempty_canonical_identity_values_must_be_unique(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    _write_roster(
        package_dir / "roster.xlsx",
        rows=(
            ("SUBJECT-ALPHA", "SYNTHETIC-IDENTITY-DUPLICATE"),
            ("SUBJECT-BETA", "SYNTHETIC-IDENTITY-DUPLICATE"),
            ("SUBJECT-GAMMA", ""),
            ("SUBJECT-DELTA", ""),
        ),
    )
    _refresh_artifact(package_dir, manifest, "roster")

    report = _validate(package_dir)

    assert "roster-identity-duplicate" in report.errors
    assert _check(report, "roster-identity-duplicate").evidence_refs == [
        "artifact-roster#row=2",
        "artifact-roster#row=3",
    ]


def test_prepared_package_allows_an_open_warning_only_exception(tmp_path):
    package_dir = tmp_path / "package"
    _write_package(
        package_dir,
        exception_items=[
            {
                "exceptionId": "exception-synthetic-warning",
                "code": "artifact-outside-package",
                "severity": "warning",
                "evidenceRefs": ["artifact-pdf"],
                "explanation": "Synthetic warning for contract coverage.",
                "requiredAction": "Review the synthetic warning.",
                "resolution": "open",
            }
        ],
    )

    report = _validate(package_dir)

    assert report.outcome == "valid"
    assert report.package_status == "prepared"
    assert report.errors == []
    assert report.warnings == ["artifact-outside-package"]


@pytest.mark.parametrize(
    ("failure", "artifact_error"),
    [
        ("missing", "artifact-missing"),
        ("malformed", "exceptions-invalid"),
        ("metadata-mismatch", "artifact-size-mismatch"),
    ],
)
def test_manifest_exception_ids_remain_unknown_when_exceptions_are_unusable(
    tmp_path, failure, artifact_error
):
    package_dir = tmp_path / "package"
    manifest = _write_package(
        package_dir,
        exception_items=[
            {
                "exceptionId": "exception-synthetic-unresolved",
                "code": "unassigned-page",
                "severity": "blocking",
                "evidenceRefs": ["source-pdf#page=2"],
                "explanation": "Synthetic exception document failure.",
                "requiredAction": "Restore the synthetic exception document.",
                "resolution": "open",
            }
        ],
    )
    exceptions_path = package_dir / "exceptions.json"
    if failure == "missing":
        exceptions_path.unlink()
    elif failure == "malformed":
        exceptions_path.write_bytes(b"{")
        _refresh_artifact(package_dir, manifest, "exceptions")
    else:
        exceptions_path.write_bytes(b"{}")

    report = _validate(package_dir)

    assert artifact_error in report.errors
    assert "exception-reference-unknown" in report.errors
    assert _check(report, "exception-reference-unknown").evidence_refs == [
        "exception-synthetic-unresolved"
    ]


@pytest.mark.parametrize("valid_document", [True, False])
def test_declared_validation_report_is_parsed_but_never_used_as_current_output(
    tmp_path, valid_document
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    document = _historical_validation_report()
    if not valid_document:
        document["schemaVersion"] = "invalid"
    _add_artifact(
        package_dir,
        manifest,
        artifact_id="artifact-validation-report",
        kind="validation-report",
        filename="historical-validation-report.json",
        content=json.dumps(document).encode("utf-8"),
    )

    report = _validate(package_dir)

    assert ("validation-report-invalid" in report.errors) is not valid_document
    if valid_document:
        assert report.outcome == "valid"
        assert report.errors == []


def test_blocking_exception_and_unresolved_coverage_are_rejected(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(
        package_dir,
        status="partially_prepared",
        exception_items=[
            {
                "exceptionId": "exception-synthetic-blocking",
                "code": "unresolved-coverage",
                "severity": "blocking",
                "evidenceRefs": ["source-pdf#page=2"],
                "explanation": "Synthetic page remains unresolved.",
                "requiredAction": "Resolve the synthetic page.",
                "resolution": "open",
            }
        ],
    )
    manifest["pdfPages"][1]["coverageState"] = "unresolved"
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert report.outcome == "invalid"
    assert report.package_status == "partially_prepared"
    assert "blocking-exception" in report.errors
    assert "unresolved-coverage" in report.errors


def test_prepared_manifest_with_unresolved_coverage_keeps_the_semantic_code(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["pdfPages"][1]["coverageState"] = "unresolved"
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert report.outcome == "invalid"
    assert report.package_status == "partially_prepared"
    assert "manifest-invalid" in report.errors
    assert "unresolved-coverage" in report.errors
    assert _check(report, "unresolved-coverage").evidence_refs == [
        "source-pdf#page=2"
    ]


def test_symlinked_artifact_is_rejected_before_file_parsing(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    outside = tmp_path / "outside.pdf"
    outside.write_bytes(b"not a pdf")
    (package_dir / "input.pdf").unlink()
    (package_dir / "input.pdf").symlink_to(outside)
    _refresh_artifact(package_dir, manifest, "input-pdf")

    report = _validate(package_dir)

    assert "symlink-not-allowed" in report.errors
    assert "pdf-unreadable" not in report.errors


def test_symlinked_package_root_uses_private_synthetic_evidence(tmp_path):
    package_dir = tmp_path / "package"
    _write_package(package_dir)
    package_alias = tmp_path / "package-alias"
    package_alias.symlink_to(package_dir, target_is_directory=True)

    report = _validate(package_alias)

    assert "symlink-not-allowed" in report.errors
    assert _check(report, "symlink-not-allowed").evidence_refs == ["package-root"]


def test_internal_open_reader_validation_uses_and_does_not_close_owned_descriptor(
    tmp_path,
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["artifacts"] = [
        artifact for artifact in manifest["artifacts"] if artifact["kind"] != "roster"
    ]
    _save_manifest(package_dir, manifest)
    reader, failure = validator._PackageReader.open(package_dir)
    assert failure is None
    assert reader is not None
    opened_original = tmp_path / "opened-original"
    package_dir.rename(opened_original)
    _write_package(package_dir)

    try:
        report = validator._validate_package_reader(reader)

        assert report.outcome == "invalid"
        assert "missing-required-artifact" in report.errors
        assert reader.root_fd is not None
        assert os.fstat(reader.root_fd)
    finally:
        reader.close()

    assert reader.root_fd is None
    closed_report = validator._validate_package_reader(reader)
    assert closed_report.errors == ["secure-open-unavailable"]
    assert _check(closed_report, "secure-open-unavailable").evidence_refs == [
        "package-root"
    ]


def test_package_reader_open_at_requires_one_exact_child_and_expected_identity(tmp_path):
    import intake_package_validator as validator

    parent = tmp_path / "parent"
    package = parent / "package"
    parent.mkdir()
    _write_package(package)
    parent_fd = os.open(parent, os.O_RDONLY | os.O_DIRECTORY)
    try:
        identity = os.stat(package, follow_symlinks=False)
        reader, failure = validator._PackageReader.open_at(
            parent_fd,
            "package",
            expected_identity=(identity.st_dev, identity.st_ino),
        )
        assert failure is None and reader is not None
        reader.close()

        for name in ("", ".", "..", "package/child", "package\\child", "/package"):
            rejected, code = validator._PackageReader.open_at(parent_fd, name)
            assert rejected is None
            assert code == "unsafe"

        rejected, code = validator._PackageReader.open_at(
            parent_fd,
            "package",
            expected_identity=(identity.st_dev, identity.st_ino + 1),
        )
        assert rejected is None
        assert code == "changed"
    finally:
        os.close(parent_fd)


def test_package_reader_open_at_rejects_symlink_closed_and_file_descriptors(tmp_path):
    import intake_package_validator as validator

    parent = tmp_path / "parent"
    package = parent / "package"
    parent.mkdir()
    _write_package(package)
    (parent / "alias").symlink_to(package, target_is_directory=True)
    parent_fd = os.open(parent, os.O_RDONLY | os.O_DIRECTORY)
    file_fd = os.open(package / "case-manifest.json", os.O_RDONLY)
    closed_fd = os.dup(parent_fd)
    os.close(closed_fd)
    try:
        reader, failure = validator._PackageReader.open_at(parent_fd, "alias")
        assert reader is None and failure == "symlink"
        reader, failure = validator._PackageReader.open_at(file_fd, "package")
        assert reader is None and failure == "secure-open-unavailable"
        reader, failure = validator._PackageReader.open_at(closed_fd, "package")
        assert reader is None and failure == "secure-open-unavailable"
    finally:
        os.close(file_fd)
        os.close(parent_fd)


def test_package_reader_tree_snapshot_is_sorted_allowlisted_and_portable(tmp_path):
    import intake_package_validator as validator

    package = tmp_path / "package"
    package.mkdir()
    (package / "z.json").write_bytes(b"z\n")
    evidence = package / "evidence"
    evidence.mkdir()
    (evidence / "a.png").write_bytes(b"a\n")
    reader, failure = validator._PackageReader.open(package)
    assert failure is None and reader is not None
    try:
        snapshot, tree_failure = reader.snapshot_tree(
            {"z.json", "evidence/a.png"},
            max_bytes_by_path={"z.json": 8, "evidence/a.png": 8},
            max_total_bytes=16,
        )
        assert tree_failure is None and snapshot is not None
        assert snapshot.paths == ("evidence/a.png", "z.json")
        lines = b"".join(
            f"{digest}  {path}\n".encode()
            for path, digest in snapshot.file_sha256
        )
        assert snapshot.tree_sha256 == hashlib.sha256(lines).hexdigest()
        assert reader.charged_bytes == 4
    finally:
        reader.close()


@pytest.mark.parametrize("kind", ["extra", "nested", "symlink", "fifo", "changed"])
def test_package_reader_tree_snapshot_rejects_non_allowlisted_or_changed_members(
    tmp_path, kind, monkeypatch
):
    import intake_package_validator as validator

    package = tmp_path / kind
    package.mkdir()
    (package / "case-manifest.json").write_bytes(b"{}\n")
    allowed = {"case-manifest.json"}
    limits = {"case-manifest.json": 16}
    if kind == "extra":
        (package / "extra.bin").write_bytes(b"extra")
    elif kind == "nested":
        (package / "nested").mkdir()
        (package / "nested" / "file.bin").write_bytes(b"nested")
    elif kind == "symlink":
        (package / "alias").symlink_to(package / "case-manifest.json")
    elif kind == "fifo":
        os.mkfifo(package / "pipe")

    reader, failure = validator._PackageReader.open(package)
    assert failure is None and reader is not None
    try:
        if kind == "changed":
            real_read = reader.read_cached

            def mutate_after_read(self, path, *, max_bytes, **kwargs):
                result = real_read(path, max_bytes=max_bytes, **kwargs)
                (package / path).write_bytes(b'{"changed":true}\n')
                return result

            monkeypatch.setattr(type(reader), "read_cached", mutate_after_read)
        snapshot, tree_failure = reader.snapshot_tree(
            allowed,
            max_bytes_by_path=limits,
            max_total_bytes=16,
        )
        assert snapshot is None
        assert tree_failure in {"extra", "nested", "symlink", "not-regular", "changed"}
    finally:
        reader.close()


def test_package_reader_binds_snapshot_bytes_to_opened_descriptor_identity(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package = tmp_path / "package"
    package.mkdir()
    target = package / "artifact.bin"
    target.write_bytes(b"original")
    saved = tmp_path / "saved.bin"
    replacement = tmp_path / "replacement.bin"
    replacement.write_bytes(b"replaced")
    real_read = validator._read_relative_to_fd
    metadata = target.stat()
    original_identity = (
        metadata.st_dev,
        metadata.st_ino,
        metadata.st_size,
        metadata.st_mtime_ns,
        metadata.st_ctime_ns,
    )

    def swap_read_swap_back(root_fd, parts, **kwargs):
        target.rename(saved)
        replacement.rename(target)
        try:
            return real_read(root_fd, parts, **kwargs)
        finally:
            target.rename(replacement)
            saved.rename(target)

    def stable_path_identity(_root_fd, _parts):
        return original_identity, None

    monkeypatch.setattr(validator, "_read_relative_to_fd", swap_read_swap_back)
    monkeypatch.setattr(
        validator, "_regular_identity_relative_to_fd", stable_path_identity
    )
    reader, failure = validator._PackageReader.open(package)
    assert failure is None and reader is not None
    try:
        content, read_failure = reader.read_cached("artifact.bin", max_bytes=16)

        assert content is None
        assert read_failure == "changed"
    finally:
        reader.close()


def test_package_reader_reserves_aggregate_budget_before_crossing_artifact_io(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package = tmp_path / "package"
    package.mkdir()
    (package / "a.bin").write_bytes(b"aaaa")
    (package / "b.bin").write_bytes(b"bbbbb")
    real_read = validator._read_relative_to_fd
    reads = []

    def counted(root_fd, parts, **kwargs):
        reads.append(parts)
        return real_read(root_fd, parts, **kwargs)

    monkeypatch.setattr(validator, "_read_relative_to_fd", counted)
    reader, failure = validator._PackageReader.open(package)
    assert failure is None and reader is not None
    try:
        snapshot, tree_failure = reader.snapshot_tree(
            {"a.bin", "b.bin"},
            max_bytes_by_path={"a.bin": 8, "b.bin": 8},
            max_total_bytes=8,
        )

        assert snapshot is None
        assert tree_failure == "too-large"
        assert reads == [("a.bin",)]
        assert reader.charged_bytes == 9

        content, read_failure = reader.read_cached("b.bin", max_bytes=8)
        assert content == b"bbbbb" and read_failure is None
        assert reads == [("a.bin",), ("b.bin",)]
    finally:
        reader.close()


def test_package_reader_rejects_growth_after_scan_before_file_io_and_charges(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package = tmp_path / "package"
    package.mkdir()
    target = package / "artifact.bin"
    target.write_bytes(b"four")
    reader, failure = validator._PackageReader.open(package)
    assert failure is None and reader is not None
    real_read_cached = reader.read_cached
    real_fdopen = validator.os.fdopen
    fdopen_calls = 0
    grown = False

    def grow_before_descriptor_read(self, path, **kwargs):
        nonlocal grown
        if not grown:
            grown = True
            target.write_bytes(b"nine-byte")
        return real_read_cached(path, **kwargs)

    def counted_fdopen(*args, **kwargs):
        nonlocal fdopen_calls
        fdopen_calls += 1
        return real_fdopen(*args, **kwargs)

    monkeypatch.setattr(type(reader), "read_cached", grow_before_descriptor_read)
    monkeypatch.setattr(validator.os, "fdopen", counted_fdopen)
    try:
        snapshot, tree_failure = reader.snapshot_tree(
            {"artifact.bin"},
            max_bytes_by_path={"artifact.bin": 16},
            max_total_bytes=8,
        )

        assert snapshot is None
        assert tree_failure == "changed"
        assert fdopen_calls == 0
        assert reader.charged_bytes == 9
    finally:
        reader.close()


def test_package_reader_charges_attempted_bytes_even_when_read_fails(tmp_path):
    import intake_package_validator as validator

    package = tmp_path / "package"
    package.mkdir()
    (package / "large.bin").write_bytes(b"x" * 17)
    reader, failure = validator._PackageReader.open(package)
    assert failure is None and reader is not None
    try:
        content, read_failure = reader.read_cached("large.bin", max_bytes=16)
        assert content is None and read_failure == "too-large"
        assert reader.charged_bytes == 17
        content, read_failure = reader.read_cached("large.bin", max_bytes=16)
        assert content is None and read_failure == "too-large"
        assert reader.charged_bytes == 17
    finally:
        reader.close()


def test_abandoned_factory_reader_does_not_leak_registry_or_descriptor(tmp_path):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    reader, failure = validator._PackageReader.open(package_dir)
    assert failure is None
    assert reader is not None
    descriptor = reader.root_fd
    assert descriptor is not None
    reader_reference = weakref.ref(reader)

    del reader
    gc.collect()

    assert reader_reference() is None
    with pytest.raises(OSError):
        os.fstat(descriptor)


def test_reader_gc_cleanup_does_not_close_another_registered_identity(tmp_path):
    import intake_package_validator as validator

    first_package = tmp_path / "first-package"
    second_package = tmp_path / "second-package"
    _write_package(first_package)
    _write_package(second_package)
    first, first_failure = validator._PackageReader.open(first_package)
    second, second_failure = validator._PackageReader.open(second_package)
    assert first_failure is None
    assert second_failure is None
    assert first is not None
    assert second is not None
    first_descriptor = first.root_fd
    second_descriptor = second.root_fd
    assert first_descriptor is not None
    assert second_descriptor is not None

    del first
    gc.collect()

    with pytest.raises(OSError):
        os.fstat(first_descriptor)
    assert os.fstat(second_descriptor)
    source_reader, source_failure = validator._SourceReader.open(
        _source_root(second_package)
    )
    assert source_failure is None
    assert source_reader is not None
    try:
        assert validator._validate_package_reader(
            second, source_reader=source_reader
        ).outcome == "valid"
    finally:
        source_reader.close()

    second.close()
    with pytest.raises(OSError):
        os.fstat(second_descriptor)


def test_package_reader_rejects_direct_construction_with_an_arbitrary_descriptor(
    tmp_path,
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    descriptor = os.open(package_dir, os.O_RDONLY)
    try:
        with pytest.raises(TypeError):
            validator._PackageReader(root_path=package_dir, root_fd=descriptor)

        assert os.fstat(descriptor)
    finally:
        os.close(descriptor)


def test_visible_factory_token_cannot_authorize_direct_reader_construction(tmp_path):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    descriptor = os.open(package_dir, os.O_RDONLY)
    visible_token = getattr(validator, "_PACKAGE_READER_FACTORY_TOKEN", object())
    try:
        with pytest.raises(TypeError):
            validator._PackageReader(
                root_path=package_dir,
                root_fd=descriptor,
                _factory_token=visible_token,
            )

        assert os.fstat(descriptor)
    finally:
        os.close(descriptor)


def test_fabricated_package_reader_fails_closed_without_artifact_io(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    fabricated = object.__new__(validator._PackageReader)
    fabricated.root_path = package_dir

    def reject_artifact_io(*_args, **_kwargs):
        pytest.fail("artifact I/O was attempted through a fabricated reader")

    monkeypatch.setattr(validator, "_read_relative_to_fd", reject_artifact_io)

    report = validator._validate_package_reader(fabricated)

    assert report.errors == ["secure-open-unavailable"]
    assert _check(report, "secure-open-unavailable").evidence_refs == ["package-root"]


def test_fully_populated_fabrication_cannot_copy_or_steal_reader_identity(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    reader, failure = validator._PackageReader.open(package_dir)
    assert failure is None
    assert reader is not None
    descriptor = reader.root_fd
    assert descriptor is not None
    fabricated = object.__new__(validator._PackageReader)
    fabricated.root_path = reader.root_path
    object.__setattr__(fabricated, "_PackageReader__root_fd", descriptor)
    if hasattr(reader, "_PackageReader__root_identity"):
        object.__setattr__(
            fabricated,
            "_PackageReader__root_identity",
            getattr(reader, "_PackageReader__root_identity"),
        )
    if hasattr(validator, "_SECURE_OPEN_PROVENANCE"):
        object.__setattr__(
            fabricated,
            "_PackageReader__secure_open_provenance",
            validator._SECURE_OPEN_PROVENANCE,
        )

    def reject_artifact_io(*_args, **_kwargs):
        pytest.fail("artifact I/O was attempted through copied visible trust data")

    def reject_parser(*_args, **_kwargs):
        pytest.fail("an artifact parser ran through copied visible trust data")

    monkeypatch.setattr(validator, "_read_relative_to_fd", reject_artifact_io)
    monkeypatch.setattr(validator.fitz, "open", reject_parser)
    monkeypatch.setattr(validator.openpyxl, "load_workbook", reject_parser)

    try:
        report = validator._validate_package_reader(fabricated)

        assert report.errors == ["secure-open-unavailable"]
        assert _check(report, "secure-open-unavailable").evidence_refs == [
            "package-root"
        ]
        fabricated.close()
        assert os.fstat(descriptor)
    finally:
        reader.close()

    with pytest.raises(OSError):
        os.fstat(descriptor)


def test_equality_forged_subclass_cannot_validate_or_close_registered_reader(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    reader, failure = validator._PackageReader.open(package_dir)
    assert failure is None
    assert reader is not None
    descriptor = reader.root_fd
    assert descriptor is not None
    authority_comparisons = {"eq": 0, "hash": 0}

    class EqualityForgedReader(validator._PackageReader):
        __slots__ = ()

        def __eq__(self, _other):
            authority_comparisons["eq"] += 1
            return True

        def __hash__(self):
            authority_comparisons["hash"] += 1
            return hash(reader)

    fabricated = object.__new__(EqualityForgedReader)
    fabricated.root_path = reader.root_path
    object.__setattr__(fabricated, "_PackageReader__root_fd", descriptor)

    def reject_artifact_io(*_args, **_kwargs):
        pytest.fail("artifact I/O was attempted through equality-based authority")

    def reject_parser(*_args, **_kwargs):
        pytest.fail("an artifact parser ran through equality-based authority")

    with monkeypatch.context() as guarded:
        guarded.setattr(validator, "_read_relative_to_fd", reject_artifact_io)
        guarded.setattr(validator.fitz, "open", reject_parser)
        guarded.setattr(validator.openpyxl, "load_workbook", reject_parser)

        forged_report = validator._validate_package_reader(fabricated)

        assert forged_report.errors == ["secure-open-unavailable"]
        assert _check(
            forged_report, "secure-open-unavailable"
        ).evidence_refs == ["package-root"]
        fabricated.close()

    assert authority_comparisons == {"eq": 0, "hash": 0}
    assert os.fstat(descriptor)
    source_reader, source_failure = validator._SourceReader.open(
        _source_root(package_dir)
    )
    assert source_failure is None
    assert source_reader is not None
    try:
        legitimate_report = validator._validate_package_reader(
            reader, source_reader=source_reader
        )
    finally:
        source_reader.close()
    assert legitimate_report.outcome == "valid"

    reader.close()
    with pytest.raises(OSError):
        os.fstat(descriptor)


def test_open_reader_fails_closed_if_secure_capability_changes_without_io(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    reader, failure = validator._PackageReader.open(package_dir)
    assert failure is None
    assert reader is not None
    descriptor = reader.root_fd
    assert descriptor is not None

    def reject_artifact_io(*_args, **_kwargs):
        pytest.fail("artifact I/O was attempted after secure capability changed")

    def reject_parser(*_args, **_kwargs):
        pytest.fail("an artifact parser ran after secure capability changed")

    monkeypatch.setattr(validator, "_SUPPORTS_SECURE_RELATIVE_OPEN", False)
    monkeypatch.setattr(validator, "_read_relative_to_fd", reject_artifact_io)
    monkeypatch.setattr(validator.fitz, "open", reject_parser)
    monkeypatch.setattr(validator.openpyxl, "load_workbook", reject_parser)

    try:
        assert reader.root_fd is None
        report = validator._validate_package_reader(reader)

        assert report.errors == ["secure-open-unavailable"]
        assert _check(report, "secure-open-unavailable").evidence_refs == [
            "package-root"
        ]
        assert os.fstat(descriptor)
    finally:
        reader.close()

    with pytest.raises(OSError):
        os.fstat(descriptor)


def test_validation_fails_closed_when_secure_relative_open_is_unavailable(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)

    def reject_pathname_fallback(*_args, **_kwargs):
        pytest.fail("an insecure package pathname reader was invoked")

    def reject_parser(*_args, **_kwargs):
        pytest.fail("an artifact parser was invoked")

    monkeypatch.setattr(validator, "_SUPPORTS_SECURE_RELATIVE_OPEN", False)
    monkeypatch.setattr(Path, "open", reject_pathname_fallback)
    monkeypatch.setattr(validator.fitz, "open", reject_parser)
    monkeypatch.setattr(validator.openpyxl, "load_workbook", reject_parser)

    report = _validate(package_dir)

    assert report.errors == ["secure-open-unavailable"]
    assert _check(report, "secure-open-unavailable").evidence_refs == ["package-root"]


def test_pdf_parser_uses_the_same_opened_bytes_that_were_digest_checked(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    pdf_path = package_dir / "input.pdf"
    replacement = tmp_path / "replacement.pdf"
    replacement.write_bytes(b"not a pdf")
    real_open = validator.fitz.open
    derived_content = pdf_path.read_bytes()
    swapped = False

    def swap_path_before_real_parse(*args, **kwargs):
        nonlocal swapped
        if not swapped and kwargs.get("stream") == derived_content:
            replacement.replace(pdf_path)
            swapped = True
        return real_open(*args, **kwargs)

    monkeypatch.setattr(validator.fitz, "open", swap_path_before_real_parse)

    report = _validate(package_dir)

    assert swapped
    assert report.outcome == "valid"
    assert "pdf-unreadable" not in report.errors


def test_roster_parser_uses_the_same_opened_bytes_that_were_preflighted(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    roster_path = package_dir / "roster.xlsx"
    replacement = tmp_path / "replacement.xlsx"
    replacement.write_bytes(b"not a workbook")
    real_load_workbook = validator.openpyxl.load_workbook
    swapped = False

    def swap_path_before_real_parse(*args, **kwargs):
        nonlocal swapped
        if not swapped:
            replacement.replace(roster_path)
            swapped = True
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(
        validator.openpyxl, "load_workbook", swap_path_before_real_parse
    )

    report = _validate(package_dir)

    assert swapped
    assert report.outcome == "valid"
    assert "roster-unreadable" not in report.errors


@pytest.mark.parametrize("declared_path", ["../outside.pdf", "bad\x00.pdf"])
def test_package_relative_path_escaping_the_package_is_rejected(
    tmp_path, declared_path
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["artifacts"][0]["path"] = declared_path
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "unsafe-artifact-path" in report.errors


def test_failed_size_check_precedes_and_suppresses_pdf_parsing(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    (package_dir / "input.pdf").write_bytes(b"not a pdf")
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "artifact-size-mismatch" in report.errors
    assert "pdf-unreadable" not in report.errors


def test_errors_and_evidence_are_deterministic_and_code_sorted(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["artifacts"] = []
    manifest["pdfPages"] = []
    manifest["sources"][0]["decisionId"] = "decision-zeta"
    manifest["sources"][1]["decisionId"] = "decision-alpha"
    _save_manifest(package_dir, manifest)

    first = _validate(package_dir)
    second = _validate(package_dir)

    assert first.errors == sorted(first.errors)
    assert first.errors == second.errors
    assert first.model_dump(exclude={"validated_at"}) == second.model_dump(
        exclude={"validated_at"}
    )
    assert _check(first, "decision-reference-unknown").evidence_refs == [
        "decision-alpha",
        "decision-zeta",
    ]
    assert _check(first, "missing-required-artifact").evidence_refs == [
        "exceptions",
        "input-pdf",
        "roster",
    ]


def test_assigned_source_page_without_target_is_not_falsely_valid(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["pdfPages"][1].pop("targetPage")
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "coverage-state-inconsistent" in report.errors
    assert "source-pdf#page=2" in _check(
        report, "coverage-state-inconsistent"
    ).evidence_refs


def test_assigned_non_pdf_source_requires_derived_artifact_provenance(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    roster_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "roster"
    )
    roster_artifact["sourceIds"] = []
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "source-provenance-missing" in report.errors
    assert _check(report, "source-provenance-missing").evidence_refs == [
        "source-roster"
    ]


@pytest.mark.parametrize("approval_mode", ["missing", "stale"])
def test_current_package_version_requires_approve_preview(tmp_path, approval_mode):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    if approval_mode == "missing":
        manifest["decisions"] = []
    else:
        manifest["decisions"][0]["proposalVersion"] = "0.9"
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "approval-missing" in report.errors
    assert _check(report, "approval-missing").evidence_refs == ["1.0"]


@pytest.mark.parametrize(
    ("owner", "state", "wrong_type"),
    [
        ("source", "shared", "assign-source"),
        ("source", "duplicate", "share-source"),
        ("source", "excluded-by-user", "assign-source"),
        ("page", "shared", "assign-page"),
        ("page", "excluded-by-user", "assign-page"),
    ],
)
def test_coverage_decision_ids_must_use_the_state_transition_type(
    tmp_path, owner, state, wrong_type
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    decision_id = "decision-state"
    manifest["decisions"].append(
        {
            "decisionId": decision_id,
            "proposalVersion": "1.0",
            "type": wrong_type,
            "actor": "user",
            "timestamp": SYNTHETIC_TIMESTAMP,
            "evidenceRefs": ["source-pdf"],
        }
    )
    if owner == "source":
        source = manifest["sources"][0]
        source["coverageState"] = state
        source["decisionId"] = decision_id
        if state == "duplicate":
            source["duplicateSourceId"] = "source-roster"
    else:
        page = manifest["pdfPages"][0]
        page["coverageState"] = state
        page["decisionId"] = decision_id
        if state == "excluded-by-user":
            page.pop("targetPage")
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "decision-type-mismatch" in report.errors


def _accepted_partial_exception() -> dict:
    return {
        "exceptionId": "exception-partial",
        "code": "unassigned-page",
        "severity": "blocking",
        "evidenceRefs": ["source-pdf#page=2"],
        "explanation": "A synthetic page is intentionally left partial.",
        "requiredAction": "Accept the visible partial package.",
        "resolution": "accepted-partial",
    }


def test_prepared_package_rejects_blocking_accepted_partial_exception(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(
        package_dir, status="prepared", exception_items=[_accepted_partial_exception()]
    )
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "partial-status-required" in report.errors


def test_partial_accepted_exception_requires_current_exact_accept_partial_decision(
    tmp_path,
):
    package_dir = tmp_path / "package"
    manifest = _write_package(
        package_dir,
        status="partially_prepared",
        exception_items=[_accepted_partial_exception()],
    )
    _save_manifest(package_dir, manifest)

    missing = _validate(package_dir)
    assert "accept-partial-decision-missing" in missing.errors

    manifest["decisions"].append(
        {
            "decisionId": "decision-accept-partial",
            "proposalVersion": "1.0",
            "type": "accept-partial",
            "actor": "user",
            "timestamp": SYNTHETIC_TIMESTAMP,
            "evidenceRefs": ["exception-partial"],
        }
    )
    _save_manifest(package_dir, manifest)

    accepted = _validate(package_dir)
    assert "accept-partial-decision-missing" not in accepted.errors
    assert "partial-status-required" not in accepted.errors
    assert "unassigned-page" in accepted.warnings


def test_raw_unsupported_compatibility_target_has_a_stable_error(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["compatibilityTarget"] = "anything-that-mentions-v1"
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "compatibility-target-unsupported" in report.errors
    assert _check(report, "compatibility-target-unsupported").evidence_refs == [
        "anything-that-mentions-v1"
    ]


@pytest.mark.parametrize("page_count", [10_001, 1_000_000_000])
def test_per_source_page_limit_rejects_without_unbounded_error_expansion(
    tmp_path, page_count
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0]["pageCount"] = page_count
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "page-count-limit-exceeded" in report.errors
    assert len(_check(report, "page-count-limit-exceeded").evidence_refs) <= 2
    assert "page-coverage-missing" not in report.errors


def test_aggregate_page_limit_rejects_without_materializing_declared_ranges(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0]["pageCount"] = 9_000
    input_artifact = next(
        artifact
        for artifact in manifest["artifacts"]
        if artifact["kind"] == "input-pdf"
    )
    for index in (2, 3):
        source_id = f"source-pdf-{index}"
        manifest["sources"].append(
            {
                "sourceId": source_id,
                "path": f"workspace/{source_id}.pdf",
                "mediaType": "application/pdf",
                "pageCount": 9_000,
                "size": 1,
                "sha256": f"{index}" * 64,
                "coverageState": "assigned",
            }
        )
        input_artifact["sourceIds"].append(source_id)
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "page-count-limit-exceeded" in report.errors
    assert _check(report, "page-count-limit-exceeded").evidence_refs == [
        "package#declared-pdf-pages=27000"
    ]
    assert "page-coverage-missing" not in report.errors


def test_contradictory_declared_validation_report_is_rejected(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    contradictory = {
        "schemaVersion": "1.0",
        "outcome": "valid",
        "packageStatus": "partially_prepared",
        "checks": [{"code": "historical-error", "passed": False, "evidenceRefs": []}],
        "errors": ["historical-error"],
        "warnings": [],
        "validatedAt": SYNTHETIC_TIMESTAMP,
        "validatorVersion": "0.9.0",
    }
    _add_artifact(
        package_dir,
        manifest,
        artifact_id="artifact-validation-report",
        kind="validation-report",
        filename="validation-report.json",
        content=json.dumps(contradictory).encode(),
    )

    report = _validate(package_dir)

    assert "validation-report-invalid" in report.errors


def test_valid_report_lists_positive_checks_for_every_executed_gate(tmp_path):
    package_dir = tmp_path / "package"
    _write_package(package_dir)

    report = _validate(package_dir)

    assert report.outcome == "valid"
    assert report.checks
    assert all(check.passed for check in report.checks)
    assert {"manifest-valid", "approval-valid", "coverage-valid"} <= {
        check.code for check in report.checks
    }


@pytest.mark.parametrize("filename", ["cccd.xlsx", "cccd.pdf"])
def test_cccd_artifact_requires_a_readable_xlsx_workbook(tmp_path, filename):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    cccd_path = package_dir / filename
    cccd_path.write_bytes(b"not an xlsx workbook")
    manifest["sources"].append(
        {
            "sourceId": "source-cccd",
            "path": "workspace/cccd.xlsx",
            "mediaType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "size": len(b"not an xlsx workbook"),
            "sha256": hashlib.sha256(b"not an xlsx workbook").hexdigest(),
            "coverageState": "assigned",
        }
    )
    manifest["artifacts"].append(
        _artifact("artifact-cccd", "cccd", cccd_path, ["source-cccd"])
    )
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "cccd-unreadable" in report.errors


def test_duck_typed_reader_cannot_override_internal_authority_check():
    import intake_package_validator as validator

    class DuckReader:
        def __init__(self):
            self.read_called = False

        def has_secure_open_provenance(self):
            return True

        def read(self, *_args, **_kwargs):
            self.read_called = True
            pytest.fail("an unauthorized duck reader performed package I/O")

    reader = DuckReader()

    report = validator._validate_package_reader(reader)

    assert report.errors == ["secure-open-unavailable"]
    assert not reader.read_called


def test_extreme_page_evidence_reference_is_bounded_and_reported_unknown(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["decisions"].append(
        {
            "decisionId": "decision-extreme-page-reference",
            "proposalVersion": "1.0",
            "type": "assign-page",
            "actor": "user",
            "timestamp": SYNTHETIC_TIMESTAMP,
            "evidenceRefs": [f"source-pdf#page={'9' * 5_000}"],
        }
    )
    _save_manifest(package_dir, manifest)

    report = _validate(package_dir)

    assert "evidence-reference-unknown" in report.errors
    assert _check(report, "evidence-reference-unknown").evidence_refs == [
        "decision-extreme-page-reference#evidence-ref=0"
    ]


def test_understated_declared_source_page_count_cannot_hide_an_original_page(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    _write_pdf(package_dir / "input.pdf", page_count=1)
    _refresh_artifact(package_dir, manifest, "input-pdf")
    manifest["sources"][0]["pageCount"] = 1
    manifest["pdfPages"] = [manifest["pdfPages"][0]]
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert "source-page-count-mismatch" in report.errors
    assert _check(report, "source-page-count-mismatch").evidence_refs == [
        "source-pdf"
    ]


def test_declared_sources_cannot_validate_without_source_root(tmp_path):
    from intake_package_validator import validate_package

    package_dir = tmp_path / "package"
    _write_package(package_dir)

    report = validate_package(package_dir)

    assert report.outcome == "invalid"
    assert "source-verification-unavailable" in report.errors
    assert _check(report, "source-verification-unavailable").evidence_refs == [
        "source-root"
    ]


@pytest.mark.parametrize("root_kind", ["missing", "symlink", "file"])
def test_unsafe_source_root_is_a_private_validation_error(tmp_path, root_kind):
    package_dir = tmp_path / "package"
    _write_package(package_dir)
    real_source_root = _source_root(package_dir)
    supplied_root = tmp_path / "supplied-source-root"
    if root_kind == "symlink":
        supplied_root.symlink_to(real_source_root, target_is_directory=True)
    elif root_kind == "file":
        supplied_root.write_bytes(b"not a source directory")

    report = _validate_with_source(package_dir, supplied_root)

    expected = (
        "source-symlink-not-allowed"
        if root_kind == "symlink"
        else "source-verification-unavailable"
    )
    assert expected in report.errors
    assert _check(report, expected).evidence_refs == ["source-root"]
    assert str(supplied_root) not in report.model_dump_json()


def test_unsafe_source_path_has_a_stable_synthetic_error(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0]["path"] = "../outside.pdf"
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert "unsafe-source-path" in report.errors
    assert _check(report, "unsafe-source-path").evidence_refs == ["source-pdf"]
    assert "../outside.pdf" not in report.model_dump_json()


def test_symlinked_and_missing_source_files_are_independent_errors(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    source_root = _source_root(package_dir)
    source_pdf = source_root / "workspace" / "source.pdf"
    source_pdf.unlink()
    source_pdf.symlink_to(source_root / "workspace" / "source.xlsx")
    (source_root / "workspace" / "source.xlsx").unlink()
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert "source-symlink-not-allowed" in report.errors
    assert "source-missing" in report.errors
    assert _check(report, "source-symlink-not-allowed").evidence_refs == [
        "source-pdf"
    ]
    assert _check(report, "source-missing").evidence_refs == ["source-roster"]


@pytest.mark.parametrize(
    ("field", "value", "expected_code"),
    [
        ("size", 1, "source-size-mismatch"),
        ("sha256", "f" * 64, "source-digest-mismatch"),
    ],
)
def test_source_size_and_digest_are_verified_from_opened_bytes(
    tmp_path, field, value, expected_code
):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0][field] = value
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert expected_code in report.errors
    assert _check(report, expected_code).evidence_refs == ["source-pdf"]


def test_unreadable_source_pdf_is_rejected_from_the_verified_snapshot(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    source_pdf = _source_root(package_dir) / "workspace" / "source.pdf"
    source_pdf.write_bytes(b"not a source pdf")
    manifest["sources"][0]["size"] = source_pdf.stat().st_size
    manifest["sources"][0]["sha256"] = _sha256(source_pdf)
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert "source-pdf-unreadable" in report.errors
    assert "source-media-type-mismatch" not in report.errors


def test_oversized_source_is_rejected_before_reading(tmp_path):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    source_pdf = _source_root(package_dir) / "workspace" / "source.pdf"
    limit = validator.MAX_ARTIFACT_BYTES_BY_KIND["input-pdf"]
    with source_pdf.open("r+b") as stream:
        stream.truncate(limit + 1)
    manifest["sources"][0]["size"] = limit + 1
    manifest["sources"][0]["sha256"] = "f" * 64
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert "source-too-large" in report.errors
    assert "source-digest-mismatch" not in report.errors


def test_repeated_source_path_uses_one_bounded_cached_read(tmp_path, monkeypatch):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    duplicate = dict(manifest["sources"][0])
    duplicate["sourceId"] = "source-pdf-repeat"
    duplicate["coverageState"] = "unsupported"
    manifest["sources"].append(duplicate)
    _save_manifest(package_dir, manifest)
    real_read = validator._read_relative_to_fd
    reads = 0

    def count_source_reads(root_fd, parts, **kwargs):
        nonlocal reads
        if parts == ("workspace", "source.pdf"):
            reads += 1
        return real_read(root_fd, parts, **kwargs)

    monkeypatch.setattr(validator, "_read_relative_to_fd", count_source_reads)

    _validate_with_source(package_dir)

    assert reads == 1


def test_source_verification_never_mutates_original_bytes_or_metadata(tmp_path):
    package_dir = tmp_path / "package"
    _write_package(package_dir)
    source_root = _source_root(package_dir)
    before = {
        path.relative_to(source_root): (
            path.read_bytes(),
            path.stat().st_mode,
            path.stat().st_mtime_ns,
        )
        for path in source_root.rglob("*")
        if path.is_file()
    }

    _validate_with_source(package_dir)

    after = {
        path.relative_to(source_root): (
            path.read_bytes(),
            path.stat().st_mode,
            path.stat().st_mtime_ns,
        )
        for path in source_root.rglob("*")
        if path.is_file()
    }
    assert after == before


def test_actual_source_pdf_page_limit_is_checked_before_page_loading(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    source_pdf = _source_root(package_dir) / "workspace" / "source.pdf"
    _write_pdf(source_pdf, page_count=1)
    manifest["sources"][0]["size"] = source_pdf.stat().st_size
    manifest["sources"][0]["sha256"] = _sha256(source_pdf)
    source_content = source_pdf.read_bytes()
    real_open = validator.fitz.open

    class OversizedDocument:
        is_pdf = True
        needs_pass = False
        page_count = 10_001

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def load_page(self, *_args):
            pytest.fail("oversized source PDF pages were loaded")

    def open_source_as_oversized(*args, **kwargs):
        if kwargs.get("stream") == source_content:
            return OversizedDocument()
        return real_open(*args, **kwargs)

    monkeypatch.setattr(validator.fitz, "open", open_source_as_oversized)

    report = _validate_with_source(package_dir)

    assert "source-page-count-mismatch" in report.errors


def test_derived_pdf_actual_page_limit_is_checked_before_page_loading(
    tmp_path, monkeypatch
):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    _write_package(package_dir)
    derived_content = (package_dir / "input.pdf").read_bytes()
    real_open = validator.fitz.open

    class OversizedDocument:
        is_pdf = True
        needs_pass = False
        page_count = 10_001

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def load_page(self, *_args):
            pytest.fail("oversized derived PDF pages were loaded")

    def open_derived_as_oversized(*args, **kwargs):
        if kwargs.get("stream") == derived_content:
            return OversizedDocument()
        return real_open(*args, **kwargs)

    monkeypatch.setattr(validator.fitz, "open", open_derived_as_oversized)

    report = _validate_with_source(package_dir)

    assert "pdf-page-limit-exceeded" in report.errors


def test_declared_historical_valid_report_with_empty_checks_is_rejected(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    empty_checks = {
        "schemaVersion": "1.0",
        "outcome": "valid",
        "packageStatus": "prepared",
        "checks": [],
        "errors": [],
        "warnings": [],
        "validatedAt": SYNTHETIC_TIMESTAMP,
        "validatorVersion": "0.9.0",
    }
    _add_artifact(
        package_dir,
        manifest,
        artifact_id="artifact-validation-report-empty-checks",
        kind="validation-report",
        filename="historical-report.json",
        content=json.dumps(empty_checks).encode(),
    )

    report = _validate_with_source(package_dir)

    assert "validation-report-invalid" in report.errors


def test_verified_pdf_bytes_cannot_hide_behind_a_generic_media_type(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    hidden_pdf = _source_root(package_dir) / "workspace" / "hidden-evidence.bin"
    _write_pdf(hidden_pdf, page_count=2)
    manifest["sources"].append(
        {
            "sourceId": "source-hidden-pdf",
            "path": "workspace/hidden-evidence.bin",
            "mediaType": "application/octet-stream",
            "size": hidden_pdf.stat().st_size,
            "sha256": _sha256(hidden_pdf),
            "coverageState": "assigned",
        }
    )
    roster_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "roster"
    )
    roster_artifact["sourceIds"].append("source-hidden-pdf")
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert report.outcome == "invalid"
    assert "source-media-type-mismatch" in report.errors
    assert _check(report, "source-media-type-mismatch").evidence_refs == [
        "source-hidden-pdf"
    ]
    assert "source-provenance-missing" in report.errors
    assert "source-hidden-pdf" in _check(
        report, "source-provenance-missing"
    ).evidence_refs
    assert "source-hidden-pdf#pages=1-2" in _check(
        report, "page-coverage-missing"
    ).evidence_refs


def test_non_pdf_bytes_remain_non_pdf_even_with_a_pdf_filename(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    opaque_source = _source_root(package_dir) / "workspace" / "not-a-pdf.pdf"
    opaque_source.write_bytes(b"SYNTHETIC OPAQUE SOURCE\n")
    manifest["sources"].append(
        {
            "sourceId": "source-opaque",
            "path": "workspace/not-a-pdf.pdf",
            "mediaType": "application/octet-stream",
            "size": opaque_source.stat().st_size,
            "sha256": _sha256(opaque_source),
            "coverageState": "assigned",
        }
    )
    roster_artifact = next(
        artifact for artifact in manifest["artifacts"] if artifact["kind"] == "roster"
    )
    roster_artifact["sourceIds"].append("source-opaque")
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert report.outcome == "valid"
    assert "source-media-type-mismatch" not in report.errors


def test_detected_pdf_membership_drives_input_pdf_provenance(tmp_path):
    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    hidden_pdf = _source_root(package_dir) / "workspace" / "hidden-page.bin"
    _write_pdf(hidden_pdf, page_count=1)
    manifest["sources"].append(
        {
            "sourceId": "source-hidden-page",
            "path": "workspace/hidden-page.bin",
            "mediaType": "application/octet-stream",
            "pageCount": 1,
            "size": hidden_pdf.stat().st_size,
            "sha256": _sha256(hidden_pdf),
            "coverageState": "assigned",
        }
    )
    manifest["pdfPages"].append(
        {
            "sourceId": "source-hidden-page",
            "sourcePage": 1,
            "coverageState": "assigned",
            "targetPage": 3,
        }
    )
    input_artifact = next(
        artifact
        for artifact in manifest["artifacts"]
        if artifact["kind"] == "input-pdf"
    )
    input_artifact["sourceIds"].append("source-hidden-page")
    _write_pdf(package_dir / "input.pdf", page_count=3)
    _refresh_artifact(package_dir, manifest, "input-pdf")
    _save_manifest(package_dir, manifest)

    report = _validate_with_source(package_dir)

    assert report.errors == ["source-media-type-mismatch"]
    assert "input-pdf-provenance-mismatch" not in report.errors
    assert "page-coverage-extra" not in report.errors


def test_detected_generic_pdf_is_capped_before_page_loading(tmp_path, monkeypatch):
    import intake_package_validator as validator

    package_dir = tmp_path / "package"
    manifest = _write_package(package_dir)
    manifest["sources"][0]["mediaType"] = "application/octet-stream"
    _save_manifest(package_dir, manifest)
    source_pdf = _source_root(package_dir) / "workspace" / "source.pdf"
    source_content = source_pdf.read_bytes()
    real_open = validator.fitz.open

    class OversizedDocument:
        is_pdf = True
        needs_pass = False
        page_count = 10_001

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def load_page(self, *_args):
            pytest.fail("misdeclared oversized PDF pages were loaded")

    def open_source_as_oversized(*args, **kwargs):
        if kwargs.get("stream") == source_content:
            return OversizedDocument()
        return real_open(*args, **kwargs)

    monkeypatch.setattr(validator.fitz, "open", open_source_as_oversized)

    report = _validate_with_source(package_dir)

    assert "source-media-type-mismatch" in report.errors
    assert "source-page-count-mismatch" in report.errors
