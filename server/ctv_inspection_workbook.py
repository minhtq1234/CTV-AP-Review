"""Bounded, byte-only OOXML workbook inspection."""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
import math
import zipfile
from xml.etree import ElementTree

import openpyxl

from ctv_inspection_classifier import (
    TextSignalContext,
    roster_header_categories_from_private_text,
    signals_from_private_text,
)
from ctv_inspection_model import (
    DEFAULT_INSPECTION_LIMITS,
    InspectionAdapterResult,
    InspectionLimits,
    InspectionUnitCountExceededError,
    InspectionUnitEvidence,
    SIGNAL_ORDER,
)
from ooxml import OoxmlRelationshipError, resolve_internal_relationship_target


_MAX_ARCHIVE_ENTRIES = 10_000
_MAX_DECOMPRESSED_BYTES = 100 * 1024 * 1024
_MAX_MEMBER_BYTES = 25 * 1024 * 1024
_MAX_COMPRESSION_RATIO = 100
_MAX_XML_ELEMENTS = 200_000
_READ_CHUNK_BYTES = 64 * 1024
_SERIALIZE_TEXT_CHARS = 4 * 1024
_CONTENT_TYPES_PART = "[Content_Types].xml"
_WORKBOOK_PART = "xl/workbook.xml"
_WORKBOOK_RELS_PART = "xl/_rels/workbook.xml.rels"
_OLE_COMPOUND_HEADER = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"


class _OutputLimitExceeded(RuntimeError):
    pass


class _CappedBytesIO(BytesIO):
    def __init__(self, limit: int) -> None:
        super().__init__()
        self.limit = limit
        self.crossed = False

    def write(self, value: bytes) -> int:
        if self.crossed:
            return len(value)
        end = self.tell() + len(value)
        if end > self.limit:
            self.crossed = True
            raise _OutputLimitExceeded()
        return super().write(value)
_CONTENT_TYPES_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/content-types"
)
_PACKAGE_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/package/2006/relationships"
)
_XML_NAMESPACE = "http://www.w3.org/XML/1998/namespace"
_XMLNS_NAMESPACE = "http://www.w3.org/2000/xmlns/"
_TRANSITIONAL_SPREADSHEET_NAMESPACE = (
    "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
_STRICT_SPREADSHEET_NAMESPACE = (
    "http://purl.oclc.org/ooxml/spreadsheetml/main"
)
_TRANSITIONAL_OFFICE_RELATIONSHIPS_NAMESPACE = (
    "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
_STRICT_OFFICE_RELATIONSHIPS_NAMESPACE = (
    "http://purl.oclc.org/ooxml/officeDocument/relationships"
)
_SPREADSHEET_NAMESPACES = frozenset({
    _TRANSITIONAL_SPREADSHEET_NAMESPACE,
    _STRICT_SPREADSHEET_NAMESPACE,
})
_OFFICE_RELATIONSHIPS_BY_SPREADSHEET_NAMESPACE = {
    _TRANSITIONAL_SPREADSHEET_NAMESPACE: (
        _TRANSITIONAL_OFFICE_RELATIONSHIPS_NAMESPACE
    ),
    _STRICT_SPREADSHEET_NAMESPACE: _STRICT_OFFICE_RELATIONSHIPS_NAMESPACE,
}
_WORKBOOK_CONTENT_TYPES = frozenset({
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
    "application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.template.main+xml",
    "application/vnd.ms-excel.template.macroEnabled.main+xml",
})
_WORKSHEET_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
)
_DRAWING_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.drawing+xml"
)
_STYLES_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"
)
_SHARED_STRINGS_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"
)
_WORKSHEET_RELATIONSHIP_TYPE_BY_SPREADSHEET_NAMESPACE = {
    namespace: f"{relationship_namespace}/worksheet"
    for namespace, relationship_namespace
    in _OFFICE_RELATIONSHIPS_BY_SPREADSHEET_NAMESPACE.items()
}
_DRAWING_RELATIONSHIP_TYPE_BY_SPREADSHEET_NAMESPACE = {
    namespace: f"{relationship_namespace}/drawing"
    for namespace, relationship_namespace
    in _OFFICE_RELATIONSHIPS_BY_SPREADSHEET_NAMESPACE.items()
}


class WorkbookParserBoundaryExceededError(RuntimeError):
    """Stable operation boundary for unsafe OOXML parser/decompression work."""

    def __init__(self) -> None:
        super().__init__("inspection-parser-boundary-exceeded")


class WorkbookWorksheetCountExceededError(RuntimeError):
    """Stable operation boundary raised before prohibited sheet iteration."""

    def __init__(self) -> None:
        super().__init__("inspection-worksheet-count-exceeded")


class WorkbookPreviewError(RuntimeError):
    """Fixed local-preview failure without workbook or parser details."""

    def __init__(self, code: str) -> None:
        if code not in {
            "preview-unavailable",
            "preview-over-limit",
            "preview-parser-boundary-exceeded",
        }:
            raise ValueError("workbook preview error code must be fixed")
        super().__init__(code)


class PackageWorkbookError(RuntimeError):
    """Fixed bounded failure for package-specific values-only extraction."""

    def __init__(self, code: str) -> None:
        if code not in {
            "package-workbook-unavailable",
            "package-workbook-over-limit",
            "package-workbook-parser-boundary-exceeded",
            "package-workbook-formula-unavailable",
        }:
            raise ValueError("package workbook error code must be fixed")
        super().__init__(code)


@dataclass(frozen=True)
class WorksheetValues:
    worksheet_index: int
    rows: tuple[tuple[object, ...], ...] = field(repr=False)


class _UnreadableWorkbookError(ValueError):
    pass


class _EncryptedWorkbookError(ValueError):
    pass


class _ActualByteBudget:
    def __init__(self) -> None:
        self.used = 0

    def consume(self, amount: int, member_used: int) -> None:
        if (
            amount < 0
            or member_used > _MAX_MEMBER_BYTES
            or self.used + amount > _MAX_DECOMPRESSED_BYTES
        ):
            raise WorkbookParserBoundaryExceededError()
        self.used += amount

    def reserve(self, amount: int) -> None:
        if amount < 0 or self.used + amount > _MAX_DECOMPRESSED_BYTES:
            raise WorkbookParserBoundaryExceededError()
        self.used += amount


class _CellElementBudget:
    def __init__(self, limit: int) -> None:
        self.limit = limit
        self.used = 0

    def consume(self) -> bool:
        if self.used >= self.limit:
            return False
        self.used += 1
        return True


def _source_problem(status: str, issue: str) -> InspectionAdapterResult:
    return InspectionAdapterResult(status, None, (issue,), ())


def _local_name(tag: object) -> str:
    if not isinstance(tag, str):
        return ""
    return tag.rsplit("}", 1)[-1]


def _expanded_name(tag: object):
    if not isinstance(tag, str) or not tag.startswith("{"):
        return None
    namespace, separator, local = tag[1:].partition("}")
    if not separator or not namespace or not local:
        return None
    return namespace, local


def _relationship_id(element, spreadsheet_namespace: str):
    relationship_namespace = _OFFICE_RELATIONSHIPS_BY_SPREADSHEET_NAMESPACE[
        spreadsheet_namespace
    ]
    expected_attribute = f"{{{relationship_namespace}}}id"
    for attribute in element.attrib:
        if _local_name(attribute) == "id" and attribute != expected_attribute:
            raise _UnreadableWorkbookError()
    relationship_id = element.attrib.get(expected_attribute)
    if not relationship_id or len(relationship_id) > 1_024:
        raise _UnreadableWorkbookError()
    return relationship_id


def _xml_declaration_text(content: bytes) -> str:
    if content.startswith((b"\x00\x00\xfe\xff", b"\xff\xfe\x00\x00")):
        encoding = "utf-32"
    elif content.startswith((b"\xfe\xff", b"\xff\xfe")):
        encoding = "utf-16"
    elif content.startswith(b"\xef\xbb\xbf"):
        encoding = "utf-8-sig"
    elif content.startswith(b"\x00\x00\x00<"):
        encoding = "utf-32-be"
    elif content.startswith(b"<\x00\x00\x00"):
        encoding = "utf-32-le"
    elif content.startswith(b"\x00<\x00"):
        encoding = "utf-16-be"
    elif content.startswith(b"<\x00?\x00"):
        encoding = "utf-16-le"
    else:
        encoding = "utf-8"
    try:
        return content.decode(encoding, errors="strict")
    except (LookupError, UnicodeError):
        raise _UnreadableWorkbookError() from None


def _reject_unsafe_xml_declarations(content: bytes) -> None:
    declaration_text = _xml_declaration_text(content).upper()
    if "<!DOCTYPE" in declaration_text or "<!ENTITY" in declaration_text:
        raise _UnreadableWorkbookError()
    declaration_text = ""


def _safe_xml_events(content: bytes):
    _reject_unsafe_xml_declarations(content)
    parser = ElementTree.XMLPullParser(events=("start", "end"))
    element_count = 0
    try:
        for offset in range(0, len(content), _READ_CHUNK_BYTES):
            parser.feed(content[offset:offset + _READ_CHUNK_BYTES])
            for event, element in parser.read_events():
                if event == "start":
                    element_count += 1
                    if element_count > _MAX_XML_ELEMENTS:
                        raise WorkbookParserBoundaryExceededError()
                yield event, element
        parser.close()
        for event, element in parser.read_events():
            if event == "start":
                element_count += 1
                if element_count > _MAX_XML_ELEMENTS:
                    raise WorkbookParserBoundaryExceededError()
            yield event, element
    except WorkbookParserBoundaryExceededError:
        raise
    except Exception:
        raise _UnreadableWorkbookError() from None
    finally:
        content = b""


def _validate_eocd(snapshot: bytes) -> int:
    minimum_eocd_bytes = 22
    if len(snapshot) < minimum_eocd_bytes:
        raise _UnreadableWorkbookError()
    search_start = max(0, len(snapshot) - (65_535 + minimum_eocd_bytes))
    eocd_offset = snapshot.rfind(b"PK\x05\x06", search_start)
    if eocd_offset < 0 or eocd_offset + minimum_eocd_bytes > len(snapshot):
        raise _UnreadableWorkbookError()
    try:
        disk_number = int.from_bytes(snapshot[eocd_offset + 4:eocd_offset + 6], "little")
        central_disk = int.from_bytes(snapshot[eocd_offset + 6:eocd_offset + 8], "little")
        disk_entries = int.from_bytes(snapshot[eocd_offset + 8:eocd_offset + 10], "little")
        total_entries = int.from_bytes(snapshot[eocd_offset + 10:eocd_offset + 12], "little")
        central_size = int.from_bytes(snapshot[eocd_offset + 12:eocd_offset + 16], "little")
        central_offset = int.from_bytes(snapshot[eocd_offset + 16:eocd_offset + 20], "little")
        comment_size = int.from_bytes(snapshot[eocd_offset + 20:eocd_offset + 22], "little")
    except Exception:
        raise _UnreadableWorkbookError() from None
    if (
        disk_number != 0
        or central_disk != 0
        or disk_entries != total_entries
        or total_entries == 0xFFFF
        or central_size == 0xFFFFFFFF
        or central_offset == 0xFFFFFFFF
    ):
        raise WorkbookParserBoundaryExceededError()
    if total_entries > _MAX_ARCHIVE_ENTRIES:
        raise WorkbookParserBoundaryExceededError()
    if (
        eocd_offset + minimum_eocd_bytes + comment_size != len(snapshot)
        or central_offset + central_size != eocd_offset
    ):
        raise _UnreadableWorkbookError()
    return total_entries


def _central_directory(archive: zipfile.ZipFile, expected_entries: int):
    try:
        infos = archive.infolist()
    except Exception:
        raise _UnreadableWorkbookError() from None
    if len(infos) != expected_entries:
        raise _UnreadableWorkbookError()
    declared_total = 0
    members = {}
    for info in infos:
        member_name = info.filename
        if member_name in members:
            raise _UnreadableWorkbookError()
        if info.flag_bits & 1:
            raise _EncryptedWorkbookError()
        declared_size = info.file_size
        compressed_size = info.compress_size
        if (
            not isinstance(declared_size, int)
            or isinstance(declared_size, bool)
            or not isinstance(compressed_size, int)
            or isinstance(compressed_size, bool)
            or declared_size < 0
            or compressed_size < 0
        ):
            raise _UnreadableWorkbookError()
        if declared_size > _MAX_MEMBER_BYTES:
            raise WorkbookParserBoundaryExceededError()
        declared_total += declared_size
        if declared_total > _MAX_DECOMPRESSED_BYTES:
            raise WorkbookParserBoundaryExceededError()
        if declared_size and (
            compressed_size == 0
            or declared_size > compressed_size * _MAX_COMPRESSION_RATIO
        ):
            raise WorkbookParserBoundaryExceededError()
        members[member_name] = info
    if _CONTENT_TYPES_PART not in members or _WORKBOOK_PART not in members:
        raise _UnreadableWorkbookError()
    return members, declared_total


def _read_member(
    archive: zipfile.ZipFile,
    info: zipfile.ZipInfo,
    budget: _ActualByteBudget,
) -> bytes:
    chunks = []
    member_used = 0
    try:
        with archive.open(info, "r") as stream:
            while True:
                chunk = stream.read(_READ_CHUNK_BYTES)
                if not chunk:
                    break
                if not isinstance(chunk, bytes):
                    raise _UnreadableWorkbookError()
                member_used += len(chunk)
                budget.consume(len(chunk), member_used)
                chunks.append(chunk)
    except WorkbookParserBoundaryExceededError:
        raise
    except RuntimeError:
        if info.flag_bits & 1:
            raise _EncryptedWorkbookError() from None
        raise _UnreadableWorkbookError() from None
    except Exception:
        raise _UnreadableWorkbookError() from None
    if member_used != info.file_size:
        raise _UnreadableWorkbookError()
    content = b"".join(chunks)
    chunks.clear()
    return content


def _content_type_map(content: bytes):
    defaults = {}
    overrides = {}
    root_seen = False
    entry_count = 0
    for event, element in _safe_xml_events(content):
        expanded = _expanded_name(element.tag)
        local = _local_name(element.tag)
        if event == "start" and not root_seen:
            if expanded != (_CONTENT_TYPES_NAMESPACE, "Types"):
                raise _UnreadableWorkbookError()
            root_seen = True
        elif event == "start" and (
            expanded is None
            or expanded[0] != _CONTENT_TYPES_NAMESPACE
            or local not in {"Default", "Override"}
        ):
            raise _UnreadableWorkbookError()
        if event == "start" and local in {"Default", "Override"}:
            if expanded != (_CONTENT_TYPES_NAMESPACE, local):
                raise _UnreadableWorkbookError()
            entry_count += 1
            if entry_count > _MAX_ARCHIVE_ENTRIES:
                raise WorkbookParserBoundaryExceededError()
            content_type = element.attrib.get("ContentType")
            if not isinstance(content_type, str) or not 1 <= len(content_type) <= 512:
                raise _UnreadableWorkbookError()
            if local == "Default":
                extension = element.attrib.get("Extension")
                if (
                    not isinstance(extension, str)
                    or not 1 <= len(extension) <= 32
                    or not extension.isascii()
                    or not extension.isalnum()
                ):
                    raise _UnreadableWorkbookError()
                extension = extension.lower()
                if extension in defaults:
                    raise _UnreadableWorkbookError()
                defaults[extension] = content_type
            else:
                part_name = element.attrib.get("PartName")
                if (
                    not isinstance(part_name, str)
                    or not 2 <= len(part_name) <= 1_024
                    or not part_name.startswith("/")
                    or "\\" in part_name
                    or any(
                        segment in {"", ".", ".."}
                        for segment in part_name[1:].split("/")
                    )
                    or part_name in overrides
                ):
                    raise _UnreadableWorkbookError()
                overrides[part_name] = content_type
        if event == "end":
            element.clear()
    if (
        not root_seen
        or overrides.get("/xl/workbook.xml") not in _WORKBOOK_CONTENT_TYPES
    ):
        raise _UnreadableWorkbookError()
    return defaults, overrides


def _part_content_type(part_name: str, content_types) -> str | None:
    defaults, overrides = content_types
    override = overrides.get(f"/{part_name}")
    if override is not None:
        return override
    _, separator, extension = part_name.rpartition(".")
    if not separator:
        return None
    return defaults.get(extension.lower())


def _strict_xml_tree(content: bytes):
    for event, element in _safe_xml_events(content):
        if event == "end":
            element.clear()
    try:
        return ElementTree.fromstring(content)
    except Exception:
        raise _UnreadableWorkbookError() from None


def _transitional_qname(name: object):
    expanded = _expanded_name(name)
    if expanded is None:
        return name
    namespace, local = expanded
    replacement_namespace = {
        _STRICT_SPREADSHEET_NAMESPACE: _TRANSITIONAL_SPREADSHEET_NAMESPACE,
        _STRICT_OFFICE_RELATIONSHIPS_NAMESPACE: (
            _TRANSITIONAL_OFFICE_RELATIONSHIPS_NAMESPACE
        ),
    }.get(namespace)
    if replacement_namespace is None:
        return name
    return f"{{{replacement_namespace}}}{local}"


class _CappedXmlOutput(BytesIO):
    def write(self, content: bytes) -> int:
        if (
            len(content) > _READ_CHUNK_BYTES
            or self.tell() + len(content) > _MAX_MEMBER_BYTES
        ):
            raise WorkbookParserBoundaryExceededError()
        return super().write(content)


def _escaped_utf8_size(text: object, *, attribute: bool) -> int:
    if not isinstance(text, str):
        raise _UnreadableWorkbookError()
    size = 0
    for offset in range(0, len(text), _SERIALIZE_TEXT_CHARS):
        chunk = text[offset:offset + _SERIALIZE_TEXT_CHARS]
        try:
            size += len(chunk.encode("utf-8", errors="xmlcharrefreplace"))
        except Exception:
            raise _UnreadableWorkbookError() from None
        size += chunk.count("&") * 4
        size += chunk.count("<") * 3
        size += chunk.count(">") * 3
        if attribute:
            size += chunk.count('"') * 5
            size += chunk.count("\r") * 4
            size += chunk.count("\n") * 4
            size += chunk.count("\t") * 4
        if size > _MAX_MEMBER_BYTES:
            raise WorkbookParserBoundaryExceededError()
    return size


def _validate_serialized_value_bytes(root) -> None:
    value_bytes = 0
    for element in root.iter():
        for text in (element.text, element.tail):
            if text is not None:
                value_bytes += _escaped_utf8_size(text, attribute=False)
        for value in element.attrib.values():
            value_bytes += _escaped_utf8_size(value, attribute=True)
        if value_bytes > _MAX_MEMBER_BYTES:
            raise WorkbookParserBoundaryExceededError()


def _xml_namespaces(root) -> dict[str, str]:
    namespaces = {}
    generated_prefix_count = 0

    def register(name: object) -> None:
        nonlocal generated_prefix_count
        if not isinstance(name, str):
            raise _UnreadableWorkbookError()
        expanded = _expanded_name(name)
        if expanded is None:
            if name.startswith("{"):
                raise _UnreadableWorkbookError()
            return
        namespace = expanded[0]
        if namespace in namespaces:
            return
        if namespace == _XMLNS_NAMESPACE:
            raise _UnreadableWorkbookError()
        if namespace == _XML_NAMESPACE:
            namespaces[namespace] = "xml"
        else:
            namespaces[namespace] = f"ns{generated_prefix_count}"
            generated_prefix_count += 1

    for element in root.iter():
        register(element.tag)
        for name in element.attrib:
            register(name)
    return namespaces


def _write_utf8(output: _CappedXmlOutput, text: object) -> None:
    if not isinstance(text, str):
        raise _UnreadableWorkbookError()
    for offset in range(0, len(text), _SERIALIZE_TEXT_CHARS):
        chunk = text[offset:offset + _SERIALIZE_TEXT_CHARS]
        try:
            encoded = chunk.encode("utf-8", errors="xmlcharrefreplace")
        except Exception:
            raise _UnreadableWorkbookError() from None
        output.write(encoded)


def _write_escaped_utf8(
    output: _CappedXmlOutput,
    text: object,
    *,
    attribute: bool,
) -> None:
    if not isinstance(text, str):
        raise _UnreadableWorkbookError()
    for offset in range(0, len(text), _SERIALIZE_TEXT_CHARS):
        chunk = text[offset:offset + _SERIALIZE_TEXT_CHARS]
        chunk = chunk.replace("&", "&amp;")
        chunk = chunk.replace("<", "&lt;")
        chunk = chunk.replace(">", "&gt;")
        if attribute:
            chunk = chunk.replace('"', "&quot;")
            chunk = chunk.replace("\r", "&#13;")
            chunk = chunk.replace("\n", "&#10;")
            chunk = chunk.replace("\t", "&#09;")
        try:
            encoded = chunk.encode("utf-8", errors="xmlcharrefreplace")
        except Exception:
            raise _UnreadableWorkbookError() from None
        output.write(encoded)


def _write_qname(
    output: _CappedXmlOutput,
    name: object,
    namespaces: dict[str, str],
) -> None:
    if not isinstance(name, str):
        raise _UnreadableWorkbookError()
    expanded = _expanded_name(name)
    if expanded is None:
        if name.startswith("{"):
            raise _UnreadableWorkbookError()
        _write_utf8(output, name)
        return
    namespace, local = expanded
    try:
        prefix = namespaces[namespace]
    except KeyError:
        raise _UnreadableWorkbookError() from None
    _write_utf8(output, prefix)
    output.write(b":")
    _write_utf8(output, local)


def _write_xml_element(
    output: _CappedXmlOutput,
    element,
    namespaces: dict[str, str],
    *,
    root: bool,
) -> None:
    output.write(b"<")
    _write_qname(output, element.tag, namespaces)
    if root:
        for namespace, prefix in namespaces.items():
            if namespace == _XML_NAMESPACE:
                continue
            output.write(b" xmlns:")
            _write_utf8(output, prefix)
            output.write(b'=\"')
            _write_escaped_utf8(output, namespace, attribute=True)
            output.write(b'\"')
    for name, value in element.attrib.items():
        output.write(b" ")
        _write_qname(output, name, namespaces)
        output.write(b'=\"')
        _write_escaped_utf8(output, value, attribute=True)
        output.write(b'\"')
    if len(element) == 0 and not element.text:
        output.write(b" />")
    else:
        output.write(b">")
        if element.text:
            _write_escaped_utf8(output, element.text, attribute=False)
        for child in element:
            _write_xml_element(output, child, namespaces, root=False)
        output.write(b"</")
        _write_qname(output, element.tag, namespaces)
        output.write(b">")
    if element.tail:
        _write_escaped_utf8(output, element.tail, attribute=False)


def _serialized_xml(root) -> bytes:
    _validate_serialized_value_bytes(root)
    output = _CappedXmlOutput()
    try:
        namespaces = _xml_namespaces(root)
        output.write(b"<?xml version='1.0' encoding='utf-8'?>\n")
        _write_xml_element(output, root, namespaces, root=True)
        return output.getvalue()
    except WorkbookParserBoundaryExceededError:
        raise
    except Exception:
        raise _UnreadableWorkbookError() from None


def _normalized_strict_xml(content: bytes) -> bytes:
    root = _strict_xml_tree(content)
    for element in root.iter():
        element.tag = _transitional_qname(element.tag)
        converted_attributes = {}
        for name, value in element.attrib.items():
            converted_name = _transitional_qname(name)
            if converted_name in converted_attributes:
                raise _UnreadableWorkbookError()
            converted_attributes[converted_name] = value
        element.attrib.clear()
        element.attrib.update(converted_attributes)
    return _serialized_xml(root)


def _normalized_strict_relationships_xml(content: bytes) -> bytes:
    root = _strict_xml_tree(content)
    strict_prefix = f"{_STRICT_OFFICE_RELATIONSHIPS_NAMESPACE}/"
    transitional_prefix = f"{_TRANSITIONAL_OFFICE_RELATIONSHIPS_NAMESPACE}/"
    for element in root:
        relationship_type = element.attrib.get("Type")
        if isinstance(relationship_type, str) and relationship_type.startswith(
            strict_prefix
        ):
            element.attrib["Type"] = (
                transitional_prefix + relationship_type[len(strict_prefix):]
            )
    return _serialized_xml(root)


def _loader_snapshot(parts) -> tuple[bytes, int]:
    declared_total = 0
    for content in parts.values():
        if type(content) is not bytes or len(content) > _MAX_MEMBER_BYTES:
            raise WorkbookParserBoundaryExceededError()
        declared_total += len(content)
        if declared_total > _MAX_DECOMPRESSED_BYTES:
            raise WorkbookParserBoundaryExceededError()
    stream = BytesIO()
    try:
        with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for part_name in tuple(parts):
                content = parts.pop(part_name)
                archive.writestr(part_name, content)
                content = b""
        return stream.getvalue(), declared_total
    except WorkbookParserBoundaryExceededError:
        raise
    except Exception:
        raise _UnreadableWorkbookError() from None


def _loader_content_types(parts, workbook_content_type: str) -> bytes:
    root = ElementTree.Element(f"{{{_CONTENT_TYPES_NAMESPACE}}}Types")
    ElementTree.SubElement(
        root,
        f"{{{_CONTENT_TYPES_NAMESPACE}}}Default",
        {
            "Extension": "rels",
            "ContentType": (
                "application/vnd.openxmlformats-package.relationships+xml"
            ),
        },
    )
    content_types = {
        _WORKBOOK_PART: workbook_content_type,
        "xl/styles.xml": _STYLES_CONTENT_TYPE,
        "xl/sharedStrings.xml": _SHARED_STRINGS_CONTENT_TYPE,
    }
    for part_name in parts:
        if part_name.startswith("xl/worksheets/"):
            content_type = _WORKSHEET_CONTENT_TYPE
        else:
            content_type = content_types.get(part_name)
        if content_type is None:
            continue
        ElementTree.SubElement(
            root,
            f"{{{_CONTENT_TYPES_NAMESPACE}}}Override",
            {"PartName": f"/{part_name}", "ContentType": content_type},
        )
    return _serialized_xml(root)


def _workbook_sheet_relationship_ids(
    content: bytes,
    max_worksheets: int,
):
    relationship_ids = []
    seen_ids = set()
    root_seen = False
    spreadsheet_namespace = None
    for event, element in _safe_xml_events(content):
        expanded = _expanded_name(element.tag)
        local = _local_name(element.tag)
        if event == "start" and not root_seen:
            if (
                expanded is None
                or expanded[0] not in _SPREADSHEET_NAMESPACES
                or expanded[1] != "workbook"
            ):
                raise _UnreadableWorkbookError()
            spreadsheet_namespace = expanded[0]
            root_seen = True
        elif event == "start" and (
            expanded is None or expanded[0] != spreadsheet_namespace
        ):
            raise _UnreadableWorkbookError()
        if event == "start" and local == "sheet":
            if expanded != (spreadsheet_namespace, "sheet"):
                raise _UnreadableWorkbookError()
            relationship_id = _relationship_id(element, spreadsheet_namespace)
            if not relationship_id or relationship_id in seen_ids:
                raise _UnreadableWorkbookError()
            relationship_ids.append(relationship_id)
            seen_ids.add(relationship_id)
            if len(relationship_ids) > max_worksheets:
                raise WorkbookWorksheetCountExceededError()
        if event == "end":
            element.clear()
    if not root_seen:
        raise _UnreadableWorkbookError()
    return tuple(relationship_ids), spreadsheet_namespace


def _relationships(content: bytes, wanted_ids):
    wanted_ids = frozenset(wanted_ids)
    relationships = {}
    root_seen = False
    for event, element in _safe_xml_events(content):
        expanded = _expanded_name(element.tag)
        local = _local_name(element.tag)
        if event == "start" and not root_seen:
            if expanded != (_PACKAGE_RELATIONSHIPS_NAMESPACE, "Relationships"):
                raise _UnreadableWorkbookError()
            root_seen = True
        elif event == "start" and expanded != (
            _PACKAGE_RELATIONSHIPS_NAMESPACE,
            "Relationship",
        ):
            raise _UnreadableWorkbookError()
        if event == "start" and local == "Relationship":
            if expanded != (_PACKAGE_RELATIONSHIPS_NAMESPACE, "Relationship"):
                raise _UnreadableWorkbookError()
            relationship_id = element.attrib.get("Id")
            if relationship_id not in wanted_ids:
                continue
            relationship_type = element.attrib.get("Type")
            target = element.attrib.get("Target")
            target_mode = element.attrib.get("TargetMode")
            if (
                not relationship_id
                or relationship_id in relationships
                or not relationship_type
                or not target
                or target_mode not in {None, "External"}
            ):
                raise _UnreadableWorkbookError()
            relationships[relationship_id] = (
                relationship_type,
                target,
                target_mode == "External",
            )
        if event == "end":
            element.clear()
    if not root_seen:
        raise _UnreadableWorkbookError()
    return relationships


def _cell_position(reference: object):
    if not isinstance(reference, str) or not reference:
        return None
    split_at = 0
    while split_at < len(reference) and "A" <= reference[split_at] <= "Z":
        split_at += 1
    letters = reference[:split_at]
    digits = reference[split_at:]
    if (
        not letters
        or len(letters) > 3
        or not digits
        or not digits.isascii()
        or not digits.isdigit()
    ):
        return None
    column = 0
    for letter in letters:
        column = column * 26 + ord(letter) - ord("A") + 1
    row = int(digits)
    if not 1 <= column <= 16_384 or not 1 <= row <= 1_048_576:
        return None
    return row, column


def _dimension_extent(reference: object):
    if not isinstance(reference, str) or reference.count(":") > 1:
        return None
    first_reference, separator, last_reference = reference.partition(":")
    first = _cell_position(first_reference)
    last = _cell_position(last_reference if separator else first_reference)
    if first is None or last is None or first[0] > last[0] or first[1] > last[1]:
        return None
    return last


def _worksheet_metadata(
    content: bytes,
    cell_budget: _CellElementBudget | None = None,
    spreadsheet_namespace: str | None = None,
):
    if cell_budget is None:
        cell_budget = _CellElementBudget(100_000)
    drawing_id = None
    root_seen = False
    dimension_seen = False
    dimension_extent = None
    cells_are_bounded = True
    seen_cell_positions = set()
    actual_max_row = 0
    actual_max_column = 0
    for event, element in _safe_xml_events(content):
        expanded = _expanded_name(element.tag)
        local = _local_name(element.tag)
        if event == "start" and not root_seen:
            if (
                expanded is None
                or expanded[0] not in _SPREADSHEET_NAMESPACES
                or expanded[1] != "worksheet"
                or (
                    spreadsheet_namespace is not None
                    and expanded[0] != spreadsheet_namespace
                )
            ):
                raise _UnreadableWorkbookError()
            spreadsheet_namespace = expanded[0]
            root_seen = True
        elif event == "start" and (
            expanded is None or expanded[0] != spreadsheet_namespace
        ):
            raise _UnreadableWorkbookError()
        if event == "start" and local == "dimension":
            if expanded != (spreadsheet_namespace, "dimension"):
                raise _UnreadableWorkbookError()
            if dimension_seen:
                cells_are_bounded = False
            dimension_seen = True
            dimension_extent = _dimension_extent(element.attrib.get("ref"))
            if dimension_extent is None:
                cells_are_bounded = False
        if event == "start" and local == "c":
            if expanded != (spreadsheet_namespace, "c"):
                raise _UnreadableWorkbookError()
            position = _cell_position(element.attrib.get("r"))
            within_budget = cell_budget.consume()
            if (
                position is None
                or not within_budget
                or position in seen_cell_positions
            ):
                cells_are_bounded = False
            else:
                seen_cell_positions.add(position)
                actual_max_row = max(actual_max_row, position[0])
                actual_max_column = max(actual_max_column, position[1])
        if event == "start" and local == "drawing":
            if expanded != (spreadsheet_namespace, "drawing"):
                raise _UnreadableWorkbookError()
            relationship_id = _relationship_id(element, spreadsheet_namespace)
            if drawing_id is None:
                drawing_id = relationship_id
        if event == "end":
            element.clear()
    if not root_seen:
        raise _UnreadableWorkbookError()
    if (
        not dimension_seen
        or dimension_extent is None
        or actual_max_row > dimension_extent[0]
        or actual_max_column > dimension_extent[1]
    ):
        cells_are_bounded = False
    cell_bound = (
        dimension_extent[0] * dimension_extent[1]
        if cells_are_bounded and dimension_extent is not None
        else None
    )
    return ((drawing_id,) if drawing_id is not None else ()), cell_bound


def _relationship_part(source_part: str) -> str:
    parent, _, name = source_part.rpartition("/")
    return f"{parent}/_rels/{name}.rels"


def _resolve_part(source_part: str, target: str, external: bool) -> str:
    try:
        return resolve_internal_relationship_target(
            source_part,
            target,
            external=external,
        )
    except OoxmlRelationshipError:
        raise _UnreadableWorkbookError() from None


def _worksheet_package_metadata(
    archive: zipfile.ZipFile,
    members,
    budget: _ActualByteBudget,
    max_worksheets: int,
    max_cells: int,
) -> tuple[bool, ...]:
    content_types = _read_member(archive, members[_CONTENT_TYPES_PART], budget)
    content_type_map = _content_type_map(content_types)

    workbook_xml = _read_member(archive, members[_WORKBOOK_PART], budget)
    relationship_ids, spreadsheet_namespace = _workbook_sheet_relationship_ids(
        workbook_xml,
        max_worksheets,
    )
    normalize_strict = spreadsheet_namespace == _STRICT_SPREADSHEET_NAMESPACE
    loader_parts = {
        _WORKBOOK_PART: (
            _normalized_strict_xml(workbook_xml)
            if normalize_strict
            else workbook_xml
        ),
    }
    content_types = b""
    workbook_xml = b""
    if _WORKBOOK_RELS_PART not in members:
        raise _UnreadableWorkbookError()
    rels_xml = _read_member(archive, members[_WORKBOOK_RELS_PART], budget)
    workbook_relationships = _relationships(rels_xml, relationship_ids)
    loader_parts[_WORKBOOK_RELS_PART] = (
        (
            _normalized_strict_relationships_xml(rels_xml)
        )
        if normalize_strict
        else rels_xml
    )

    worksheet_parts = []
    for relationship_id in relationship_ids:
        relationship = workbook_relationships.get(relationship_id)
        if relationship is None:
            raise _UnreadableWorkbookError()
        relationship_type, target, external = relationship
        if (
            external
            or relationship_type
            != _WORKSHEET_RELATIONSHIP_TYPE_BY_SPREADSHEET_NAMESPACE[
                spreadsheet_namespace
            ]
        ):
            raise _UnreadableWorkbookError()
        worksheet_part = _resolve_part(_WORKBOOK_PART, target, external)
        if (
            not worksheet_part.startswith("xl/worksheets/")
            or not worksheet_part.endswith(".xml")
            or worksheet_part not in members
            or worksheet_part in worksheet_parts
            or _part_content_type(worksheet_part, content_type_map)
            != _WORKSHEET_CONTENT_TYPE
        ):
            raise _UnreadableWorkbookError()
        worksheet_parts.append(worksheet_part)
    workbook_relationships.clear()

    metadata = []
    cell_budget = _CellElementBudget(max_cells)
    for worksheet_part in worksheet_parts:
        worksheet_xml = _read_member(archive, members[worksheet_part], budget)
        drawing_ids, cell_bound = _worksheet_metadata(
            worksheet_xml,
            cell_budget,
            spreadsheet_namespace,
        )
        loader_parts[worksheet_part] = (
            _normalized_strict_xml(worksheet_xml)
            if normalize_strict
            else worksheet_xml
        )
        worksheet_xml = b""
        if not drawing_ids:
            metadata.append((False, cell_bound))
            continue
        rels_part = _relationship_part(worksheet_part)
        if rels_part not in members:
            raise _UnreadableWorkbookError()
        sheet_rels_xml = _read_member(archive, members[rels_part], budget)
        sheet_relationships = _relationships(sheet_rels_xml, drawing_ids)
        sheet_rels_xml = b""
        for drawing_id in drawing_ids:
            relationship = sheet_relationships.get(drawing_id)
            if relationship is None:
                raise _UnreadableWorkbookError()
            relationship_type, target, external = relationship
            if (
                external
                or relationship_type
                != _DRAWING_RELATIONSHIP_TYPE_BY_SPREADSHEET_NAMESPACE[
                    spreadsheet_namespace
                ]
            ):
                raise _UnreadableWorkbookError()
            drawing_part = _resolve_part(worksheet_part, target, external)
            if (
                drawing_part not in members
                or _part_content_type(drawing_part, content_type_map)
                != _DRAWING_CONTENT_TYPE
            ):
                raise _UnreadableWorkbookError()
            drawing_part = ""
        sheet_relationships.clear()
        metadata.append((True, cell_bound))
    worksheet_parts.clear()

    auxiliary_parts = (
        ("xl/styles.xml", _STYLES_CONTENT_TYPE, "styleSheet"),
        ("xl/sharedStrings.xml", _SHARED_STRINGS_CONTENT_TYPE, "sst"),
    )
    for part_name, content_type, expected_root in auxiliary_parts:
        member_exists = part_name in members
        exact_override = (
            content_type_map[1].get(f"/{part_name}") == content_type
        )
        if member_exists != exact_override:
            raise _UnreadableWorkbookError()
        if not member_exists:
            continue
        auxiliary_xml = _read_member(archive, members[part_name], budget)
        root_seen = False
        for event, element in _safe_xml_events(auxiliary_xml):
            if event == "start" and not root_seen:
                if _expanded_name(element.tag) != (
                    spreadsheet_namespace,
                    expected_root,
                ):
                    raise _UnreadableWorkbookError()
                root_seen = True
            elif event == "start" and (
                _expanded_name(element.tag) is None
                or _expanded_name(element.tag)[0] != spreadsheet_namespace
            ):
                raise _UnreadableWorkbookError()
            if event == "end":
                element.clear()
        if not root_seen:
            raise _UnreadableWorkbookError()
        loader_parts[part_name] = (
            _normalized_strict_xml(auxiliary_xml)
            if normalize_strict
            else auxiliary_xml
        )
        auxiliary_xml = b""
    loader_parts[_CONTENT_TYPES_PART] = _loader_content_types(
        loader_parts,
        content_type_map[1]["/xl/workbook.xml"],
    )
    loader_snapshot, loader_declared_total = _loader_snapshot(loader_parts)
    content_type_map[0].clear()
    content_type_map[1].clear()
    return (
        tuple(metadata),
        spreadsheet_namespace,
        loader_snapshot,
        loader_declared_total,
    )


def _preflight(snapshot: bytes, limits: InspectionLimits):
    expected_entries = _validate_eocd(snapshot)
    try:
        with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
            members, _declared_total = _central_directory(archive, expected_entries)
            budget = _ActualByteBudget()
            (
                metadata,
                spreadsheet_namespace,
                loader_snapshot,
                loader_declared_total,
            ) = _worksheet_package_metadata(
                archive,
                members,
                budget,
                limits.max_worksheets_per_workbook,
                limits.max_cells_per_workbook,
            )
            budget.reserve(loader_declared_total * 2)
            return metadata, spreadsheet_namespace, loader_snapshot
    except (WorkbookParserBoundaryExceededError, WorkbookWorksheetCountExceededError):
        raise
    except (_UnreadableWorkbookError, _EncryptedWorkbookError):
        raise
    except Exception:
        raise _UnreadableWorkbookError() from None


def _private_scalar_text(value: object, character_limit: int):
    if value is None:
        return "", False
    if isinstance(value, str):
        text = value
    elif isinstance(value, (bool, int, float, date)):
        try:
            text = str(value)
        except Exception:
            return "", True
    else:
        return "", True
    if len(text) > character_limit:
        text = ""
        return "", True
    return text[:character_limit], False


def _canonical_signals(signal_set) -> tuple[str, ...]:
    return tuple(code for code in SIGNAL_ORDER if code in signal_set)


def _structural_signals(*, embedded_media: bool, worksheet_hidden: bool):
    signals = set()
    if embedded_media:
        signals.add("embedded-media-present")
    if worksheet_hidden:
        signals.add("worksheet-hidden")
    return _canonical_signals(signals)


def _unit_issues(
    *, over_limit: bool, embedded_media: bool, worksheet_hidden: bool
) -> tuple[str, ...]:
    issues = set()
    if over_limit:
        issues.add("unit-over-limit")
    if embedded_media:
        issues.add("embedded-media-present")
    if worksheet_hidden:
        issues.add("worksheet-hidden")
    return tuple(
        code
        for code in (
            "unit-over-limit",
            "embedded-media-present",
            "worksheet-hidden",
        )
        if code in issues
    )


def _worksheet_cell_bound(worksheet) -> int | None:
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    if max_row is None or max_column is None:
        return None
    if (
        not isinstance(max_row, int)
        or isinstance(max_row, bool)
        or not isinstance(max_column, int)
        or isinstance(max_column, bool)
        or max_row < 0
        or max_column < 0
    ):
        raise _UnreadableWorkbookError()
    return max_row * max_column


def _inspect_worksheet(
    worksheet,
    unit_index: int,
    *,
    embedded_media: bool,
    expected_cell_bound: int | None,
    remaining_cells: int,
    character_limit: int,
):
    state = worksheet.sheet_state
    if state not in {"visible", "hidden", "veryHidden"}:
        raise _UnreadableWorkbookError()
    worksheet_hidden = state != "visible"
    runtime_cell_bound = _worksheet_cell_bound(worksheet)
    cell_bound = (
        runtime_cell_bound
        if expected_cell_bound is not None and runtime_cell_bound == expected_cell_bound
        else None
    )
    forced_over_limit = cell_bound is None or cell_bound > remaining_cells
    if forced_over_limit:
        return (
            InspectionUnitEvidence(
                "worksheet",
                unit_index,
                "none",
                _structural_signals(
                    embedded_media=embedded_media,
                    worksheet_hidden=worksheet_hidden,
                ),
                _unit_issues(
                    over_limit=True,
                    embedded_media=embedded_media,
                    worksheet_hidden=worksheet_hidden,
                ),
            ),
            0,
        )
    quota = cell_bound
    consumed = 0
    text_signals = set()
    roster_column_pattern = False
    roster_identity_or_name_columns: set[int] = set()
    roster_payment_columns: set[int] = set()
    qualifying_data_rows = 0
    row_pattern = False
    scalar_over_limit = False

    if quota:
        for row in worksheet.iter_rows():
            if consumed >= quota:
                break
            populated_columns = set()
            row_signals = set()
            row_header_categories: dict[int, tuple[str, ...]] = {}
            for column_index, cell in enumerate(row, start=1):
                if consumed >= quota:
                    break
                consumed += 1
                private_text, truncated = _private_scalar_text(
                    cell.value,
                    character_limit,
                )
                if truncated:
                    scalar_over_limit = True
                    private_text = ""
                    break
                if private_text:
                    populated_columns.add(column_index)
                    header_categories = roster_header_categories_from_private_text(
                        private_text
                    )
                    if header_categories:
                        row_header_categories[column_index] = header_categories
                    cell_signals = signals_from_private_text(
                        private_text,
                        TextSignalContext(
                            "worksheet",
                            mostly_image=False,
                            embedded_media=False,
                            worksheet_hidden=False,
                            row_pattern=False,
                        ),
                    )
                    private_text = ""
                    row_signals.update(cell_signals)
            if scalar_over_limit:
                break
            payment_columns = {
                column_index
                for column_index, categories in row_header_categories.items()
                if "payment" in categories
            }
            identity_or_name_columns = {
                column_index
                for column_index, categories in row_header_categories.items()
                if "identity" in categories or "name" in categories
            }
            if payment_columns and identity_or_name_columns:
                roster_column_pattern = True
                roster_payment_columns = payment_columns
                roster_identity_or_name_columns = identity_or_name_columns
                qualifying_data_rows = 0
            elif roster_column_pattern and (
                populated_columns & roster_payment_columns
                and populated_columns & roster_identity_or_name_columns
            ):
                qualifying_data_rows += 1
                row_pattern = qualifying_data_rows >= 2
            text_signals.update(row_signals)

    over_limit = scalar_over_limit
    if over_limit:
        return (
            InspectionUnitEvidence(
                "worksheet",
                unit_index,
                "none",
                _structural_signals(
                    embedded_media=embedded_media,
                    worksheet_hidden=worksheet_hidden,
                ),
                _unit_issues(
                    over_limit=True,
                    embedded_media=embedded_media,
                    worksheet_hidden=worksheet_hidden,
                ),
            ),
            consumed,
        )

    structural_signals = signals_from_private_text(
        "",
        TextSignalContext(
            "worksheet",
            mostly_image=False,
            embedded_media=embedded_media,
            worksheet_hidden=worksheet_hidden,
            row_pattern=row_pattern,
            roster_column_pattern=roster_column_pattern,
        ),
    )
    text_signals.update(structural_signals)
    return (
        InspectionUnitEvidence(
            "worksheet",
            unit_index,
            "worksheet-structure",
            _canonical_signals(text_signals),
            _unit_issues(
                over_limit=False,
                embedded_media=embedded_media,
                worksheet_hidden=worksheet_hidden,
            ),
        ),
        consumed,
    )


def inspect_workbook(
    snapshot: bytes,
    *,
    limits: InspectionLimits,
    remaining_units: int | None = None,
) -> InspectionAdapterResult:
    """Inspect each actual worksheet from one immutable in-memory snapshot."""
    if type(snapshot) is not bytes:
        raise TypeError("inspection snapshot must be bytes")
    if type(limits) is not InspectionLimits:
        raise TypeError("inspection limits must be valid")
    if remaining_units is None:
        remaining_units = limits.max_units
    if (
        type(remaining_units) is not int
        or remaining_units < 0
        or remaining_units > limits.max_units
    ):
        raise ValueError("remaining_units must be within the inspection unit budget")
    if len(snapshot) > limits.max_workbook_source_bytes:
        return _source_problem("over-limit", "document-over-limit")
    if snapshot.startswith(_OLE_COMPOUND_HEADER):
        return _source_problem("encrypted", "document-encrypted")

    try:
        (
            worksheet_metadata,
            spreadsheet_namespace,
            loader_snapshot,
        ) = _preflight(snapshot, limits)
    except (WorkbookParserBoundaryExceededError, WorkbookWorksheetCountExceededError):
        raise
    except _EncryptedWorkbookError:
        return _source_problem("encrypted", "document-encrypted")
    except Exception:
        return _source_problem("unreadable", "document-unreadable")

    if len(worksheet_metadata) > remaining_units:
        raise InspectionUnitCountExceededError()

    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            BytesIO(loader_snapshot),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        worksheets = workbook.worksheets
        worksheet_count = len(worksheets)
        if worksheet_count > limits.max_worksheets_per_workbook:
            raise WorkbookWorksheetCountExceededError()
        if worksheet_count != len(worksheet_metadata):
            raise _UnreadableWorkbookError()

        units = []
        remaining_cells = limits.max_cells_per_workbook
        for unit_index, (worksheet, sheet_metadata) in enumerate(
            zip(worksheets, worksheet_metadata),
            start=1,
        ):
            embedded_media, expected_cell_bound = sheet_metadata
            try:
                worksheet_hidden = worksheet.sheet_state in {"hidden", "veryHidden"}
            except Exception:
                worksheet_hidden = False
            try:
                unit, consumed = _inspect_worksheet(
                    worksheet,
                    unit_index,
                    embedded_media=embedded_media,
                    expected_cell_bound=expected_cell_bound,
                    remaining_cells=remaining_cells,
                    character_limit=limits.max_cell_text_characters,
                )
            except Exception:
                unit = InspectionUnitEvidence(
                    "worksheet",
                    unit_index,
                    "none",
                    _structural_signals(
                        embedded_media=embedded_media,
                        worksheet_hidden=worksheet_hidden,
                    ),
                    _unit_issues(
                        over_limit=True,
                        embedded_media=embedded_media,
                        worksheet_hidden=worksheet_hidden,
                    ),
                )
                consumed = remaining_cells
            remaining_cells -= consumed
            units.append(unit)
        return InspectionAdapterResult(
            "inspected",
            worksheet_count,
            (),
            tuple(units),
        )
    except (
        InspectionUnitCountExceededError,
        WorkbookParserBoundaryExceededError,
        WorkbookWorksheetCountExceededError,
    ):
        raise
    except Exception:
        return _source_problem("unreadable", "document-unreadable")
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def worksheet_preview(
    snapshot: bytes,
    unit_index: int,
    *,
    limits: InspectionLimits = DEFAULT_INSPECTION_LIMITS,
) -> dict:
    """Return one proved worksheet's bounded local-only cell preview."""
    if type(snapshot) is not bytes or type(limits) is not InspectionLimits:
        raise TypeError("preview input must use bounded snapshot bytes and limits")
    if (
        type(unit_index) is not int
        or not 1 <= unit_index <= limits.max_worksheets_per_workbook
    ):
        raise WorkbookPreviewError("preview-unavailable")
    if len(snapshot) > limits.max_workbook_source_bytes:
        raise WorkbookPreviewError("preview-over-limit")
    if snapshot.startswith(_OLE_COMPOUND_HEADER):
        raise WorkbookPreviewError("preview-unavailable")
    try:
        worksheet_metadata, _spreadsheet_namespace, loader_snapshot = _preflight(
            snapshot, limits
        )
    except (WorkbookParserBoundaryExceededError, WorkbookWorksheetCountExceededError):
        raise WorkbookPreviewError("preview-parser-boundary-exceeded") from None
    except Exception:
        raise WorkbookPreviewError("preview-unavailable") from None
    if unit_index > len(worksheet_metadata):
        raise WorkbookPreviewError("preview-unavailable")

    workbook = None
    try:
        workbook = openpyxl.load_workbook(
            BytesIO(loader_snapshot),
            read_only=True,
            data_only=False,
            keep_links=False,
        )
        worksheets = workbook.worksheets
        if len(worksheets) != len(worksheet_metadata) or unit_index > len(worksheets):
            raise WorkbookPreviewError("preview-unavailable")
        worksheet = worksheets[unit_index - 1]
        max_row = worksheet.max_row
        max_column = worksheet.max_column
        if (
            type(max_row) is not int
            or type(max_column) is not int
            or max_row < 0
            or max_column < 0
            or max_row * max_column > limits.max_cells_per_workbook
        ):
            raise WorkbookPreviewError("preview-over-limit")
        row_limit = min(max_row, 200)
        column_limit = min(max_column, 50)
        rows = []
        if row_limit and column_limit:
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=row_limit,
                min_col=1,
                max_col=column_limit,
            ):
                values = []
                for cell in row:
                    text, over_limit = _private_scalar_text(cell.value, 256)
                    if over_limit:
                        raise WorkbookPreviewError("preview-over-limit")
                    values.append(text)
                rows.append(values)
        return {
            "rows": rows,
            "rowCount": len(rows),
            "columnCount": max((len(row) for row in rows), default=0),
            "truncated": max_row > 200 or max_column > 50,
        }
    except WorkbookPreviewError:
        raise
    except Exception:
        raise WorkbookPreviewError("preview-unavailable") from None
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def _package_scalar(value: object, character_limit: int) -> object:
    if value is None or type(value) in {bool, int}:
        return value
    if type(value) is float:
        if not math.isfinite(value):
            raise PackageWorkbookError("package-workbook-over-limit")
        return value
    if type(value) is str:
        if len(value) > character_limit:
            raise PackageWorkbookError("package-workbook-over-limit")
        return value
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(timespec="microseconds").rstrip("0").rstrip(".")
    if isinstance(value, date):
        return value.isoformat()
    raise PackageWorkbookError("package-workbook-unavailable")


def _trim_package_rows(rows: list[list[object]]) -> tuple[tuple[object, ...], ...]:
    while rows and not any(value is not None for value in rows[-1]):
        rows.pop()
    width = max(
        (index + 1 for row in rows for index, value in enumerate(row) if value is not None),
        default=0,
    )
    return tuple(tuple(row[:width]) for row in rows)


def selected_worksheet_values(
    snapshot: bytes,
    worksheet_indexes: tuple[int, ...],
    *,
    limits: InspectionLimits,
) -> tuple[WorksheetValues, ...]:
    """Return bounded values-only rows after the existing OOXML preflight."""
    if type(snapshot) is not bytes or type(limits) is not InspectionLimits:
        raise TypeError("package workbook input must use bounded snapshot bytes and limits")
    if (
        type(worksheet_indexes) is not tuple
        or not worksheet_indexes
        or any(type(index) is not int for index in worksheet_indexes)
        or len(set(worksheet_indexes)) != len(worksheet_indexes)
        or any(not 1 <= index <= limits.max_worksheets_per_workbook for index in worksheet_indexes)
    ):
        raise PackageWorkbookError("package-workbook-unavailable")
    if len(snapshot) > limits.max_workbook_source_bytes:
        raise PackageWorkbookError("package-workbook-over-limit")
    if snapshot.startswith(_OLE_COMPOUND_HEADER):
        raise PackageWorkbookError("package-workbook-unavailable")
    try:
        metadata, _spreadsheet_namespace, loader_snapshot = _preflight(snapshot, limits)
    except (WorkbookParserBoundaryExceededError, WorkbookWorksheetCountExceededError):
        raise PackageWorkbookError("package-workbook-parser-boundary-exceeded") from None
    except Exception:
        raise PackageWorkbookError("package-workbook-unavailable") from None
    if any(index > len(metadata) for index in worksheet_indexes):
        raise PackageWorkbookError("package-workbook-unavailable")

    formula_book = None
    value_book = None
    try:
        formula_book = openpyxl.load_workbook(
            BytesIO(loader_snapshot), read_only=True, data_only=False, keep_links=False
        )
        value_book = openpyxl.load_workbook(
            BytesIO(loader_snapshot), read_only=True, data_only=True, keep_links=False
        )
        if (
            len(formula_book.worksheets) != len(metadata)
            or len(value_book.worksheets) != len(metadata)
        ):
            raise PackageWorkbookError("package-workbook-unavailable")
        consumed = 0
        result = []
        for worksheet_index in worksheet_indexes:
            formula_sheet = formula_book.worksheets[worksheet_index - 1]
            value_sheet = value_book.worksheets[worksheet_index - 1]
            if (
                formula_sheet.max_row != value_sheet.max_row
                or formula_sheet.max_column != value_sheet.max_column
            ):
                raise PackageWorkbookError("package-workbook-unavailable")
            rows = []
            formula_rows = formula_sheet.iter_rows()
            value_rows = value_sheet.iter_rows()
            for formula_row, value_row in zip(formula_rows, value_rows):
                if len(formula_row) != len(value_row):
                    raise PackageWorkbookError("package-workbook-unavailable")
                output_row = []
                for formula_cell, value_cell in zip(formula_row, value_row):
                    consumed += 1
                    if consumed > limits.max_cells_per_workbook:
                        raise PackageWorkbookError("package-workbook-over-limit")
                    if formula_cell.data_type == "f" and value_cell.value is None:
                        raise PackageWorkbookError("package-workbook-formula-unavailable")
                    output_row.append(
                        _package_scalar(
                            value_cell.value,
                            limits.max_cell_text_characters,
                        )
                    )
                rows.append(output_row)
            result.append(WorksheetValues(worksheet_index, _trim_package_rows(rows)))
        return tuple(result)
    except PackageWorkbookError:
        raise
    except Exception:
        raise PackageWorkbookError("package-workbook-unavailable") from None
    finally:
        for workbook in (formula_book, value_book):
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass


def _canonical_package_workbook_bytes(
    workbook: openpyxl.Workbook,
    *,
    max_bytes: int,
) -> bytes:
    """Privately repack one generated workbook with fixed archive metadata."""
    raw = _CappedBytesIO(max_bytes)
    try:
        workbook.save(raw)
    except _OutputLimitExceeded:
        raise PackageWorkbookError("package-workbook-over-limit") from None
    except Exception:
        raise PackageWorkbookError("package-workbook-unavailable") from None
    finally:
        try:
            workbook.close()
        except Exception:
            pass
    output = _CappedBytesIO(max_bytes)
    try:
        raw.seek(0)
        with zipfile.ZipFile(raw, "r") as source:
            names = sorted(source.namelist())
            if len(names) != len(set(names)):
                raise PackageWorkbookError("package-workbook-unavailable")
            with zipfile.ZipFile(
                output,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                compresslevel=9,
                strict_timestamps=True,
            ) as target:
                for name in names:
                    if name == "docProps/core.xml":
                        content = source.read(name)
                        opening = b"<dcterms:modified"
                        closing = b"</dcterms:modified>"
                        start = content.find(opening)
                        value_start = content.find(b">", start) + 1
                        value_end = content.find(closing, value_start)
                        if (
                            start < 0
                            or value_start <= 0
                            or value_end < value_start
                            or content.find(opening, start + 1) >= 0
                        ):
                            raise PackageWorkbookError("package-workbook-unavailable")
                        content = (
                            content[:value_start]
                            + b"1980-01-01T00:00:00Z"
                            + content[value_end:]
                        )
                    info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
                    info.compress_type = zipfile.ZIP_DEFLATED
                    info.create_system = 3
                    info.external_attr = 0o100644 << 16
                    if name == "docProps/core.xml":
                        target.writestr(
                            info,
                            content,
                            compress_type=zipfile.ZIP_DEFLATED,
                            compresslevel=9,
                        )
                    else:
                        with source.open(name, "r") as reader:
                            with target.open(info, "w", force_zip64=False) as writer:
                                while True:
                                    chunk = reader.read(_READ_CHUNK_BYTES)
                                    if not chunk:
                                        break
                                    writer.write(chunk)
        return output.getvalue()
    except _OutputLimitExceeded:
        raise PackageWorkbookError("package-workbook-over-limit") from None
    except PackageWorkbookError:
        raise
    except Exception:
        raise PackageWorkbookError("package-workbook-unavailable") from None
