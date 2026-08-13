import ast
import builtins
from collections import Counter
from datetime import date
from io import BytesIO
from pathlib import Path
import socket
import struct
import tempfile
import zipfile

from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorkbookImage
from PIL import Image
import pytest

from ctv_inspection_model import InspectionLimits


PRIVATE_SHEET_NAMES = (
    "Bảng kê CTV riêng 079123456789",
    "Hỗ trợ nội bộ có dấu",
    "Ảnh đính kèm bí mật",
)
PRIVATE_CELL_VALUES = (
    "NGUYEN VAN KIEM THU 079123456789",
    "=SUM(C3:C4)+987654321",
    "13/08/2026",
    "1250000",
)
PRIVATE_MEMBER_NAME = "xl/media/private-identity-079123456789.png"
PRIVATE_EXTERNAL_VALUE = "EXTERNAL-PRIVATE-079123456789"
TRANSITIONAL_SPREADSHEET_NAMESPACE = (
    b"http://schemas.openxmlformats.org/spreadsheetml/2006/main"
)
STRICT_SPREADSHEET_NAMESPACE = b"http://purl.oclc.org/ooxml/spreadsheetml/main"
TRANSITIONAL_OFFICE_RELATIONSHIPS = (
    b"http://schemas.openxmlformats.org/officeDocument/2006/relationships"
)
STRICT_OFFICE_RELATIONSHIPS = (
    b"http://purl.oclc.org/ooxml/officeDocument/relationships"
)


def _save(workbook):
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _workbook_with_sheets(sheet_specs):
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, state, rows in sheet_specs:
        sheet = workbook.create_sheet(name)
        sheet.sheet_state = state
        for row in rows:
            sheet.append(row)
    return workbook


def _synthetic_png():
    stream = BytesIO()
    with Image.new("RGB", (2, 2), (12, 34, 56)) as image:
        image.save(stream, format="PNG")
    stream.seek(0)
    return stream


def _roster_workbook(*, with_image=False):
    workbook = _workbook_with_sheets(
        (
            (
                PRIVATE_SHEET_NAMES[0],
                "visible",
                (
                    ("DANH SACH CHI TRA", None, None, None),
                    ("Ho ten", "Ma so nhan vien", "So tien", "Cong thuc"),
                    (
                        PRIVATE_CELL_VALUES[0],
                        "CTV-001",
                        1_250_000,
                        PRIVATE_CELL_VALUES[1],
                    ),
                    ("Nguoi thu hai", "CTV-002", date(2026, 8, 13), 250_000),
                ),
            ),
            (
                PRIVATE_SHEET_NAMES[1],
                "hidden",
                (("MA SO NHAN VIEN",),),
            ),
            (
                PRIVATE_SHEET_NAMES[2],
                "veryHidden",
                (("TAI LIEU KEM THEO", PRIVATE_CELL_VALUES[2], PRIVATE_CELL_VALUES[3]),),
            ),
        )
    )
    if with_image:
        workbook.worksheets[2].add_image(WorkbookImage(_synthetic_png()), "D4")
    return workbook


def _inspect(snapshot, *, limits=None):
    from ctv_inspection_workbook import inspect_workbook

    return inspect_workbook(snapshot, limits=limits or InspectionLimits())


def _rewrite_package(snapshot, *, replacements=None, additions=None, drop=()):
    replacements = replacements or {}
    additions = additions or {}
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(snapshot), "r") as source:
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for info in source.infolist():
                if info.filename in drop:
                    continue
                content = source.read(info)
                target.writestr(info.filename, replacements.get(info.filename, content))
            for name, content in additions.items():
                target.writestr(name, content)
    return output.getvalue()


def _rewrite_package_with_stored_member(snapshot, member_name, content):
    output = BytesIO()
    with zipfile.ZipFile(BytesIO(snapshot), "r") as source:
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as target:
            for info in source.infolist():
                member_content = content if info.filename == member_name else source.read(info)
                compression = (
                    zipfile.ZIP_STORED
                    if info.filename == member_name
                    else zipfile.ZIP_DEFLATED
                )
                target.writestr(
                    info.filename,
                    member_content,
                    compress_type=compression,
                )
    return output.getvalue()


def _strict_package(snapshot):
    replacements = {}
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        for info in archive.infolist():
            if (
                info.filename in {
                    "xl/workbook.xml",
                    "xl/styles.xml",
                    "xl/sharedStrings.xml",
                }
                or (
                    info.filename.startswith("xl/worksheets/")
                    and info.filename.endswith(".xml")
                )
            ):
                content = archive.read(info)
                content = content.replace(
                    TRANSITIONAL_SPREADSHEET_NAMESPACE,
                    STRICT_SPREADSHEET_NAMESPACE,
                ).replace(
                    TRANSITIONAL_OFFICE_RELATIONSHIPS,
                    STRICT_OFFICE_RELATIONSHIPS,
                )
                if info.filename == "xl/workbook.xml":
                    content = content.replace(
                        b"<workbook ",
                        b'<workbook conformance="strict" ',
                        1,
                    )
                replacements[info.filename] = content
            elif info.filename.endswith(".rels"):
                content = archive.read(info).replace(
                    TRANSITIONAL_OFFICE_RELATIONSHIPS,
                    STRICT_OFFICE_RELATIONSHIPS,
                )
                replacements[info.filename] = content
    return _rewrite_package(snapshot, replacements=replacements)


def _encoded_xml(content, encoding):
    text = content.decode("utf-8")
    if text.startswith("<?xml"):
        text = text[text.index("?>") + 2:]
    declaration = '<?xml version="1.0" encoding="UTF-16"?>'
    byte_order_mark = {
        "utf-16-le": b"\xff\xfe",
        "utf-16-be": b"\xfe\xff",
    }[encoding]
    return byte_order_mark + (declaration + text).encode(encoding)


def _patch_central_sizes(snapshot, patches):
    data = bytearray(snapshot)
    cursor = 0
    seen = set()
    signature = b"PK\x01\x02"
    while True:
        cursor = data.find(signature, cursor)
        if cursor < 0:
            break
        name_length, extra_length, comment_length = struct.unpack_from(
            "<HHH", data, cursor + 28
        )
        name_start = cursor + 46
        name = bytes(data[name_start:name_start + name_length]).decode("utf-8")
        if name in patches:
            compressed_size, uncompressed_size = patches[name]
            struct.pack_into("<II", data, cursor + 20, compressed_size, uncompressed_size)
            seen.add(name)
        cursor = name_start + name_length + extra_length + comment_length
    assert seen == set(patches)
    return bytes(data)


def _patch_encrypted_flag(snapshot, member_name):
    data = bytearray(snapshot)
    central = data.find(b"PK\x01\x02")
    patched_local = False
    patched_central = False
    while central >= 0:
        name_length, extra_length, comment_length = struct.unpack_from(
            "<HHH", data, central + 28
        )
        name_start = central + 46
        name = bytes(data[name_start:name_start + name_length]).decode("utf-8")
        if name == member_name:
            flag = struct.unpack_from("<H", data, central + 8)[0] | 1
            struct.pack_into("<H", data, central + 8, flag)
            local_offset = struct.unpack_from("<I", data, central + 42)[0]
            local_flag = struct.unpack_from("<H", data, local_offset + 6)[0] | 1
            struct.pack_into("<H", data, local_offset + 6, local_flag)
            patched_local = patched_central = True
            break
        central = data.find(
            b"PK\x01\x02",
            name_start + name_length + extra_length + comment_length,
        )
    assert patched_local and patched_central
    return bytes(data)


def _assert_private_values_absent(value, *extra_fragments):
    public = f"{value!s}\n{value!r}"
    forbidden = (
        *PRIVATE_SHEET_NAMES,
        *PRIVATE_CELL_VALUES,
        PRIVATE_MEMBER_NAME,
        PRIVATE_EXTERNAL_VALUE,
        "079123456789",
        "987654321",
        "1250000",
        "13/08/2026",
        *extra_fragments,
    )
    assert not any(fragment in public for fragment in forbidden)


def test_workbook_emits_one_ordered_unit_for_every_sheet_with_fixed_signals_only():
    snapshot = _save(_roster_workbook(with_image=True))

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3
    assert result.source_issue_codes == ()
    assert [unit.unit_index for unit in result.units] == [1, 2, 3]
    assert [unit.unit_kind for unit in result.units] == ["worksheet"] * 3
    assert [unit.inspection_method for unit in result.units] == [
        "worksheet-structure",
        "worksheet-structure",
        "worksheet-structure",
    ]
    assert result.units[0].signal_codes == (
        "roster-column-pattern",
        "roster-row-pattern",
        "identity-number-pattern-present",
        "mostly-text-page",
    )
    assert result.units[0].issue_codes == ()
    assert result.units[1].signal_codes == (
        "roster-column-pattern",
        "worksheet-hidden",
        "mostly-text-page",
    )
    assert result.units[1].issue_codes == ("worksheet-hidden",)
    assert result.units[2].signal_codes == (
        "supporting-document-heading",
        "embedded-media-present",
        "worksheet-hidden",
        "mostly-text-page",
    )
    assert result.units[2].issue_codes == (
        "embedded-media-present",
        "worksheet-hidden",
    )
    _assert_private_values_absent(result)


def test_workbook_passes_only_bytesio_and_exact_safe_openpyxl_options(monkeypatch):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook())
    real_loader = workbook_adapter.openpyxl.load_workbook
    calls = []

    def recording_loader(source, **kwargs):
        calls.append((source, kwargs))
        return real_loader(source, **kwargs)

    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", recording_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert len(calls) == 1
    assert isinstance(calls[0][0], BytesIO)
    with zipfile.ZipFile(BytesIO(calls[0][0].getvalue()), "r") as loader_archive:
        assert set(loader_archive.namelist()) == {
            "[Content_Types].xml",
            "xl/workbook.xml",
            "xl/_rels/workbook.xml.rels",
            "xl/worksheets/sheet1.xml",
            "xl/worksheets/sheet2.xml",
            "xl/worksheets/sheet3.xml",
            "xl/styles.xml",
        }
    assert calls[0][1] == {
        "read_only": True,
        "data_only": False,
        "keep_links": False,
    }


@pytest.mark.parametrize("sheet_count", [100, 101])
def test_actual_worksheet_count_has_exact_operation_boundary(sheet_count):
    workbook = Workbook()
    workbook.active.title = "sheet-000"
    for index in range(1, sheet_count):
        workbook.create_sheet(f"sheet-{index:03d}")
    snapshot = _save(workbook)

    if sheet_count == 100:
        result = _inspect(snapshot)
        assert result.inspection_status == "inspected"
        assert result.unit_count == 100
        assert [unit.unit_index for unit in result.units] == list(range(1, 101))
        return

    from ctv_inspection_workbook import WorkbookWorksheetCountExceededError

    with pytest.raises(WorkbookWorksheetCountExceededError) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-worksheet-count-exceeded"
    _assert_private_values_absent(raised.value)


def test_global_100000_cell_budget_keeps_later_sheets_as_known_unknown_units():
    workbook = Workbook()
    first = workbook.active
    first.title = "private-budget-first"
    first["A1"] = "DANH SACH CHI TRA"
    first["A100000"] = "bounded-tail"
    second = workbook.create_sheet("private-budget-second")
    second["A1"] = "TAI LIEU KEM THEO"
    snapshot = _save(workbook)

    result = _inspect(snapshot)

    assert result.unit_count == 2
    assert result.units[0].inspection_method == "worksheet-structure"
    assert result.units[0].issue_codes == ()
    assert result.units[1].inspection_method == "none"
    assert result.units[1].signal_codes == ()
    assert result.units[1].issue_codes == ("unit-over-limit",)
    _assert_private_values_absent(result, "private-budget-first", "private-budget-second")


@pytest.mark.parametrize("character_count", [256, 257])
def test_per_cell_text_boundary_is_exact_and_never_returns_a_prefix(character_count):
    marker = "DANH SACH CHI TRA"
    private_value = marker + " " + "Z" * (character_count - len(marker) - 1)
    workbook = Workbook()
    workbook.active["A1"] = private_value
    snapshot = _save(workbook)

    result = _inspect(snapshot)

    unit = result.units[0]
    if character_count == 256:
        assert unit.inspection_method == "worksheet-structure"
        assert "roster-column-pattern" in unit.signal_codes
        assert unit.issue_codes == ()
    else:
        assert unit.inspection_method == "none"
        assert unit.signal_codes == ()
        assert unit.issue_codes == ("unit-over-limit",)
    assert private_value not in repr(result)


def test_workbook_source_byte_limit_is_inclusive_and_source_only_when_exceeded():
    snapshot = _save(_roster_workbook())

    accepted = _inspect(
        snapshot,
        limits=InspectionLimits(max_workbook_source_bytes=len(snapshot)),
    )
    rejected = _inspect(
        snapshot,
        limits=InspectionLimits(max_workbook_source_bytes=len(snapshot) - 1),
    )

    assert accepted.inspection_status == "inspected"
    assert rejected.inspection_status == "over-limit"
    assert rejected.unit_count is None
    assert rejected.source_issue_codes == ("document-over-limit",)
    assert rejected.units == ()


@pytest.mark.parametrize(
    "snapshot",
    (
        bytearray(b"private-mutable-079123456789"),
        memoryview(b"private-view-079123456789"),
        "private-text-079123456789",
    ),
)
def test_workbook_accepts_only_exact_immutable_bytes(snapshot):
    with pytest.raises(TypeError) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection snapshot must be bytes"
    _assert_private_values_absent(
        raised.value,
        "private-mutable",
        "private-view",
        "private-text",
    )


@pytest.mark.parametrize(
    ("snapshot", "status", "issue"),
    (
        (b"not-a-private-workbook-079123456789", "unreadable", "document-unreadable"),
        (
            b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1encrypted-like-private",
            "encrypted",
            "document-encrypted",
        ),
    ),
)
def test_corrupt_and_encrypted_like_sources_return_only_safe_source_status(
    snapshot, status, issue
):
    result = _inspect(snapshot)

    assert result.inspection_status == status
    assert result.unit_count is None
    assert result.source_issue_codes == (issue,)
    assert result.units == ()
    _assert_private_values_absent(result)


def test_zip_encryption_flag_returns_encrypted_without_opening_members(monkeypatch):
    snapshot = _patch_encrypted_flag(
        _save(_roster_workbook()), "[Content_Types].xml"
    )
    opened = []
    original_open = zipfile.ZipFile.open

    def recording_open(self, name, *args, **kwargs):
        opened.append(name)
        return original_open(self, name, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", recording_open)

    result = _inspect(snapshot)

    assert result.inspection_status == "encrypted"
    assert result.source_issue_codes == ("document-encrypted",)
    assert opened == []


def test_generic_zip_and_missing_required_member_are_not_treated_as_workbooks():
    generic = BytesIO()
    with zipfile.ZipFile(generic, "w") as archive:
        archive.writestr("private-member-079123456789.txt", b"private bytes")
    missing = _rewrite_package(
        _save(_roster_workbook()), drop=("xl/workbook.xml",)
    )

    for snapshot in (generic.getvalue(), missing):
        result = _inspect(snapshot)
        assert result.inspection_status == "unreadable"
        assert result.unit_count is None
        assert result.source_issue_codes == ("document-unreadable",)
        assert result.units == ()
        _assert_private_values_absent(result, "private-member-079123456789.txt")


@pytest.mark.parametrize("declaration", [b"<!DOCTYPE workbook>", b"<!ENTITY private 'x'>"])
def test_malformed_or_entity_declaring_ooxml_is_safely_unreadable(declaration):
    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
    mutated = declaration + workbook_xml + b"PRIVATE-PARSER-TOKEN-079123456789"
    snapshot = _rewrite_package(
        snapshot,
        replacements={"xl/workbook.xml": mutated},
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "unreadable"
    assert result.source_issue_codes == ("document-unreadable",)
    _assert_private_values_absent(result, "PRIVATE-PARSER-TOKEN-079123456789")


@pytest.mark.parametrize(
    ("encoding", "byte_order_mark"),
    (
        ("utf-16-le", b"\xff\xfe"),
        ("utf-16-be", b"\xfe\xff"),
        ("utf-32-le", b"\xff\xfe\x00\x00"),
        ("utf-32-be", b"\x00\x00\xfe\xff"),
    ),
)
def test_encoded_dtd_is_rejected_before_elementtree(
    monkeypatch, encoding, byte_order_mark
):
    import ctv_inspection_workbook as workbook_adapter

    declaration = (
        '<?xml version="1.0"?><!DOCTYPE workbook '
        '[<!ENTITY private "PRIVATE-ENTITY-079123456789">]>'
        '<workbook xmlns="http://schemas.openxmlformats.org/'
        'spreadsheetml/2006/main">&private;</workbook>'
    )
    content = byte_order_mark + declaration.encode(encoding)
    parser_calls = []

    def forbidden_parser(*args, **kwargs):
        parser_calls.append((args, kwargs))
        raise AssertionError("declaration scan must run before ElementTree")

    monkeypatch.setattr(
        workbook_adapter.ElementTree,
        "XMLPullParser",
        forbidden_parser,
    )

    with pytest.raises(workbook_adapter._UnreadableWorkbookError):
        list(workbook_adapter._safe_xml_events(content))
    assert parser_calls == []


def test_xml_pull_parser_is_fed_only_bounded_chunks(monkeypatch):
    import ctv_inspection_workbook as workbook_adapter

    workbook = Workbook()
    sheet = workbook.active
    bounded_value = "Z" * 256
    for row in range(1, 401):
        sheet.cell(row=row, column=1, value=bounded_value)
    snapshot = _save(workbook)
    original_parser = workbook_adapter.ElementTree.XMLPullParser
    feed_sizes = []

    class TrackingPullParser:
        def __init__(self, *args, **kwargs):
            self.parser = original_parser(*args, **kwargs)

        def feed(self, data):
            feed_sizes.append(len(data))
            return self.parser.feed(data)

        def read_events(self):
            return self.parser.read_events()

        def close(self):
            return self.parser.close()

    monkeypatch.setattr(
        workbook_adapter.ElementTree,
        "XMLPullParser",
        TrackingPullParser,
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert feed_sizes
    assert max(feed_sizes) <= 64 * 1024


def test_missing_worksheet_dimension_is_not_silently_treated_as_an_empty_sheet():
    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    dimension_start = sheet_xml.index(b"<dimension")
    dimension_end = sheet_xml.index(b"/>", dimension_start) + 2
    sheet_xml = sheet_xml[:dimension_start] + sheet_xml[dimension_end:]
    snapshot = _rewrite_package(
        snapshot,
        replacements={"xl/worksheets/sheet1.xml": sheet_xml},
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3
    assert result.units[0].inspection_method == "none"
    assert result.units[0].signal_codes == ()
    assert result.units[0].issue_codes == ("unit-over-limit",)
    _assert_private_values_absent(result)


def test_underreported_worksheet_dimension_cannot_hide_private_cells():
    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    dimension_start = sheet_xml.index(b"<dimension")
    dimension_end = sheet_xml.index(b"/>", dimension_start) + 2
    sheet_xml = (
        sheet_xml[:dimension_start]
        + b'<dimension ref="A1:A1"/>'
        + sheet_xml[dimension_end:]
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={"xl/worksheets/sheet1.xml": sheet_xml},
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3
    assert result.units[0].inspection_method == "none"
    assert result.units[0].signal_codes == ()
    assert result.units[0].issue_codes == ("unit-over-limit",)
    _assert_private_values_absent(result)


def test_duplicate_cell_coordinates_make_the_sheet_unknown_before_iteration(
    monkeypatch,
):
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    workbook = Workbook()
    workbook.active["A1"] = "DANH SACH CHI TRA"
    snapshot = _save(workbook)
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    cell_start = sheet_xml.index(b'<c r="A1"')
    cell_end = sheet_xml.index(b"</c>", cell_start) + len(b"</c>")
    cell_xml = sheet_xml[cell_start:cell_end]
    snapshot = _rewrite_package(
        snapshot,
        replacements={
            "xl/worksheets/sheet1.xml": (
                sheet_xml[:cell_start]
                + cell_xml
                + cell_xml
                + sheet_xml[cell_end:]
            ),
        },
    )
    iteration_calls = []
    original_iter_rows = ReadOnlyWorksheet.iter_rows

    def recording_iter_rows(self, *args, **kwargs):
        iteration_calls.append(True)
        return original_iter_rows(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", recording_iter_rows)

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.units[0].inspection_method == "none"
    assert result.units[0].signal_codes == ()
    assert result.units[0].issue_codes == ("unit-over-limit",)
    assert iteration_calls == []
    _assert_private_values_absent(result)


def test_cell_element_budget_is_decided_before_lazy_iteration(monkeypatch):
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(("one", "two", "three"))
    snapshot = _save(workbook)
    iteration_calls = []
    original_iter_rows = ReadOnlyWorksheet.iter_rows

    def recording_iter_rows(self, *args, **kwargs):
        iteration_calls.append(True)
        return original_iter_rows(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", recording_iter_rows)

    result = _inspect(
        snapshot,
        limits=InspectionLimits(max_cells_per_workbook=2),
    )

    assert result.inspection_status == "inspected"
    assert result.units[0].inspection_method == "none"
    assert result.units[0].issue_codes == ("unit-over-limit",)
    assert iteration_calls == []


def test_sparse_wide_sheet_is_not_materialized_when_its_dimension_is_over_limit(
    monkeypatch,
):
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "private"
    worksheet["XFD1"] = None
    snapshot = _save(workbook)
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        sheet_xml = archive.read("xl/worksheets/sheet1.xml")
    sheet_xml = sheet_xml.replace(b'<dimension ref="A1:A1"', b'<dimension ref="A1:XFD1"')
    snapshot = _rewrite_package(
        snapshot,
        replacements={"xl/worksheets/sheet1.xml": sheet_xml},
    )
    iteration_calls = []
    original_iter_rows = ReadOnlyWorksheet.iter_rows

    def recording_iter_rows(self, *args, **kwargs):
        iteration_calls.append(True)
        return original_iter_rows(self, *args, **kwargs)

    monkeypatch.setattr(ReadOnlyWorksheet, "iter_rows", recording_iter_rows)

    result = _inspect(
        snapshot,
        limits=InspectionLimits(max_cells_per_workbook=10),
    )

    assert result.inspection_status == "inspected"
    assert result.units[0].inspection_method == "none"
    assert result.units[0].issue_codes == ("unit-over-limit",)
    assert iteration_calls == []


def test_lazy_cell_parse_failure_retains_hidden_sheet_signal_without_diagnostics():
    workbook = _workbook_with_sheets(
        (
            ("visible-private", "visible", (("safe",),)),
            ("hidden-private", "hidden", ((123,),)),
        )
    )
    snapshot = _save(workbook)
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        sheet_xml = archive.read("xl/worksheets/sheet2.xml")
    sheet_xml = sheet_xml.replace(b"<v>123</v>", b"<v>PRIVATE-NOT-A-NUMBER</v>")
    snapshot = _rewrite_package(
        snapshot,
        replacements={"xl/worksheets/sheet2.xml": sheet_xml},
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.units[1].inspection_method == "none"
    assert result.units[1].signal_codes == ("worksheet-hidden",)
    assert result.units[1].issue_codes == ("unit-over-limit", "worksheet-hidden")
    _assert_private_values_absent(
        result,
        "visible-private",
        "hidden-private",
        "PRIVATE-NOT-A-NUMBER",
        "invalid literal",
    )


def test_spoofed_worksheet_relationship_type_is_rejected_before_openpyxl(monkeypatch):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        rels = archive.read("xl/_rels/workbook.xml.rels")
    rels = rels.replace(
        b"http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
        b"https://private.invalid/worksheet",
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={"xl/_rels/workbook.xml.rels": rels},
    )
    loader_calls = []

    def forbidden_loader(*args, **kwargs):
        loader_calls.append((args, kwargs))
        raise AssertionError("preflight must reject spoofed relationship types")

    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", forbidden_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "unreadable"
    assert result.source_issue_codes == ("document-unreadable",)
    assert loader_calls == []
    _assert_private_values_absent(result, "private.invalid")


@pytest.mark.parametrize("dialect", ["transitional", "strict"])
@pytest.mark.parametrize("relationship_kind", ["worksheet", "drawing"])
def test_relationship_type_must_match_the_selected_workbook_dialect(
    monkeypatch, dialect, relationship_kind
):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook(with_image=relationship_kind == "drawing"))
    if dialect == "strict":
        snapshot = _strict_package(snapshot)
        selected_prefix = STRICT_OFFICE_RELATIONSHIPS
        mismatched_prefix = TRANSITIONAL_OFFICE_RELATIONSHIPS
    else:
        selected_prefix = TRANSITIONAL_OFFICE_RELATIONSHIPS
        mismatched_prefix = STRICT_OFFICE_RELATIONSHIPS
    relationship_part = (
        "xl/_rels/workbook.xml.rels"
        if relationship_kind == "worksheet"
        else "xl/worksheets/_rels/sheet3.xml.rels"
    )
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        relationships = archive.read(relationship_part)
    selected_type = selected_prefix + b"/" + relationship_kind.encode("ascii")
    mismatched_type = mismatched_prefix + b"/" + relationship_kind.encode("ascii")
    assert selected_type in relationships
    snapshot = _rewrite_package(
        snapshot,
        replacements={
            relationship_part: relationships.replace(
                selected_type,
                mismatched_type,
                1,
            ),
        },
    )
    loader_calls = []

    def forbidden_loader(*args, **kwargs):
        loader_calls.append(True)
        raise AssertionError("relationship dialect mismatch must fail in preflight")

    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", forbidden_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "unreadable"
    assert result.source_issue_codes == ("document-unreadable",)
    assert loader_calls == []


@pytest.mark.parametrize(
    ("part_name", "original_namespace", "attacker_namespace"),
    (
        (
            "[Content_Types].xml",
            b"http://schemas.openxmlformats.org/package/2006/content-types",
            b"https://attacker.invalid/content-types",
        ),
        (
            "xl/workbook.xml",
            TRANSITIONAL_SPREADSHEET_NAMESPACE,
            b"https://attacker.invalid/spreadsheet-main",
        ),
        (
            "xl/_rels/workbook.xml.rels",
            b"http://schemas.openxmlformats.org/package/2006/relationships",
            b"https://attacker.invalid/package-relationships",
        ),
        (
            "xl/worksheets/sheet1.xml",
            TRANSITIONAL_SPREADSHEET_NAMESPACE,
            b"https://attacker.invalid/worksheet-main",
        ),
    ),
)
def test_attacker_qnames_are_rejected_before_openpyxl(
    monkeypatch, part_name, original_namespace, attacker_namespace
):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        content = archive.read(part_name)
    assert original_namespace in content
    snapshot = _rewrite_package(
        snapshot,
        replacements={
            part_name: content.replace(original_namespace, attacker_namespace),
        },
    )
    loader_calls = []
    real_loader = workbook_adapter.openpyxl.load_workbook

    def recording_loader(*args, **kwargs):
        loader_calls.append(True)
        return real_loader(*args, **kwargs)

    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", recording_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "unreadable"
    assert result.source_issue_codes == ("document-unreadable",)
    assert loader_calls == []
    _assert_private_values_absent(result, "attacker.invalid")


@pytest.mark.parametrize(
    ("part_name", "with_image"),
    (
        ("xl/workbook.xml", False),
        ("xl/worksheets/sheet3.xml", True),
    ),
)
def test_attacker_relationship_id_qname_is_rejected_before_openpyxl(
    monkeypatch, part_name, with_image
):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook(with_image=with_image))
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        content = archive.read(part_name)
    assert TRANSITIONAL_OFFICE_RELATIONSHIPS in content
    snapshot = _rewrite_package(
        snapshot,
        replacements={
            part_name: content.replace(
                TRANSITIONAL_OFFICE_RELATIONSHIPS,
                b"https://attacker.invalid/office-relationships",
            ),
        },
    )
    loader_calls = []

    def forbidden_loader(*args, **kwargs):
        loader_calls.append(True)
        raise AssertionError("attacker r:id QName must fail in preflight")

    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", forbidden_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "unreadable"
    assert result.source_issue_codes == ("document-unreadable",)
    assert loader_calls == []


def test_strict_namespace_workbook_is_inspected_with_its_cell_signals():
    workbook = Workbook()
    workbook.active["A1"] = "DANH SACH CHI TRA"
    snapshot = _strict_package(_save(workbook))

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 1
    assert result.units[0].inspection_method == "worksheet-structure"
    assert "roster-column-pattern" in result.units[0].signal_codes


def test_strict_qname_conversion_preserves_cell_text_and_formula(monkeypatch):
    import ctv_inspection_workbook as workbook_adapter

    private_text = STRICT_SPREADSHEET_NAMESPACE.decode("ascii")
    private_formula = f'="{private_text}"'
    workbook = Workbook()
    workbook.active["A1"] = private_text
    workbook.active["A2"] = private_formula
    snapshot = _strict_package(_save(workbook))
    reduced_text = []
    real_reducer = workbook_adapter.signals_from_private_text

    def recording_reducer(text, context):
        reduced_text.append(text)
        return real_reducer(text, context)

    monkeypatch.setattr(
        workbook_adapter,
        "signals_from_private_text",
        recording_reducer,
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert private_text in reduced_text
    assert private_formula in reduced_text
    assert TRANSITIONAL_SPREADSHEET_NAMESPACE.decode("ascii") not in reduced_text
    _assert_private_values_absent(result, private_text, private_formula)


@pytest.mark.parametrize("encoding", ["utf-16-le", "utf-16-be"])
def test_utf16_strict_workbook_is_structurally_converted_and_inspected(encoding):
    workbook = Workbook()
    workbook.active["A1"] = "DANH SACH CHI TRA"
    snapshot = _strict_package(_save(workbook))
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        workbook_xml = _encoded_xml(archive.read("xl/workbook.xml"), encoding)
        worksheet_xml = _encoded_xml(
            archive.read("xl/worksheets/sheet1.xml"),
            encoding,
        )
    snapshot = _rewrite_package(
        snapshot,
        replacements={
            "xl/workbook.xml": workbook_xml,
            "xl/worksheets/sheet1.xml": worksheet_xml,
        },
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 1
    assert result.units[0].inspection_method == "worksheet-structure"
    assert "roster-column-pattern" in result.units[0].signal_codes


@pytest.mark.parametrize(
    "attacker_content_type",
    (
        b"application/vnd.openxmlformats-officedocument."
        b"spreadsheetml.styles+xml",
        b"application/vnd.openxmlformats-officedocument."
        b"spreadsheetml.sharedStrings+xml",
    ),
)
def test_strict_auxiliary_selection_never_opens_same_mime_attacker_member(
    monkeypatch, attacker_content_type
):
    attacker_part = "xl/media/private-same-mime-079123456789.xml"
    snapshot = _strict_package(_save(_roster_workbook()))
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        content_types = archive.read("[Content_Types].xml")
    content_types = content_types.replace(
        b"</Types>",
        b'<Override PartName="/xl/media/private-same-mime-079123456789.xml" '
        b'ContentType="' + attacker_content_type + b'"/></Types>',
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={"[Content_Types].xml": content_types},
        additions={attacker_part: b"PRIVATE-AUXILIARY-079123456789"},
    )
    opened = []
    original_open = zipfile.ZipFile.open

    def guarded_open(self, name, *args, **kwargs):
        member_name = name.filename if isinstance(name, zipfile.ZipInfo) else name
        if member_name == attacker_part:
            opened.append(member_name)
            raise AssertionError("same-MIME attacker member was opened")
        return original_open(self, name, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", guarded_open)

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3
    assert attacker_part not in opened
    _assert_private_values_absent(result, "PRIVATE-AUXILIARY-079123456789")


@pytest.mark.parametrize(
    "attacker_content_type",
    (
        b"application/vnd.openxmlformats-officedocument."
        b"spreadsheetml.styles+xml",
        b"application/vnd.openxmlformats-officedocument."
        b"spreadsheetml.sharedStrings+xml",
    ),
)
def test_transitional_loader_never_opens_same_mime_attacker_member(
    monkeypatch, attacker_content_type
):
    import ctv_inspection_workbook as workbook_adapter

    attacker_part = "xl/media/private-transitional-079123456789.xml"
    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        content_types = archive.read("[Content_Types].xml")
    first_override = content_types.index(b"<Override")
    attacker_override = (
        b'<Override PartName="/xl/media/private-transitional-079123456789.xml" '
        b'ContentType="' + attacker_content_type + b'"/>'
    )
    content_types = (
        content_types[:first_override]
        + attacker_override
        + content_types[first_override:]
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={"[Content_Types].xml": content_types},
        additions={attacker_part: b"PRIVATE-TRANSITIONAL-AUX-079123456789"},
    )
    opened = []
    original_open = zipfile.ZipFile.open
    real_loader = workbook_adapter.openpyxl.load_workbook

    def guarded_open(self, name, *args, **kwargs):
        member_name = name.filename if isinstance(name, zipfile.ZipInfo) else name
        if member_name == attacker_part:
            opened.append(member_name)
            raise AssertionError("same-MIME transitional attacker member was opened")
        return original_open(self, name, *args, **kwargs)

    def guarded_loader(source, **kwargs):
        with zipfile.ZipFile(BytesIO(source.getvalue()), "r") as loader_archive:
            assert attacker_part not in loader_archive.namelist()
        return real_loader(source, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", guarded_open)
    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", guarded_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3
    assert attacker_part not in opened
    _assert_private_values_absent(result, "PRIVATE-TRANSITIONAL-AUX-079123456789")


def test_xml_element_ceiling_rejects_before_full_tree_or_openpyxl(monkeypatch):
    import ctv_inspection_workbook as workbook_adapter

    strict_namespace = STRICT_SPREADSHEET_NAMESPACE
    worksheet_xml = (
        b'<worksheet xmlns="'
        + strict_namespace
        + b'"><dimension ref="A1:A1"/>'
        + b"<x/>" * 199_999
        + b"</worksheet>"
    )
    snapshot = _strict_package(_save(Workbook()))
    snapshot = _rewrite_package_with_stored_member(
        snapshot,
        "xl/worksheets/sheet1.xml",
        worksheet_xml,
    )
    tree_calls = []
    loader_calls = []
    real_fromstring = workbook_adapter.ElementTree.fromstring

    def guarded_tree(content, *args, **kwargs):
        if content == worksheet_xml:
            tree_calls.append(True)
            raise AssertionError("full tree must not be built past the element ceiling")
        return real_fromstring(content, *args, **kwargs)

    def forbidden_loader(*args, **kwargs):
        loader_calls.append((args, kwargs))
        raise AssertionError("OpenPyXL must not receive over-element XML")

    monkeypatch.setattr(workbook_adapter.ElementTree, "fromstring", guarded_tree)
    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", forbidden_loader)

    with pytest.raises(
        workbook_adapter.WorkbookParserBoundaryExceededError
    ) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"
    assert tree_calls == []
    assert loader_calls == []


def test_normalized_member_output_cap_precedes_loader_zip_and_openpyxl(monkeypatch):
    import ctv_inspection_workbook as workbook_adapter

    repeated_element = b"<x>" + b"A" * 117 + b"</x>"
    worksheet_xml = (
        b'<worksheet xmlns="'
        + STRICT_SPREADSHEET_NAMESPACE
        + b'"><dimension ref="A1:A1"/><sheetData/>'
        + repeated_element * 199_997
        + b"</worksheet>"
    )
    assert len(worksheet_xml) < 25 * 1024 * 1024
    snapshot = _strict_package(_save(Workbook()))
    snapshot = _rewrite_package_with_stored_member(
        snapshot,
        "xl/worksheets/sheet1.xml",
        worksheet_xml,
    )
    loader_zip_calls = []
    loader_calls = []
    real_zipfile = workbook_adapter.zipfile.ZipFile

    def recording_zipfile(source, mode="r", *args, **kwargs):
        if mode == "w":
            loader_zip_calls.append(True)
        return real_zipfile(source, mode, *args, **kwargs)

    def forbidden_loader(*args, **kwargs):
        loader_calls.append((args, kwargs))
        raise AssertionError("OpenPyXL must not receive oversized normalized XML")

    monkeypatch.setattr(workbook_adapter.zipfile, "ZipFile", recording_zipfile)
    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", forbidden_loader)

    with pytest.raises(
        workbook_adapter.WorkbookParserBoundaryExceededError
    ) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"
    assert loader_zip_calls == []
    assert loader_calls == []


def test_normalized_text_expansion_is_rejected_before_large_writer_chunk(
    monkeypatch,
):
    import ctv_inspection_workbook as workbook_adapter

    content = (
        b'<workbook xmlns="'
        + STRICT_SPREADSHEET_NAMESPACE
        + b'">'
        + b">" * (7 * 1024 * 1024)
        + b"</workbook>"
    )
    write_sizes = []
    real_write = workbook_adapter._CappedXmlOutput.write

    def recording_write(self, output):
        write_sizes.append(len(output))
        return real_write(self, output)

    monkeypatch.setattr(workbook_adapter._CappedXmlOutput, "write", recording_write)

    with pytest.raises(
        workbook_adapter.WorkbookParserBoundaryExceededError
    ) as raised:
        workbook_adapter._normalized_strict_xml(content)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"
    assert not write_sizes or max(write_sizes) <= 64 * 1024


@pytest.mark.parametrize("target_kind", ["worksheet", "drawing"])
def test_wrong_target_content_type_fails_safe_before_openpyxl(
    monkeypatch, target_kind
):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook(with_image=target_kind == "drawing"))
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        content_types = archive.read("[Content_Types].xml")
    expected_type = {
        "worksheet": (
            b"application/vnd.openxmlformats-officedocument."
            b"spreadsheetml.worksheet+xml"
        ),
        "drawing": b"application/vnd.openxmlformats-officedocument.drawing+xml",
    }[target_kind]
    assert expected_type in content_types
    content_types = content_types.replace(
        expected_type,
        b"application/private-wrong-target+xml",
        1,
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={"[Content_Types].xml": content_types},
    )
    loader_calls = []

    def forbidden_loader(*args, **kwargs):
        loader_calls.append(True)
        raise AssertionError("wrong content type must fail in preflight")

    monkeypatch.setattr(workbook_adapter.openpyxl, "load_workbook", forbidden_loader)

    result = _inspect(snapshot)

    assert result.inspection_status == "unreadable"
    assert result.source_issue_codes == ("document-unreadable",)
    assert loader_calls == []


def test_worksheet_content_type_may_use_the_bounded_default_map():
    workbook = Workbook()
    workbook.active["A1"] = "DANH SACH CHI TRA"
    snapshot = _save(workbook)
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        content_types = archive.read("[Content_Types].xml")
    worksheet_override_start = content_types.index(
        b'<Override PartName="/xl/worksheets/sheet1.xml"'
    )
    worksheet_override_end = content_types.index(
        b"/>", worksheet_override_start
    ) + 2
    content_types = (
        content_types[:worksheet_override_start]
        + content_types[worksheet_override_end:]
    ).replace(
        b'<Default Extension="xml" ContentType="application/xml"/>',
        b'<Default Extension="xml" ContentType="application/vnd.'
        b'openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>',
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={"[Content_Types].xml": content_types},
    )

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.units[0].inspection_method == "worksheet-structure"
    assert "roster-column-pattern" in result.units[0].signal_codes


def test_relationship_pull_parser_retains_only_requested_ids():
    import ctv_inspection_workbook as workbook_adapter

    unrelated = b"".join(
        (
            b'<Relationship Id="private-%d" Type="private-type" '
            b'Target="private-target-%d"/>'
        ) % (index, index)
        for index in range(1_000)
    )
    content = (
        b'<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        + unrelated
        + b'<Relationship Id="rIdWanted" Type="wanted-type" Target="wanted-target"/>'
        + b"</Relationships>"
    )

    relationships = workbook_adapter._relationships(content, {"rIdWanted"})

    assert relationships == {
        "rIdWanted": ("wanted-type", "wanted-target", False),
    }
    assert "private-target" not in repr(relationships)


def test_worksheet_pull_parser_retains_only_one_drawing_presence_reference():
    import ctv_inspection_workbook as workbook_adapter

    drawings = b"".join(
        b'<drawing r:id="rId%d"/>' % index
        for index in range(1_000)
    )
    content = (
        b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        b'<dimension ref="A1:A1"/>'
        + drawings
        + b"</worksheet>"
    )

    drawing_ids, cell_bound = workbook_adapter._worksheet_metadata(content)

    assert drawing_ids == ("rId0",)
    assert cell_bound == 1


def test_workbook_xml_sheet_limit_raises_while_relationship_ids_are_parsed():
    import ctv_inspection_workbook as workbook_adapter

    content = b"""<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
        <sheets><sheet name="private-one" sheetId="1" r:id="rId1"/>
        <sheet name="private-two" sheetId="2" r:id="rId2"/></sheets></workbook>"""

    with pytest.raises(
        workbook_adapter.WorkbookWorksheetCountExceededError
    ) as raised:
        workbook_adapter._workbook_sheet_relationship_ids(content, 1)
    assert str(raised.value) == "inspection-worksheet-count-exceeded"
    _assert_private_values_absent(raised.value, "private-one", "private-two")


@pytest.mark.parametrize("strict", [False, True])
def test_macro_external_link_and_image_members_are_never_read(monkeypatch, strict):
    snapshot = _save(_roster_workbook(with_image=True))
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        workbook_xml = archive.read("xl/workbook.xml")
        workbook_rels = archive.read("xl/_rels/workbook.xml.rels")
        content_types = archive.read("[Content_Types].xml")
        media_names = [
            info.filename for info in archive.infolist()
            if info.filename.startswith("xl/media/")
        ]
        drawing_names = [
            info.filename for info in archive.infolist()
            if info.filename.startswith("xl/drawings/")
        ]
    assert len(media_names) == 1
    workbook_xml = workbook_xml.replace(
        b"</workbook>",
        b'<externalReferences><externalReference xmlns:r="http://schemas.'
        b'openxmlformats.org/officeDocument/2006/relationships" r:id="rId99"/>'
        b'</externalReferences></workbook>',
    )
    workbook_rels = workbook_rels.replace(
        b"</Relationships>",
        b'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        b'relationships/externalLink" Target="externalLinks/externalLink1.xml" '
        b'Id="rId99"/></Relationships>',
    )
    content_types = content_types.replace(
        b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml",
        b"application/vnd.ms-excel.sheet.macroEnabled.main+xml",
    ).replace(
        b"</Types>",
        b'<Override PartName="/xl/vbaProject.bin" '
        b'ContentType="application/vnd.ms-office.vbaProject"/></Types>',
    )
    snapshot = _rewrite_package(
        snapshot,
        replacements={
            "xl/workbook.xml": workbook_xml,
            "xl/_rels/workbook.xml.rels": workbook_rels,
            "[Content_Types].xml": content_types,
        },
        additions={
            "xl/vbaProject.bin": b"PRIVATE-MACRO-079123456789",
            "xl/externalLinks/externalLink1.xml": (
                b"<externalLink>" + PRIVATE_EXTERNAL_VALUE.encode() + b"</externalLink>"
            ),
        },
    )
    if strict:
        snapshot = _strict_package(snapshot)
    prohibited = {
        *media_names,
        *drawing_names,
        "xl/vbaProject.bin",
        "xl/externalLinks/externalLink1.xml",
    }
    original_open = zipfile.ZipFile.open
    opened = []

    def guarded_open(self, name, *args, **kwargs):
        member_name = name.filename if isinstance(name, zipfile.ZipInfo) else name
        if member_name in prohibited:
            raise AssertionError("non-worksheet workbook member was read")
        opened.append(member_name)
        return original_open(self, name, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", guarded_open)

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert "embedded-media-present" in result.units[2].signal_codes
    assert prohibited.isdisjoint(opened)
    _assert_private_values_absent(result, "PRIVATE-MACRO-079123456789")


def test_zip_entry_count_boundary_raises_before_zipfile_allocation(monkeypatch):
    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        existing_count = len(archive.infolist())
    additions = {
        f"private-extra/{index:05d}.bin": b""
        for index in range(10_000 - existing_count)
    }
    snapshot = _rewrite_package(snapshot, additions=additions)

    accepted = _inspect(snapshot)

    assert accepted.inspection_status == "inspected"
    assert accepted.unit_count == 3
    snapshot = _rewrite_package(
        snapshot,
        additions={"private-extra/over-limit.bin": b""},
    )

    def forbidden_zipfile(*_args, **_kwargs):
        raise AssertionError("oversized central directory was allocated")

    monkeypatch.setattr(zipfile, "ZipFile", forbidden_zipfile)

    from ctv_inspection_workbook import WorkbookParserBoundaryExceededError

    with pytest.raises(WorkbookParserBoundaryExceededError) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"
    _assert_private_values_absent(raised.value, "private-extra")


def test_sanitized_loader_work_is_reserved_for_both_openpyxl_worksheet_reads(
    monkeypatch,
):
    import ctv_inspection_workbook as workbook_adapter

    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        manual_total = sum(
            archive.getinfo(name).file_size
            for name in (
                "[Content_Types].xml",
                "xl/workbook.xml",
                "xl/_rels/workbook.xml.rels",
                "xl/worksheets/sheet1.xml",
                "xl/worksheets/sheet2.xml",
                "xl/worksheets/sheet3.xml",
                "xl/styles.xml",
            )
        )
    loader_snapshots = []
    real_loader = workbook_adapter.openpyxl.load_workbook

    def capturing_loader(source, **kwargs):
        loader_snapshots.append(source.getvalue())
        return real_loader(source, **kwargs)

    monkeypatch.setattr(
        workbook_adapter.openpyxl,
        "load_workbook",
        capturing_loader,
    )
    characterized = _inspect(snapshot)
    assert characterized.inspection_status == "inspected"
    assert len(loader_snapshots) == 1
    with zipfile.ZipFile(BytesIO(loader_snapshots[0]), "r") as loader_archive:
        loader_total = sum(info.file_size for info in loader_archive.infolist())
    exact_operation_limit = manual_total + 2 * loader_total
    monkeypatch.setattr(
        workbook_adapter,
        "_MAX_DECOMPRESSED_BYTES",
        exact_operation_limit,
    )
    opened = []
    original_open = zipfile.ZipFile.open

    def recording_open(self, name, *args, **kwargs):
        member_name = name.filename if isinstance(name, zipfile.ZipInfo) else name
        if self.mode == "r":
            opened.append(member_name)
        return original_open(self, name, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", recording_open)

    loader_snapshots.clear()
    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3
    open_counts = Counter(opened)
    assert open_counts["[Content_Types].xml"] == 2
    assert open_counts["xl/workbook.xml"] == 2
    assert open_counts["xl/_rels/workbook.xml.rels"] == 2
    assert open_counts["xl/worksheets/sheet1.xml"] == 3
    assert open_counts["xl/worksheets/sheet2.xml"] == 3
    assert open_counts["xl/worksheets/sheet3.xml"] == 3
    assert open_counts["xl/styles.xml"] == 2

    opened.clear()
    loader_snapshots.clear()
    monkeypatch.setattr(
        workbook_adapter,
        "_MAX_DECOMPRESSED_BYTES",
        exact_operation_limit - 1,
    )
    from ctv_inspection_workbook import WorkbookParserBoundaryExceededError

    with pytest.raises(WorkbookParserBoundaryExceededError) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"
    assert loader_snapshots == []


@pytest.mark.parametrize("adversary", ["member", "aggregate", "ratio"])
def test_declared_zip_resource_boundaries_raise_only_stable_parser_error(adversary):
    snapshot = _save(_roster_workbook())
    with zipfile.ZipFile(BytesIO(snapshot), "r") as archive:
        names = [info.filename for info in archive.infolist()]
    if adversary == "member":
        patches = {names[0]: (1024 * 1024, 25 * 1024 * 1024 + 1)}
    elif adversary == "aggregate":
        patches = {
            name: (1024 * 1024, 25 * 1024 * 1024)
            for name in names[:5]
        }
    else:
        patches = {names[0]: (1, 101)}
    snapshot = _patch_central_sizes(snapshot, patches)

    from ctv_inspection_workbook import WorkbookParserBoundaryExceededError

    with pytest.raises(WorkbookParserBoundaryExceededError) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"
    _assert_private_values_absent(raised.value)


def test_actual_member_decompression_is_counted_and_bounded(monkeypatch):
    snapshot = _save(_roster_workbook())
    original_open = zipfile.ZipFile.open

    class ExpandingReader:
        def __init__(self):
            self.remaining = 25 * 1024 * 1024 + 1

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return None

        def read(self, size=-1):
            if self.remaining <= 0:
                return b""
            amount = self.remaining if size < 0 else min(size, self.remaining)
            self.remaining -= amount
            return b"X" * amount

    def expanding_open(self, name, *args, **kwargs):
        member_name = name.filename if isinstance(name, zipfile.ZipInfo) else name
        if member_name == "[Content_Types].xml":
            return ExpandingReader()
        return original_open(self, name, *args, **kwargs)

    monkeypatch.setattr(zipfile.ZipFile, "open", expanding_open)

    from ctv_inspection_workbook import WorkbookParserBoundaryExceededError

    with pytest.raises(WorkbookParserBoundaryExceededError) as raised:
        _inspect(snapshot)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"


def test_actual_aggregate_decompression_budget_is_exact():
    import ctv_inspection_workbook as workbook_adapter

    budget = workbook_adapter._ActualByteBudget()
    for _ in range(4):
        budget.consume(25 * 1024 * 1024, 25 * 1024 * 1024)
    assert budget.used == 100 * 1024 * 1024

    with pytest.raises(
        workbook_adapter.WorkbookParserBoundaryExceededError
    ) as raised:
        budget.consume(1, 1)
    assert str(raised.value) == "inspection-parser-boundary-exceeded"


def test_inspection_never_extracts_writes_uses_temp_network_or_ocr(monkeypatch):
    snapshot = _save(_roster_workbook(with_image=True))

    def forbidden(*_args, **_kwargs):
        raise AssertionError("forbidden side effect")

    monkeypatch.setattr(zipfile.ZipFile, "extract", forbidden)
    monkeypatch.setattr(zipfile.ZipFile, "extractall", forbidden)
    monkeypatch.setattr(builtins, "open", forbidden)
    monkeypatch.setattr(Path, "write_bytes", forbidden)
    monkeypatch.setattr(Path, "write_text", forbidden)
    monkeypatch.setattr(tempfile, "NamedTemporaryFile", forbidden)
    monkeypatch.setattr(tempfile, "TemporaryDirectory", forbidden)
    monkeypatch.setattr(socket, "socket", forbidden)

    result = _inspect(snapshot)

    assert result.inspection_status == "inspected"
    assert result.unit_count == 3


def test_workbook_module_has_a_narrow_import_and_archive_surface():
    module_path = Path(__file__).with_name("ctv_inspection_workbook.py")
    tree = ast.parse(module_path.read_text())
    imported_roots = set()
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            imported_roots.update(alias.name.split(".")[0] for alias in node.names)
        elif isinstance(node, ast.ImportFrom) and node.module:
            imported_roots.add(node.module.split(".")[0])
    assert imported_roots <= {
        "__future__",
        "datetime",
        "io",
        "openpyxl",
        "zipfile",
        "xml",
        "ctv_inspection_classifier",
        "ctv_inspection_model",
        "ooxml",
    }
    assert imported_roots.isdisjoint({
        "os",
        "pathlib",
        "tempfile",
        "socket",
        "subprocess",
        "tarfile",
        "rarfile",
        "pytesseract",
    })
    source = module_path.read_text()
    assert ".extract(" not in source
    assert ".extractall(" not in source

    inspection_modules = sorted(
        candidate
        for candidate in Path(__file__).parent.glob("ctv_inspection*.py")
        if not candidate.name.endswith("_test.py")
    )
    archive_importers = []
    for candidate in inspection_modules:
        candidate_tree = ast.parse(candidate.read_text())
        if any(
            (
                isinstance(node, ast.Import)
                and any(alias.name == "zipfile" for alias in node.names)
            )
            or (
                isinstance(node, ast.ImportFrom)
                and node.module == "zipfile"
            )
            for node in ast.walk(candidate_tree)
        ):
            archive_importers.append(candidate.name)
    assert archive_importers == ["ctv_inspection_workbook.py"]
