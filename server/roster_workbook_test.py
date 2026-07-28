import io
import zipfile

import openpyxl
import pytest

import roster_workbook
from roster_workbook import (
    RosterWorkbookError,
    load_roster_rows,
    preflight_roster_workbook,
)

_XLSX_MAIN_CONTENT_TYPE = (
    b"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
)


def _workbook_bytes() -> bytes:
    content = io.BytesIO()
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.append(["Họ và tên", "Số CCCD"])
    worksheet.append(["Synthetic A", "079123456789"])
    workbook.save(content)
    workbook.close()
    return content.getvalue()


def _replace_part(content: bytes, member: str, replacement: bytes) -> bytes:
    target_bytes = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content)) as source,
        zipfile.ZipFile(target_bytes, "w") as target,
    ):
        for info in source.infolist():
            target.writestr(
                info,
                replacement if info.filename == member else source.read(info),
            )
    return target_bytes.getvalue()


def _remove_part(content: bytes, member: str) -> bytes:
    target_bytes = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content)) as source,
        zipfile.ZipFile(target_bytes, "w") as target,
    ):
        for info in source.infolist():
            if info.filename != member:
                target.writestr(info, source.read(info))
    return target_bytes.getvalue()


def _move_worksheet_to_custom_target(content: bytes, worksheet: bytes) -> bytes:
    target_bytes = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content)) as source,
        zipfile.ZipFile(target_bytes, "w") as target,
    ):
        for info in source.infolist():
            member = source.read(info)
            if info.filename == "xl/_rels/workbook.xml.rels":
                member = member.replace(
                    b'Target="/xl/worksheets/sheet1.xml"',
                    b'Target="/xl/custom.xml"',
                )
            elif info.filename == "[Content_Types].xml":
                member = member.replace(
                    b'PartName="/xl/worksheets/sheet1.xml"',
                    b'PartName="/xl/custom.xml"',
                )
            target.writestr(info, member)
        target.writestr("xl/custom.xml", worksheet)
    return target_bytes.getvalue()


def _move_workbook_main_to_custom_target(
    content: bytes,
    worksheet: bytes,
) -> bytes:
    target_bytes = io.BytesIO()
    with (
        zipfile.ZipFile(io.BytesIO(content)) as source,
        zipfile.ZipFile(target_bytes, "w") as target,
    ):
        workbook = source.read("xl/workbook.xml")
        relationships = source.read("xl/_rels/workbook.xml.rels").replace(
            b'Target="/xl/worksheets/sheet1.xml"',
            b'Target="../custom-sheet.xml"',
        ).replace(
            b'Target="styles.xml"',
            b'Target="/xl/styles.xml"',
        ).replace(
            b'Target="theme/theme1.xml"',
            b'Target="/xl/theme/theme1.xml"',
        )
        for info in source.infolist():
            member = source.read(info)
            if info.filename == "[Content_Types].xml":
                member = member.replace(
                    b'PartName="/xl/worksheets/sheet1.xml"',
                    b'PartName="/xl/custom-sheet.xml"',
                ).replace(
                    b'PartName="/xl/workbook.xml"',
                    b'PartName="/xl/custom/main.xml"',
                )
            target.writestr(info, member)
        target.writestr("xl/custom/main.xml", workbook)
        target.writestr(
            "xl/custom/_rels/main.xml.rels",
            relationships,
        )
        target.writestr("xl/custom-sheet.xml", worksheet)
    return target_bytes.getvalue()


def test_valid_roster_is_preflighted_and_loaded():
    rows = load_roster_rows(io.BytesIO(_workbook_bytes()))

    assert rows == [
        ["Họ và tên", "Số CCCD"],
        ["Synthetic A", "079123456789"],
    ]


def test_declared_custom_workbook_main_sheet_is_bounded_before_openpyxl(
    monkeypatch,
):
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        expanded = archive.read("xl/worksheets/sheet1.xml").replace(
            b"A1:B2",
            b"A1:B200001",
        )
    custom_main = _move_workbook_main_to_custom_target(original, expanded)
    loader_calls = 0

    def unbounded_loader(*_args, **_kwargs):
        nonlocal loader_calls
        loader_calls += 1
        raise AssertionError("OpenPyXL must not run before bounded preflight")

    monkeypatch.setattr(roster_workbook.openpyxl, "load_workbook", unbounded_loader)

    with pytest.raises(RosterWorkbookError, match="worksheet-row-limit"):
        load_roster_rows(io.BytesIO(custom_main))

    assert loader_calls == 0


def test_declared_custom_workbook_main_loads_when_within_limits():
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")

    rows = load_roster_rows(
        io.BytesIO(_move_workbook_main_to_custom_target(original, worksheet))
    )

    assert rows == [
        ["Họ và tên", "Số CCCD"],
        ["Synthetic A", "079123456789"],
    ]


def test_multiple_workbook_main_declarations_are_rejected():
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")
    custom_main = _move_workbook_main_to_custom_target(original, worksheet)
    with zipfile.ZipFile(io.BytesIO(custom_main)) as archive:
        content_types = archive.read("[Content_Types].xml")
    duplicate = (
        b'<Override PartName="/xl/workbook.xml" ContentType="'
        + _XLSX_MAIN_CONTENT_TYPE
        + b'"/>'
    )
    malformed = _replace_part(
        custom_main,
        "[Content_Types].xml",
        content_types.replace(b"</Types>", duplicate + b"</Types>"),
    )

    with pytest.raises(RosterWorkbookError, match="multiple-workbook-parts"):
        preflight_roster_workbook(io.BytesIO(malformed))


def test_missing_workbook_main_declaration_is_rejected():
    content = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        content_types = archive.read("[Content_Types].xml")
    malformed = _replace_part(
        content,
        "[Content_Types].xml",
        content_types.replace(
            _XLSX_MAIN_CONTENT_TYPE,
            b"application/vnd.example.not-a-workbook+xml",
        ),
    )

    with pytest.raises(RosterWorkbookError, match="missing-workbook-part"):
        preflight_roster_workbook(io.BytesIO(malformed))


def test_traversal_workbook_main_declaration_is_rejected():
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")
    custom_main = _move_workbook_main_to_custom_target(original, worksheet)
    with zipfile.ZipFile(io.BytesIO(custom_main)) as archive:
        content_types = archive.read("[Content_Types].xml")
    malformed = _replace_part(
        custom_main,
        "[Content_Types].xml",
        content_types.replace(
            b'PartName="/xl/custom/main.xml"',
            b'PartName="/../xl/custom/main.xml"',
        ),
    )

    with pytest.raises(RosterWorkbookError, match="invalid-workbook-part"):
        preflight_roster_workbook(io.BytesIO(malformed))


def test_declared_custom_workbook_requires_its_own_relationship_part():
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")
    custom_main = _move_workbook_main_to_custom_target(original, worksheet)
    malformed = _remove_part(
        custom_main,
        "xl/custom/_rels/main.xml.rels",
    )

    with pytest.raises(RosterWorkbookError, match="missing-required-part"):
        preflight_roster_workbook(io.BytesIO(malformed))


def test_declared_custom_workbook_rejects_external_worksheet_relationship():
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")
    custom_main = _move_workbook_main_to_custom_target(original, worksheet)
    with zipfile.ZipFile(io.BytesIO(custom_main)) as archive:
        relationships = archive.read("xl/custom/_rels/main.xml.rels")
    malformed = _replace_part(
        custom_main,
        "xl/custom/_rels/main.xml.rels",
        relationships.replace(
            b'Id="rId1"',
            b'TargetMode="External" Id="rId1"',
        ),
    )

    with pytest.raises(RosterWorkbookError, match="external-relationship"):
        preflight_roster_workbook(io.BytesIO(malformed))


def test_openpyxl_compatible_default_workbook_content_type_is_supported():
    content = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        content_types = archive.read("[Content_Types].xml")
    override = (
        b'<Override PartName="/xl/workbook.xml" ContentType="'
        + _XLSX_MAIN_CONTENT_TYPE
        + b'"/>'
    )
    assert override in content_types
    modified_content_types = content_types.replace(override, b"").replace(
        b'<Default Extension="xml" ContentType="application/xml"/>',
        (
            b'<Default Extension="xml" ContentType="'
            + _XLSX_MAIN_CONTENT_TYPE
            + b'"/>'
        ),
    )
    modified = _replace_part(
        content,
        "[Content_Types].xml",
        modified_content_types,
    )

    assert load_roster_rows(io.BytesIO(modified)) == [
        ["Họ và tên", "Số CCCD"],
        ["Synthetic A", "079123456789"],
    ]


def test_relationship_targeted_custom_worksheet_is_bounded_before_openpyxl(
    monkeypatch,
):
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        expanded = archive.read("xl/worksheets/sheet1.xml").replace(
            b'A1:B2',
            b'A1:B200001',
        )
    custom_target = _move_worksheet_to_custom_target(original, expanded)
    loader_calls = 0

    def unbounded_loader(*_args, **_kwargs):
        nonlocal loader_calls
        loader_calls += 1
        raise AssertionError("OpenPyXL must not run before bounded preflight")

    monkeypatch.setattr(roster_workbook.openpyxl, "load_workbook", unbounded_loader)

    with pytest.raises(RosterWorkbookError, match="worksheet-row-limit"):
        load_roster_rows(io.BytesIO(custom_target))

    assert loader_calls == 0


def test_relationship_targeted_custom_worksheet_loads_when_within_limits():
    original = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(original)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml")

    rows = load_roster_rows(
        io.BytesIO(_move_worksheet_to_custom_target(original, worksheet))
    )

    assert rows == [
        ["Họ và tên", "Số CCCD"],
        ["Synthetic A", "079123456789"],
    ]


@pytest.mark.parametrize(
    ("old", "new", "error_code"),
    [
        (
            b'Target="/xl/worksheets/sheet1.xml"',
            b'Target="/xl/missing.xml"',
            "invalid-workbook-relationships",
        ),
        (
            b'Target="/xl/worksheets/sheet1.xml"',
            b'Target="../../outside.xml"',
            "invalid-target",
        ),
        (
            b'Id="rId1"',
            b'TargetMode="External" Id="rId1"',
            "external-relationship",
        ),
        (
            b"</Relationships>",
            (
                b'<Relationship Type="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships/worksheet" '
                b'Target="/xl/worksheets/sheet1.xml" Id="rId1"/>'
                b"</Relationships>"
            ),
            "invalid-workbook-relationships",
        ),
    ],
)
def test_invalid_worksheet_relationships_are_rejected(old, new, error_code):
    content = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        relationships = archive.read("xl/_rels/workbook.xml.rels")
    assert old in relationships
    malformed = _replace_part(
        content,
        "xl/_rels/workbook.xml.rels",
        relationships.replace(old, new, 1),
    )

    with pytest.raises(RosterWorkbookError, match=error_code):
        preflight_roster_workbook(io.BytesIO(malformed))


def test_two_sheets_cannot_ambiguously_reference_one_worksheet_target():
    content = _workbook_bytes()
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
        workbook = archive.read("xl/workbook.xml").replace(
            b"</sheets>",
            (
                b'<sheet xmlns:r="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships" name="Duplicate" '
                b'sheetId="2" state="visible" r:id="rId4"/></sheets>'
            ),
        )
        relationships = archive.read("xl/_rels/workbook.xml.rels").replace(
            b"</Relationships>",
            (
                b'<Relationship Type="http://schemas.openxmlformats.org/'
                b'officeDocument/2006/relationships/worksheet" '
                b'Target="/xl/worksheets/sheet1.xml" Id="rId4"/>'
                b"</Relationships>"
            ),
        )
    malformed = _replace_part(content, "xl/workbook.xml", workbook)
    malformed = _replace_part(
        malformed,
        "xl/_rels/workbook.xml.rels",
        relationships,
    )

    with pytest.raises(
        RosterWorkbookError,
        match="invalid-workbook-relationships",
    ):
        preflight_roster_workbook(io.BytesIO(malformed))


@pytest.mark.parametrize(
    ("member", "replacement"),
    [
        ("xl/workbook.xml", b"<workbook"),
        ("xl/worksheets/sheet1.xml", b"<worksheet"),
    ],
)
def test_malformed_workbook_or_worksheet_xml_is_rejected(member, replacement):
    malformed = _replace_part(_workbook_bytes(), member, replacement)

    with pytest.raises(RosterWorkbookError, match="invalid-workbook"):
        preflight_roster_workbook(io.BytesIO(malformed))


@pytest.mark.parametrize(
    ("limit_name", "error_code"),
    [
        ("MAX_WORKBOOK_BYTES", "workbook-too-large"),
        ("MAX_ARCHIVE_MEMBERS", "archive-member-limit"),
        ("MAX_ARCHIVE_MEMBER_BYTES", "archive-member-too-large"),
        ("MAX_ARCHIVE_UNCOMPRESSED_BYTES", "archive-uncompressed-too-large"),
        ("MAX_XML_BYTES", "xml-too-large"),
    ],
)
def test_container_and_xml_limits_are_hard_failures(
    monkeypatch,
    limit_name,
    error_code,
):
    monkeypatch.setattr(roster_workbook, limit_name, 1)

    with pytest.raises(RosterWorkbookError, match=error_code):
        preflight_roster_workbook(io.BytesIO(_workbook_bytes()))


@pytest.mark.parametrize(
    ("limit_name", "error_code"),
    [
        ("MAX_WORKSHEET_ROWS", "worksheet-row-limit"),
        ("MAX_WORKSHEET_COLUMNS", "worksheet-column-limit"),
        ("MAX_WORKSHEET_CELLS", "worksheet-cell-limit"),
        ("MAX_STRING_CHARACTERS", "string-too-large"),
    ],
)
def test_worksheet_shape_and_string_limits_are_hard_failures(
    monkeypatch,
    limit_name,
    error_code,
):
    monkeypatch.setattr(roster_workbook, limit_name, 1)

    with pytest.raises(RosterWorkbookError, match=error_code):
        preflight_roster_workbook(io.BytesIO(_workbook_bytes()))
