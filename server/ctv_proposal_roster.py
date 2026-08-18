"""Immutable, privacy-safe roster candidates from retained workbook snapshots."""

from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass, field
from io import BytesIO
import re
from typing import Literal

from openpyxl import load_workbook

from ctv_inspection_classifier import roster_header_categories_from_private_text
from ctv_inspection_model import DEFAULT_INSPECTION_LIMITS, InspectionResult
from ctv_inspection_workbook import (
    PackageWorkbookError,
    PrivateRosterCandidateFacts,
    worksheet_nonblank_row_indexes,
)


_MAX_ROWS = 10_000
_MAX_CELLS = 100_000
_MAX_CELL_TEXT = 256
_MAX_WORKBOOK_BYTES = 25 * 1024 * 1024
_MAX_CAPTURED_CHARACTERS = 16 * 1024 * 1024
_UNIT_ID = re.compile(r"^unit-[0-9]{4,}$")
_EVIDENCE_ID = re.compile(r"^evidence-[0-9]{4,}$")
_ROSTER_ISSUE_ORDER = (
    "roster-header-missing",
    "roster-row-invalid",
    "roster-identity-duplicate",
    "roster-over-limit",
    "roster-unreadable",
)
_PACKAGE_ISSUE_ORDER = (
    "roster-fa-code-blank",
    "roster-fa-code-conflict",
    "roster-fa-code-missing",
    "roster-header-duplicate",
)
_CANONICAL_ROSTER_FIELDS = (
    "name",
    "identity",
    "faCode",
    "taxId",
    "birthDate",
    "bankAccount",
    "serviceFee",
    "product",
)
_CANONICAL_ROSTER_FIELD_SET = frozenset(_CANONICAL_ROSTER_FIELDS)
_SELECTION_ISSUE_BY_STATUS = {
    "selected": (),
    "missing": ("roster-missing",),
    "ambiguous": ("roster-ambiguous",),
    "invalid": ("roster-invalid",),
}


def _exact_private_pairs(value, field_name: str) -> tuple[tuple[str, str], ...]:
    if type(value) is not tuple:
        raise TypeError(f"{field_name} must be an exact tuple")
    if any(type(pair) is not tuple or len(pair) != 2 for pair in value):
        raise TypeError(f"{field_name} must contain exact pairs")
    if any(
        type(key) is not str or type(private_value) is not str
        for key, private_value in value
    ):
        raise TypeError(f"{field_name} pairs must contain exact strings")
    keys = tuple(key for key, _private_value in value)
    if (
        any(key not in _CANONICAL_ROSTER_FIELD_SET for key in keys)
        or len(keys) != len(set(keys))
        or keys != tuple(sorted(keys))
    ):
        raise ValueError(f"{field_name} must use unique canonical fields")
    if any(len(private_value) > _MAX_CELL_TEXT for _key, private_value in value):
        raise ValueError(f"{field_name} values must stay within the cell limit")
    return value


def _exact_issue_codes(value, allowed: tuple[str, ...], field_name: str) -> None:
    if type(value) is not tuple:
        raise TypeError(f"{field_name} must be an exact tuple")
    if any(type(code) is not str for code in value):
        raise TypeError(f"{field_name} must contain exact strings")
    positions = {code: index for index, code in enumerate(allowed)}
    if (
        any(code not in positions for code in value)
        or len(value) != len(set(value))
        or value != tuple(sorted(value, key=positions.__getitem__))
    ):
        raise ValueError(f"{field_name} must use approved ordered codes")


@dataclass(frozen=True)
class RosterCandidateRow:
    row_index: int
    name: str = field(repr=False)
    identity: str = field(repr=False)
    values: tuple[tuple[str, str], ...] = field(repr=False)

    def __post_init__(self) -> None:
        if type(self.row_index) is not int:
            raise TypeError("row_index must be an exact integer")
        if not 1 <= self.row_index <= _MAX_ROWS:
            raise ValueError("row_index must be within the roster row limit")
        if type(self.name) is not str or type(self.identity) is not str:
            raise TypeError("private roster identity fields must be exact strings")
        if (
            not self.name
            or not self.identity
            or len(self.name) > _MAX_CELL_TEXT
            or len(self.identity) > _MAX_CELL_TEXT
        ):
            raise ValueError("private roster identity fields must be bounded")
        pairs = _exact_private_pairs(self.values, "values")
        private_values = dict(pairs)
        if (
            private_values.get("name") != self.name
            or private_values.get("identity") != self.identity
        ):
            raise ValueError("private roster fields must agree")


@dataclass(frozen=True)
class RosterCandidate:
    unit_id: str
    evidence_id: str
    worksheet_index: int
    rows: tuple[RosterCandidateRow, ...] = field(repr=False)
    blocking_issue_codes: tuple[str, ...]
    package_issue_codes: tuple[str, ...]
    canonical_to_source_columns: tuple[tuple[str, str], ...] = field(repr=False)
    score: tuple[int, int, int]

    def __post_init__(self) -> None:
        if type(self.unit_id) is not str or not _UNIT_ID.fullmatch(self.unit_id):
            raise TypeError("unit_id must be a valid opaque ID")
        if (
            type(self.evidence_id) is not str
            or not _EVIDENCE_ID.fullmatch(self.evidence_id)
        ):
            raise TypeError("evidence_id must be a valid opaque ID")
        if type(self.worksheet_index) is not int:
            raise TypeError("worksheet_index must be an exact integer")
        if not 1 <= self.worksheet_index <= 100:
            raise ValueError("worksheet_index must be within the workbook limit")
        if type(self.rows) is not tuple:
            raise TypeError("rows must be an exact tuple")
        if any(type(row) is not RosterCandidateRow for row in self.rows):
            raise TypeError("rows must contain exact roster candidate rows")
        for row in self.rows:
            row.__post_init__()
        row_indexes = tuple(row.row_index for row in self.rows)
        if (
            len(self.rows) > _MAX_ROWS
            or len(row_indexes) != len(set(row_indexes))
            or row_indexes != tuple(sorted(row_indexes))
        ):
            raise ValueError("rows must be unique and ordered within the row limit")
        _exact_issue_codes(
            self.blocking_issue_codes,
            _ROSTER_ISSUE_ORDER,
            "blocking_issue_codes",
        )
        _exact_issue_codes(
            self.package_issue_codes,
            _PACKAGE_ISSUE_ORDER,
            "package_issue_codes",
        )
        columns = _exact_private_pairs(
            self.canonical_to_source_columns,
            "canonical_to_source_columns",
        )
        column_fields = tuple(field_name for field_name, _source_name in columns)
        for row in self.rows:
            if tuple(field_name for field_name, _value in row.values) != column_fields:
                raise ValueError("roster row fields must match source columns")
        if type(self.score) is not tuple or len(self.score) != 3:
            raise TypeError("score must be an exact three-item tuple")
        if any(type(value) is not int for value in self.score):
            raise TypeError("score must contain exact integers")
        expected_score = (
            int(not self.blocking_issue_codes),
            int(not self.package_issue_codes),
            len(self.rows),
        )
        if self.score != expected_score:
            raise ValueError("score must agree with candidate eligibility facts")


@dataclass(frozen=True)
class RosterSelection:
    status: Literal["selected", "missing", "ambiguous", "invalid"]
    roster_unit_id: str | None
    candidate_unit_ids: tuple[str, ...]
    issue_codes: tuple[str, ...]

    def __post_init__(self) -> None:
        if type(self.status) is not str:
            raise TypeError("status must be an exact string")
        if self.status not in _SELECTION_ISSUE_BY_STATUS:
            raise ValueError("status must be an approved roster selection status")
        if self.roster_unit_id is not None and (
            type(self.roster_unit_id) is not str
            or not _UNIT_ID.fullmatch(self.roster_unit_id)
        ):
            raise TypeError("roster_unit_id must be a valid opaque ID or null")
        if type(self.candidate_unit_ids) is not tuple:
            raise TypeError("candidate_unit_ids must be an exact tuple")
        if any(type(unit_id) is not str for unit_id in self.candidate_unit_ids):
            raise TypeError("candidate_unit_ids must contain exact strings")
        if any(not _UNIT_ID.fullmatch(unit_id) for unit_id in self.candidate_unit_ids):
            raise ValueError("candidate_unit_ids must contain valid opaque IDs")
        numeric_ids = tuple(
            int(unit_id.rsplit("-", 1)[1]) for unit_id in self.candidate_unit_ids
        )
        if (
            len(self.candidate_unit_ids) != len(set(self.candidate_unit_ids))
            or numeric_ids != tuple(sorted(numeric_ids))
        ):
            raise ValueError("candidate_unit_ids must be unique and ordered")
        if type(self.issue_codes) is not tuple:
            raise TypeError("issue_codes must be an exact tuple")
        if any(type(code) is not str for code in self.issue_codes):
            raise TypeError("issue_codes must contain exact strings")
        if self.issue_codes != _SELECTION_ISSUE_BY_STATUS[self.status]:
            raise ValueError("issue_codes must agree with selection status")
        if self.status == "selected":
            if (
                self.roster_unit_id is None
                or self.roster_unit_id not in self.candidate_unit_ids
            ):
                raise ValueError("selected roster must identify one candidate")
        elif self.roster_unit_id is not None:
            raise ValueError("nonselected roster status must not identify a roster")
        if self.status == "missing" and self.candidate_unit_ids:
            raise ValueError("missing roster status must have no candidates")
        if self.status == "ambiguous" and len(self.candidate_unit_ids) < 2:
            raise ValueError("ambiguous roster status requires multiple candidates")
        if self.status == "invalid" and not self.candidate_unit_ids:
            raise ValueError("invalid roster status requires candidates")


def _private_cell_text(value: object) -> str:
    if value is None or type(value) not in {str, int, float}:
        return ""
    text = str(value).strip()
    if len(text) > _MAX_CELL_TEXT:
        return ""
    return text


def _numeric_unit_key(value) -> int:
    return int(value.unit_id.rsplit("-", 1)[1])


def _candidate(
    unit,
    *,
    rows: tuple[RosterCandidateRow, ...] = (),
    blocking_issue_codes: tuple[str, ...] = (),
    package_issue_codes: tuple[str, ...] = (),
    canonical_to_source_columns: tuple[tuple[str, str], ...] = (),
) -> RosterCandidate:
    return RosterCandidate(
        unit_id=unit.unit_id,
        evidence_id=unit.evidence_id,
        worksheet_index=unit.unit_index,
        rows=rows,
        blocking_issue_codes=blocking_issue_codes,
        package_issue_codes=package_issue_codes,
        canonical_to_source_columns=canonical_to_source_columns,
        score=(
            int(not blocking_issue_codes),
            int(not package_issue_codes),
            len(rows),
        ),
    )


def _fixed_snapshot_failure(unit, code: str = "roster-unreadable") -> RosterCandidate:
    return _candidate(unit, blocking_issue_codes=(code,))


def _candidate_from_private_facts(unit, facts) -> RosterCandidate:
    if type(facts) is not PrivateRosterCandidateFacts:
        return _fixed_snapshot_failure(unit)
    try:
        facts.__post_init__()
        if facts.worksheet_index != unit.unit_index:
            return _fixed_snapshot_failure(unit)
        return _candidate(
            unit,
            rows=tuple(
                RosterCandidateRow(
                    row_index=row.row_index,
                    name=row.name,
                    identity=row.identity,
                    values=row.values,
                )
                for row in facts.rows
            ),
            blocking_issue_codes=facts.blocking_issue_codes,
            package_issue_codes=facts.package_issue_codes,
            canonical_to_source_columns=facts.canonical_to_source_columns,
        )
    except Exception:
        return _fixed_snapshot_failure(unit)


class RosterCandidateEvidence:
    """Bounded caller-owned private facts captured by the original parse."""

    def __init__(
        self,
        *,
        max_worksheets: int = 10_000,
        max_rows: int = _MAX_ROWS,
        max_values: int = _MAX_CELLS,
        max_characters: int = _MAX_CAPTURED_CHARACTERS,
    ) -> None:
        for value in (max_worksheets, max_rows, max_values, max_characters):
            if type(value) is not int:
                raise TypeError("roster candidate evidence limits must be integers")
            if value <= 0:
                raise ValueError("roster candidate evidence limits must be positive")
        self._limits = (max_worksheets, max_rows, max_values, max_characters)
        self._facts = {}
        self._seen = set()
        self._rows = 0
        self._values = 0
        self._characters = 0
        self._cleared = False
        self.complete = True

    def capture(self, evidence_id, worksheet_index, facts) -> None:
        if self._cleared:
            raise ValueError("roster candidate evidence is cleared")
        if type(evidence_id) is not str or _EVIDENCE_ID.fullmatch(evidence_id) is None:
            raise TypeError("evidence_id must be a valid opaque ID")
        if type(worksheet_index) is not int or not 1 <= worksheet_index <= 100:
            raise TypeError("worksheet_index must be a bounded exact integer")
        if type(facts) is not PrivateRosterCandidateFacts:
            raise TypeError("roster facts must be exact private records")
        facts.__post_init__()
        if facts.worksheet_index != worksheet_index:
            raise ValueError("roster facts must match their worksheet")
        key = (evidence_id, worksheet_index)
        if key in self._seen:
            raise ValueError("roster facts already captured")
        self._seen.add(key)
        rows = len(facts.rows)
        values = sum(len(row.values) for row in facts.rows)
        characters = sum(
            len(value)
            for row in facts.rows
            for _field_name, value in row.values
        ) + sum(len(value) for _field_name, value in facts.canonical_to_source_columns)
        max_worksheets, max_rows, max_values, max_characters = self._limits
        if (
            len(self._seen) > max_worksheets
            or self._rows + rows > max_rows
            or self._values + values > max_values
            or self._characters + characters > max_characters
        ):
            self.complete = False
            return
        self._facts[key] = facts
        self._rows += rows
        self._values += values
        self._characters += characters

    def candidates_for(self, inspection: InspectionResult) -> tuple[RosterCandidate, ...]:
        if self._cleared:
            raise ValueError("roster candidate evidence is cleared")
        if type(inspection) is not InspectionResult:
            raise TypeError("inspection must be an inspection result")
        units = tuple(
            sorted(
                (
                    unit
                    for unit in inspection.units
                    if unit.unit_kind == "worksheet"
                    and unit.suggested_role == "payment-roster"
                ),
                key=_numeric_unit_key,
            )
        )
        candidates = []
        for unit in units:
            facts = self._facts.get((unit.evidence_id, unit.unit_index))
            if facts is None:
                candidates.append(
                    _fixed_snapshot_failure(
                        unit,
                        "roster-over-limit" if not self.complete else "roster-unreadable",
                    )
                )
            else:
                candidates.append(_candidate_from_private_facts(unit, facts))
        return validate_roster_candidates(inspection, tuple(candidates))

    def clear(self) -> None:
        self._facts.clear()
        self._seen.clear()
        self._rows = 0
        self._values = 0
        self._characters = 0
        self._cleared = True
        self.complete = False


def validate_roster_candidates(
    inspection: InspectionResult,
    candidates: tuple[RosterCandidate, ...],
) -> tuple[RosterCandidate, ...]:
    if type(inspection) is not InspectionResult:
        raise TypeError("inspection must be an inspection result")
    if type(candidates) is not tuple or any(
        type(candidate) is not RosterCandidate for candidate in candidates
    ):
        raise TypeError("preloaded candidates must be exact roster candidates")
    units = tuple(
        sorted(
            (
                unit
                for unit in inspection.units
                if unit.unit_kind == "worksheet"
                and unit.suggested_role == "payment-roster"
            ),
            key=_numeric_unit_key,
        )
    )
    if len(candidates) != len(units):
        raise ValueError("preloaded candidates must cover every roster unit")
    for candidate, unit in zip(candidates, units):
        candidate.__post_init__()
        if (
            candidate.unit_id != unit.unit_id
            or candidate.evidence_id != unit.evidence_id
            or candidate.worksheet_index != unit.unit_index
        ):
            raise ValueError("preloaded candidate must match its inspection unit")
    return candidates


def _parse_candidate(unit, snapshot: bytes) -> RosterCandidate:
    try:
        physically_nonblank_rows = frozenset(
            worksheet_nonblank_row_indexes(
                snapshot,
                unit.unit_index,
                limits=DEFAULT_INSPECTION_LIMITS,
            )
        )
    except PackageWorkbookError as error:
        if str(error) == "package-workbook-over-limit":
            return _fixed_snapshot_failure(unit, "roster-over-limit")
        return _fixed_snapshot_failure(unit)
    except Exception:
        return _fixed_snapshot_failure(unit)

    workbook = None
    try:
        workbook = load_workbook(
            BytesIO(snapshot), read_only=True, data_only=True, keep_links=False
        )
        worksheets = workbook.worksheets
        if unit.unit_index > len(worksheets):
            return _fixed_snapshot_failure(unit)
        worksheet = worksheets[unit.unit_index - 1]
        header = None
        package_issues = set()
        candidate_rows = []
        cells = 0
        issues = set()
        for row_index, workbook_row in enumerate(worksheet.iter_rows(), start=1):
            if row_index > _MAX_ROWS:
                issues.add("roster-over-limit")
                break
            values = []
            for cell in workbook_row:
                cells += 1
                if cells > _MAX_CELLS:
                    issues.add("roster-over-limit")
                    break
                values.append(_private_cell_text(cell.value))
            if "roster-over-limit" in issues:
                break
            categories = [
                roster_header_categories_from_private_text(value) if value else ()
                for value in values
            ]
            columns = {
                field_name: [
                    index
                    for index, value in enumerate(categories)
                    if field_name in value
                ]
                for field_name in _CANONICAL_ROSTER_FIELDS
            }
            name_columns = columns["name"]
            identity_columns = columns["identity"]
            if header is None and name_columns and identity_columns:
                header = {
                    field_name: (positions[0], values[positions[0]])
                    for field_name, positions in columns.items()
                    if positions
                }
                if any(len(positions) > 1 for positions in columns.values()):
                    package_issues.add("roster-header-duplicate")
                if "faCode" not in header:
                    package_issues.add("roster-fa-code-missing")
                continue
            if header is None or row_index not in physically_nonblank_rows:
                continue
            row_values = {
                field_name: values[index] if index < len(values) else ""
                for field_name, (index, _column_name) in header.items()
            }
            name = row_values.get("name", "")
            identity = row_values.get("identity", "")
            if not name or not identity:
                issues.add("roster-row-invalid")
                continue
            candidate_rows.append(
                RosterCandidateRow(
                    row_index=row_index,
                    name=name,
                    identity=identity,
                    values=tuple(sorted(row_values.items())),
                )
            )
        if header is None:
            issues.add("roster-header-missing")
        if not candidate_rows:
            issues.add("roster-row-invalid")
        if len(candidate_rows) != len({row.identity for row in candidate_rows}):
            issues.add("roster-identity-duplicate")
        fa_codes = {
            dict(row.values).get("faCode", "") for row in candidate_rows
        }
        if not fa_codes or "" in fa_codes:
            package_issues.add("roster-fa-code-blank")
        if len(fa_codes - {""}) > 1:
            package_issues.add("roster-fa-code-conflict")
        source_columns = (
            tuple(
                (field_name, column_name)
                for field_name, (_index, column_name) in sorted(header.items())
            )
            if header is not None
            else ()
        )
        blocking_issue_codes = tuple(
            code for code in _ROSTER_ISSUE_ORDER if code in issues
        )
        return _candidate(
            unit,
            rows=tuple(candidate_rows),
            blocking_issue_codes=blocking_issue_codes,
            package_issue_codes=tuple(sorted(package_issues)),
            canonical_to_source_columns=source_columns,
        )
    except Exception:
        return _fixed_snapshot_failure(unit)
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def load_roster_candidates(
    inspection: InspectionResult,
    snapshot_source: Callable[..., bytes],
) -> tuple[RosterCandidate, ...]:
    if type(inspection) is not InspectionResult:
        raise TypeError("inspection must be an inspection result")
    if not callable(snapshot_source):
        raise TypeError("snapshot source must be callable")
    sources_by_id = {source.evidence_id: source for source in inspection.sources}
    units = sorted(
        (
            unit
            for unit in inspection.units
            if unit.unit_kind == "worksheet"
            and unit.suggested_role == "payment-roster"
        ),
        key=_numeric_unit_key,
    )
    snapshot_cache: dict[str, bytes | None] = {}
    candidates = []
    for unit in units:
        source = sources_by_id[unit.evidence_id]
        if source.detected_type != "xlsx" or source.inspection_status != "inspected":
            candidates.append(_fixed_snapshot_failure(unit))
            continue
        if unit.evidence_id not in snapshot_cache:
            try:
                snapshot = snapshot_source(
                    unit.evidence_id, max_bytes=_MAX_WORKBOOK_BYTES
                )
            except Exception:
                snapshot = None
            snapshot_cache[unit.evidence_id] = snapshot
        snapshot = snapshot_cache[unit.evidence_id]
        if type(snapshot) is not bytes:
            candidates.append(_fixed_snapshot_failure(unit))
            continue
        if len(snapshot) > _MAX_WORKBOOK_BYTES:
            candidates.append(_fixed_snapshot_failure(unit, "roster-over-limit"))
            continue
        candidates.append(_parse_candidate(unit, snapshot))
    return validate_roster_candidates(inspection, tuple(candidates))


def choose_automatic_roster(
    candidates: tuple[RosterCandidate, ...],
) -> RosterSelection:
    if type(candidates) is not tuple or any(
        type(candidate) is not RosterCandidate for candidate in candidates
    ):
        raise TypeError("candidates must be immutable roster candidates")
    ordered = tuple(sorted(candidates, key=_numeric_unit_key))
    candidate_unit_ids = tuple(candidate.unit_id for candidate in ordered)
    if not ordered:
        return RosterSelection(
            status="missing",
            roster_unit_id=None,
            candidate_unit_ids=(),
            issue_codes=("roster-missing",),
        )
    highest_score = max(candidate.score for candidate in ordered)
    highest = tuple(
        candidate for candidate in ordered if candidate.score == highest_score
    )
    eligible = (
        highest_score[0] == 1
        and highest_score[1] == 1
        and highest_score[2] >= 1
    )
    if not eligible:
        return RosterSelection(
            status="invalid",
            roster_unit_id=None,
            candidate_unit_ids=candidate_unit_ids,
            issue_codes=("roster-invalid",),
        )
    if len(highest) != 1:
        return RosterSelection(
            status="ambiguous",
            roster_unit_id=None,
            candidate_unit_ids=candidate_unit_ids,
            issue_codes=("roster-ambiguous",),
        )
    return RosterSelection(
        status="selected",
        roster_unit_id=highest[0].unit_id,
        candidate_unit_ids=candidate_unit_ids,
        issue_codes=(),
    )
