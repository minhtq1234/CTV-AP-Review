"""Pure planning and bounded rendering for CTV intake v2 packages."""

from __future__ import annotations

import hashlib
import json
from io import BytesIO
from datetime import datetime
from dataclasses import dataclass, field
import platform
from types import MappingProxyType
from typing import Mapping
import zlib

import fitz
import openpyxl
from PIL import __version__ as pillow_version
from PIL import features as pillow_features
from pydantic import ValidationError

from ctv_inspection_model import InspectionResult
from ctv_inspection_model import InspectionLimits
from ctv_inspection_media import (
    PackageImageError,
    PdfParserBoundaryExceededError,
    _normalize_package_image,
    _prove_pdf_page_bounds,
)
from ctv_inspection_workbook import (
    PackageWorkbookError,
    _canonical_package_workbook_bytes,
    selected_worksheet_values,
)
from ctv_inventory import InventoryObservation
from ctv_package_assignment import AssignmentBuildResult, build_assignments
from ctv_proposal import ApprovedProposalSnapshot
from intake_contract_v2 import (
    AssignmentsArtifactV2,
    CanonicalRosterRowV2,
    CanonicalRosterValuesV2,
    CanonicalSourceColumnsV2,
    EvidenceArtifactV2,
    ExceptionsArtifactV2,
    ExceptionsDocumentV2,
    ImageLocatorV2,
    MAX_EVIDENCE_ARTIFACTS,
    MAX_INPUT_PDF_BYTES,
    MAX_JSON_BYTES,
    MAX_PACKAGE_BYTES,
    MAX_PACKAGE_PDF_PAGES,
    MAX_ROSTER_OR_EVIDENCE_BYTES,
    InputPdfArtifactV2,
    PackageManifestV2,
    PdfPageV2,
    PdfPageLocatorV2,
    RosterArtifactV2,
    RosterLocatorV2,
    RosterMappingV2,
    UnacquiredSourceV2,
    VerifiedSourceV2,
    WorksheetLocatorV2,
)


_VALIDATOR_VERSION = "ctv-intake-v2-validator/1.0"
_MAX_SOURCE_WORK_BYTES = MAX_PACKAGE_BYTES
_ROSTER_HEADERS = (
    "Roster Row ID",
    "Name",
    "Identity",
    "FA Code",
    "Tax ID",
    "Birth Date",
    "Bank Account",
    "Service Fee",
    "Product",
)
_ROSTER_FIELDS = (
    "name",
    "identity",
    "fa_code",
    "tax_id",
    "birth_date",
    "bank_account",
    "service_fee",
    "product",
)
class _OutputLimitExceeded(RuntimeError):
    def __init__(self, code: str = "package-artifact-over-limit") -> None:
        self.code = code
        super().__init__(code)


class _CappedBytesIO(BytesIO):
    def __init__(self, limit: int, code: str = "package-artifact-over-limit") -> None:
        super().__init__()
        self.limit = limit
        self.code = code
        self.crossed = False

    def write(self, value: bytes) -> int:
        end = self.tell() + len(value)
        if end > self.limit:
            self.crossed = True
            raise _OutputLimitExceeded(self.code)
        return super().write(value)


def _identity_digest(domain: str, payload: bytes) -> str:
    return hashlib.sha256(f"ctv-{domain}-identity-v2".encode() + b"\0" + payload).hexdigest()


@dataclass(frozen=True)
class PackageIdentity:
    digest: str
    package_id: str
    batch_id: str
    case_id: str
    final_directory: str

    @classmethod
    def derive(
        cls,
        observation_id: str,
        proposal_digest: str,
        writer_version: str,
        schema_version: str,
        compatibility_target: str,
    ) -> "PackageIdentity":
        payload = json.dumps(
            {
                "compatibilityTarget": compatibility_target,
                "proposalDigest": proposal_digest,
                "schemaVersion": schema_version,
                "sourceObservationId": observation_id,
                "writerVersion": writer_version,
            },
            ensure_ascii=False,
            separators=(",", ":"),
            sort_keys=True,
        ).encode("utf-8")
        package_digest = _identity_digest("package", payload)
        batch_digest = _identity_digest("batch", payload)
        case_digest = _identity_digest("case", payload)
        return cls(
            digest=package_digest,
            package_id="package-" + package_digest,
            batch_id="batch-" + batch_digest,
            case_id="case-" + case_digest,
            final_directory="ctv-package-" + package_digest[:24],
        )


def _mupdf_fingerprint() -> object:
    return {
        "binding": fitz.VersionBind,
        "native": getattr(fitz, "VersionFitz", "unknown"),
    }


def _lxml_fingerprint() -> object:
    try:
        import lxml.etree as etree
    except ImportError:
        return {"enabled": False, "version": "absent"}
    return {
        "enabled": bool(getattr(__import__("openpyxl.xml.functions", fromlist=["LXML"]), "LXML")),
        "lxml": tuple(etree.LXML_VERSION),
        "libxml-runtime": tuple(etree.LIBXML_VERSION),
        "libxml-compiled": tuple(etree.LIBXML_COMPILED_VERSION),
        "libxslt-runtime": tuple(etree.LIBXSLT_VERSION),
    }


def _zlib_fingerprint() -> object:
    return {
        "compiled": zlib.ZLIB_VERSION,
        "runtime": getattr(zlib, "ZLIB_RUNTIME_VERSION", "unknown"),
    }


def _pillow_fingerprint() -> object:
    libraries = {}
    for name in ("jpg", "jpg_2000", "zlib", "libtiff", "webp", "libjpeg_turbo"):
        libraries[name] = {
            "enabled": bool(pillow_features.check(name)),
            "version": pillow_features.version(name) or "absent",
        }
    return {"pillow": pillow_version, "libraries": libraries}


def writer_version_string() -> str:
    """Return a path-free digest of every effective serializer dependency."""
    dependency_payload = {
        "lxml": _lxml_fingerprint(),
        "mupdf": _mupdf_fingerprint(),
        "openpyxl": openpyxl.__version__,
        "pillow": _pillow_fingerprint(),
        "python": {
            "implementation": platform.python_implementation(),
            "version": platform.python_version(),
        },
        "zlib": _zlib_fingerprint(),
    }
    fingerprint = hashlib.sha256(json.dumps(
        dependency_payload,
        ensure_ascii=True,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("ascii")).hexdigest()
    return f"ctv-package-writer/2.0;deps={fingerprint}"


def _opaque_id(prefix: str, digest: str, value: str) -> str:
    token = hashlib.sha256(f"{digest}:{value}".encode("utf-8")).hexdigest()[:32]
    return f"{prefix}-{token}"


def _source_id(proposal_digest: str, evidence_id: str) -> str:
    return _opaque_id("source", proposal_digest, evidence_id)


@dataclass(frozen=True)
class PlannedPdfPage:
    unit_id: str
    evidence_id: str
    source_id: str
    source_page: int
    target_page: int
    coverage_state: str

    @property
    def source_page_key(self) -> tuple[str, int]:
        return self.source_id, self.source_page


@dataclass(frozen=True)
class ArtifactRecipe:
    artifact_id: str
    kind: str
    path: str
    source_ids: tuple[str, ...]
    evidence_id: str | None = None
    source_unit_indexes: tuple[int, ...] = ()


@dataclass(frozen=True)
class RenderedArtifact:
    artifact_id: str
    kind: str
    path: str
    source_ids: tuple[str, ...]
    content: bytes = field(repr=False)

    @property
    def size(self) -> int:
        return len(self.content)


@dataclass(frozen=True)
class ArtifactReceipt:
    artifact_id: str
    kind: str
    path: str
    source_ids: tuple[str, ...]
    size: int
    sha256: str

    @classmethod
    def from_rendered(cls, artifact: RenderedArtifact) -> "ArtifactReceipt":
        if type(artifact) is not RenderedArtifact:
            raise TypeError("artifact receipt requires one rendered artifact")
        return cls(
            artifact_id=artifact.artifact_id,
            kind=artifact.kind,
            path=artifact.path,
            source_ids=artifact.source_ids,
            size=len(artifact.content),
            sha256=hashlib.sha256(artifact.content).hexdigest(),
        )


class PackageBuildError(RuntimeError):
    """Stable package planning/rendering boundary with no private details."""

    def __init__(self, code: str) -> None:
        if code not in {
            "package-source-over-limit",
            "package-artifact-over-limit",
            "package-aggregate-over-limit",
            "package-pdf-unavailable",
            "package-roster-unavailable",
            "package-evidence-unavailable",
            "package-receipt-invalid",
            "package-plan-invalid",
        }:
            raise ValueError("package build error code must be fixed")
        super().__init__(code)


@dataclass(frozen=True)
class PackageBuildPlan:
    identity: PackageIdentity
    observation_id: str
    proposal_digest: str
    writer_version: str
    schema_version: str
    compatibility_target: str
    pdf_pages: tuple[PlannedPdfPage, ...]
    manifest_pdf_pages: tuple[PdfPageV2, ...]
    pdf_trailer_id: str
    recipes: tuple[ArtifactRecipe, ...]
    locators: Mapping[str, object]
    assignments: AssignmentBuildResult
    sources: tuple[object, ...]
    roster_mapping: RosterMappingV2
    approved: ApprovedProposalSnapshot
    inspection: InspectionResult


def _artifact_recipe(
    identity: PackageIdentity,
    kind: str,
    path: str,
    source_ids: tuple[str, ...],
    *,
    evidence_id: str | None = None,
    source_unit_indexes: tuple[int, ...] = (),
) -> ArtifactRecipe:
    return ArtifactRecipe(
        artifact_id=_opaque_id("artifact", identity.digest, f"{kind}:{path}"),
        kind=kind,
        path=path,
        source_ids=source_ids,
        evidence_id=evidence_id,
        source_unit_indexes=source_unit_indexes,
    )


def _pdf_order(item, source_order: Mapping[str, int], participant_order: Mapping[str, int]):
    scope_rank = {"case": 0, "shared": 1, "individual": 2}[item.scope]
    participant_rank = (
        participant_order[item.participant_handles[0]]
        if item.scope == "individual"
        else -1
    )
    return scope_rank, participant_rank, source_order[item.evidence_id], item.unit_index


def _source_path(observed) -> str:
    suffix = observed.extension if observed.extension.startswith(".") else ""
    return f"sources/{observed.evidence_id}{suffix}"


def _source_coverage(items) -> str:
    included = [item for item in items if item.decision in {"accepted", "reassigned"}]
    if included:
        return "shared" if all(item.scope == "shared" for item in included) else "assigned"
    return "duplicate" if items and all(item.reason == "duplicate" for item in items) else "excluded-by-user"


def _original_media_type(source, _observed) -> str:
    if source.detected_type == "pdf":
        return "application/pdf"
    # The reviewed inventory/inspection interface retains only a broad family,
    # not the exact image codec or OOXML content type. Do not promote an
    # extension guess to verified MIME provenance.
    return "application/octet-stream"


def _approved_shapes_are_bounded(approved: ApprovedProposalSnapshot) -> bool:
    if (
        type(approved.fa_code) is not str
        or not 1 <= len(approved.fa_code) <= 128
        or not any(not character.isspace() for character in approved.fa_code)
    ):
        return False
    limits = {
        "name": 256,
        "identity": 256,
        "fa_code": 128,
        "tax_id": 128,
        "birth_date": 128,
        "bank_account": 128,
        "service_fee": 128,
        "product": 256,
    }
    required = {"name", "identity", "fa_code"}
    for row in approved.roster_rows:
        for attribute, maximum in limits.items():
            value = getattr(row, attribute)
            if type(value) is not str or len(value) > maximum:
                return False
            if attribute in required and not value:
                return False
    for field_name, source_column in approved.canonical_to_source_columns:
        if (
            type(field_name) is not str
            or type(source_column) is not str
            or not 1 <= len(source_column) <= 128
        ):
            return False
    return True


def _pdf_trailer_id(identity: PackageIdentity, pages: tuple[PlannedPdfPage, ...]) -> str:
    payload = _canonical_json_bytes([
        {
            "sourceId": page.source_id,
            "sourcePage": page.source_page,
            "targetPage": page.target_page,
        }
        for page in pages
    ])
    return hashlib.sha256(b"ctv-pdf-trailer-v2\0" + identity.digest.encode() + payload).hexdigest()[:32]


def _decision_by_subject(assignments: AssignmentBuildResult) -> dict[str, object]:
    result = {}
    for decision in assignments.decisions:
        if len(decision.subject_refs) == 1 and decision.type in {
            "accept-unit", "reassign-unit", "exclude-unit", "exclude-source"
        }:
            result[decision.subject_refs[0]] = decision
    return result


def _create_build_plan(
    observation: InventoryObservation,
    inspection: InspectionResult,
    approved: ApprovedProposalSnapshot,
) -> PackageBuildPlan:
    """Fix all identities, artifact recipes, ordering, and output locators without I/O."""
    if type(observation) is not InventoryObservation:
        raise TypeError("build planning requires a retained inventory observation")
    if type(inspection) is not InspectionResult or type(approved) is not ApprovedProposalSnapshot:
        raise TypeError("build planning requires reviewed inspection and approval values")
    if observation.observation_id != inspection.observation_id or approved.observation_id != inspection.observation_id:
        raise ValueError("build inputs must share one observation identity")
    if not _approved_shapes_are_bounded(approved):
        raise PackageBuildError("package-plan-invalid")

    writer_version = writer_version_string()
    identity = PackageIdentity.derive(
        observation_id=observation.observation_id,
        proposal_digest=approved.proposal_digest,
        writer_version=writer_version,
        schema_version="2.0",
        compatibility_target="ctv-intake-v2",
    )
    source_order = {source.evidence_id: index for index, source in enumerate(inspection.sources)}
    participant_order = {
        row.participant_handle: index for index, row in enumerate(approved.roster_rows)
    }
    included = tuple(
        item for item in approved.unit_decisions
        if item.decision in {"accepted", "reassigned"}
    )
    if any(item.evidence_id not in source_order for item in included):
        raise ValueError("approved units must resolve to inspected sources")
    pdf_units = sorted(
        (item for item in included if item.unit_kind == "pdf-page"),
        key=lambda item: _pdf_order(item, source_order, participant_order),
    )
    if not pdf_units:
        raise ValueError("prepared package requires an included PDF page")
    if len(pdf_units) > MAX_PACKAGE_PDF_PAGES:
        raise ValueError("prepared package exceeds the PDF page ceiling")
    pdf_keys = [(item.evidence_id, item.unit_index) for item in pdf_units]
    if len(pdf_keys) != len(set(pdf_keys)):
        raise ValueError("included PDF source pages must be unique")

    included_pdf_evidence = tuple(
        source.evidence_id for source in inspection.sources
        if any(item.evidence_id == source.evidence_id for item in pdf_units)
    )
    input_recipe = _artifact_recipe(
        identity,
        "input-pdf",
        "input.pdf",
        tuple(_source_id(approved.proposal_digest, value) for value in included_pdf_evidence),
    )
    roster_source_id = _source_id(approved.proposal_digest, approved.roster_evidence_id)
    roster_recipe = _artifact_recipe(
        identity,
        "roster",
        "roster.xlsx",
        (roster_source_id,),
        evidence_id=approved.roster_evidence_id,
        source_unit_indexes=(approved.roster_worksheet_index,),
    )

    locators: dict[str, object] = {}
    pdf_pages = []
    for target_page, item in enumerate(pdf_units, start=1):
        locators[item.unit_id] = PdfPageLocatorV2(
            kind="pdf-page", artifactId=input_recipe.artifact_id, targetPage=target_page
        )
        pdf_pages.append(PlannedPdfPage(
            unit_id=item.unit_id,
            evidence_id=item.evidence_id,
            source_id=_source_id(approved.proposal_digest, item.evidence_id),
            source_page=item.unit_index,
            target_page=target_page,
            coverage_state="shared" if item.scope == "shared" else "assigned",
        ))
    locators[approved.roster_unit_id] = RosterLocatorV2(
        kind="roster", artifactId=roster_recipe.artifact_id, worksheetIndex=1
    )

    evidence_recipes = []
    evidence_number = 0
    for source in inspection.sources:
        source_units = sorted(
            (
                item for item in included
                if item.evidence_id == source.evidence_id
                and item.unit_id != approved.roster_unit_id
                and item.unit_kind in {"image", "worksheet"}
            ),
            key=lambda item: item.unit_index,
        )
        if not source_units:
            continue
        unit_kinds = {item.unit_kind for item in source_units}
        if len(unit_kinds) != 1:
            raise ValueError("one evidence source must use one transformation kind")
        evidence_number += 1
        if evidence_number > MAX_EVIDENCE_ARTIFACTS:
            raise ValueError("prepared package exceeds the evidence artifact ceiling")
        unit_kind = source_units[0].unit_kind
        suffix = "png" if unit_kind == "image" else "xlsx"
        recipe = _artifact_recipe(
            identity,
            "evidence",
            f"evidence/evidence-{evidence_number:04d}.{suffix}",
            (_source_id(approved.proposal_digest, source.evidence_id),),
            evidence_id=source.evidence_id,
            source_unit_indexes=tuple(item.unit_index for item in source_units),
        )
        evidence_recipes.append(recipe)
        for output_index, item in enumerate(source_units, start=1):
            if unit_kind == "image":
                if len(source_units) != 1 or item.unit_index != 1:
                    raise ValueError("included image source must contain exactly one image unit")
                locators[item.unit_id] = ImageLocatorV2(
                    kind="image", artifactId=recipe.artifact_id
                )
            else:
                locators[item.unit_id] = WorksheetLocatorV2(
                    kind="worksheet",
                    artifactId=recipe.artifact_id,
                    worksheetIndex=output_index,
                )
    if set(locators) != {item.unit_id for item in included}:
        raise ValueError("included units must all have normative output transformations")

    assignments = build_assignments(
        approved,
        package_id=identity.package_id,
        locators=locators,
    )
    assignments_recipe = _artifact_recipe(
        identity, "assignments", "assignments.json", ()
    )
    exceptions_recipe = _artifact_recipe(
        identity, "exceptions", "exceptions.json", ()
    )
    recipes = (
        input_recipe,
        roster_recipe,
        *evidence_recipes,
        assignments_recipe,
        exceptions_recipe,
    )
    decisions_by_subject = _decision_by_subject(assignments)
    observed_by_id = {source.evidence_id: source for source in observation.sources}
    inventory_by_id = {item.evidence_id: item for item in observation.result.items}
    units_by_source = {
        source.evidence_id: tuple(
            item for item in approved.unit_decisions if item.evidence_id == source.evidence_id
        )
        for source in inspection.sources
    }
    dispositions = {item.evidence_id: item for item in approved.source_dispositions}
    source_models = []
    for source in inspection.sources:
        source_id = _source_id(approved.proposal_digest, source.evidence_id)
        observed = observed_by_id[source.evidence_id]
        inventory = inventory_by_id[source.evidence_id]
        if source.evidence_id in dispositions:
            disposition = dispositions[source.evidence_id]
            exclusion = decisions_by_subject.get(source_id)
            if exclusion is None:
                raise ValueError("source disposition decision must resolve")
            source_models.append(UnacquiredSourceV2(
                bindingStatus="unacquired-exclusion",
                sourceId=source_id,
                path=_source_path(observed),
                acquisitionStatus=disposition.acquisition_status,
                issueCodes=list(disposition.issue_codes),
                coverageState=disposition.coverage_state,
                decisionId=exclusion.decision_id,
            ))
            continue
        if (
            source.inspection_status != "inspected"
            or type(inventory.size) is not int
            or inventory.sha256 is None
            or inventory.hash_status != "computed"
        ):
            raise ValueError("materialized sources require verified inventory content")
        media_type = _original_media_type(source, observed)
        source_models.append(VerifiedSourceV2(
            bindingStatus="verified-content",
            sourceId=source_id,
            path=_source_path(observed),
            mediaType=media_type,
            size=inventory.size,
            sha256=inventory.sha256,
            pageCount=source.unit_count if source.detected_type == "pdf" else None,
            coverageState=_source_coverage(units_by_source[source.evidence_id]),
            decisionId=None,
        ))

    planned_by_key = {page.source_page_key: page for page in pdf_pages}
    manifest_pdf_pages = []
    for item in approved.unit_decisions:
        if item.unit_kind != "pdf-page":
            continue
        source_id = _source_id(approved.proposal_digest, item.evidence_id)
        planned = planned_by_key.get((source_id, item.unit_index))
        decision = decisions_by_subject.get(item.unit_id)
        if decision is None:
            raise ValueError("PDF decision must resolve")
        manifest_pdf_pages.append(PdfPageV2(
            sourceId=source_id,
            sourcePage=item.unit_index,
            targetPage=None if planned is None else planned.target_page,
            coverageState=(
                planned.coverage_state if planned is not None
                else "duplicate" if item.reason == "duplicate"
                else "excluded-by-user"
            ),
            decisionId=decision.decision_id,
        ))
    canonical_columns = CanonicalSourceColumnsV2.model_validate(
        dict(approved.canonical_to_source_columns)
    )
    roster_mapping = RosterMappingV2(
        sourceId=roster_source_id,
        sheetName="Roster",
        canonicalToSourceColumns=canonical_columns,
    )
    return PackageBuildPlan(
        identity=identity,
        observation_id=observation.observation_id,
        proposal_digest=approved.proposal_digest,
        writer_version=writer_version,
        schema_version="2.0",
        compatibility_target="ctv-intake-v2",
        pdf_pages=tuple(pdf_pages),
        manifest_pdf_pages=tuple(manifest_pdf_pages),
        pdf_trailer_id=_pdf_trailer_id(identity, tuple(pdf_pages)),
        recipes=tuple(recipes),
        locators=MappingProxyType(dict(locators)),
        assignments=assignments,
        sources=tuple(source_models),
        roster_mapping=roster_mapping,
        approved=approved,
        inspection=inspection,
    )


_CONTRACT_FAILURE = object()


def _contract_result(operation):
    try:
        return operation()
    except ValidationError:
        return _CONTRACT_FAILURE


def _render_result(operation):
    try:
        return "ok", operation()
    except _OutputLimitExceeded as error:
        return "limit", error.code
    except (PackageImageError, PackageWorkbookError) as error:
        return "helper-error", str(error)
    except ValidationError:
        return "helper-error", None


def _manifest_result(operation):
    try:
        return "ok", operation()
    except (ValidationError, ValueError):
        return "contract-error", None
    except _OutputLimitExceeded as error:
        return "limit", error.code


def create_build_plan(
    observation: InventoryObservation,
    inspection: InspectionResult,
    approved: ApprovedProposalSnapshot,
) -> PackageBuildPlan:
    """Fix identities and recipes while suppressing contract diagnostics."""
    result = _contract_result(lambda: _create_build_plan(observation, inspection, approved))
    if result is _CONTRACT_FAILURE:
        raise PackageBuildError("package-plan-invalid")
    return result


def _canonical_json_bytes(
    value: object,
    *,
    max_bytes: int = MAX_JSON_BYTES,
    limit_code: str = "package-artifact-over-limit",
) -> bytes:
    output = _CappedBytesIO(max_bytes, limit_code)
    encoder = json.JSONEncoder(
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    )
    for chunk in encoder.iterencode(value):
        output.write(chunk.encode("utf-8"))
    output.write(b"\n")
    return output.getvalue()


def _clean_workbook() -> openpyxl.Workbook:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    fixed = datetime(1980, 1, 1)
    workbook.properties.creator = ""
    workbook.properties.lastModifiedBy = ""
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    workbook.properties.title = ""
    workbook.properties.subject = ""
    workbook.properties.description = ""
    workbook.properties.keywords = ""
    workbook.properties.category = ""
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    return workbook


def _set_value(cell, value: object) -> None:
    cell.value = value
    if type(value) is str and value.startswith("="):
        cell.data_type = "s"


def _roster_snapshot_values(
    plan: PackageBuildPlan,
    snapshot: bytes,
) -> tuple[dict[str, str | None], ...]:
    values = selected_worksheet_values(
        snapshot,
        (plan.approved.roster_worksheet_index,),
        limits=InspectionLimits(),
    )[0]
    rows = values.rows
    if not rows or not plan.approved.roster_rows:
        raise PackageBuildError("package-roster-unavailable")
    source_columns = dict(plan.approved.canonical_to_source_columns)
    first_data_index = min(row.row_index for row in plan.approved.roster_rows) - 1
    header_index = None
    positions = None
    for index, row in enumerate(rows[:first_data_index]):
        candidate = {}
        for field_name, label in source_columns.items():
            matches = [position for position, value in enumerate(row) if str(value) == label]
            if len(matches) != 1:
                break
            candidate[field_name] = matches[0]
        else:
            header_index = index
            positions = candidate
            break
    if header_index is None or positions is None:
        raise PackageBuildError("package-roster-unavailable")
    attribute_by_field = {
        "name": "name", "identity": "identity", "faCode": "fa_code",
        "taxId": "tax_id", "birthDate": "birth_date", "bankAccount": "bank_account",
        "serviceFee": "service_fee", "product": "product",
    }
    normalized_rows = []
    for roster_row in plan.approved.roster_rows:
        if roster_row.row_index > len(rows):
            raise PackageBuildError("package-roster-unavailable")
        source_row = rows[roster_row.row_index - 1]
        normalized = {}
        for field_name, position in positions.items():
            actual = (
                ""
                if position >= len(source_row) or source_row[position] is None
                else str(source_row[position]).strip()
            )
            approved_value = getattr(roster_row, attribute_by_field[field_name])
            if (
                field_name in {"name", "identity", "faCode"}
                or approved_value
            ) and actual != approved_value:
                raise PackageBuildError("package-roster-unavailable")
            normalized[field_name] = actual or None
        normalized_rows.append(normalized)
    return tuple(normalized_rows)


def _render_roster_unchecked(
    plan: PackageBuildPlan,
    snapshot: bytes,
    max_output_bytes: int,
) -> bytes:
    normalized_rows = _roster_snapshot_values(plan, snapshot)
    workbook = _clean_workbook()
    worksheet = workbook.create_sheet("Roster")
    for column, header in enumerate(_ROSTER_HEADERS, start=1):
        _set_value(worksheet.cell(1, column), header)
    if len(plan.assignments.document.participants) != len(plan.approved.roster_rows):
        raise PackageBuildError("package-roster-unavailable")
    for output_row, (participant, source_row, normalized) in enumerate(
        zip(
            plan.assignments.document.participants,
            plan.approved.roster_rows,
            normalized_rows,
        ),
        start=2,
    ):
        canonical_values = {
            "name": normalized["name"],
            "identity": normalized["identity"],
            "fa_code": normalized["faCode"],
            "tax_id": normalized.get("taxId"),
            "birth_date": normalized.get("birthDate"),
            "bank_account": normalized.get("bankAccount"),
            "service_fee": normalized.get("serviceFee"),
            "product": normalized.get("product"),
        }
        values = (participant.roster_row_id,) + tuple(
            canonical_values[field_name] for field_name in _ROSTER_FIELDS
        )
        CanonicalRosterRowV2(
            rosterRowId=participant.roster_row_id,
            values=CanonicalRosterValuesV2(
                name=canonical_values["name"],
                identity=canonical_values["identity"],
                faCode=canonical_values["fa_code"],
                taxId=canonical_values["tax_id"],
                birthDate=canonical_values["birth_date"],
                bankAccount=canonical_values["bank_account"],
                serviceFee=canonical_values["service_fee"],
                product=canonical_values["product"],
            ),
        )
        for column, value in enumerate(values, start=1):
            _set_value(worksheet.cell(output_row, column), value)
    return _canonical_package_workbook_bytes(workbook, max_bytes=max_output_bytes)


def _render_roster(
    plan: PackageBuildPlan,
    snapshot: bytes,
    max_output_bytes: int,
) -> bytes:
    result = _contract_result(
        lambda: _render_roster_unchecked(plan, snapshot, max_output_bytes)
    )
    if result is _CONTRACT_FAILURE:
        raise PackageBuildError("package-roster-unavailable")
    return result


def _render_evidence_workbook(
    snapshot: bytes,
    indexes: tuple[int, ...],
    max_output_bytes: int,
) -> bytes:
    selected = selected_worksheet_values(snapshot, indexes, limits=InspectionLimits())
    workbook = _clean_workbook()
    for output_index, worksheet_values in enumerate(selected, start=1):
        worksheet = workbook.create_sheet(f"Worksheet {output_index:04d}")
        for row_index, row in enumerate(worksheet_values.rows, start=1):
            for column_index, value in enumerate(row, start=1):
                _set_value(worksheet.cell(row_index, column_index), value)
    return _canonical_package_workbook_bytes(workbook, max_bytes=max_output_bytes)


def _render_pdf(
    plan: PackageBuildPlan,
    observation: InventoryObservation,
    source_charge,
    max_output_bytes: int,
    limit_code: str,
) -> bytes:
    limits = InspectionLimits()
    snapshots = {}
    documents = {}
    output = fitz.open()
    rendered = None
    try:
        for source in plan.inspection.sources:
            if not any(page.evidence_id == source.evidence_id for page in plan.pdf_pages):
                continue
            snapshots[source.evidence_id] = source_charge(
                source.evidence_id, limits.max_pdf_source_bytes
            )
            document = fitz.open(stream=snapshots[source.evidence_id], filetype="pdf")
            if document.needs_pass or document.page_count != source.unit_count:
                raise PackageBuildError("package-pdf-unavailable")
            documents[source.evidence_id] = document
        for page_plan in plan.pdf_pages:
            document = documents[page_plan.evidence_id]
            page = document.load_page(page_plan.source_page - 1)
            _prove_pdf_page_bounds(document, page, limits)
            output.insert_pdf(
                document,
                from_page=page_plan.source_page - 1,
                to_page=page_plan.source_page - 1,
                links=True,
                annots=True,
            )
        output.set_metadata({})
        output.xref_set_key(
            -1,
            "ID",
            f"[<{plan.pdf_trailer_id}><{plan.pdf_trailer_id}>]",
        )
        rendered = _CappedBytesIO(max_output_bytes, limit_code)
        output.save(
            rendered,
            garbage=4,
            clean=1,
            deflate=1,
            no_new_id=1,
        )
        return rendered.getvalue()
    except PackageBuildError:
        raise
    except PdfParserBoundaryExceededError:
        raise PackageBuildError("package-pdf-unavailable") from None
    except _OutputLimitExceeded:
        raise
    except Exception:
        if rendered is not None and rendered.crossed:
            raise _OutputLimitExceeded(limit_code) from None
        raise PackageBuildError("package-pdf-unavailable") from None
    finally:
        output.close()
        for document in documents.values():
            document.close()
        snapshots.clear()


def _artifact_limit(kind: str) -> int:
    if kind == "input-pdf":
        return MAX_INPUT_PDF_BYTES
    if kind in {"roster", "evidence"}:
        return MAX_ROSTER_OR_EVIDENCE_BYTES
    return MAX_JSON_BYTES


def iter_rendered_artifacts(
    plan: PackageBuildPlan,
    observation: InventoryObservation,
    *,
    _snapshot_source=None,
):
    """Yield one bounded artifact at a time from the retained observation."""
    if type(plan) is not PackageBuildPlan or type(observation) is not InventoryObservation:
        raise TypeError("rendering requires a build plan and retained observation")
    if observation.observation_id != plan.observation_id:
        raise ValueError("rendering observation must match its build plan")
    if _snapshot_source is not None and not callable(_snapshot_source):
        raise TypeError("snapshot source must be callable")
    snapshot_source = observation.snapshot if _snapshot_source is None else _snapshot_source
    observed_by_id = {source.evidence_id: source for source in observation.sources}
    source_work = 0
    output_work = 0

    def source_snapshot(evidence_id: str, maximum: int) -> bytes:
        nonlocal source_work
        source = observed_by_id.get(evidence_id)
        if source is None or type(source.size) is not int or source.size > maximum:
            raise PackageBuildError("package-source-over-limit")
        if source_work + source.size > _MAX_SOURCE_WORK_BYTES:
            raise PackageBuildError("package-aggregate-over-limit")
        source_work += source.size
        return snapshot_source(evidence_id, max_bytes=maximum)

    for recipe in plan.recipes:
        per_file_limit = _artifact_limit(recipe.kind)
        remaining_package_bytes = MAX_PACKAGE_BYTES - output_work
        if remaining_package_bytes <= 0:
            raise PackageBuildError("package-aggregate-over-limit")
        max_output_bytes = min(per_file_limit, remaining_package_bytes)
        limit_code = (
            "package-aggregate-over-limit"
            if remaining_package_bytes < per_file_limit
            else "package-artifact-over-limit"
        )
        if recipe.kind == "input-pdf":
            rendered_result = _render_result(lambda: _render_pdf(
                plan, observation, source_snapshot, max_output_bytes, limit_code
            ))
        elif recipe.kind == "roster":
            assert recipe.evidence_id is not None
            rendered_result = _render_result(lambda: _render_roster(
                plan,
                source_snapshot(recipe.evidence_id, InspectionLimits().max_workbook_source_bytes),
                max_output_bytes,
            ))
        elif recipe.kind == "evidence" and recipe.path.endswith(".png"):
            assert recipe.evidence_id is not None
            rendered_result = _render_result(lambda: _normalize_package_image(
                source_snapshot(recipe.evidence_id, InspectionLimits().max_image_source_bytes),
                limits=InspectionLimits(),
                max_output_bytes=max_output_bytes,
            ))
        elif recipe.kind == "evidence" and recipe.path.endswith(".xlsx"):
            assert recipe.evidence_id is not None
            rendered_result = _render_result(lambda: _render_evidence_workbook(
                source_snapshot(recipe.evidence_id, InspectionLimits().max_workbook_source_bytes),
                recipe.source_unit_indexes,
                max_output_bytes,
            ))
        elif recipe.kind == "assignments":
            rendered_result = _render_result(lambda: _canonical_json_bytes(
                plan.assignments.document.model_dump(by_alias=True),
                max_bytes=max_output_bytes,
                limit_code=limit_code,
            ))
        elif recipe.kind == "exceptions":
            rendered_result = _render_result(lambda: _canonical_json_bytes(
                ExceptionsDocumentV2(schemaVersion="2.0", items=[]).model_dump(by_alias=True),
                max_bytes=max_output_bytes,
                limit_code=limit_code,
            ))
        else:
            raise PackageBuildError("package-evidence-unavailable")
        if rendered_result[0] == "limit":
            raise PackageBuildError(rendered_result[1])
        if rendered_result[0] == "helper-error":
            if rendered_result[1] is not None and rendered_result[1].endswith("over-limit"):
                raise PackageBuildError(limit_code)
            raise PackageBuildError(
                "package-roster-unavailable"
                if recipe.kind == "roster"
                else "package-evidence-unavailable"
            )
        content = rendered_result[1]
        output_work += len(content)
        rendered = RenderedArtifact(
            artifact_id=recipe.artifact_id,
            kind=recipe.kind,
            path=recipe.path,
            source_ids=recipe.source_ids,
            content=content,
        )
        yield rendered
        del rendered
        content = b""


def _build_manifest_bytes(
    plan: PackageBuildPlan,
    receipts: tuple[ArtifactReceipt, ...],
) -> bytes:
    """Bind exact transaction receipts into one canonical v2 manifest."""
    if type(plan) is not PackageBuildPlan or type(receipts) is not tuple:
        raise TypeError("manifest construction requires a build plan and exact receipts")
    if len(receipts) != len(plan.recipes):
        raise PackageBuildError("package-receipt-invalid")
    artifacts = []
    model_by_kind = {
        "input-pdf": InputPdfArtifactV2,
        "roster": RosterArtifactV2,
        "assignments": AssignmentsArtifactV2,
        "exceptions": ExceptionsArtifactV2,
        "evidence": EvidenceArtifactV2,
    }
    total_size = 0
    for recipe, receipt in zip(plan.recipes, receipts):
        if (
            type(receipt) is not ArtifactReceipt
            or (
                receipt.artifact_id,
                receipt.kind,
                receipt.path,
                receipt.source_ids,
            ) != (
                recipe.artifact_id,
                recipe.kind,
                recipe.path,
                recipe.source_ids,
            )
            or type(receipt.size) is not int
            or not 0 <= receipt.size <= _artifact_limit(receipt.kind)
            or type(receipt.sha256) is not str
            or len(receipt.sha256) != 64
            or any(character not in "0123456789abcdef" for character in receipt.sha256)
        ):
            raise PackageBuildError("package-receipt-invalid")
        total_size += receipt.size
        if total_size > MAX_PACKAGE_BYTES:
            raise PackageBuildError("package-receipt-invalid")
        artifacts.append(model_by_kind[receipt.kind](
            artifactId=receipt.artifact_id,
            kind=receipt.kind,
            formatVersion="2.0",
            path=receipt.path,
            size=receipt.size,
            sha256=receipt.sha256,
            sourceIds=list(receipt.source_ids),
        ))
    manifest = PackageManifestV2(
        schemaVersion="2.0",
        compatibilityTarget="ctv-intake-v2",
        packageId=plan.identity.package_id,
        sourceObservationId=plan.observation_id,
        proposalDigest=plan.proposal_digest,
        batchId=plan.identity.batch_id,
        caseId=plan.identity.case_id,
        faCode=plan.approved.fa_code,
        packageVersion=plan.writer_version,
        status="prepared",
        validatorVersion=_VALIDATOR_VERSION,
        sources=list(plan.sources),
        pdfPages=list(plan.manifest_pdf_pages),
        artifacts=artifacts,
        rosterMapping=plan.roster_mapping,
        decisions=list(plan.assignments.decisions),
        exceptionIds=[],
    )
    plan.assignments.document.validate_against_manifest(manifest)
    remaining_package_bytes = MAX_PACKAGE_BYTES - total_size
    if remaining_package_bytes <= 0:
        raise PackageBuildError("package-aggregate-over-limit")
    return _canonical_json_bytes(
        manifest.model_dump(by_alias=True),
        max_bytes=min(MAX_JSON_BYTES, remaining_package_bytes),
        limit_code=(
            "package-aggregate-over-limit"
            if remaining_package_bytes < MAX_JSON_BYTES
            else "package-artifact-over-limit"
        ),
    )


def build_manifest_bytes(
    plan: PackageBuildPlan,
    receipts: tuple[ArtifactReceipt, ...],
) -> bytes:
    """Bind receipts while suppressing all contract diagnostics and input echoes."""
    status, result = _manifest_result(lambda: _build_manifest_bytes(plan, receipts))
    if status == "contract-error":
        raise PackageBuildError("package-receipt-invalid")
    if status == "limit":
        raise PackageBuildError(result)
    return result
