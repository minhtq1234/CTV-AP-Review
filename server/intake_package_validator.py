"""Read-only semantic validation for prepared CTV intake packages."""
from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
import hashlib
import io
import json
import os
from pathlib import Path, PurePosixPath
import re
import stat
import weakref

import fitz
import openpyxl
from pydantic import ValidationError

from intake_contract import ExceptionsDocument, PackageManifest, ValidationCheck, ValidationReport
from roster_workbook import MAX_WORKBOOK_BYTES as MAX_ROSTER_WORKBOOK_BYTES
from roster_workbook import preflight_roster_workbook


VALIDATOR_VERSION = "1.0.0"
_MIB = 1024 * 1024
MAX_MANIFEST_BYTES = 16 * _MIB
# V1 package ceilings bound in-memory snapshots. Workbook limits mirror the
# existing app gates; the PDF ceiling is deliberately higher because the
# current upload path has no lower PDF limit, and JSON artifacts stay compact.
MAX_ARTIFACT_BYTES_BY_KIND = {
    "input-pdf": 256 * _MIB,
    "roster": MAX_ROSTER_WORKBOOK_BYTES,
    "cccd": 100 * _MIB,
    "exceptions": 16 * _MIB,
    "validation-report": 16 * _MIB,
}
# Source reads are bounded independently from package artifacts. PDF originals
# use the input-PDF ceiling, workbook originals use the app's workbook ceiling,
# and every other v1 source is capped at the CCCD artifact ceiling.
MAX_OTHER_SOURCE_BYTES = MAX_ARTIFACT_BYTES_BY_KIND["cccd"]
MAX_PDF_PAGES_PER_SOURCE = 10_000
MAX_PDF_PAGES_PER_PACKAGE = 25_000
_MANIFEST_NAME = "case-manifest.json"
_REQUIRED_ARTIFACT_KINDS = frozenset({"input-pdf", "roster", "exceptions"})
_WINDOWS_ABSOLUTE_PATH_RE = re.compile(r"^[A-Za-z]:[\\/]")
_SOURCE_PAGE_EVIDENCE_RE = re.compile(r"^(.+)#page=([1-9][0-9]*)$")
_PDF_HEADER_RE = re.compile(br"%PDF-[0-9]\.[0-9]")
_CLOSE_ON_EXEC = getattr(os, "O_CLOEXEC", 0)
_DIRECTORY = getattr(os, "O_DIRECTORY", 0)
_NO_FOLLOW = getattr(os, "O_NOFOLLOW", 0)
_NON_BLOCKING = getattr(os, "O_NONBLOCK", 0)
_SUPPORTS_SECURE_RELATIVE_OPEN = os.open in os.supports_dir_fd and bool(_NO_FOLLOW)


@dataclass(frozen=True)
class PackageTreeSnapshot:
    """A private-content-free digest of one exact opened package tree."""

    paths: tuple[str, ...]
    file_sha256: tuple[tuple[str, str], ...]
    tree_sha256: str
    total_bytes: int


@dataclass(frozen=True)
class _TreeEntry:
    path: str
    identity: tuple[int, int, int, int, int]


def _make_package_reader_type():
    registry: dict[int, dict[str, object]] = {}
    manifest_not_read = object()

    def close_state(state: dict[str, object], *, suppress_errors: bool) -> None:
        if state["closed"]:
            return
        state["closed"] = True
        try:
            os.close(state["descriptor"])
        except OSError:
            if not suppress_errors:
                raise

    def lookup(reader):
        state = registry.get(id(reader))
        if state is None:
            return None
        registered_ref = state["reader_ref"]
        if registered_ref() is not reader:
            return None
        return state

    def register(reader, descriptor: int, root_identity: tuple[int, int]) -> None:
        reader_id = id(reader)
        prior_state = registry.get(reader_id)
        if prior_state is not None:
            prior_ref = prior_state["reader_ref"]
            if prior_ref() is not None:
                raise RuntimeError("live package reader identity collision")
            registry.pop(reader_id, None)
            close_state(prior_state, suppress_errors=True)

        state: dict[str, object] = {
            "descriptor": descriptor,
            "root_identity": root_identity,
            "reader_ref": None,
            "closed": False,
            "manifest_snapshot": manifest_not_read,
            "read_cache": {},
            "read_identities": {},
            "charged_bytes": 0,
        }

        def release(dead_ref) -> None:
            current = registry.get(reader_id)
            if current is not state or current["reader_ref"] is not dead_ref:
                return
            registry.pop(reader_id, None)
            close_state(state, suppress_errors=True)

        reader_ref = weakref.ref(reader, release)
        state["reader_ref"] = reader_ref
        registry[reader_id] = state

    class _PackageReader:
        """A package root whose authority exists only in the factory registry."""

        __slots__ = ("root_path", "__root_fd", "__weakref__")

        def __init__(self, *_args, **_kwargs) -> None:
            raise TypeError("_PackageReader instances must be created by open()")

        @classmethod
        def open(
            cls, package_dir: Path
        ) -> tuple["_PackageReader | None", str | None]:
            root_path = Path(package_dir)
            if not _SUPPORTS_SECURE_RELATIVE_OPEN:
                return None, "secure-open-unavailable"
            try:
                root_fd = os.open(
                    root_path,
                    os.O_RDONLY | _CLOSE_ON_EXEC | _DIRECTORY | _NO_FOLLOW,
                )
            except OSError:
                return None, "symlink" if root_path.is_symlink() else "missing"
            try:
                metadata = os.fstat(root_fd)
                if not stat.S_ISDIR(metadata.st_mode):
                    raise ValueError("package root descriptor is not a directory")
                reader = object.__new__(cls)
                reader.root_path = root_path
                reader.__root_fd = root_fd
                register(
                    reader,
                    root_fd,
                    (metadata.st_dev, metadata.st_ino),
                )
            except (OSError, RuntimeError, ValueError):
                os.close(root_fd)
                return None, "missing"
            return reader, None

        @classmethod
        def open_at(
            cls,
            parent_fd: int,
            child_name: str,
            expected_identity: tuple[int, int] | None = None,
        ) -> tuple["_PackageReader | None", str | None]:
            """Open one exact directory child without resolving a pathname."""
            if (
                not _SUPPORTS_SECURE_RELATIVE_OPEN
                or type(parent_fd) is not int
                or type(child_name) is not str
            ):
                return None, "secure-open-unavailable"
            if (
                not child_name
                or child_name in {".", ".."}
                or "/" in child_name
                or "\\" in child_name
                or "\x00" in child_name
            ):
                return None, "unsafe"
            try:
                parent_metadata = os.fstat(parent_fd)
            except OSError:
                return None, "secure-open-unavailable"
            if not stat.S_ISDIR(parent_metadata.st_mode):
                return None, "secure-open-unavailable"
            try:
                root_fd = os.open(
                    child_name,
                    os.O_RDONLY | _CLOSE_ON_EXEC | _DIRECTORY | _NO_FOLLOW,
                    dir_fd=parent_fd,
                )
            except OSError:
                return (
                    None,
                    "symlink" if _is_symlink_at(parent_fd, child_name) else "missing",
                )
            try:
                metadata = os.fstat(root_fd)
                identity = (metadata.st_dev, metadata.st_ino)
                if not stat.S_ISDIR(metadata.st_mode):
                    raise ValueError("package child is not a directory")
                if expected_identity is not None and identity != expected_identity:
                    os.close(root_fd)
                    return None, "changed"
                reader = object.__new__(cls)
                reader.root_path = Path(child_name)
                reader.__root_fd = root_fd
                register(reader, root_fd, identity)
            except (OSError, RuntimeError, ValueError):
                try:
                    os.close(root_fd)
                except OSError:
                    pass
                return None, "missing"
            return reader, None

        def has_secure_open_provenance(self) -> bool:
            return has_authority(self)

        @property
        def root_fd(self) -> int | None:
            if not self.has_secure_open_provenance():
                return None
            state = lookup(self)
            return state["descriptor"] if state is not None else None

        def close(self) -> None:
            state = lookup(self)
            if hasattr(self, "_PackageReader__root_fd"):
                self.__root_fd = None
            if state is None:
                return
            reader_id = id(self)
            current = registry.get(reader_id)
            if current is not state or current["reader_ref"]() is not self:
                return
            registry.pop(reader_id, None)
            close_state(state, suppress_errors=False)

        def read(
            self,
            declared_path: str,
            *,
            expected_size: int | None = None,
            max_bytes: int,
            _expected_identity: tuple[int, int, int, int, int] | None = None,
            _aggregate_allowance: int | None = None,
        ) -> tuple[bytes | None, str | None]:
            if not self.has_secure_open_provenance():
                return None, "secure-open-unavailable"
            parts = _relative_path_parts(declared_path)
            if parts is None:
                return None, "unsafe"
            state = lookup(self)
            if state is None:
                return None, "secure-open-unavailable"
            identity_before, identity_failure = _regular_identity_relative_to_fd(
                state["descriptor"], parts
            )
            identity_expected = (
                identity_before
                if _expected_identity is None
                else _expected_identity
            )
            attempt_charged = False

            def reserve_attempt(size: int) -> None:
                nonlocal attempt_charged
                if not attempt_charged:
                    state["charged_bytes"] += size
                    attempt_charged = True

            content, failure, opened_identity = _read_relative_to_fd(
                state["descriptor"],
                parts,
                expected_size=expected_size,
                max_bytes=max_bytes,
                expected_identity=identity_expected,
                aggregate_allowance=_aggregate_allowance,
                reserve_attempt=reserve_attempt,
            )
            if not attempt_charged and identity_before is not None:
                reserve_attempt(identity_before[2])
            identity_after, after_failure = _regular_identity_relative_to_fd(
                state["descriptor"], parts
            )
            if (
                failure is None
                and identity_before is not None
                and opened_identity is not None
                and identity_after is not None
                and not (
                    identity_before == opened_identity == identity_after
                )
            ):
                content, failure = None, "changed"
            elif failure is None and (identity_failure or after_failure):
                content, failure = None, "changed"
            state["read_identities"][declared_path] = identity_expected
            return content, failure

        def read_manifest(self) -> tuple[bytes | None, str | None]:
            state = lookup(self)
            if state is None or not has_authority(self):
                return None, "secure-open-unavailable"
            cached = state["manifest_snapshot"]
            if cached is manifest_not_read:
                cached = self.read_cached(
                    _MANIFEST_NAME, max_bytes=MAX_MANIFEST_BYTES
                )
                state["manifest_snapshot"] = cached
            return cached

        def read_cached(
            self,
            declared_path: str,
            *,
            max_bytes: int,
            _expected_identity: tuple[int, int, int, int, int] | None = None,
            _aggregate_allowance: int | None = None,
        ) -> tuple[bytes | None, str | None]:
            """Read one relative path at most once for this opened root."""
            state = lookup(self)
            if state is None or not has_authority(self):
                return None, "secure-open-unavailable"
            cache = state["read_cache"]
            cached = cache.get(declared_path)
            if cached is None:
                cached = self.read(
                    declared_path,
                    expected_size=None,
                    max_bytes=max_bytes,
                    _expected_identity=_expected_identity,
                    _aggregate_allowance=_aggregate_allowance,
                )
                cache[declared_path] = cached
            return cached

        def discard_cached(self, declared_path: str) -> None:
            """Release a source snapshot after all declarations using it finish."""
            state = lookup(self)
            if state is not None and has_authority(self):
                state["read_cache"].pop(declared_path, None)
                state["read_identities"].pop(declared_path, None)

        @property
        def charged_bytes(self) -> int:
            state = lookup(self)
            if state is None or not has_authority(self):
                return 0
            return state["charged_bytes"]

        def snapshot_tree(
            self,
            required_paths: set[str] | frozenset[str],
            *,
            max_bytes_by_path: dict[str, int],
            max_total_bytes: int,
            optional_paths: set[str] | frozenset[str] = frozenset(),
            exclude_from_digest: set[str] | frozenset[str] = frozenset(),
        ) -> tuple[PackageTreeSnapshot | None, str | None]:
            """Snapshot one exact regular allowlisted tree through this reader."""
            state = lookup(self)
            if state is None or not has_authority(self):
                return None, "secure-open-unavailable"
            if (
                not isinstance(required_paths, (set, frozenset))
                or not isinstance(optional_paths, (set, frozenset))
                or not isinstance(max_bytes_by_path, dict)
                or type(max_total_bytes) is not int
                or max_total_bytes <= 0
            ):
                return None, "unsafe"
            required = frozenset(required_paths)
            optional = frozenset(optional_paths)
            allowed = required | optional
            excluded = frozenset(exclude_from_digest)
            if (
                required & optional
                or excluded - allowed
                or set(max_bytes_by_path) != set(allowed)
            ):
                return None, "unsafe"
            if any(not _is_v2_tree_path(path) for path in allowed):
                return None, "unsafe"

            before, failure = _scan_package_tree(state["descriptor"])
            if failure is not None:
                return None, failure
            before_by_path = {entry.path: entry.identity for entry in before}
            actual_paths = set(before_by_path)
            if actual_paths - allowed:
                return None, "extra"
            if required - actual_paths:
                return None, "missing"
            for path, identity in state["read_identities"].items():
                if (
                    path in allowed
                    and identity is not None
                    and before_by_path.get(path) != identity
                ):
                    return None, "changed"

            total = 0
            digests: list[tuple[str, str]] = []
            for path in sorted(actual_paths, key=lambda value: value.encode("utf-8")):
                size = before_by_path[path][2]
                remaining = max_total_bytes - total
                if size > remaining:
                    if path not in state["read_cache"]:
                        state["charged_bytes"] += size
                        state["read_identities"][path] = before_by_path[path]
                    return None, "too-large"
                total += size
                content, read_failure = self.read_cached(
                    path,
                    max_bytes=max_bytes_by_path[path],
                    _expected_identity=before_by_path[path],
                    _aggregate_allowance=remaining,
                )
                if read_failure is not None or content is None:
                    return None, read_failure or "missing"
                if path not in excluded:
                    digests.append((path, hashlib.sha256(content).hexdigest()))

            after, failure = _scan_package_tree(state["descriptor"])
            if failure is not None:
                return None, failure
            if before != after:
                return None, "changed"
            lines = b"".join(
                f"{digest}  {path}\n".encode("utf-8")
                for path, digest in digests
            )
            return PackageTreeSnapshot(
                paths=tuple(path for path, _digest in digests),
                file_sha256=tuple(digests),
                tree_sha256=hashlib.sha256(lines).hexdigest(),
                total_bytes=total,
            ), None

    def has_authority(reader: object) -> bool:
        state = lookup(reader)
        if state is None or not _SUPPORTS_SECURE_RELATIVE_OPEN:
            return False
        descriptor = state["descriptor"]
        root_identity = state["root_identity"]
        if getattr(reader, "_PackageReader__root_fd", None) != descriptor:
            return False
        try:
            metadata = os.fstat(descriptor)
        except OSError:
            return False
        return (
            stat.S_ISDIR(metadata.st_mode)
            and (metadata.st_dev, metadata.st_ino) == root_identity
        )

    return _PackageReader, has_authority


_PackageReader, _has_package_reader_authority = _make_package_reader_type()
_SourceReader, _has_source_reader_authority = _make_package_reader_type()
del _make_package_reader_type


def _relative_path_parts(declared_path: str) -> tuple[str, ...] | None:
    parts = declared_path.split("/")
    pure_path = PurePosixPath(declared_path)
    if (
        not declared_path
        or "\x00" in declared_path
        or pure_path.is_absolute()
        or _WINDOWS_ABSOLUTE_PATH_RE.match(declared_path)
        or "\\" in declared_path
        or any(part in {"", ".", ".."} for part in parts)
    ):
        return None
    return tuple(parts)


def _is_v2_tree_path(path: str) -> bool:
    parts = _relative_path_parts(path)
    return parts is not None and (
        len(parts) == 1 or (len(parts) == 2 and parts[0] == "evidence")
    )


def _identity(metadata: os.stat_result) -> tuple[int, int, int, int, int]:
    return (
        metadata.st_dev,
        metadata.st_ino,
        metadata.st_size,
        metadata.st_mtime_ns,
        metadata.st_ctime_ns,
    )


def _regular_identity_relative_to_fd(
    root_fd: int, parts: tuple[str, ...]
) -> tuple[tuple[int, int, int, int, int] | None, str | None]:
    parent_fd = os.dup(root_fd)
    try:
        for part in parts[:-1]:
            try:
                child_fd = os.open(
                    part,
                    os.O_RDONLY | _CLOSE_ON_EXEC | _DIRECTORY | _NO_FOLLOW,
                    dir_fd=parent_fd,
                )
            except OSError:
                return None, "symlink" if _is_symlink_at(parent_fd, part) else "missing"
            os.close(parent_fd)
            parent_fd = child_fd
        try:
            metadata = os.stat(parts[-1], dir_fd=parent_fd, follow_symlinks=False)
        except OSError:
            return None, "missing"
        if stat.S_ISLNK(metadata.st_mode):
            return None, "symlink"
        if not stat.S_ISREG(metadata.st_mode):
            return None, "not-regular"
        return _identity(metadata), None
    finally:
        os.close(parent_fd)


def _scan_package_tree(
    root_fd: int,
) -> tuple[tuple[_TreeEntry, ...], str | None]:
    entries: list[_TreeEntry] = []
    try:
        root_names = sorted(os.listdir(root_fd), key=os.fsencode)
    except OSError:
        return (), "changed"
    for name in root_names:
        try:
            metadata = os.stat(name, dir_fd=root_fd, follow_symlinks=False)
        except OSError:
            return (), "changed"
        if stat.S_ISLNK(metadata.st_mode):
            return (), "symlink"
        if stat.S_ISDIR(metadata.st_mode):
            if name != "evidence":
                return (), "nested"
            try:
                evidence_fd = os.open(
                    name,
                    os.O_RDONLY | _CLOSE_ON_EXEC | _DIRECTORY | _NO_FOLLOW,
                    dir_fd=root_fd,
                )
            except OSError:
                return (), "changed"
            try:
                for child in sorted(os.listdir(evidence_fd), key=os.fsencode):
                    try:
                        child_metadata = os.stat(
                            child, dir_fd=evidence_fd, follow_symlinks=False
                        )
                    except OSError:
                        return (), "changed"
                    if stat.S_ISLNK(child_metadata.st_mode):
                        return (), "symlink"
                    if stat.S_ISDIR(child_metadata.st_mode):
                        return (), "nested"
                    if not stat.S_ISREG(child_metadata.st_mode):
                        return (), "not-regular"
                    entries.append(
                        _TreeEntry(f"evidence/{child}", _identity(child_metadata))
                    )
            finally:
                os.close(evidence_fd)
            continue
        if not stat.S_ISREG(metadata.st_mode):
            return (), "not-regular"
        entries.append(_TreeEntry(name, _identity(metadata)))
    return tuple(sorted(entries, key=lambda entry: entry.path.encode("utf-8"))), None


def _read_relative_to_fd(
    root_fd: int,
    parts: tuple[str, ...],
    *,
    expected_size: int | None,
    max_bytes: int,
    expected_identity: tuple[int, int, int, int, int] | None = None,
    aggregate_allowance: int | None = None,
    reserve_attempt=None,
) -> tuple[
    bytes | None,
    str | None,
    tuple[int, int, int, int, int] | None,
]:
    parent_fd = os.dup(root_fd)
    try:
        for part in parts[:-1]:
            try:
                child_fd = os.open(
                    part,
                    os.O_RDONLY | _CLOSE_ON_EXEC | _DIRECTORY | _NO_FOLLOW,
                    dir_fd=parent_fd,
                )
            except OSError:
                return (
                    None,
                    "symlink" if _is_symlink_at(parent_fd, part) else "missing",
                    None,
                )
            os.close(parent_fd)
            parent_fd = child_fd

        try:
            file_fd = os.open(
                parts[-1],
                os.O_RDONLY | _CLOSE_ON_EXEC | _NO_FOLLOW | _NON_BLOCKING,
                dir_fd=parent_fd,
            )
        except OSError:
            return (
                None,
                "symlink" if _is_symlink_at(parent_fd, parts[-1]) else "missing",
                None,
            )
        try:
            metadata = os.fstat(file_fd)
            opened_identity = _identity(metadata)
            if reserve_attempt is not None:
                reserve_attempt(metadata.st_size)
            if not stat.S_ISREG(metadata.st_mode):
                return None, "not-regular", opened_identity
            if (
                expected_identity is not None
                and opened_identity != expected_identity
            ):
                return None, "changed", opened_identity
            if (
                aggregate_allowance is not None
                and metadata.st_size > aggregate_allowance
            ):
                return None, "too-large", opened_identity
            if metadata.st_size > max_bytes:
                return None, "too-large", opened_identity
            if expected_size is not None and metadata.st_size != expected_size:
                return None, "size-mismatch", opened_identity
            with os.fdopen(file_fd, "rb", closefd=False) as stream:
                content = stream.read(max_bytes + 1)
            if len(content) > max_bytes:
                return None, "too-large", opened_identity
            if expected_size is not None and len(content) != expected_size:
                return None, "size-mismatch", opened_identity
            return content, None, opened_identity
        except OSError:
            return None, "missing", None
        finally:
            os.close(file_fd)
    finally:
        os.close(parent_fd)


def _is_symlink_at(parent_fd: int, name: str) -> bool:
    try:
        metadata = os.stat(name, dir_fd=parent_fd, follow_symlinks=False)
    except OSError:
        return False
    return stat.S_ISLNK(metadata.st_mode)


@dataclass
class _Issues:
    errors: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    warnings: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))
    passed: dict[str, set[str]] = field(default_factory=lambda: defaultdict(set))

    def error(self, code: str, *evidence_refs: str) -> None:
        self.errors[code].update(evidence_refs)

    def warning(self, code: str, *evidence_refs: str) -> None:
        self.warnings[code].update(evidence_refs)

    def pass_check(self, code: str, *evidence_refs: str) -> None:
        if code not in self.errors and code not in self.warnings:
            self.passed[code].update(evidence_refs)

    def issue_count(self) -> int:
        return sum(len(values) or 1 for values in self.errors.values()) + sum(
            len(values) or 1 for values in self.warnings.values()
        )

    def pass_gate(self, code: str, evidence_ref: str, before: int) -> None:
        if self.issue_count() == before:
            self.pass_check(code, evidence_ref)

    def report(self, manifest: PackageManifest | None) -> ValidationReport:
        error_codes = sorted(self.errors)
        warning_codes = sorted(self.warnings)
        failed_codes = set(error_codes) | set(warning_codes)
        checks = [
            ValidationCheck(
                code=code,
                passed=False,
                evidenceRefs=sorted(self.errors[code] | self.warnings[code]),
            )
            for code in sorted(failed_codes)
        ]
        checks.extend(
            ValidationCheck(code=code, passed=True, evidenceRefs=sorted(evidence))
            for code, evidence in sorted(self.passed.items())
            if code not in failed_codes
        )
        checks.sort(key=lambda check: check.code)
        invalid = bool(error_codes)
        return ValidationReport(
            schemaVersion="1.0",
            outcome="invalid" if invalid else "valid",
            packageStatus=(
                "partially_prepared"
                if invalid or manifest is None
                else manifest.status
            ),
            checks=checks,
            errors=error_codes,
            warnings=warning_codes,
            validatedAt=datetime.now(timezone.utc),
            validatorVersion=VALIDATOR_VERSION,
        )


@dataclass(frozen=True)
class _SourceVerification:
    detected_pdf_source_ids: frozenset[str]
    pdf_source_ids: frozenset[str]
    actual_pdf_page_counts: tuple[tuple[str, int], ...]


def validate_package(
    package_dir: Path, source_root: Path | None = None
) -> ValidationReport:
    """Validate one package and, when available, its immutable source workspace."""
    reader, root_failure = _PackageReader.open(Path(package_dir))
    if reader is None:
        return _report_package_root_failure(root_failure)
    source_reader = None
    source_root_failure = None
    try:
        if source_root is not None:
            source_reader, source_root_failure = _SourceReader.open(Path(source_root))
        return _validate_package_reader(
            reader,
            source_reader=source_reader,
            source_root_failure=source_root_failure,
        )
    finally:
        if source_reader is not None:
            source_reader.close()
        reader.close()


def _report_package_root_failure(root_failure: str | None) -> ValidationReport:
    """Build the stable report for a package root that could not be opened."""
    issues = _Issues()
    if root_failure == "symlink":
        issues.error("symlink-not-allowed", "package-root")
        return issues.report(None)
    if root_failure == "secure-open-unavailable":
        issues.error("secure-open-unavailable", "package-root")
        return issues.report(None)
    issues.error("manifest-invalid", _MANIFEST_NAME)
    return issues.report(None)


def _validate_package_reader(
    reader: _PackageReader,
    *,
    source_reader: _SourceReader | None = None,
    source_root_failure: str | None = None,
) -> ValidationReport:
    """Validate through a caller-owned open reader without closing it."""
    if not _has_package_reader_authority(reader):
        return _report_package_root_failure("secure-open-unavailable")
    if source_reader is not None and not _has_source_reader_authority(source_reader):
        source_reader = None
        source_root_failure = "secure-open-unavailable"
    return _validate_open_package(
        reader,
        _Issues(),
        source_reader=source_reader,
        source_root_failure=source_root_failure,
    )


def _validate_open_package(
    reader: _PackageReader,
    issues: _Issues,
    *,
    source_reader: _SourceReader | None,
    source_root_failure: str | None,
) -> ValidationReport:
    manifest_content, manifest_failure = reader.read_manifest()
    if manifest_failure == "symlink":
        issues.error("symlink-not-allowed", _MANIFEST_NAME)
        return issues.report(None)
    if manifest_content is None:
        issues.error("manifest-invalid", _MANIFEST_NAME)
        return issues.report(None)
    try:
        raw_manifest = _read_json_bytes(manifest_content)
    except (UnicodeError, json.JSONDecodeError):
        issues.error("manifest-invalid", _MANIFEST_NAME)
        return issues.report(None)

    _inspect_raw_artifact_paths(raw_manifest, issues)
    _inspect_raw_source_paths(raw_manifest, issues)
    _inspect_raw_unresolved_coverage(raw_manifest, issues)
    _inspect_raw_compatibility_target(raw_manifest, issues)
    try:
        manifest = PackageManifest.model_validate(raw_manifest)
    except ValidationError:
        issues.error("manifest-invalid", _MANIFEST_NAME)
        return issues.report(None)
    issues.pass_check("manifest-valid", _MANIFEST_NAME)

    source_verification = _verify_sources(
        manifest,
        source_reader=source_reader,
        source_root_failure=source_root_failure,
        issues=issues,
    )

    page_limits_exceeded = _check_page_count_limits(
        manifest, source_verification, issues
    )

    artifact_gate_before = issues.issue_count()
    artifacts_by_kind: dict[str, list] = defaultdict(list)
    for artifact in manifest.artifacts:
        artifacts_by_kind[artifact.kind].append(artifact)
    for kind in sorted(_REQUIRED_ARTIFACT_KINDS - set(artifacts_by_kind)):
        issues.error("missing-required-artifact", kind)
    duplicate_artifact_kinds: set[str] = set()
    for kind, artifacts in sorted(artifacts_by_kind.items()):
        if len(artifacts) > 1:
            duplicate_artifact_kinds.add(kind)
            issues.error(
                "duplicate-artifact-kind",
                *(artifact.artifact_id for artifact in sorted(
                    artifacts, key=lambda item: item.artifact_id
                )),
            )

    usable_artifacts: dict[str, bytes] = {}
    for artifact in sorted(manifest.artifacts, key=lambda item: item.artifact_id):
        if artifact.kind in duplicate_artifact_kinds:
            continue
        content, failure = reader.read(
            artifact.path,
            expected_size=artifact.size,
            max_bytes=MAX_ARTIFACT_BYTES_BY_KIND[artifact.kind],
        )
        if failure == "unsafe":
            issues.error("unsafe-artifact-path", artifact.artifact_id)
            continue
        if failure == "symlink":
            issues.error("symlink-not-allowed", artifact.artifact_id)
            continue
        if failure == "too-large":
            issues.error("artifact-too-large", artifact.artifact_id)
            continue
        if failure == "size-mismatch":
            issues.error("artifact-size-mismatch", artifact.artifact_id)
            continue
        if content is None:
            issues.error("artifact-missing", artifact.artifact_id)
            continue
        digest_matches = hashlib.sha256(content).hexdigest() == artifact.sha256
        if not digest_matches:
            issues.error("artifact-digest-mismatch", artifact.artifact_id)
        if digest_matches:
            usable_artifacts[artifact.artifact_id] = content
    issues.pass_gate("artifacts-valid", "artifacts", artifact_gate_before)

    exceptions_ready = _unique_artifact_is_usable(
        "exceptions", artifacts_by_kind, usable_artifacts
    )
    exceptions_gate_before = issues.issue_count()
    exceptions = _load_exceptions(
        manifest, artifacts_by_kind, usable_artifacts, issues
    )
    if exceptions_ready:
        issues.pass_gate("exceptions-valid", "exceptions.json", exceptions_gate_before)
    pdf_ready = _unique_artifact_is_usable(
        "input-pdf", artifacts_by_kind, usable_artifacts
    )
    pdf_gate_before = issues.issue_count()
    derived_pdf_page_counts = _inspect_pdfs(
        artifacts_by_kind, usable_artifacts, issues
    )
    if pdf_ready:
        issues.pass_gate("pdf-valid", "input.pdf", pdf_gate_before)
    roster_ready = _unique_artifact_is_usable(
        "roster", artifacts_by_kind, usable_artifacts
    )
    roster_gate_before = issues.issue_count()
    _inspect_rosters(
        manifest, artifacts_by_kind, usable_artifacts, issues
    )
    if roster_ready:
        issues.pass_gate("roster-valid", "roster.xlsx", roster_gate_before)
    cccd_gate_before = issues.issue_count()
    _inspect_cccd_workbooks(artifacts_by_kind, usable_artifacts, issues)
    issues.pass_gate("cccd-valid", "cccd", cccd_gate_before)
    report_gate_before = issues.issue_count()
    _inspect_validation_reports(artifacts_by_kind, usable_artifacts, issues)
    issues.pass_gate("validation-report-valid", "validation-report", report_gate_before)
    reference_gate_before = issues.issue_count()
    _check_references(manifest, exceptions, source_verification, issues)
    issues.pass_gate("references-valid", "references", reference_gate_before)
    state_gate_before = issues.issue_count()
    _check_coverage_state_and_decisions(manifest, issues)
    issues.pass_gate("coverage-state-valid", "coverage-state", state_gate_before)
    provenance_gate_before = issues.issue_count()
    _check_input_pdf_provenance(
        manifest, artifacts_by_kind, source_verification.pdf_source_ids, issues
    )
    _check_source_provenance(
        manifest, source_verification.pdf_source_ids, issues
    )
    issues.pass_gate("provenance-valid", "sources", provenance_gate_before)
    coverage_gate_before = issues.issue_count()
    if not page_limits_exceeded:
        _check_pdf_coverage(
            manifest,
            artifacts_by_kind,
            derived_pdf_page_counts,
            source_verification,
            issues,
        )
        issues.pass_gate("coverage-valid", "pdfPages", coverage_gate_before)
    approval_gate_before = issues.issue_count()
    _check_current_approval(manifest, issues)
    issues.pass_gate("approval-valid", manifest.package_version, approval_gate_before)
    exception_status_gate_before = issues.issue_count()
    _check_exceptions(manifest, exceptions, issues)
    _check_unresolved_coverage(manifest, issues)
    issues.pass_gate("package-status-valid", manifest.status, exception_status_gate_before)
    return issues.report(manifest)


def _unique_artifact_is_usable(
    kind: str, artifacts_by_kind: dict[str, list], usable_artifacts: dict[str, bytes]
) -> bool:
    artifacts = artifacts_by_kind.get(kind, [])
    return len(artifacts) == 1 and artifacts[0].artifact_id in usable_artifacts


def _read_json_bytes(content: bytes) -> object:
    return json.loads(content.decode("utf-8"))


def _inspect_raw_artifact_paths(raw_manifest: object, issues: _Issues) -> None:
    if not isinstance(raw_manifest, dict):
        return
    artifacts = raw_manifest.get("artifacts")
    if not isinstance(artifacts, list):
        return
    for index, artifact in enumerate(artifacts):
        if not isinstance(artifact, dict):
            continue
        evidence = artifact.get("artifactId")
        if not isinstance(evidence, str) or not evidence:
            evidence = f"artifact-index-{index}"
        path = artifact.get("path")
        if not isinstance(path, str):
            issues.error("unsafe-artifact-path", evidence)
            continue
        if _relative_path_parts(path) is None:
            issues.error("unsafe-artifact-path", evidence)


def _inspect_raw_source_paths(raw_manifest: object, issues: _Issues) -> None:
    if not isinstance(raw_manifest, dict):
        return
    sources = raw_manifest.get("sources")
    if not isinstance(sources, list):
        return
    for index, source in enumerate(sources):
        if not isinstance(source, dict):
            continue
        evidence = source.get("sourceId")
        if not isinstance(evidence, str) or not evidence:
            evidence = f"source-index-{index}"
        declared_path = source.get("path")
        if not isinstance(declared_path, str) or _relative_path_parts(declared_path) is None:
            issues.error("unsafe-source-path", evidence)


def _source_byte_limit(media_type: str) -> int:
    if media_type == "application/pdf":
        return MAX_ARTIFACT_BYTES_BY_KIND["input-pdf"]
    if media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return MAX_ROSTER_WORKBOOK_BYTES
    return MAX_OTHER_SOURCE_BYTES


def _verify_sources(
    manifest: PackageManifest,
    *,
    source_reader: _SourceReader | None,
    source_root_failure: str | None,
    issues: _Issues,
) -> _SourceVerification:
    declared_pdf_source_ids = {
        source.source_id
        for source in manifest.sources
        if source.media_type == "application/pdf"
    }
    detected_pdf_source_ids: set[str] = set()
    actual_pdf_page_counts: dict[str, int] = {}

    def result() -> _SourceVerification:
        return _SourceVerification(
            detected_pdf_source_ids=frozenset(detected_pdf_source_ids),
            pdf_source_ids=frozenset(
                declared_pdf_source_ids | detected_pdf_source_ids
            ),
            actual_pdf_page_counts=tuple(sorted(actual_pdf_page_counts.items())),
        )

    gate_before = issues.issue_count()
    if not manifest.sources:
        issues.pass_check("sources-valid", "sources")
        return result()
    if source_reader is None:
        code = (
            "source-symlink-not-allowed"
            if source_root_failure == "symlink"
            else "source-verification-unavailable"
        )
        issues.error(code, "source-root")
        return result()

    sources_by_path: dict[str, list] = defaultdict(list)
    for source in manifest.sources:
        sources_by_path[source.path].append(source)

    for declared_path, sources in sorted(sources_by_path.items()):
        read_limit = max(
            _source_byte_limit(source.media_type) for source in sources
        )
        try:
            for source in sorted(sources, key=lambda item: item.source_id):
                content, failure = source_reader.read_cached(
                    declared_path,
                    max_bytes=read_limit,
                )
                if failure == "unsafe":
                    issues.error("unsafe-source-path", source.source_id)
                    continue
                if failure == "symlink":
                    issues.error("source-symlink-not-allowed", source.source_id)
                    continue
                if failure == "too-large":
                    issues.error("source-too-large", source.source_id)
                    continue
                if content is None:
                    issues.error("source-missing", source.source_id)
                    continue

                if len(content) > _source_byte_limit(source.media_type):
                    issues.error("source-too-large", source.source_id)
                    continue
                size_matches = len(content) == source.size
                if not size_matches:
                    issues.error("source-size-mismatch", source.source_id)
                digest_matches = hashlib.sha256(content).hexdigest() == source.sha256
                if not digest_matches:
                    issues.error("source-digest-mismatch", source.source_id)
                if not size_matches or not digest_matches:
                    if source.media_type == "application/pdf":
                        _verify_source_pdf(source, content, issues)
                    continue

                detected_pdf = _snapshot_is_pdf(content)
                if detected_pdf:
                    detected_pdf_source_ids.add(source.source_id)
                    if source.media_type != "application/pdf":
                        issues.error(
                            "source-media-type-mismatch", source.source_id
                        )
                if detected_pdf or source.media_type == "application/pdf":
                    actual_page_count = _verify_source_pdf(source, content, issues)
                    if actual_page_count is not None:
                        detected_pdf_source_ids.add(source.source_id)
                        actual_pdf_page_counts[source.source_id] = actual_page_count
        finally:
            source_reader.discard_cached(declared_path)

    issues.pass_gate("sources-valid", "sources", gate_before)
    return result()


def _snapshot_is_pdf(content: bytes) -> bool:
    try:
        with fitz.open(stream=content) as document:
            if document.is_pdf:
                return True
    except (OSError, RuntimeError, ValueError, fitz.FileDataError):
        pass
    return _PDF_HEADER_RE.search(content[:1032]) is not None


def _verify_source_pdf(source, content: bytes, issues: _Issues) -> int | None:
    try:
        with fitz.open(stream=content, filetype="pdf") as document:
            if not document.is_pdf or document.needs_pass:
                raise ValueError("PDF is encrypted or has the wrong format")
            actual_page_count = document.page_count
            if actual_page_count > MAX_PDF_PAGES_PER_SOURCE:
                issues.error("source-page-count-mismatch", source.source_id)
                return actual_page_count
            if actual_page_count != source.page_count:
                issues.error("source-page-count-mismatch", source.source_id)
            for page_index in range(actual_page_count):
                document.load_page(page_index)
            return actual_page_count
    except (OSError, RuntimeError, ValueError, fitz.FileDataError):
        issues.error("source-pdf-unreadable", source.source_id)
        return None


def _inspect_raw_unresolved_coverage(raw_manifest: object, issues: _Issues) -> None:
    if not isinstance(raw_manifest, dict):
        return
    sources = raw_manifest.get("sources")
    if isinstance(sources, list):
        for index, source in enumerate(sources):
            if not isinstance(source, dict) or source.get("coverageState") != "unresolved":
                continue
            source_id = source.get("sourceId")
            evidence = source_id if isinstance(source_id, str) else f"source-index-{index}"
            issues.error("unresolved-coverage", evidence)


def _inspect_raw_compatibility_target(raw_manifest: object, issues: _Issues) -> None:
    if not isinstance(raw_manifest, dict):
        return
    target = raw_manifest.get("compatibilityTarget")
    if isinstance(target, str) and target != "ctv-intake-v1":
        issues.error("compatibility-target-unsupported", target)
    pages = raw_manifest.get("pdfPages")
    if isinstance(pages, list):
        for index, page in enumerate(pages):
            if not isinstance(page, dict) or page.get("coverageState") != "unresolved":
                continue
            source_id = page.get("sourceId")
            page_number = page.get("sourcePage")
            evidence = (
                f"{source_id}#page={page_number}"
                if isinstance(source_id, str) and isinstance(page_number, int)
                else f"page-index-{index}"
            )
            issues.error("unresolved-coverage", evidence)


def _load_exceptions(
    manifest: PackageManifest,
    artifacts_by_kind: dict[str, list],
    usable_artifacts: dict[str, bytes],
    issues: _Issues,
) -> ExceptionsDocument | None:
    documents: list[ExceptionsDocument] = []
    for artifact in sorted(
        artifacts_by_kind.get("exceptions", []), key=lambda item: item.artifact_id
    ):
        content = usable_artifacts.get(artifact.artifact_id)
        if content is None:
            continue
        try:
            documents.append(
                ExceptionsDocument.model_validate(_read_json_bytes(content))
            )
        except (UnicodeError, json.JSONDecodeError, ValidationError):
            issues.error("exceptions-invalid", artifact.artifact_id)
    if not documents:
        return None
    if len(documents) == 1:
        return documents[0]
    all_items = [item for document in documents for item in document.items]
    try:
        return ExceptionsDocument(schemaVersion="1.0", items=all_items)
    except ValidationError:
        issues.error(
            "exceptions-invalid",
            *(artifact.artifact_id for artifact in artifacts_by_kind["exceptions"]),
        )
        return None


def _inspect_pdfs(
    artifacts_by_kind: dict[str, list],
    usable_artifacts: dict[str, bytes],
    issues: _Issues,
) -> dict[str, int]:
    page_counts: dict[str, int] = {}
    for artifact in sorted(
        artifacts_by_kind.get("input-pdf", []), key=lambda item: item.artifact_id
    ):
        content = usable_artifacts.get(artifact.artifact_id)
        if content is None:
            continue
        try:
            with fitz.open(stream=content, filetype="pdf") as document:
                if not document.is_pdf or document.needs_pass:
                    raise ValueError("PDF is encrypted or has the wrong format")
                page_count = document.page_count
                if page_count > MAX_PDF_PAGES_PER_SOURCE:
                    issues.error("pdf-page-limit-exceeded", artifact.artifact_id)
                    continue
                for page_index in range(page_count):
                    document.load_page(page_index)
        except (OSError, RuntimeError, ValueError, fitz.FileDataError):
            issues.error("pdf-unreadable", artifact.artifact_id)
            continue
        page_counts[artifact.artifact_id] = page_count
    return page_counts


def _inspect_rosters(
    manifest: PackageManifest,
    artifacts_by_kind: dict[str, list],
    usable_artifacts: dict[str, bytes],
    issues: _Issues,
) -> None:
    mapping = manifest.roster_mapping
    mapping_valid = _mapping_is_unambiguous(mapping)
    if not mapping_valid:
        issues.error("roster-mapping-missing", "rosterMapping")

    roster_artifacts = sorted(
        artifacts_by_kind.get("roster", []), key=lambda item: item.artifact_id
    )
    if mapping is not None and not any(
        mapping.source_id in artifact.source_ids for artifact in roster_artifacts
    ):
        issues.error("roster-mapping-missing", mapping.source_id)

    for artifact in roster_artifacts:
        content = usable_artifacts.get(artifact.artifact_id)
        if content is None:
            continue
        try:
            preflight_roster_workbook(io.BytesIO(content))
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
        except Exception:
            issues.error("roster-unreadable", artifact.artifact_id)
            continue
        try:
            if not mapping_valid or mapping is None:
                continue
            if mapping.sheet_name not in workbook.sheetnames:
                issues.error("roster-sheet-missing", mapping.sheet_name)
                continue
            worksheet = workbook[mapping.sheet_name]
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ()
            )
            header_positions: dict[object, list[int]] = defaultdict(list)
            for index, value in enumerate(header_row):
                header_positions[value].append(index)
            column_names = mapping.canonical_to_source_columns
            missing_columns = sorted(
                name
                for name in (column_names["name"], column_names["identity"])
                if len(header_positions[name]) != 1
            )
            if missing_columns:
                issues.error("roster-column-missing", *missing_columns)
                continue
            identity_index = header_positions[column_names["identity"]][0]
            identity_rows: dict[object, list[int]] = defaultdict(list)
            for row_number, row in enumerate(
                worksheet.iter_rows(min_row=2, values_only=True), start=2
            ):
                value = row[identity_index] if identity_index < len(row) else None
                if value is not None and value != "":
                    identity_rows[value].append(row_number)
            duplicate_rows = sorted(
                row_number
                for rows in identity_rows.values()
                if len(rows) > 1
                for row_number in rows
            )
            if duplicate_rows:
                issues.error(
                    "roster-identity-duplicate",
                    *(
                        f"{artifact.artifact_id}#row={row_number}"
                        for row_number in duplicate_rows
                    ),
                )
        except (OSError, RuntimeError, ValueError):
            issues.error("roster-unreadable", artifact.artifact_id)
        finally:
            workbook.close()


def _inspect_validation_reports(
    artifacts_by_kind: dict[str, list],
    usable_artifacts: dict[str, bytes],
    issues: _Issues,
) -> None:
    for artifact in sorted(
        artifacts_by_kind.get("validation-report", []),
        key=lambda item: item.artifact_id,
    ):
        content = usable_artifacts.get(artifact.artifact_id)
        if content is None:
            continue
        try:
            ValidationReport.model_validate(_read_json_bytes(content))
        except (UnicodeError, json.JSONDecodeError, ValidationError):
            issues.error("validation-report-invalid", artifact.artifact_id)


def _inspect_cccd_workbooks(
    artifacts_by_kind: dict[str, list],
    usable_artifacts: dict[str, bytes],
    issues: _Issues,
) -> None:
    for artifact in sorted(
        artifacts_by_kind.get("cccd", []), key=lambda item: item.artifact_id
    ):
        content = usable_artifacts.get(artifact.artifact_id)
        if content is None:
            continue
        if not artifact.path.casefold().endswith(".xlsx"):
            issues.error("cccd-unreadable", artifact.artifact_id)
            continue
        workbook = None
        try:
            preflight_roster_workbook(io.BytesIO(content))
            workbook = openpyxl.load_workbook(
                io.BytesIO(content), read_only=True, data_only=True
            )
            if not workbook.sheetnames:
                raise ValueError("CCCD workbook has no readable worksheets")
            for sheet_name in workbook.sheetnames:
                next(workbook[sheet_name].iter_rows(values_only=True), ())
        except Exception:
            issues.error("cccd-unreadable", artifact.artifact_id)
        finally:
            if workbook is not None:
                workbook.close()


def _check_page_count_limits(
    manifest: PackageManifest,
    source_verification: _SourceVerification,
    issues: _Issues,
) -> bool:
    pdf_sources = [
        source
        for source in manifest.sources
        if source.source_id in source_verification.pdf_source_ids
    ]
    exceeded = False
    declared_total = 0
    for source in sorted(pdf_sources, key=lambda item: item.source_id):
        if source.page_count is None:
            continue
        declared_total += source.page_count
        if source.page_count > MAX_PDF_PAGES_PER_SOURCE:
            exceeded = True
            issues.error(
                "page-count-limit-exceeded",
                f"{source.source_id}#page-count={source.page_count}",
            )
    if declared_total > MAX_PDF_PAGES_PER_PACKAGE:
        exceeded = True
        issues.error(
            "page-count-limit-exceeded",
            f"package#declared-pdf-pages={declared_total}",
        )
    actual_page_counts = dict(source_verification.actual_pdf_page_counts)
    actual_total = sum(actual_page_counts.values())
    if any(
        page_count > MAX_PDF_PAGES_PER_SOURCE
        for page_count in actual_page_counts.values()
    ):
        exceeded = True
    if actual_total > MAX_PDF_PAGES_PER_PACKAGE:
        exceeded = True
        issues.error(
            "page-count-limit-exceeded",
            f"package#actual-pdf-pages={actual_total}",
        )
    return exceeded


def _check_coverage_state_and_decisions(
    manifest: PackageManifest, issues: _Issues
) -> None:
    decisions = {decision.decision_id: decision for decision in manifest.decisions}
    source_decision_types = {
        "assigned": "assign-source",
        "shared": "share-source",
        "duplicate": "mark-duplicate",
        "excluded-by-user": "exclude-source",
    }
    page_decision_types = {
        "assigned": "assign-page",
        "shared": "share-source",
        "duplicate": "mark-duplicate",
        "excluded-by-user": "exclude-source",
    }
    decision_required_states = {"shared", "duplicate", "excluded-by-user"}
    source_ids = {source.source_id for source in manifest.sources}

    for source in sorted(manifest.sources, key=lambda item: item.source_id):
        if source.coverage_state == "duplicate" and (
            source.duplicate_source_id is None
            or source.duplicate_source_id == source.source_id
            or source.duplicate_source_id not in source_ids
        ):
            issues.error("coverage-state-inconsistent", source.source_id)
        expected_type = source_decision_types.get(source.coverage_state)
        if source.coverage_state in decision_required_states and not source.decision_id:
            issues.error("decision-type-mismatch", source.source_id)
        elif source.decision_id and expected_type:
            decision = decisions.get(source.decision_id)
            if decision is not None and decision.type != expected_type:
                issues.error("decision-type-mismatch", source.source_id)

    for page in sorted(
        manifest.pdf_pages, key=lambda item: (item.source_id, item.source_page)
    ):
        evidence = f"{page.source_id}#page={page.source_page}"
        has_target = page.target_page is not None
        if page.coverage_state in {"assigned", "shared"}:
            if not has_target:
                issues.error("coverage-state-inconsistent", evidence)
        elif has_target:
            issues.error("coverage-state-inconsistent", evidence)
        expected_type = page_decision_types.get(page.coverage_state)
        if page.coverage_state in decision_required_states and not page.decision_id:
            issues.error("decision-type-mismatch", evidence)
        elif page.decision_id and expected_type:
            decision = decisions.get(page.decision_id)
            if decision is not None and decision.type != expected_type:
                issues.error("decision-type-mismatch", evidence)


def _check_source_provenance(
    manifest: PackageManifest,
    pdf_source_ids: frozenset[str],
    issues: _Issues,
) -> None:
    artifact_source_ids = {
        source_id
        for artifact in manifest.artifacts
        for source_id in artifact.source_ids
    }
    mapped_pdf_sources = {
        page.source_id
        for page in manifest.pdf_pages
        if page.coverage_state in {"assigned", "shared"}
        and page.target_page is not None
    }
    for source in sorted(manifest.sources, key=lambda item: item.source_id):
        if source.coverage_state not in {"assigned", "shared"}:
            continue
        represented = (
            source.source_id in mapped_pdf_sources
            if source.source_id in pdf_source_ids
            else source.source_id in artifact_source_ids
        )
        if not represented:
            issues.error("source-provenance-missing", source.source_id)


def _check_current_approval(manifest: PackageManifest, issues: _Issues) -> None:
    if not any(
        decision.type == "approve-preview"
        and decision.actor == "user"
        and decision.proposal_version == manifest.package_version
        for decision in manifest.decisions
    ):
        issues.error("approval-missing", manifest.package_version)


def _mapping_is_unambiguous(mapping) -> bool:
    if mapping is None:
        return False
    columns = mapping.canonical_to_source_columns
    name = columns.get("name")
    identity = columns.get("identity")
    return (
        isinstance(name, str)
        and bool(name)
        and isinstance(identity, str)
        and bool(identity)
        and name != identity
    )


def _check_references(
    manifest: PackageManifest,
    exceptions: ExceptionsDocument | None,
    source_verification: _SourceVerification,
    issues: _Issues,
) -> None:
    source_ids = {source.source_id for source in manifest.sources}
    artifact_ids = {artifact.artifact_id for artifact in manifest.artifacts}
    decision_ids = {decision.decision_id for decision in manifest.decisions}
    exception_ids = (
        {item.exception_id for item in exceptions.items} if exceptions is not None else set()
    )
    known_evidence_refs = source_ids | artifact_ids | decision_ids | exception_ids
    known_evidence_refs.update(
        f"{page.source_id}#page={page.source_page}"
        for page in manifest.pdf_pages
    )
    actual_pdf_page_counts = dict(source_verification.actual_pdf_page_counts)
    pdf_page_counts: dict[str, int] = {}
    for source in manifest.sources:
        if source.source_id not in source_verification.pdf_source_ids:
            continue
        page_count = actual_pdf_page_counts.get(source.source_id, source.page_count)
        if page_count is not None:
            pdf_page_counts[source.source_id] = page_count

    unknown_sources: set[str] = set()
    unknown_decisions: set[str] = set()
    for source in manifest.sources:
        if source.duplicate_source_id and source.duplicate_source_id not in source_ids:
            unknown_sources.add(source.duplicate_source_id)
        if source.decision_id and source.decision_id not in decision_ids:
            unknown_decisions.add(source.decision_id)
    for page in manifest.pdf_pages:
        if page.source_id not in source_ids:
            unknown_sources.add(page.source_id)
        if page.decision_id and page.decision_id not in decision_ids:
            unknown_decisions.add(page.decision_id)
    for artifact in manifest.artifacts:
        unknown_sources.update(
            source_id for source_id in artifact.source_ids if source_id not in source_ids
        )
    if manifest.roster_mapping and manifest.roster_mapping.source_id not in source_ids:
        unknown_sources.add(manifest.roster_mapping.source_id)

    unknown_evidence_refs: set[str] = set()
    for decision in manifest.decisions:
        for index, evidence_ref in enumerate(decision.evidence_refs):
            if not _evidence_reference_is_known(
                evidence_ref, known_evidence_refs, pdf_page_counts
            ):
                unknown_evidence_refs.add(
                    f"{decision.decision_id}#evidence-ref={index}"
                )
    if exceptions is not None:
        for item in exceptions.items:
            for index, evidence_ref in enumerate(item.evidence_refs):
                if not _evidence_reference_is_known(
                    evidence_ref, known_evidence_refs, pdf_page_counts
                ):
                    unknown_evidence_refs.add(
                        f"{item.exception_id}#evidence-ref={index}"
                    )

    if unknown_sources:
        issues.error("source-reference-unknown", *sorted(unknown_sources))
    if unknown_decisions:
        issues.error("decision-reference-unknown", *sorted(unknown_decisions))
    if unknown_evidence_refs:
        issues.error(
            "evidence-reference-unknown", *sorted(unknown_evidence_refs)
        )


def _evidence_reference_is_known(
    evidence_ref: str,
    known_evidence_refs: set[str],
    pdf_page_counts: dict[str, int],
) -> bool:
    if evidence_ref in known_evidence_refs:
        return True
    match = _SOURCE_PAGE_EVIDENCE_RE.fullmatch(evidence_ref)
    if match is None:
        return False
    source_id, raw_page = match.groups()
    page_count = pdf_page_counts.get(source_id)
    if page_count is None:
        return False
    maximum = str(page_count)
    return len(raw_page) < len(maximum) or (
        len(raw_page) == len(maximum) and raw_page <= maximum
    )


def _check_input_pdf_provenance(
    manifest: PackageManifest,
    artifacts_by_kind: dict[str, list],
    pdf_source_ids: frozenset[str],
    issues: _Issues,
) -> None:
    sources_by_id = {source.source_id: source for source in manifest.sources}
    represented_source_ids = {
        page.source_id for page in manifest.pdf_pages if page.target_page is not None
    }
    for artifact in sorted(
        artifacts_by_kind.get("input-pdf", []), key=lambda item: item.artifact_id
    ):
        declared_source_ids = set(artifact.source_ids)
        omitted_source_ids = sorted(represented_source_ids - declared_source_ids)
        extra_source_ids = sorted(declared_source_ids - represented_source_ids)
        non_pdf_source_ids = sorted(
            source_id
            for source_id in declared_source_ids
            if source_id in sources_by_id
            and source_id not in pdf_source_ids
        )
        evidence_refs = [
            f"{artifact.artifact_id}#omitted-source={index}"
            for index, _source_id in enumerate(omitted_source_ids, start=1)
        ]
        evidence_refs.extend(
            f"{artifact.artifact_id}#extra-source={index}"
            for index, _source_id in enumerate(extra_source_ids, start=1)
        )
        evidence_refs.extend(
            f"{artifact.artifact_id}#non-pdf-source={index}"
            for index, _source_id in enumerate(non_pdf_source_ids, start=1)
        )
        if evidence_refs:
            issues.error("input-pdf-provenance-mismatch", *evidence_refs)


def _check_pdf_coverage(
    manifest: PackageManifest,
    artifacts_by_kind: dict[str, list],
    derived_page_counts: dict[str, int],
    source_verification: _SourceVerification,
    issues: _Issues,
) -> None:
    source_coverage = Counter(
        (page.source_id, page.source_page) for page in manifest.pdf_pages
    )
    source_ids = {source.source_id for source in manifest.sources}
    pdf_sources = sorted(
        (
            source
            for source in manifest.sources
            if source.source_id in source_verification.pdf_source_ids
        ),
        key=lambda source: source.source_id,
    )
    pdf_source_ids = {source.source_id for source in pdf_sources}
    actual_pdf_page_counts = dict(source_verification.actual_pdf_page_counts)
    for source in pdf_sources:
        source_id = source.source_id
        if source.page_count is None:
            issues.error("page-coverage-missing", f"{source_id}#page-count")
        page_count = actual_pdf_page_counts.get(source_id, source.page_count)
        if page_count is None:
            continue
        observed = sorted(
            page_number
            for covered_source, page_number in source_coverage
            if covered_source == source_id and 1 <= page_number <= page_count
        )
        for start, end in _missing_ranges(observed, page_count):
            issues.error(
                "page-coverage-missing",
                _page_range_evidence(source_id, start, end),
            )
        for page_number in observed:
            if source_coverage[(source_id, page_number)] > 1:
                issues.error(
                    "page-coverage-extra", f"{source_id}#page={page_number}"
                )
        for covered_source, page_number in sorted(source_coverage):
            if covered_source == source_id and page_number > page_count:
                issues.error("page-coverage-extra", f"{source_id}#page={page_number}")
    for source_id, page_number in sorted(source_coverage):
        if source_id not in pdf_source_ids and source_id in source_ids:
            issues.error(
                "page-coverage-extra", f"{source_id}#page={page_number}"
            )

    target_coverage = Counter(
        page.target_page
        for page in manifest.pdf_pages
        if page.target_page is not None
    )
    for artifact in sorted(
        artifacts_by_kind.get("input-pdf", []), key=lambda item: item.artifact_id
    ):
        page_count = derived_page_counts.get(artifact.artifact_id)
        if page_count is None:
            continue
        observed_targets = sorted(
            page_number for page_number in target_coverage if page_number <= page_count
        )
        for start, end in _missing_ranges(observed_targets, page_count):
            evidence = (
                f"{artifact.artifact_id}#target-page={start}"
                if start == end
                else f"{artifact.artifact_id}#target-pages={start}-{end}"
            )
            issues.error("page-coverage-missing", evidence)
        for page_number in observed_targets:
            if target_coverage[page_number] > 1:
                issues.error(
                    "page-coverage-extra",
                    f"{artifact.artifact_id}#target-page={page_number}",
                )
        for page_number in sorted(target_coverage):
            if page_number > page_count:
                issues.error(
                    "page-coverage-extra",
                    f"{artifact.artifact_id}#target-page={page_number}",
                )


def _missing_ranges(observed: list[int], expected_count: int):
    cursor = 1
    for page_number in observed:
        if page_number > cursor:
            yield cursor, page_number - 1
        cursor = max(cursor, page_number + 1)
    if cursor <= expected_count:
        yield cursor, expected_count


def _page_range_evidence(source_id: str, start: int, end: int) -> str:
    if start == end:
        return f"{source_id}#page={start}"
    return f"{source_id}#pages={start}-{end}"


def _check_exceptions(
    manifest: PackageManifest,
    exceptions: ExceptionsDocument | None,
    issues: _Issues,
) -> None:
    declared_ids = set(manifest.exception_ids)
    document_ids = (
        {item.exception_id for item in exceptions.items}
        if exceptions is not None
        else set()
    )
    unresolved_references = declared_ids ^ document_ids
    if unresolved_references:
        issues.error("exception-reference-unknown", *sorted(unresolved_references))
    if exceptions is None:
        return
    for item in sorted(exceptions.items, key=lambda value: value.exception_id):
        if item.resolution == "resolved":
            continue
        if item.resolution == "open" and item.severity == "blocking":
            issues.error("blocking-exception", item.exception_id)
            continue
        if item.resolution == "open":
            issues.warning(item.code, item.exception_id)
            continue

        # Accepted partial evidence remains visible and is never compatible
        # with a package claiming the complete prepared state.
        issues.warning(item.code, item.exception_id)
        if item.severity != "blocking":
            continue
        if manifest.status != "partially_prepared":
            issues.error("partial-status-required", item.exception_id)
        accepted = any(
            decision.type == "accept-partial"
            and decision.proposal_version == manifest.package_version
            and item.exception_id in decision.evidence_refs
            for decision in manifest.decisions
        )
        if not accepted:
            issues.error("accept-partial-decision-missing", item.exception_id)


def _check_unresolved_coverage(manifest: PackageManifest, issues: _Issues) -> None:
    for source in manifest.sources:
        if source.coverage_state == "unresolved":
            issues.error("unresolved-coverage", source.source_id)
    for page in manifest.pdf_pages:
        if page.coverage_state == "unresolved":
            issues.error(
                "unresolved-coverage",
                f"{page.source_id}#page={page.source_page}",
            )
