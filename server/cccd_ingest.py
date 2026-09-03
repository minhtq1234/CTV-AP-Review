"""Exact-only, case-local CCCD planning and v1 evidence attachment."""

from __future__ import annotations

from copy import deepcopy
from dataclasses import dataclass
import hashlib
import json
import os
import shutil
import tempfile
from typing import Callable, TypedDict

from cccd_matching import (
    CardResolution,
    ResolutionResult,
    resolve_candidates,
)
from cccd_ocr import EvidenceWriteBudget, analyze_drawing
from cccd_pairing import (
    AnalyzedDrawing,
    CardCandidate,
    pair_drawings,
)
from cccd_workbook import extract_drawings


ProgressCallback = Callable[[str, int, int, str], None]


class CccdIngestResult(TypedDict):
    packets: list[dict]
    cccdWorkbook: dict


@dataclass(frozen=True)
class PlannedMapping:
    candidate: CardCandidate
    resolution: CardResolution
    target_packet_index: int | None
    mapping: dict


def _digits(value: object) -> str:
    return (
        "".join(character for character in value if character.isdigit())
        if isinstance(value, str)
        else ""
    )


def _roster_index(roster_key: str | None) -> int | None:
    if not roster_key or not roster_key.startswith("roster-"):
        return None
    raw = roster_key.removeprefix("roster-")
    return int(raw) if raw.isdigit() else None


def _case_relative(case_dir: str, path: str) -> str:
    root = os.path.realpath(case_dir)
    candidate = os.path.realpath(path)
    try:
        shared_path = os.path.commonpath([root, candidate])
    except ValueError as error:
        raise ValueError("CCCD asset escaped case directory") from error
    if shared_path != root:
        raise ValueError("CCCD asset escaped case directory")
    return os.path.relpath(candidate, root).replace(os.sep, "/")


def _serialize_side(analyzed, case_dir: str) -> dict | None:
    if analyzed is None:
        return None
    drawing = analyzed.drawing
    anchor = drawing.anchor
    return {
        "drawingId": drawing.id,
        "mediaType": drawing.media_type,
        "width": drawing.width,
        "height": drawing.height,
        "sha256": drawing.sha256,
        "sourcePath": _case_relative(case_dir, drawing.stored_path),
        "packetPath": None,
        "anchor": {
            "sheet": anchor.sheet,
            "fromRow": anchor.from_row,
            "fromCol": anchor.from_col,
            "toRow": anchor.to_row,
            "toCol": anchor.to_col,
            "fromRowOffset": anchor.from_row_offset,
            "fromColOffset": anchor.from_col_offset,
            "toRowOffset": anchor.to_row_offset,
            "toColOffset": anchor.to_col_offset,
        },
    }


def _append_issue(issues: list[str], issue: str) -> None:
    if issue not in issues:
        issues.append(issue)


def _packet_target_index(
    packet: object,
    roster_cccd: str,
) -> int | None:
    if not isinstance(packet, dict):
        return None
    index = packet.get("index")
    if (
        not isinstance(index, int)
        or isinstance(index, bool)
        or index < 0
    ):
        return None
    identity = packet.get("rosterIdentity")
    if not isinstance(identity, dict):
        return None
    return (
        index
        if _digits(identity.get("cccd")) == roster_cccd
        else None
    )


def _validated_candidate_resolution_maps(
    candidates: list[CardCandidate],
    resolution_result: ResolutionResult,
) -> tuple[dict[str, CardCandidate], dict[str, CardResolution]]:
    candidate_by_id = {
        candidate.id: candidate for candidate in candidates
    }
    resolution_by_id = {
        resolution.candidate_id: resolution
        for resolution in resolution_result.resolutions
    }
    if (
        len(candidate_by_id) != len(candidates)
        or len(resolution_by_id) != len(resolution_result.resolutions)
        or set(candidate_by_id) != set(resolution_by_id)
    ):
        raise ValueError("candidate-resolution-mismatch")
    return candidate_by_id, resolution_by_id


def _target_packet_index(
    resolution: CardResolution,
    roster_rows: list[dict[str, str]],
    packets: list[dict],
    issues: list[str],
) -> int | None:
    if resolution.state != "exact":
        return None

    roster_index = _roster_index(resolution.roster_key)
    if roster_index is None or roster_index >= len(roster_rows):
        _append_issue(issues, "invalid-roster-key")
        return None

    roster_cccd = _digits(roster_rows[roster_index].get("cccd"))
    if len(roster_cccd) != 12:
        _append_issue(issues, "non-12-digit-roster-cccd")
        return None

    targets = [
        index
        for packet in packets
        if (
            index := _packet_target_index(packet, roster_cccd)
        ) is not None
    ]
    if len(targets) == 1:
        return targets[0]
    _append_issue(
        issues,
        (
            "packet-target-not-found"
            if not targets
            else "non-unique-packet-target"
        ),
    )
    return None


def plan_candidate_mappings(
    candidates: list[CardCandidate],
    resolution_result: ResolutionResult,
    roster_rows: list[dict[str, str]],
    packets: list[dict],
    case_dir: str,
) -> list[PlannedMapping]:
    """Plan exact roster-to-packet matches without mutating packet state."""
    candidate_by_id, resolution_by_id = (
        _validated_candidate_resolution_maps(candidates, resolution_result)
    )
    planned = []
    for candidate_id in sorted(candidate_by_id):
        candidate = candidate_by_id[candidate_id]
        resolution = resolution_by_id[candidate_id]
        front_ocr = (
            candidate.front.ocr
            if candidate.front is not None
            else None
        )
        provenance_ocr = (
            front_ocr
            if front_ocr is not None
            else (
                candidate.unknown.ocr
                if candidate.unknown is not None
                else None
            )
        )
        issues = list(
            dict.fromkeys((*candidate.issues, *resolution.issues))
        )
        target_packet_index = _target_packet_index(
            resolution,
            roster_rows,
            packets,
            issues,
        )
        if candidate.front is None:
            _append_issue(issues, "missing-front")
            target_packet_index = None
        if candidate.back is None:
            # Recorded so the reviewer knows the back is absent, but a front
            # alone is still evidence worth showing -- it no longer blocks
            # attachment. A missing FRONT still does (above).
            _append_issue(issues, "missing-back")
        mapping = {
            "candidateId": candidate.id,
            "front": _serialize_side(candidate.front, case_dir),
            "back": _serialize_side(candidate.back, case_dir),
            "unknown": _serialize_side(candidate.unknown, case_dir),
            "ocrIdentity": {
                "cccd": (
                    provenance_ocr.cccd
                    if provenance_ocr is not None
                    else ""
                ),
                "name": (
                    provenance_ocr.name
                    if provenance_ocr is not None
                    else ""
                ),
            },
            "ocrConfidence": {
                "cccd": (
                    provenance_ocr.cccd_confidence
                    if provenance_ocr is not None
                    else 0.0
                ),
                "name": (
                    provenance_ocr.name_confidence
                    if provenance_ocr is not None
                    else 0.0
                ),
            },
            "numberBbox": (
                provenance_ocr.number_bbox
                if provenance_ocr is not None
                else None
            ),
            "state": resolution.state,
            "attachedPacketIndex": None,
            "matchMethod": resolution.matched_by,
            "issues": issues,
        }
        planned.append(PlannedMapping(
            candidate=candidate,
            resolution=resolution,
            target_packet_index=target_packet_index,
            mapping=mapping,
        ))
    return planned


def _atomic_json_write(path: str, payload: dict) -> None:
    directory = os.path.dirname(path)
    descriptor, temporary_path = tempfile.mkstemp(
        prefix=".manifest-",
        suffix=".json",
        dir=directory,
    )
    try:
        with os.fdopen(
            descriptor,
            "w",
            encoding="utf-8",
        ) as handle:
            json.dump(
                payload,
                handle,
                ensure_ascii=False,
                indent=2,
            )
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_path, path)
    finally:
        if os.path.exists(temporary_path):
            os.unlink(temporary_path)


def _owned_doc_id(candidate_id: str, side: str) -> str:
    return f"cccd-excel-{candidate_id}-{side}"


#: Document id prefixes this ingest owns, and therefore cleans up when the
#: workbook is read again. Two families, because two different things in the
#: workbook become evidence: a card side, paired and OCR'd and resolved by the
#: number on its face; and a sheet screenshot, which has no face and is
#: attributed by the identity written on its own row (`sheet_identity`). Both
#: are written by this ingest and must not survive it going stale.
_OWNED_PREFIXES = ("cccd-excel-", "sheet-excel-")


def _sheet_doc_id(drawing_id: str, kind: str) -> str:
    return f"sheet-excel-{drawing_id}-{kind}"


def _packet_filename(
    plan: PlannedMapping,
    analyzed,
    side: str,
) -> str:
    candidate_token = hashlib.sha256(
        plan.candidate.id.encode("utf-8")
    ).hexdigest()[:12]
    image_token = analyzed.drawing.sha256[:12]
    orientation_token = (
        "-upright"
        if analyzed.ocr.evidence_path is not None
        else ""
    )
    return (
        f"cccd-{candidate_token}-{image_token}{orientation_token}-{side}."
        f"{analyzed.drawing.extension}"
    )


#: Which sheet image kind becomes which manifest document kind, and its label.
#:
#: `bank` is deliberately absent. `evaluate.DOC_KINDS` has no bank entry and
#: `EvidenceKind` in the frontend has no bank member, so a bank document would
#: be written into the manifest and then reach no criterion and render nowhere
#: -- storage with no reader. Wiring it needs a criterion that consumes it
#: first. `tax` has both halves already: DOC_KINDS maps MST_LOOKUP to `pit`,
#: which is what #6 reads.
_SHEET_DOC_KINDS: dict[str, tuple[str, str]] = {
    "tax": ("pit", "Tra cứu MST (Excel)"),
}


def _sheet_packet_filename(drawing, kind: str) -> str:
    token = hashlib.sha256(drawing.id.encode("utf-8")).hexdigest()[:12]
    return f"sheet-{kind}-{token}-{drawing.sha256[:12]}.{drawing.extension}"


def _sheet_rows(xlsx_path: str) -> dict[str, list]:
    """Every sheet's rows, for reading the identity beside an image.

    Values-only and read-only: this needs the text in the name/CCCD/MST
    columns and nothing else. Returns `{}` rather than raising -- a workbook
    whose drawings extracted fine but whose cells cannot be read should cost
    the sheet evidence, not the whole card ingest.
    """
    try:
        import openpyxl

        book = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        try:
            return {
                name: [list(row) for row in book[name].iter_rows(values_only=True)]
                for name in book.sheetnames
            }
        finally:
            book.close()
    except Exception:
        return {}


def attach_sheet_evidence(
    drawings: list,
    sheet_rows: dict[str, list],
    roster_rows: list[dict[str, str]],
    packets: list[dict],
    case_dir: str,
    packet_manifest_paths: dict[int, str],
) -> tuple[dict[int, set[str]], dict[str, str]]:
    """Write each attributable sheet screenshot into its packet as evidence.

    Returns `({packet index: doc ids written}, {drawing id: why not})`. The
    first is what `reconcile_owned_evidence` must keep; the second is a reason
    per screenshot that stayed unattached, because "we could not tell whose
    this is" is a different thing from "there was none" and only the second is
    a finding.

    The join is `sheet_identity`'s, not a positional one: on the real combined
    workbook the MST sheet lists the same 25 people as the roster in a
    different order, agreeing on 8 of 25 positions, so row N of that sheet is
    not person N. See that module for the measurements.
    """
    import pipeline
    import sheet_identity

    wanted = [
        drawing for drawing in drawings
        if getattr(drawing, "kind", None) in _SHEET_DOC_KINDS
    ]
    if not wanted:
        return {}, {}

    by_cccd, by_name, by_mst = pipeline.index_roster_rows(roster_rows)
    matched, refused = sheet_identity.attribute(
        wanted, sheet_rows, by_cccd, by_name, by_mst,
    )

    written: dict[int, set[str]] = {}
    for drawing in wanted:
        row = matched.get(drawing.id)
        if row is None:
            continue
        roster_cccd = _digits(row.get("cccd"))
        targets = [
            index
            for packet in packets
            if (index := _packet_target_index(packet, roster_cccd)) is not None
        ]
        if len(targets) != 1:
            # Zero means the roster row has no packet; more than one means two
            # packets claim the same identity. Neither may be guessed at.
            refused[drawing.id] = (
                "no-packet-for-person" if not targets
                else "several-packets-for-person"
            )
            continue
        packet_index = targets[0]
        manifest_path = packet_manifest_paths.get(packet_index)
        if not isinstance(manifest_path, str):
            refused[drawing.id] = "no-manifest"
            continue

        kind, label = _SHEET_DOC_KINDS[drawing.kind]
        doc_id = _sheet_doc_id(drawing.id, kind)
        packet_dir = os.path.dirname(manifest_path)
        destination = os.path.join(
            packet_dir, _sheet_packet_filename(drawing, kind),
        )
        # Path validation is a precondition, not an I/O failure, and it gets
        # its own reason: folded into "attachment-failed" it reads as a disk
        # problem when it means the extraction directory sits outside the case.
        try:
            _case_relative(case_dir, drawing.stored_path)
            _case_relative(case_dir, destination)
        except ValueError:
            refused[drawing.id] = "asset-outside-case-dir"
            continue
        try:
            if not os.path.exists(destination):
                shutil.copyfile(drawing.stored_path, destination)
            if not _write_sheet_doc(
                manifest_path, doc_id, kind, label, destination, drawing,
            ):
                refused[drawing.id] = "manifest-write-failed"
                continue
        except Exception:
            refused[drawing.id] = "attachment-failed"
            continue
        written.setdefault(packet_index, set()).add(doc_id)
    return written, refused


def _write_sheet_doc(
    manifest_path: str,
    doc_id: str,
    kind: str,
    label: str,
    destination: str,
    drawing,
) -> bool:
    """Put one sheet-sourced document into a manifest, replacing its own.

    Keyed on the doc id, which is derived from the drawing, so re-reading the
    same workbook rewrites the same entry rather than accumulating copies.
    """
    try:
        with open(manifest_path, "r", encoding="utf-8") as handle:
            manifest = json.load(handle)
        if not isinstance(manifest, dict):
            return False
        docs = manifest.get("docs")
        if not isinstance(docs, list):
            return False
        manifest["docs"] = [
            document for document in docs
            if not (isinstance(document, dict)
                    and document.get("id") == doc_id)
        ] + [{
            "id": doc_id,
            "kind": kind,
            "label": label,
            "pages": [{
                "src": destination,
                "width": drawing.width,
                "height": drawing.height,
            }],
        }]
        _atomic_json_write(manifest_path, manifest)
        return True
    except Exception:
        return False


def _attachment_failure(plan: PlannedMapping) -> dict:
    mapping = deepcopy(plan.mapping)
    issues = list(mapping.get("issues", []))
    _append_issue(issues, "attachment-failed")
    mapping["issues"] = issues
    mapping["attachedPacketIndex"] = None
    return mapping


def _remove_stale_owned_files(
    old_documents,
    packet_dir: str,
    new_paths: set[str],
) -> None:
    real_packet_dir = os.path.realpath(packet_dir)
    for document in old_documents:
        if not isinstance(document, dict):
            continue
        pages = document.get("pages")
        if not isinstance(pages, list):
            continue
        for page in pages:
            source = (
                page.get("src")
                if isinstance(page, dict)
                else None
            )
            if not isinstance(source, str):
                continue
            old_path = os.path.realpath(source)
            if (
                os.path.dirname(old_path) == real_packet_dir
                and old_path not in new_paths
                and os.path.basename(old_path).startswith("cccd-")
                and os.path.isfile(old_path)
            ):
                os.unlink(old_path)


def _is_owned_doc_id(value: object) -> bool:
    return isinstance(value, str) and value.startswith(_OWNED_PREFIXES)


def _kept_owned_ids_by_packet(mappings: list[dict]) -> dict[int, set[str]]:
    kept: dict[int, set[str]] = {}
    for mapping in mappings:
        packet_index = mapping.get("attachedPacketIndex")
        candidate_id = mapping.get("candidateId")
        if (
            not isinstance(packet_index, int)
            or isinstance(packet_index, bool)
            or packet_index < 0
            or not isinstance(candidate_id, str)
        ):
            continue
        kept.setdefault(packet_index, set()).update({
            _owned_doc_id(candidate_id, "front"),
            _owned_doc_id(candidate_id, "back"),
        })
    return kept


def _reconcile_manifest_owned_evidence(
    manifest_path: str,
    case_dir: str,
    keep_ids: set[str],
) -> bool:
    try:
        _case_relative(case_dir, manifest_path)
        with open(manifest_path, "r", encoding="utf-8") as handle:
            original = json.load(handle)
        if not isinstance(original, dict):
            raise ValueError("manifest must be an object")
        updated = deepcopy(original)
        docs = updated.get("docs")
        fields = updated.get("fields")
        if not isinstance(docs, list) or not isinstance(fields, list):
            raise ValueError("manifest structure is invalid")
        removed_docs = [
            document
            for document in docs
            if (
                isinstance(document, dict)
                and _is_owned_doc_id(document.get("id"))
                and document.get("id") not in keep_ids
            )
        ]
        updated["docs"] = [
            document
            for document in docs
            if not (
                isinstance(document, dict)
                and _is_owned_doc_id(document.get("id"))
                and document.get("id") not in keep_ids
            )
        ]
        changed = bool(removed_docs)
        for field in fields:
            if not isinstance(field, dict):
                continue
            sources = field.get("sources")
            if not isinstance(sources, list):
                continue
            kept_sources = [
                source
                for source in sources
                if not (
                    isinstance(source, dict)
                    and _is_owned_doc_id(source.get("docId"))
                    and source.get("docId") not in keep_ids
                )
            ]
            if len(kept_sources) != len(sources):
                field["sources"] = kept_sources
                changed = True
        if changed:
            _atomic_json_write(manifest_path, updated)
            _remove_stale_owned_files(
                removed_docs,
                os.path.dirname(manifest_path),
                set(),
            )
        return True
    except Exception:
        return False


def reconcile_owned_evidence(
    manifest_paths: dict[int, str],
    case_dir: str,
    mappings: list[dict],
    sheet_keep: dict[int, set[str]] | None = None,
) -> bool:
    """Drop every owned document a fresh read of the workbook did not produce.

    `sheet_keep` is the sheet-evidence half: `{packet index: doc ids}` from
    `attach_sheet_evidence`. Without it a re-ingest would delete the pit
    documents it had just written, since they are owned (`_OWNED_PREFIXES`)
    and would not appear in the card keep-set. Defaults to none kept, which is
    what the two failure paths want -- nothing was written, so nothing stays.
    """
    kept_by_packet = _kept_owned_ids_by_packet(mappings)
    for packet_index, ids in (sheet_keep or {}).items():
        kept_by_packet.setdefault(packet_index, set()).update(ids)
    successful = True
    for packet_index, manifest_path in manifest_paths.items():
        if (
            not isinstance(packet_index, int)
            or isinstance(packet_index, bool)
            or packet_index < 0
            or not isinstance(manifest_path, str)
        ):
            successful = False
            continue
        successful = (
            _reconcile_manifest_owned_evidence(
                manifest_path,
                case_dir,
                kept_by_packet.get(packet_index, set()),
            )
            and successful
        )
    return successful


def _cleanup_attempt_files(paths: list[str]) -> bool:
    failed = False
    for path in paths:
        try:
            if os.path.isfile(path):
                os.unlink(path)
        except Exception:
            failed = True
    return failed


# States whose plan may write card evidence into a packet. "exact" is the
# automatic resolver's unique-roster-hit; "assigned" is a reviewer naming the
# packet by eye (see cccd_manual). Both attach identically -- `matchMethod`
# records which one it was, so the provenance stays visible.
_ATTACHABLE_STATES = frozenset({"exact", "assigned"})


def _is_exact_attachment_plan(plan: PlannedMapping) -> bool:
    target = plan.target_packet_index
    sides_present = [
        side
        for side in ("front", "back")
        if getattr(plan.candidate, side) is not None
    ]
    return (
        isinstance(target, int)
        and not isinstance(target, bool)
        and target >= 0
        and getattr(plan.resolution, "state", None) in _ATTACHABLE_STATES
        and isinstance(plan.mapping, dict)
        and plan.mapping.get("state") in _ATTACHABLE_STATES
        and bool(sides_present)
        and all(
            isinstance(plan.mapping.get(side), dict)
            for side in sides_present
        )
    )


def _report_progress(
    progress_cb: ProgressCallback,
    done: int,
    total: int,
) -> None:
    try:
        progress_cb("cccd", done, total, "")
    except Exception:
        pass


def _error_result(
    packets: list[dict],
    error_code: str,
) -> CccdIngestResult:
    return {
        "packets": packets,
        "cccdWorkbook": {
            "status": "error",
            "errorCode": error_code,
            "summary": {
                "candidates": 0,
                "attached": 0,
                "unresolved": 0,
            },
            "mappings": [],
        },
    }


def _reconciled_error_result(
    packets: list[dict],
    error_code: str,
    manifest_paths: dict[int, str],
    case_dir: str,
) -> CccdIngestResult:
    if not reconcile_owned_evidence(manifest_paths, case_dir, []):
        error_code = "attachment-failed"
    return _error_result(packets, error_code)


def _packet_by_index(packets: list[dict]) -> dict[int, dict]:
    indexed = {}
    for packet in packets:
        if not isinstance(packet, dict):
            continue
        index = packet.get("index")
        if (
            isinstance(index, int)
            and not isinstance(index, bool)
            and index >= 0
        ):
            indexed[index] = packet
    return indexed


def ingest_cccd_workbook(
    xlsx_path: str,
    roster_rows: list[dict[str, str]],
    packets: list[dict],
    case_dir: str,
    packet_manifest_paths: dict[int, str],
    assets_dir: str,
    progress_cb: ProgressCallback,
    analyze=None,
) -> CccdIngestResult:
    """Run CCCD ingestion and preserve partial usable results.

    `analyze` reads one card. It defaults to the local Tesseract reader; pass
    `cccd_idp.reader(...)` to use GreenNode IDP instead. Everything downstream
    -- pairing, roster resolution, conflict detection, attachment and
    reconciliation -- is unchanged either way, because both speak CccdImageOcr.
    """
    # Resolved here, not as a default argument, so the module attribute stays
    # monkeypatchable in tests.
    analyze = analyze or analyze_drawing
    manifest_paths = (
        packet_manifest_paths
        if isinstance(packet_manifest_paths, dict)
        else {}
    )
    try:
        extraction = extract_drawings(
            xlsx_path,
            os.path.join(assets_dir, "extracted"),
        )
    except Exception:
        return _reconciled_error_result(
            packets,
            "invalid-workbook",
            manifest_paths,
            case_dir,
        )
    if not extraction.drawings:
        return _reconciled_error_result(
            packets,
            "no-supported-images",
            manifest_paths,
            case_dir,
        )

    analyzed = []
    ocr_failures = 0
    evidence_budget = EvidenceWriteBudget()
    for drawing in extraction.drawings:
        try:
            analyzed.append(
                AnalyzedDrawing(
                    drawing,
                    analyze(drawing, evidence_budget),
                )
            )
        except Exception:
            ocr_failures += 1
    if not analyzed:
        return _reconciled_error_result(
            packets,
            "ocr-unavailable",
            manifest_paths,
            case_dir,
        )

    try:
        candidates = pair_drawings(analyzed)
        resolution_result = resolve_candidates(
            candidates,
            roster_rows,
        )
        plans = plan_candidate_mappings(
            candidates,
            resolution_result,
            roster_rows,
            packets,
            case_dir,
        )
    except Exception:
        return _reconciled_error_result(
            packets,
            "invalid-workbook",
            manifest_paths,
            case_dir,
        )

    packet_by_index = _packet_by_index(packets)
    mappings = []
    for done, planned in enumerate(plans, start=1):
        if planned.target_packet_index is None:
            mapping = deepcopy(planned.mapping)
        else:
            target_packet = packet_by_index.get(
                planned.target_packet_index
            )
            manifest_path = manifest_paths.get(
                planned.target_packet_index
            )
            mapping = (
                attach_planned_mapping(
                    planned,
                    target_packet,
                    manifest_path,
                    case_dir,
                )
                if target_packet is not None
                and isinstance(manifest_path, str)
                else _attachment_failure(planned)
            )
        mappings.append(mapping)
        _report_progress(progress_cb, done, len(plans))

    # The sheet screenshots: evidence the workbook has always carried and
    # nothing could reach. Classified at upload and then used only to keep
    # themselves out of the card candidate pool, so on the combined template
    # 25 tax lookups sat in the file while #6 read "Hồ sơ thiếu Website tra
    # cứu MST" on all 25 packets -- missing, when the document was right there.
    sheet_written, sheet_refused = attach_sheet_evidence(
        extraction.drawings,
        _sheet_rows(xlsx_path),
        roster_rows,
        packets,
        case_dir,
        manifest_paths,
    )
    reconciliation_failed = not reconcile_owned_evidence(
        manifest_paths,
        case_dir,
        mappings,
        sheet_keep=sheet_written,
    )
    attached = sum(
        mapping.get("attachedPacketIndex") is not None
        for mapping in mappings
    )
    summary = {
        "candidates": len(mappings),
        "attached": attached,
        "unresolved": len(mappings) - attached,
    }
    if reconciliation_failed or any(
        "attachment-failed" in mapping.get("issues", [])
        for mapping in mappings
    ):
        error_code = "attachment-failed"
    elif extraction.issues:
        error_code = "extraction-incomplete"
    elif ocr_failures:
        error_code = "ocr-unavailable"
    else:
        error_code = None
    workbook = {
        "status": "partial" if error_code else "ready",
        "summary": summary,
        "mappings": mappings,
    }
    if error_code:
        workbook["errorCode"] = error_code
    return {"packets": packets, "cccdWorkbook": workbook}


def attach_planned_mapping(
    plan: PlannedMapping,
    packet: dict,
    manifest_path: str,
    case_dir: str,
) -> dict:
    """Attach one exact plan without exposing partial packet state."""
    if plan.target_packet_index is None:
        return deepcopy(plan.mapping)
    if (
        not _is_exact_attachment_plan(plan)
        or not isinstance(packet, dict)
        or not isinstance(packet.get("index"), int)
        or isinstance(packet.get("index"), bool)
        or packet.get("index") != plan.target_packet_index
        or not isinstance(manifest_path, str)
        or not isinstance(case_dir, str)
    ):
        return _attachment_failure(plan)

    created_files: list[str] = []
    try:
        _case_relative(case_dir, manifest_path)
        with open(
            manifest_path,
            "r",
            encoding="utf-8",
        ) as handle:
            original = json.load(handle)
        if not isinstance(original, dict):
            raise ValueError("manifest must be an object")
        updated = deepcopy(original)
        packet_dir = os.path.dirname(manifest_path)
        docs = updated.get("docs")
        fields = updated.get("fields")
        if not isinstance(docs, list) or not isinstance(fields, list):
            raise ValueError("manifest structure is invalid")
        owned_ids = {
            _owned_doc_id(plan.candidate.id, "front"),
            _owned_doc_id(plan.candidate.id, "back"),
        }
        old_owned_docs = [
            document
            for document in docs
            if (
                isinstance(document, dict)
                and document.get("id") in owned_ids
            )
        ]
        mapping = deepcopy(plan.mapping)
        new_docs = []
        new_paths: set[str] = set()
        for side in ("front", "back"):
            analyzed = getattr(plan.candidate, side)
            if analyzed is None:
                continue
            source = (
                analyzed.ocr.evidence_path
                or analyzed.drawing.stored_path
            )
            _case_relative(case_dir, source)
            destination = os.path.join(
                packet_dir,
                _packet_filename(plan, analyzed, side),
            )
            _case_relative(case_dir, destination)
            if not os.path.exists(destination):
                created_files.append(destination)
                shutil.copyfile(source, destination)
            new_paths.add(os.path.realpath(destination))
            side_mapping = mapping.get(side)
            if not isinstance(side_mapping, dict):
                raise ValueError("missing mapped side")
            side_mapping["packetPath"] = _case_relative(
                case_dir,
                destination,
            )
            new_docs.append({
                "id": _owned_doc_id(plan.candidate.id, side),
                "kind": (
                    "id_front" if side == "front" else "id_back"
                ),
                "label": (
                    "CCCD (Excel) · Mặt trước"
                    if side == "front"
                    else "CCCD (Excel) · Mặt sau"
                ),
                "pages": [{
                    "src": destination,
                    "width": (
                        analyzed.ocr.evidence_width
                        or analyzed.drawing.width
                    ),
                    "height": (
                        analyzed.ocr.evidence_height
                        or analyzed.drawing.height
                    ),
                }],
            })
        updated["docs"] = [
            document
            for document in docs
            if not (
                isinstance(document, dict)
                and document.get("id") in owned_ids
            )
        ] + new_docs
        cccd_field = next(
            (
                field
                for field in fields
                if (
                    isinstance(field, dict)
                    and field.get("key") == "cccd"
                )
            ),
            None,
        )
        if cccd_field is not None:
            sources = cccd_field.get("sources", [])
            if not isinstance(sources, list):
                raise ValueError("CCCD sources are invalid")
            cccd_field["sources"] = [
                source
                for source in sources
                if not (
                    isinstance(source, dict)
                    and source.get("docId") in owned_ids
                )
            ]
        front = plan.candidate.front
        located = front is not None and front.ocr.number_bbox is not None
        if plan.mapping.get("state") == "exact" and not located:
            # An automatic match is only "exact" because a number was read out
            # of a located region -- without that there is nothing exact about
            # it. A reviewer's manual assignment rests on their eyes instead,
            # so it attaches the card as evidence and claims no read.
            raise ValueError("exact attachment requires located front")
        if cccd_field is not None and located:
            cccd_field["sources"].append({
                "docId": _owned_doc_id(
                    plan.candidate.id,
                    "front",
                ),
                "page": 0,
                "value": front.ocr.cccd,
                "bbox": front.ocr.number_bbox,
                "confidence": front.ocr.cccd_confidence,
            })
        _atomic_json_write(manifest_path, updated)
    except Exception:
        failure = _attachment_failure(plan)
        if _cleanup_attempt_files(created_files):
            _append_issue(failure["issues"], "cleanup-failed")
        return failure

    mapping["attachedPacketIndex"] = plan.target_packet_index
    try:
        _remove_stale_owned_files(
            old_owned_docs,
            packet_dir,
            new_paths,
        )
    except Exception:
        pass
    return mapping
