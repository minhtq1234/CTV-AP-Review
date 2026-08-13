"""Memory-only, privacy-safe proposal state for the local CTV review."""

import hashlib
import hmac
import json
import re
from io import BytesIO

from openpyxl import load_workbook

from ctv_inspection_classifier import roster_header_categories_from_private_text
from ctv_inspection_model import InspectionResult
from ctv_inventory import InventoryObservation


_VERSION = "1.0"
_MAX_ROWS = 10_000
_MAX_CELLS = 100_000
_MAX_CELL_TEXT = 256
_MAX_WORKBOOK_BYTES = 25 * 1024 * 1024
_UNIT_ID = re.compile(r"^unit-[0-9]{4,}$")
_EVIDENCE_ID = re.compile(r"^evidence-[0-9]{4,}$")
_PARTICIPANT_HANDLE = re.compile(r"^participant-[0-9]{4,}$")
_ROLES_BY_KIND = {
    "pdf-page": frozenset({
        "payment-roster", "service-contract", "acceptance-record", "payment-tax-form",
        "identity-front", "identity-back", "shared-supporting-evidence",
        "other-supporting-evidence",
    }),
    "worksheet": frozenset({"payment-roster", "other-supporting-evidence"}),
    "image": frozenset({
        "identity-front", "identity-back", "shared-supporting-evidence",
        "other-supporting-evidence",
    }),
}
_UNIT_DECISIONS = frozenset({"accepted", "reassigned", "excluded", "unresolved"})
_SOURCE_DECISIONS = frozenset({"excluded", "unresolved"})
_SCOPES = frozenset({"individual", "shared", "case"})
_EXCLUSION_REASONS = frozenset({
    "duplicate", "irrelevant", "unreadable-replacement-available",
    "intentionally-omitted", "other",
})
_ROSTER_ISSUE_ORDER = (
    "roster-header-missing", "roster-row-invalid", "roster-identity-duplicate",
    "roster-over-limit", "roster-unreadable",
)


def _mapping(value, keys):
    if type(value) is not dict or set(value) != set(keys):
        raise ValueError("proposal request must use its exact object shape")
    return value


def _string(value, pattern, name):
    if type(value) is not str or not pattern.fullmatch(value):
        raise ValueError(f"{name} must be a valid opaque ID")
    return value


def _enum(value, allowed, name):
    if type(value) is not str or value not in allowed:
        raise ValueError(f"{name} must be an approved value")
    return value


def _canonical_digest(value):
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _private_cell_text(value):
    if value is None:
        return ""
    if type(value) not in {str, int, float}:
        return ""
    text = str(value).strip()
    if len(text) > _MAX_CELL_TEXT:
        return ""
    return text


class ProposalState:
    """Trusted proposal records constructed after strict local API conversion."""

    def __init__(self, observation, inspection):
        self._observation = observation
        self._inspection = inspection
        self._units_by_id = {unit.unit_id: unit for unit in inspection.units}
        self._sources_by_id = {source.evidence_id: source for source in inspection.sources}
        self._unit_decisions = {}
        self._source_dispositions = {}
        self._roster_unit_id = None
        self._participant_handles = ()
        self._participant_display = ()
        self._roster_issues = ()
        self.units = tuple(
            {
                "unitId": unit.unit_id, "evidenceId": unit.evidence_id,
                "unitKind": unit.unit_kind, "suggestedRole": unit.suggested_role,
                "issueCodes": list(unit.issue_codes),
            }
            for unit in inspection.units
        )
        self.sources = tuple(
            {
                "evidenceId": source.evidence_id, "detectedType": source.detected_type,
                "inspectionStatus": source.inspection_status, "unitCount": source.unit_count,
                "issueCodes": list(source.issue_codes),
            }
            for source in inspection.sources
        )

    @classmethod
    def from_inspection(cls, observation, inspection):
        if type(observation) is not InventoryObservation:
            raise TypeError("observation must be a live inventory observation")
        if type(inspection) is not InspectionResult:
            raise TypeError("inspection must be an inspection result")
        if inspection.observation_id != observation.observation_id:
            raise ValueError("inspection must belong to its observation")
        return cls(observation, inspection)

    def _roster_rows(self, unit):
        source = self._sources_by_id[unit.evidence_id]
        if source.detected_type != "xlsx" or source.inspection_status != "inspected":
            return (), ("roster-unreadable",)
        try:
            snapshot = self._observation.snapshot(unit.evidence_id, max_bytes=_MAX_WORKBOOK_BYTES)
            workbook = load_workbook(
                BytesIO(snapshot), read_only=True, data_only=True, keep_links=False
            )
            try:
                worksheets = workbook.worksheets
                if unit.unit_index > len(worksheets):
                    return (), ("roster-unreadable",)
                worksheet = worksheets[unit.unit_index - 1]
                header = None
                rows = []
                cells = 0
                issues = set()
                for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                    if row_index > _MAX_ROWS:
                        issues.add("roster-over-limit")
                        break
                    values = []
                    for cell in row:
                        cells += 1
                        if cells > _MAX_CELLS:
                            issues.add("roster-over-limit")
                            break
                        values.append(_private_cell_text(cell.value))
                    if "roster-over-limit" in issues:
                        break
                    categories = [roster_header_categories_from_private_text(value) if value else () for value in values]
                    name_columns = [index for index, value in enumerate(categories) if "name" in value]
                    identity_columns = [index for index, value in enumerate(categories) if "identity" in value]
                    if header is None and name_columns and identity_columns:
                        header = (name_columns[0], identity_columns[0])
                        continue
                    if header is None:
                        continue
                    name = values[header[0]] if header[0] < len(values) else ""
                    identity = values[header[1]] if header[1] < len(values) else ""
                    if not name and not identity:
                        issues.add("roster-row-invalid")
                    elif not name or not identity:
                        issues.add("roster-row-invalid")
                    else:
                        rows.append((name, identity))
                if header is None:
                    issues.add("roster-header-missing")
                if not rows:
                    issues.add("roster-row-invalid")
                if len(rows) != len({identity for _name, identity in rows}):
                    issues.add("roster-identity-duplicate")
                return tuple(rows), tuple(code for code in _ROSTER_ISSUE_ORDER if code in issues)
            finally:
                workbook.close()
        except Exception:
            return (), ("roster-unreadable",)

    def select_roster(self, mapping):
        mapping = _mapping(mapping, {"rosterUnitId"})
        unit_id = _string(mapping["rosterUnitId"], _UNIT_ID, "rosterUnitId")
        unit = self._units_by_id.get(unit_id)
        if unit is None or unit.unit_kind != "worksheet" or unit.suggested_role != "payment-roster":
            raise ValueError("rosterUnitId must identify an inspected roster worksheet")
        rows, issues = self._roster_rows(unit)
        if self._roster_unit_id is not None and self._roster_unit_id != unit_id:
            self._unit_decisions = {
                decision_unit_id: record
                for decision_unit_id, record in self._unit_decisions.items()
                if not (
                    record["decision"] in {"accepted", "reassigned"}
                    and record["target"]["participantHandles"]
                )
            }
        self._roster_unit_id = unit_id
        self._participant_handles = tuple(
            f"participant-{index:04d}" for index, _row in enumerate(rows, start=1)
        )
        self._participant_display = tuple(
            {
                "participantHandle": handle,
                "name": name,
                "identityHint": f"***-{identity[-3:]}",
            }
            for handle, (name, identity) in zip(self._participant_handles, rows)
        )
        self._roster_issues = issues

    def participants_for_local_review(self):
        """Return private roster display fields only to the local review session."""
        return [dict(participant) for participant in self._participant_display]

    def _target(self, value):
        value = _mapping(value, {"scope", "participantHandles"})
        scope = _enum(value["scope"], _SCOPES, "scope")
        handles = value["participantHandles"]
        if type(handles) is not list:
            raise ValueError("participantHandles must be a list")
        if any(type(handle) is not str or not _PARTICIPANT_HANDLE.fullmatch(handle) for handle in handles):
            raise ValueError("participantHandles must be opaque participant handles")
        if len(handles) != len(set(handles)) or any(handle not in self._participant_handles for handle in handles):
            raise ValueError("participantHandles must be selected roster handles")
        if (scope == "individual" and len(handles) != 1) or (scope == "shared" and len(handles) < 2) or (scope == "case" and handles):
            raise ValueError("participantHandles must match the assignment scope")
        return {"scope": scope, "participantHandles": tuple(handles)}

    def set_unit_decision(self, mapping):
        if type(mapping) is not dict:
            raise ValueError("proposal request must use its exact object shape")
        decision = _enum(mapping.get("decision"), _UNIT_DECISIONS, "decision")
        required = {
            "accepted": {"unitId", "decision", "role", "target"},
            "reassigned": {"unitId", "decision", "role", "target"},
            "excluded": {"unitId", "decision", "reason"},
            "unresolved": {"unitId", "decision"},
        }[decision]
        _mapping(mapping, required)
        unit_id = _string(mapping["unitId"], _UNIT_ID, "unitId")
        unit = self._units_by_id.get(unit_id)
        if unit is None:
            raise ValueError("unitId must identify an inspected unit")
        record = {"decision": decision}
        if decision in {"accepted", "reassigned"}:
            role = _enum(mapping["role"], _ROLES_BY_KIND[unit.unit_kind], "role")
            if decision == "accepted" and (
                unit.suggested_role == "unknown" or role != unit.suggested_role
            ):
                raise ValueError("accepted role must equal a concrete suggested role")
            if decision == "reassigned" and (
                unit.suggested_role != "unknown" and role == unit.suggested_role
            ):
                raise ValueError("reassigned role must differ from a concrete suggested role")
            record.update({"role": role, "target": self._target(mapping["target"])})
        elif decision == "excluded":
            record["reason"] = _enum(mapping["reason"], _EXCLUSION_REASONS, "reason")
        self._unit_decisions[unit_id] = record

    def set_source_disposition(self, mapping):
        if type(mapping) is not dict:
            raise ValueError("proposal request must use its exact object shape")
        decision = _enum(mapping.get("decision"), _SOURCE_DECISIONS, "decision")
        required = {"excluded": {"evidenceId", "decision", "reason"}, "unresolved": {"evidenceId", "decision"}}[decision]
        _mapping(mapping, required)
        evidence_id = _string(mapping["evidenceId"], _EVIDENCE_ID, "evidenceId")
        source = self._sources_by_id.get(evidence_id)
        if source is None or any(unit.evidence_id == evidence_id for unit in self._inspection.units):
            raise ValueError("evidenceId must identify a source-only record")
        record = {"decision": decision}
        if decision == "excluded":
            record["reason"] = _enum(mapping["reason"], _EXCLUSION_REASONS, "reason")
        self._source_dispositions[evidence_id] = record

    def _issue_codes(self):
        issues = set(self._roster_issues)
        for source in self._inspection.sources:
            issues.update(source.issue_codes)
        for unit in self._inspection.units:
            issues.update(unit.issue_codes)
        return sorted(issues)

    def _ready(self):
        source_only_ids = {
            source.evidence_id for source in self._inspection.sources
            if not any(unit.evidence_id == source.evidence_id for unit in self._inspection.units)
        }
        return (
            self._roster_unit_id is not None
            and not self._roster_issues
            and set(self._unit_decisions) == set(self._units_by_id)
            and all(value["decision"] != "unresolved" for value in self._unit_decisions.values())
            and set(self._source_dispositions) == source_only_ids
            and all(value["decision"] != "unresolved" for value in self._source_dispositions.values())
        )

    def _counts(self):
        source_only_ids = {
            source.evidence_id for source in self._inspection.sources
            if not any(unit.evidence_id == source.evidence_id for unit in self._inspection.units)
        }
        return {
            "sources": len(self._inspection.sources), "units": len(self._inspection.units),
            "participants": len(self._participant_handles),
            "accepted": sum(value["decision"] == "accepted" for value in self._unit_decisions.values()),
            "reassigned": sum(value["decision"] == "reassigned" for value in self._unit_decisions.values()),
            "excluded": sum(value["decision"] == "excluded" for value in self._unit_decisions.values()) + sum(value["decision"] == "excluded" for value in self._source_dispositions.values()),
            "unresolved": sum(
                self._unit_decisions.get(unit_id, {"decision": "unresolved"})["decision"] == "unresolved"
                for unit_id in self._units_by_id
            ) + sum(
                self._source_dispositions.get(evidence_id, {"decision": "unresolved"})["decision"] == "unresolved"
                for evidence_id in source_only_ids
            ),
        }

    def _digest_input(self):
        assignments = []
        for unit_id in sorted(self._units_by_id):
            record = self._unit_decisions.get(unit_id, {"decision": "unresolved"})
            value = {"unitId": unit_id, "decision": record["decision"]}
            if "role" in record:
                value["role"] = record["role"]
                value["target"] = {"scope": record["target"]["scope"], "participantHandles": list(record["target"]["participantHandles"])}
            if "reason" in record:
                value["reason"] = record["reason"]
            assignments.append(value)
        unit_evidence_ids = {unit.evidence_id for unit in self._inspection.units}
        dispositions = [
            {
                "evidenceId": evidence_id,
                **self._source_dispositions.get(evidence_id, {"decision": "unresolved"}),
            }
            for evidence_id in sorted(
                source.evidence_id for source in self._inspection.sources
                if source.evidence_id not in unit_evidence_ids
            )
        ]
        return {
            "observationId": self._inspection.observation_id, "rosterUnitId": self._roster_unit_id,
            "participantHandles": list(self._participant_handles), "unitAssignments": assignments,
            "sourceDispositions": dispositions, "issueCodes": self._issue_codes(), "counts": self._counts(),
        }

    def approval_summary(self):
        digest = _canonical_digest(self._digest_input())
        return {
            "observationId": self._inspection.observation_id, "rosterUnitId": self._roster_unit_id,
            "participantHandles": list(self._participant_handles), "counts": self._counts(),
            "issueCodes": self._issue_codes(), "readyToPrepare": self._ready(), "proposalDigest": digest,
        }

    def _public_assignments(self):
        return self._digest_input()["unitAssignments"], self._digest_input()["sourceDispositions"]

    def draft_result(self):
        return {
            "version": _VERSION, "outcome": "draft", "observationId": self._inspection.observation_id,
            "readyToPrepare": False, "counts": self._counts(), "issueCodes": self._issue_codes(),
        }

    def cancelled_result(self):
        return {"version": _VERSION, "outcome": "cancelled", "readyToPrepare": False}

    def approve(self, expected_digest):
        if type(expected_digest) is not str or not re.fullmatch(r"[a-f0-9]{64}", expected_digest):
            raise ValueError("expected proposal digest must be a SHA-256 digest")
        summary = self.approval_summary()
        if not self._ready() or not hmac.compare_digest(expected_digest, summary["proposalDigest"]):
            raise ValueError("proposal is not ready for approval")
        assignments, dispositions = self._public_assignments()
        return {
            "version": _VERSION, "outcome": "approved", "observationId": self._inspection.observation_id,
            "proposalDigest": summary["proposalDigest"], "readyToPrepare": True,
            "rosterUnitId": self._roster_unit_id, "participantHandles": list(self._participant_handles),
            "unitAssignments": assignments, "sourceDispositions": dispositions, "counts": self._counts(),
            "issueCodes": self._issue_codes(),
            "approval": {"status": "user-approved", "approvedProposalDigest": summary["proposalDigest"]},
        }
