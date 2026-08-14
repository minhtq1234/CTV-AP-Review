"""Capability-aware semantic validation for generated intake v2 packages."""

from __future__ import annotations

from contextlib import contextmanager
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
import dataclasses
from hashlib import sha256
from io import BytesIO
import json
from pathlib import PurePosixPath
import re
import threading

import fitz
import openpyxl
import pytest
from PIL import Image

from ctv_inspection import inspect_observation
from ctv_inventory import InventoryObservation, open_inventory_observation
from ctv_package_builder import (
    ArtifactReceipt,
    build_manifest_bytes,
    create_build_plan,
    iter_rendered_artifacts,
)
from intake_fixture_factory_v2 import _approve, materialize_v2_fixture
from intake_package_validator import _PackageReader
from intake_package_validator_v2 import (
    V2ValidationExpectation,
    canonical_v2_receipt_bytes,
    validate_v2_content_reader,
    validate_v2_publication_reader,
)


def _canonical(value: object) -> bytes:
    return json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode() + b"\n"


@contextmanager
def _opened(package_dir, source_dir):
    reader, failure = _PackageReader.open(package_dir)
    assert failure is None and reader is not None
    with open_inventory_observation(source_dir) as observation:
        try:
            yield reader, observation
        finally:
            reader.close()


def _expectation(fixture):
    return V2ValidationExpectation(
        observation_id=fixture.observation_id,
        proposal_digest=fixture.proposal_digest,
        expected_manifest_sha256=fixture.manifest_sha256,
    )


def _validate_content(fixture):
    with _opened(fixture.package_dir, fixture.source_dir) as (reader, observation):
        return validate_v2_content_reader(reader, observation, _expectation(fixture))


def _validate_publication(fixture):
    with _opened(fixture.package_dir, fixture.source_dir) as (reader, observation):
        return validate_v2_publication_reader(
            reader, observation, _expectation(fixture)
        )


def _rebuild_with_changed_decision(fixture):
    original_manifest_sha256 = sha256(
        (fixture.package_dir / "case-manifest.json").read_bytes()
    ).hexdigest()
    with open_inventory_observation(fixture.source_dir) as observation:
        inspection = inspect_observation(observation)
        approved = _approve(observation, inspection)
        changed_item = next(
            item
            for item in approved.unit_decisions
            if item.unit_kind == "pdf-page" and item.decision == "accepted"
        )
        changed = dataclasses.replace(
            changed_item,
            decision="reassigned",
            role="payment-tax-form",
        )
        changed_approved = dataclasses.replace(
            approved,
            unit_decisions=tuple(
                changed if item is changed_item else item
                for item in approved.unit_decisions
            ),
        )
        plan = create_build_plan(observation, inspection, changed_approved)
        rendered = tuple(iter_rendered_artifacts(plan, observation))
        manifest_bytes = build_manifest_bytes(
            plan, tuple(ArtifactReceipt.from_rendered(item) for item in rendered)
        )
        for item in rendered:
            path = fixture.package_dir.joinpath(*item.path.split("/"))
            path.parent.mkdir(exist_ok=True)
            path.write_bytes(item.content)
        (fixture.package_dir / "case-manifest.json").write_bytes(manifest_bytes)
    assert sha256(manifest_bytes).hexdigest() != original_manifest_sha256
    return original_manifest_sha256


def _read_document(path):
    return json.loads(path.read_bytes())


def _write_manifest(fixture, document):
    (fixture.package_dir / "case-manifest.json").write_bytes(_canonical(document))


def _rewrite_artifact(fixture, relative_path, document, *, bind_digest=True):
    content = _canonical(document)
    (fixture.package_dir / relative_path).write_bytes(content)
    if bind_digest:
        manifest = _read_document(fixture.package_dir / "case-manifest.json")
        artifact = next(item for item in manifest["artifacts"] if item["path"] == relative_path)
        artifact["size"] = len(content)
        artifact["sha256"] = sha256(content).hexdigest()
        _write_manifest(fixture, manifest)


def _bind_artifact_bytes(fixture, relative_path, content):
    (fixture.package_dir / relative_path).write_bytes(content)
    manifest = _read_document(fixture.package_dir / "case-manifest.json")
    artifact = next(
        item for item in manifest["artifacts"] if item["path"] == relative_path
    )
    artifact["size"] = len(content)
    artifact["sha256"] = sha256(content).hexdigest()
    _write_manifest(fixture, manifest)


def test_complete_content_is_valid_and_binds_manifest_artifacts_and_checks(tmp_path):
    fixture = materialize_v2_fixture("complete", tmp_path / "fixture")
    content = _validate_content(fixture)

    manifest_bytes = (fixture.package_dir / "case-manifest.json").read_bytes()
    assert content.report.outcome == "valid"
    assert content.manifest_sha256 == sha256(manifest_bytes).hexdigest()
    assert content.declared_artifact_set_sha256
    assert content.tree_sha256
    assert content.report.checks
    assert all(check.passed for check in content.report.checks)
    assert canonical_v2_receipt_bytes(content).endswith(b"\n")


@pytest.mark.parametrize(
    ("record_type", "reason"),
    [
        *(('unit', reason) for reason in (
            "duplicate",
            "irrelevant",
            "unreadable-replacement-available",
            "intentionally-omitted",
            "other",
        )),
        *(('source', reason) for reason in (
            "duplicate",
            "irrelevant",
            "unreadable-replacement-available",
            "intentionally-omitted",
            "other",
        )),
    ],
)
def test_content_accepts_real_packages_for_lossy_user_exclusion_reasons(
    tmp_path, record_type, reason
):
    fixture = materialize_v2_fixture(
        "complete",
        tmp_path / f"{record_type}-{reason}",
        unit_exclusion_reason=reason if record_type == "unit" else None,
        source_exclusion_reason=reason if record_type == "source" else "irrelevant",
    )

    content = _validate_content(fixture)

    assert content.report.outcome == "valid", (record_type, reason, content.report.errors)


def test_writer_manifest_binding_rejects_coherent_rebuilt_changed_decision(tmp_path):
    fixture = materialize_v2_fixture("complete", tmp_path / "fixture")
    expected_manifest_sha256 = _rebuild_with_changed_decision(fixture)
    expectation = V2ValidationExpectation(
        observation_id=fixture.observation_id,
        proposal_digest=fixture.proposal_digest,
        expected_manifest_sha256=expected_manifest_sha256,
    )

    with _opened(fixture.package_dir, fixture.source_dir) as (reader, observation):
        content = validate_v2_content_reader(reader, observation, expectation)

    assert content.report.outcome == "invalid"
    assert content.report.errors == ["writer-manifest-binding-mismatch"]
    binding_check = next(
        check
        for check in content.report.checks
        if check.code == "writer-manifest-binding-mismatch"
    )
    assert binding_check.evidence_refs == ["manifest"]
    assert expected_manifest_sha256.encode() not in canonical_v2_receipt_bytes(content)


@pytest.mark.parametrize(
    "value", [b"0" * 64, "0" * 63, "A" * 64, "0" * 65]
)
def test_writer_manifest_binding_expectation_is_closed(value):
    with pytest.raises(ValueError, match="^v2-validation-expectation-invalid$"):
        V2ValidationExpectation(
            observation_id="observation-" + "0" * 64,
            proposal_digest="0" * 64,
            expected_manifest_sha256=value,
        )


def test_publication_without_writer_manifest_binding_is_mechanical_only(tmp_path):
    fixture = materialize_v2_fixture(
        "complete", tmp_path / "fixture", include_receipt=True
    )
    expectation = V2ValidationExpectation(
        observation_id=fixture.observation_id,
        proposal_digest=fixture.proposal_digest,
    )

    with _opened(fixture.package_dir, fixture.source_dir) as (reader, observation):
        publication = validate_v2_publication_reader(
            reader, observation, expectation
        )

    assert publication.report.outcome == "valid"
    assert all(
        not check.code.startswith("writer-manifest-binding")
        for check in publication.report.checks
    )


def test_content_acquires_each_verified_source_once_and_skips_unacquired(
    tmp_path, monkeypatch
):
    fixture = materialize_v2_fixture("complete", tmp_path / "fixture")
    calls = []
    original = InventoryObservation.snapshot

    def counted(self, evidence_id, *, max_bytes):
        calls.append(evidence_id)
        return original(self, evidence_id, max_bytes=max_bytes)

    monkeypatch.setattr(InventoryObservation, "snapshot", counted)
    content = _validate_content(fixture)

    expected = {
        PurePosixPath(item.path).stem
        for item in fixture.manifest.sources
        if item.binding_status == "verified-content"
    }
    assert content.report.outcome == "valid"
    assert Counter(calls) == Counter({evidence_id: 1 for evidence_id in expected})


def _corrupt(fixture, case):
    manifest_path = fixture.package_dir / "case-manifest.json"
    manifest = _read_document(manifest_path)
    assignments_path = fixture.package_dir / "assignments.json"
    assignments = _read_document(assignments_path)

    if case == "artifact-digest":
        next(item for item in manifest["artifacts"] if item["kind"] == "input-pdf")["sha256"] = "0" * 64
        _write_manifest(fixture, manifest)
    elif case == "assignment-digest":
        assignments["units"][0]["role"] = "acceptance-record"
        _rewrite_artifact(fixture, "assignments.json", assignments, bind_digest=False)
    elif case == "source-binding":
        manifest["sourceObservationId"] = "observation-" + "f" * 64
        _write_manifest(fixture, manifest)
    elif case == "actual-pdf-page-count":
        document = fitz.open()
        document.new_page().insert_text((72, 72), "Synthetic changed PDF")
        document.save(fixture.package_dir / "input.pdf")
        document.close()
        content = (fixture.package_dir / "input.pdf").read_bytes()
        artifact = next(item for item in manifest["artifacts"] if item["kind"] == "input-pdf")
        artifact["size"] = len(content)
        artifact["sha256"] = sha256(content).hexdigest()
        _write_manifest(fixture, manifest)
    elif case == "hidden-target-page":
        page = manifest["pdfPages"][-1]
        page["targetPage"] = None
        page["coverageState"] = "excluded-by-user"
        _write_manifest(fixture, manifest)
    elif case == "duplicate-target-page":
        manifest["pdfPages"][-1]["targetPage"] = manifest["pdfPages"][0]["targetPage"]
        _write_manifest(fixture, manifest)
    elif case == "missing-target-page":
        manifest["pdfPages"][-1]["targetPage"] += 1
        _write_manifest(fixture, manifest)
    elif case == "participant-row":
        assignments["participants"][0]["rosterRowId"] = "roster-row-deadbeef"
        _rewrite_artifact(fixture, "assignments.json", assignments)
    elif case == "decision-type":
        unit = assignments["units"][0]
        next(item for item in manifest["decisions"] if item["decisionId"] == unit["decisionId"])["type"] = "exclude-unit"
        _write_manifest(fixture, manifest)
    elif case == "accepted-role":
        unit = next(item for item in assignments["units"] if item["unitKind"] == "pdf-page" and item["decision"] == "accepted")
        unit["role"] = "acceptance-record" if unit["role"] != "acceptance-record" else "service-contract"
        _rewrite_artifact(fixture, "assignments.json", assignments)
    elif case == "scope-order":
        unit = next(item for item in assignments["units"] if item["target"]["scope"] == "shared")
        unit["target"]["participantHandles"].reverse()
        _rewrite_artifact(fixture, "assignments.json", assignments)
    elif case == "artifact-source-ids":
        evidence = next(item for item in manifest["artifacts"] if item["kind"] == "evidence")
        evidence["sourceIds"] = [manifest["rosterMapping"]["sourceId"]]
        _write_manifest(fixture, manifest)
    elif case == "output-locator":
        unit = next(item for item in assignments["units"] if item["unitKind"] == "pdf-page")
        unit["outputLocator"]["targetPage"] += 1
        _rewrite_artifact(fixture, "assignments.json", assignments)
    elif case == "evidence-index":
        unit = next(item for item in assignments["units"] if item["outputLocator"]["kind"] == "worksheet" and item["role"] != "payment-roster")
        unit["outputLocator"]["worksheetIndex"] = 2 if unit["outputLocator"]["worksheetIndex"] == 1 else 1
        _rewrite_artifact(fixture, "assignments.json", assignments)
    elif case == "fa-code":
        manifest["faCode"] = "FA-SYNTHETIC-OTHER"
        _write_manifest(fixture, manifest)
    elif case == "package-identity":
        manifest["packageId"] = "package-" + "f" * 64
        _write_manifest(fixture, manifest)
    elif case == "nonempty-exceptions":
        exceptions = {
            "schemaVersion": "2.0",
            "items": [{
                "exceptionId": "exception-synthetic",
                "code": "synthetic-review",
                "severity": "warning",
                "evidenceRefs": [],
                "explanation": "Synthetic fixture exception.",
                "requiredAction": "Review synthetic fixture.",
                "resolution": "resolved",
            }],
        }
        _rewrite_artifact(fixture, "exceptions.json", exceptions)
    elif case == "extra-file":
        (fixture.package_dir / "extra-private-name.txt").write_text("PRIVATE-079123456789")
    elif case == "content-report-present":
        (fixture.package_dir / "validation-report.json").write_bytes(b"{}\n")
    else:
        raise AssertionError(case)


@pytest.mark.parametrize(
    "case",
    [
        "artifact-digest",
        "assignment-digest",
        "source-binding",
        "actual-pdf-page-count",
        "hidden-target-page",
        "duplicate-target-page",
        "missing-target-page",
        "participant-row",
        "decision-type",
        "accepted-role",
        "scope-order",
        "artifact-source-ids",
        "output-locator",
        "evidence-index",
        "fa-code",
        "package-identity",
        "nonempty-exceptions",
        "extra-file",
        "content-report-present",
    ],
)
def test_content_independently_rejects_each_corrupted_relationship(tmp_path, case):
    fixture = materialize_v2_fixture("complete", tmp_path / case)
    _corrupt(fixture, case)

    content = _validate_content(fixture)

    assert content.report.outcome == "invalid", case
    assert content.report.errors, case
    serialized = canonical_v2_receipt_bytes(content)
    assert str(tmp_path).encode() not in serialized
    assert b"PRIVATE-079123456789" not in serialized
    assert all(
        re.fullmatch(r"[a-z0-9]+(?:-[a-z0-9]+)*", ref)
        for check in content.report.checks
        for ref in check.evidence_refs
    )


def test_invalid_assignment_fixture_fails_semantics_not_digest_binding(tmp_path):
    fixture = materialize_v2_fixture("invalid-assignment", tmp_path / "fixture")
    content = _validate_content(fixture)

    assert content.report.outcome == "invalid"
    assert "assignment-invalid" in content.report.errors
    assert "artifact-digest-mismatch" not in content.report.errors


def _replace_source_derived_artifact(fixture, case):
    if case == "same-page-count-unrelated-pdf":
        document = fitz.open()
        output = BytesIO()
        try:
            for page_number in range(1, 5):
                page = document.new_page()
                page.insert_text((72, 72), f"UNRELATED SYNTHETIC {page_number}")
            document.save(output)
        finally:
            document.close()
        _bind_artifact_bytes(fixture, "input.pdf", output.getvalue())
        return
    if case == "changed-roster-values":
        path = fixture.package_dir / "roster.xlsx"
    elif case == "changed-evidence-workbook":
        path = next((fixture.package_dir / "evidence").glob("*.xlsx"))
    elif case == "changed-normalized-png":
        path = next((fixture.package_dir / "evidence").glob("*.png"))
        output = BytesIO()
        with Image.open(path) as image:
            changed = image.convert("RGB")
            changed.putpixel((0, 0), (250, 1, 2))
            changed.save(output, format="PNG", compress_level=9, optimize=False)
        _bind_artifact_bytes(
            fixture, path.relative_to(fixture.package_dir).as_posix(), output.getvalue()
        )
        return
    else:
        raise AssertionError(case)
    workbook = openpyxl.load_workbook(path)
    try:
        worksheet = workbook.worksheets[0]
        if case == "changed-roster-values":
            worksheet.cell(2, 2).value = "Changed Synthetic Person"
        else:
            worksheet.cell(1, 1).value = "CHANGED SYNTHETIC EVIDENCE"
        output = BytesIO()
        workbook.save(output)
    finally:
        workbook.close()
    _bind_artifact_bytes(
        fixture, path.relative_to(fixture.package_dir).as_posix(), output.getvalue()
    )


@pytest.mark.parametrize(
    "case",
    [
        "same-page-count-unrelated-pdf",
        "changed-roster-values",
        "changed-normalized-png",
        "changed-evidence-workbook",
    ],
)
def test_content_rejects_digest_rebound_nonproduction_artifact_bytes(tmp_path, case):
    fixture = materialize_v2_fixture("complete", tmp_path / case)
    _replace_source_derived_artifact(fixture, case)

    content = _validate_content(fixture)

    assert content.report.outcome == "invalid", case
    assert "production-projection-mismatch" in content.report.errors


def _corrupt_production_projection(fixture, case):
    manifest = _read_document(fixture.package_dir / "case-manifest.json")
    assignments = _read_document(fixture.package_dir / "assignments.json")
    if case == "roster-sheet":
        manifest["rosterMapping"]["sheetName"] = "Wrong synthetic sheet"
    elif case == "roster-columns":
        manifest["rosterMapping"]["canonicalToSourceColumns"]["product"] = (
            "Wrong synthetic product column"
        )
    elif case == "source-coverage":
        source = next(
            item
            for item in manifest["sources"]
            if item["bindingStatus"] == "verified-content"
            and item["coverageState"] == "assigned"
        )
        source["coverageState"] = "shared"
    elif case == "source-decision":
        source = next(
            item
            for item in manifest["sources"]
            if item["bindingStatus"] == "verified-content"
        )
        source["decisionId"] = manifest["decisions"][0]["decisionId"]
    elif case == "participant-row-ids":
        assignments["participants"][0]["rosterRowId"] = "roster-row-" + "f" * 32
        roster_path = fixture.package_dir / "roster.xlsx"
        workbook = openpyxl.load_workbook(roster_path)
        try:
            workbook.active.cell(2, 1).value = assignments["participants"][0][
                "rosterRowId"
            ]
            output = BytesIO()
            workbook.save(output)
        finally:
            workbook.close()
        _bind_artifact_bytes(fixture, "roster.xlsx", output.getvalue())
        _rewrite_artifact(fixture, "assignments.json", assignments)
        return
    elif case == "pdf-decision-id":
        manifest["pdfPages"][0]["decisionId"] = "decision-" + "f" * 32
    elif case == "decision-evidence":
        unit = assignments["units"][0]
        decision = next(
            item
            for item in manifest["decisions"]
            if item["decisionId"] == unit["decisionId"]
        )
        assignments_artifact = next(
            item for item in manifest["artifacts"] if item["kind"] == "assignments"
        )
        decision["evidenceRefs"].append(assignments_artifact["artifactId"])
    else:
        raise AssertionError(case)
    _write_manifest(fixture, manifest)


@pytest.mark.parametrize(
    "case",
    [
        "roster-sheet",
        "roster-columns",
        "source-coverage",
        "source-decision",
        "participant-row-ids",
        "pdf-decision-id",
        "decision-evidence",
    ],
)
def test_content_rejects_nonproduction_complete_projection(tmp_path, case):
    fixture = materialize_v2_fixture("complete", tmp_path / case)
    _corrupt_production_projection(fixture, case)

    content = _validate_content(fixture)

    assert content.report.outcome == "invalid", case
    assert "production-projection-mismatch" in content.report.errors


def test_mechanical_content_accepts_coherent_changed_decision_without_writer_binding(
    tmp_path,
):
    fixture = materialize_v2_fixture("complete", tmp_path / "fixture")
    manifest = _read_document(fixture.package_dir / "case-manifest.json")
    assignments = _read_document(fixture.package_dir / "assignments.json")
    unit = next(
        item
        for item in assignments["units"]
        if item["unitKind"] == "pdf-page" and item["decision"] == "accepted"
    )
    unit["decision"] = "reassigned"
    unit["role"] = (
        "acceptance-record"
        if unit["role"] != "acceptance-record"
        else "payment-tax-form"
    )
    decision = next(
        item
        for item in manifest["decisions"]
        if item["decisionId"] == unit["decisionId"]
    )
    decision["type"] = "reassign-unit"
    assignment_bytes = _canonical(assignments)
    (fixture.package_dir / "assignments.json").write_bytes(assignment_bytes)
    artifact = next(
        item for item in manifest["artifacts"] if item["kind"] == "assignments"
    )
    artifact["size"] = len(assignment_bytes)
    artifact["sha256"] = sha256(assignment_bytes).hexdigest()
    _write_manifest(fixture, manifest)

    expectation = V2ValidationExpectation(
        observation_id=fixture.observation_id,
        proposal_digest=fixture.proposal_digest,
    )
    with _opened(fixture.package_dir, fixture.source_dir) as (reader, observation):
        content = validate_v2_content_reader(reader, observation, expectation)

    assert content.report.outcome == "valid"
    assert all(
        not check.code.startswith("writer-manifest-binding")
        for check in content.report.checks
    )


def test_two_validations_share_no_snapshot_cache_or_global_serialization(
    tmp_path, monkeypatch
):
    first = materialize_v2_fixture("complete", tmp_path / "first")
    second = materialize_v2_fixture("complete", tmp_path / "second")
    original = InventoryObservation.snapshot
    installed = None
    barrier = threading.Barrier(2)
    lock = threading.Lock()
    calls = Counter()

    def counted(self, evidence_id, *, max_bytes):
        with lock:
            key = (id(self), evidence_id)
            calls[key] += 1
            first_for_observation = sum(
                value for (owner, _source), value in calls.items() if owner == id(self)
            ) == 1
        assert InventoryObservation.snapshot is installed
        if first_for_observation:
            barrier.wait(timeout=5)
        return original(self, evidence_id, max_bytes=max_bytes)

    installed = counted
    monkeypatch.setattr(InventoryObservation, "snapshot", installed)
    with ThreadPoolExecutor(max_workers=2) as executor:
        results = list(executor.map(_validate_content, (first, second)))

    assert InventoryObservation.snapshot is installed
    assert all(item.report.outcome == "valid" for item in results)
    assert len(calls) == 8
    assert set(calls.values()) == {1}


def test_publication_matches_content_and_adds_only_receipt_consistency(tmp_path):
    fixture = materialize_v2_fixture(
        "complete", tmp_path / "fixture", include_receipt=True
    )
    receipt = _read_document(fixture.package_dir / "validation-report.json")
    publication = _validate_publication(fixture)

    assert publication.report.outcome == "valid"
    assert [check.code for check in publication.report.checks[:-1]] == [
        check["code"] for check in receipt["checks"]
    ]
    assert publication.report.checks[-1].model_dump(by_alias=True) == {
        "code": "validation-report-consistent",
        "passed": True,
        "evidenceRefs": ["receipt"],
    }
    assert all(
        check["code"] != "validation-report-consistent"
        for check in receipt["checks"]
    )


def _mutate_receipt(fixture, case):
    report_path = fixture.package_dir / "validation-report.json"
    if case == "missing":
        report_path.unlink()
        return
    if case == "malformed":
        report_path.write_bytes(b"{not-json\n")
        return
    receipt = _read_document(report_path)
    if case == "stale":
        manifest = _read_document(fixture.package_dir / "case-manifest.json")
        (fixture.package_dir / "case-manifest.json").write_text(
            json.dumps(manifest, ensure_ascii=False, sort_keys=True, indent=2) + "\n"
        )
        return
    if case == "self-declared":
        manifest = _read_document(fixture.package_dir / "case-manifest.json")
        manifest["artifacts"].append({
            "artifactId": "artifact-validation-report",
            "kind": "validation-report",
            "formatVersion": "2.0",
            "path": "validation-report.json",
            "size": report_path.stat().st_size,
            "sha256": sha256(report_path.read_bytes()).hexdigest(),
            "sourceIds": [],
        })
        _write_manifest(fixture, manifest)
        return
    if case == "empty-checks":
        receipt["checks"] = []
    elif case == "private-verbose":
        receipt["diagnostic"] = "PRIVATE-079123456789 /private/source/path"
    elif case == "observation":
        receipt["sourceObservationId"] = "observation-" + "f" * 64
    elif case == "proposal":
        receipt["proposalDigest"] = "f" * 64
    elif case == "package":
        receipt["packageId"] = "package-" + "f" * 64
    elif case == "manifest":
        receipt["manifestSha256"] = "f" * 64
    elif case == "artifacts":
        receipt["declaredArtifactSetSha256"] = "f" * 64
    elif case == "outcome":
        receipt["outcome"] = "invalid"
        receipt["errors"] = [receipt["checks"][0]["code"]]
        receipt["checks"][0]["passed"] = False
    elif case == "ordered-checks":
        receipt["checks"].reverse()
    else:
        raise AssertionError(case)
    report_path.write_bytes(_canonical(receipt))


@pytest.mark.parametrize(
    "case",
    [
        "missing",
        "malformed",
        "stale",
        "self-declared",
        "empty-checks",
        "private-verbose",
        "observation",
        "proposal",
        "package",
        "manifest",
        "artifacts",
        "outcome",
        "ordered-checks",
    ],
)
def test_publication_rejects_missing_malformed_stale_or_mismatched_receipt(
    tmp_path, case
):
    fixture = materialize_v2_fixture(
        "complete", tmp_path / case, include_receipt=True
    )
    _mutate_receipt(fixture, case)

    publication = _validate_publication(fixture)

    assert publication.report.outcome == "invalid", case
    assert "validation-report-consistent" in publication.report.errors
    serialized = canonical_v2_receipt_bytes(publication)
    assert b"PRIVATE-079123456789" not in serialized
    assert b"/private/source/path" not in serialized


@pytest.mark.parametrize(
    "variant", ["indented", "reordered", "missing-final-lf", "extra-final-lf"]
)
def test_publication_requires_exact_canonical_receipt_bytes(tmp_path, variant):
    fixture = materialize_v2_fixture(
        "complete", tmp_path / variant, include_receipt=True
    )
    path = fixture.package_dir / "validation-report.json"
    receipt = _read_document(path)
    if variant == "indented":
        changed = json.dumps(
            receipt, ensure_ascii=False, sort_keys=True, indent=2
        ).encode() + b"\n"
    elif variant == "reordered":
        changed = json.dumps(
            dict(reversed(list(receipt.items()))),
            ensure_ascii=False,
            separators=(",", ":"),
        ).encode() + b"\n"
    elif variant == "missing-final-lf":
        changed = path.read_bytes()[:-1]
    else:
        changed = path.read_bytes() + b"\n"
    path.write_bytes(changed)

    publication = _validate_publication(fixture)

    assert publication.report.outcome == "invalid"
    assert publication.report.errors[-1] == "validation-report-consistent"
