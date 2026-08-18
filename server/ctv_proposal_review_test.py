import http.client
import json
from io import BytesIO
from pathlib import Path
import socket
from urllib.parse import urlsplit

import fitz
from openpyxl import Workbook
import pytest

from ctv_grouping_evidence import GroupingEvidence
from ctv_inspection import inspect_observation
from ctv_inventory import open_inventory_observation
from ctv_proposal import ProposalState
from ctv_proposal_review import ReviewError, run_local_review


PRIVATE_NAME = "PRIVATE-ROSTER-079123456789"
PRIVATE_SECOND_NAME = "PRIVATE-SECOND-078987654321"
PRIVATE_PATH_PART = "private-source-079123456789"


def _workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "payment roster private"
    sheet.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    sheet.append((PRIVATE_NAME, "CTV-001", "FA-PRIVATE-001", 100))
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _ambiguous_workbook_bytes():
    workbook = Workbook()
    for index, participant in enumerate(("FIRST-PRIVATE", "SECOND-PRIVATE")):
        sheet = workbook.active if index == 0 else workbook.create_sheet()
        sheet.title = f"payment roster {index + 1}"
        sheet.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
        sheet.append(
            (participant, f"CTV-{index + 1:03d}", f"FA-PRIVATE-{index + 1:03d}", 100)
        )
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _two_participant_workbook_bytes():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "payment roster private"
    sheet.append(("Ho ten", "Ma so nhan vien", "faCode", "So tien"))
    sheet.append((PRIVATE_NAME, "CTV-001", "FA-PRIVATE-001", 100))
    sheet.append((PRIVATE_SECOND_NAME, "CTV-002", "FA-PRIVATE-001", 100))
    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _pdf_bytes(
    text="HOP DONG DICH VU BEN A BEN B NOI DUNG CHI TIET CHU KY CAC BEN",
    *,
    pages=1,
):
    document = fitz.open()
    page_texts = (text,) * pages if type(text) is str else tuple(text)
    for page_text in page_texts:
        page = document.new_page(width=300, height=200)
        page.insert_textbox(
            fitz.Rect(10, 10, 290, 190),
            page_text,
            fontsize=8,
        )
    snapshot = document.tobytes()
    document.close()
    return snapshot


def _state(tmp_path):
    source = tmp_path / PRIVATE_PATH_PART
    source.mkdir()
    (source / "private-roster.xlsx").write_bytes(_workbook_bytes())
    contract = _pdf_bytes()
    (source / "private-contract.pdf").write_bytes(contract)
    (source / "private-contract-copy.pdf").write_bytes(contract)
    (source / "private-unknown.pdf").write_bytes(
        _pdf_bytes("MYSTERY LOCAL EVIDENCE WITHOUT A SUPPORTED DOCUMENT ROLE")
    )
    (source / "private-note.txt").write_text("local private supporting note")
    context = open_inventory_observation(source)
    observation = context.__enter__()
    facts = GroupingEvidence()
    for item in observation.result.items:
        if item.duplicate_group_id is not None:
            facts.capture_source_duplicate(item.evidence_id, item.duplicate_group_id)
    inspection = inspect_observation(
        observation,
        _private_text_sink=facts.capture,
    )
    state = ProposalState.from_inspection(
        observation,
        inspection,
        _grouping_evidence=facts,
    )
    return source, context, state


def _ambiguous_state(tmp_path):
    source = tmp_path / f"{PRIVATE_PATH_PART}-ambiguous"
    source.mkdir()
    (source / "private-rosters.xlsx").write_bytes(_ambiguous_workbook_bytes())
    context = open_inventory_observation(source)
    observation = context.__enter__()
    facts = GroupingEvidence()
    inspection = inspect_observation(
        observation,
        _private_text_sink=facts.capture,
    )
    state = ProposalState.from_inspection(
        observation,
        inspection,
        _grouping_evidence=facts,
    )
    return context, state


def _missing_roster_state(tmp_path):
    source = tmp_path / f"{PRIVATE_PATH_PART}-missing-roster"
    source.mkdir()
    (source / "private-evidence.pdf").write_bytes(
        _pdf_bytes(
            "BIEN BAN NGHIEM THU THOI GIAN NGHIEM THU BEN A BEN B CHU KY"
        )
    )
    context = open_inventory_observation(source)
    observation = context.__enter__()
    facts = GroupingEvidence()
    inspection = inspect_observation(
        observation,
        _private_text_sink=facts.capture,
    )
    state = ProposalState.from_inspection(
        observation,
        inspection,
        _grouping_evidence=facts,
    )
    return context, state


def _effective_resolution_state(tmp_path):
    source = tmp_path / f"{PRIVATE_PATH_PART}-effective"
    source.mkdir()
    (source / "private-roster.xlsx").write_bytes(
        _two_participant_workbook_bytes()
    )
    (source / "private-unknown.pdf").write_bytes(
        _pdf_bytes("MYSTERY LOCAL EVIDENCE WITHOUT A SUPPORTED DOCUMENT ROLE")
    )
    (source / "private-note.txt").write_text("local private supporting note")
    context = open_inventory_observation(source)
    observation = context.__enter__()
    facts = GroupingEvidence()
    inspection = inspect_observation(
        observation,
        _private_text_sink=facts.capture,
    )
    state = ProposalState.from_inspection(
        observation,
        inspection,
        _grouping_evidence=facts,
    )
    return context, state


def _request(parsed, method, target, *, cookie=None, csrf=None, body=None, headers=None):
    connection = http.client.HTTPConnection(parsed.hostname, parsed.port, timeout=3)
    merged = dict(headers or {})
    if cookie is not None:
        merged["Cookie"] = cookie
    if csrf is not None:
        merged["X-CSRF-Token"] = csrf
    encoded = body
    if isinstance(body, (dict, list)):
        encoded = json.dumps(body, separators=(",", ":")).encode("utf-8")
        merged.setdefault("Content-Type", "application/json")
    connection.request(method, target, body=encoded, headers=merged)
    response = connection.getresponse()
    payload = response.read()
    result = response.status, response.getheaders(), payload
    connection.close()
    return result


def _header(headers, name):
    values = [value for key, value in headers if key.lower() == name.lower()]
    assert len(values) == 1
    return values[0]


def _bootstrap(url):
    parsed = urlsplit(url)
    assert parsed.hostname == "127.0.0.1"
    status, headers, body = _request(parsed, "GET", f"{parsed.path}?{parsed.query}")
    assert status == 303
    assert body == b""
    assert _header(headers, "Location") == "/"
    cookie_header = _header(headers, "Set-Cookie")
    assert "HttpOnly" in cookie_header
    assert "SameSite=Strict" in cookie_header
    assert "Path=/" in cookie_header
    return parsed, cookie_header.split(";", 1)[0], headers


def _json(payload):
    return json.loads(payload.decode("utf-8"))


def _security_headers(headers):
    assert _header(headers, "Cache-Control") == "no-store"
    assert _header(headers, "Referrer-Policy") == "no-referrer"
    assert _header(headers, "X-Content-Type-Options") == "nosniff"
    assert _header(headers, "X-Frame-Options") == "DENY"
    csp = _header(headers, "Content-Security-Policy")
    assert "default-src 'self'" in csp
    assert "script-src 'self'" in csp
    assert "style-src 'self'" in csp
    assert "img-src 'self' blob:" in csp
    assert "connect-src 'self'" in csp
    assert "frame-ancestors 'none'" in csp


def _post(parsed, route, cookie, csrf, body):
    return _request(
        parsed,
        "POST",
        route,
        cookie=cookie,
        csrf=csrf,
        body=body,
        headers={"Origin": f"http://127.0.0.1:{parsed.port}"},
    )


def _run_visible(state, drive):
    driver_errors = []

    def visible_driver(url):
        try:
            return drive(url)
        except BaseException as error:
            driver_errors.append(error)
            raise

    try:
        return run_local_review(state, browser_open=visible_driver)
    except ReviewError:
        if driver_errors:
            raise driver_errors[0]
        raise


def test_group_state_bootstrap_is_one_time_authenticated_and_exactly_projected(tmp_path):
    source, context, state = _state(tmp_path)
    captured = {}
    snapshot_calls = 0
    original_snapshot = state.local_review_snapshot

    def counted_snapshot():
        nonlocal snapshot_calls
        snapshot_calls += 1
        return original_snapshot()

    state.local_review_snapshot = counted_snapshot

    def drive(url):
        parsed, cookie, bootstrap_headers = _bootstrap(url)
        captured["port"] = parsed.port
        _security_headers(bootstrap_headers)
        for route, content_type in (("/", "text/html"), ("/review.css", "text/css"), ("/review.js", "text/javascript")):
            status, headers, body = _request(parsed, "GET", route, cookie=cookie)
            assert status == 200
            assert _header(headers, "Content-Type").startswith(content_type)
            assert body
            _security_headers(headers)

        status, _headers, _body = _request(parsed, "GET", urlsplit(url).path + "?" + urlsplit(url).query)
        assert status == 403
        status, _headers, _body = _request(parsed, "GET", "/api/state")
        assert status == 403

        status, headers, body = _request(parsed, "GET", "/api/state", cookie=cookie)
        assert status == 200
        _security_headers(headers)
        local_state = _json(body)
        assert set(local_state) == {
            "csrfToken", "participants", "roster", "review", "summary"
        }
        assert snapshot_calls == 1
        assert set(local_state["review"]) == {
            "exceptions",
            "organizedGroups",
            "resolvedExclusions",
            "coverage",
            "issueCodes",
        }
        assert "groups" not in local_state["review"]
        assert "unitDecisions" not in local_state["review"]
        assert local_state["review"]["coverage"]["unaccountedUnits"] == 0
        assert len(local_state["review"]["exceptions"]) == 2
        assert len(local_state["review"]["organizedGroups"]) == 4
        assert local_state["roster"]["status"] == "selected"
        captured["csrf"] = local_state["csrfToken"]

        status, _headers, body = _request(
            parsed,
            "GET",
            "/api/state",
            cookie=f"theme=light; {cookie}; locale=vi",
        )
        assert status == 200
        assert _json(body)["csrfToken"] == local_state["csrfToken"]
        for malformed_cookie in (
            f"{cookie}; {cookie}",
            f"theme; {cookie}",
            f"bad name=value; {cookie}",
            f"=value; {cookie}",
            f"theme=light;; {cookie}",
        ):
            status, _headers, _body = _request(
                parsed, "GET", "/api/state", cookie=malformed_cookie
            )
            assert status == 403

        raw = socket.create_connection(("127.0.0.1", parsed.port), timeout=2)
        raw.sendall(b"GET /api/state HTTP/1.1\r\nHost: 127.0.0.1\r\n")
        raw.close()

        status, _headers, _body = _post(parsed, "/api/cancel", cookie, local_state["csrfToken"], {})
        assert status == 200
        return True

    try:
        result = _run_visible(state, drive)
        assert result == {"version": "1.0", "outcome": "cancelled", "readyToPrepare": False}
        assert PRIVATE_NAME not in repr(result)
        with pytest.raises(OSError):
            socket.create_connection(("127.0.0.1", captured["port"]), timeout=0.2)
    finally:
        context.__exit__(None, None, None)


def test_missing_roster_api_projects_unassigned_fallback_and_returns_draft(tmp_path):
    context, state = _missing_roster_state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        first = _json(body)
        assert first["participants"] == []
        assert first["roster"] == {
            "status": "missing",
            "rosterUnitId": None,
            "candidateUnitIds": [],
            "candidateSummaries": [],
            "participantHandles": [],
            "issueCodes": ["roster-missing"],
        }
        assert first["review"]["exceptions"] == [
            {
                "exceptionId": "exception-0001",
                "kind": "roster",
                "issueCode": "roster-missing",
                "allowedActions": [],
                "similarityKey": first["review"]["exceptions"][0][
                    "similarityKey"
                ],
            }
        ]
        assert first["review"]["coverage"] == {
            "groups": 1,
            "automaticallyOrganizedUnits": 0,
            "exceptionClusters": 1,
            "exceptionUnits": 1,
            "unaccountedUnits": 0,
        }
        assert first["review"]["organizedGroups"][0]["target"] is None
        assert first["summary"]["readyToPrepare"] is False

        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        second = _json(body)
        assert second["summary"]["proposalDigest"] == first["summary"][
            "proposalDigest"
        ]
        assert PRIVATE_PATH_PART not in body.decode("utf-8")

        status, _headers, body = _post(
            parsed, "/api/draft", cookie, first["csrfToken"], {}
        )
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        result = _run_visible(state, drive)
        assert result["outcome"] == "draft"
    finally:
        context.__exit__(None, None, None)


def test_grouped_roster_route_is_not_found_and_cannot_stale_review_ids(tmp_path):
    _source, context, state = _state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        before = _json(body)
        csrf = before["csrfToken"]
        before_digest = before["summary"]["proposalDigest"]
        before_group_ids = [
            group["groupId"] for group in before["review"]["organizedGroups"]
        ]
        before_exception_ids = [
            item["exceptionId"] for item in before["review"]["exceptions"]
        ]

        status, _headers, body = _post(
            parsed,
            "/api/roster",
            cookie,
            csrf,
            {"rosterUnitId": before["roster"]["rosterUnitId"]},
        )
        assert status == 404
        assert _json(body) == {"error": "route-not-found"}

        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        after = _json(body)
        assert after["summary"]["proposalDigest"] == before_digest
        assert [
            group["groupId"] for group in after["review"]["organizedGroups"]
        ] == before_group_ids
        assert [
            item["exceptionId"] for item in after["review"]["exceptions"]
        ] == before_exception_ids

        status, _headers, body = _post(parsed, "/api/draft", cookie, csrf, {})
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        assert _run_visible(state, drive)["outcome"] == "draft"
    finally:
        context.__exit__(None, None, None)


def test_ambiguous_roster_choice_uses_exact_exception_action_atomically(tmp_path):
    context, state = _ambiguous_state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        before = _json(body)
        csrf = before["csrfToken"]
        assert before["roster"]["status"] == "ambiguous"
        assert len(before["roster"]["candidateUnitIds"]) == 2
        assert len(before["review"]["exceptions"]) == 1
        roster_exception = before["review"]["exceptions"][0]
        assert set(roster_exception) == {
            "exceptionId",
            "kind",
            "issueCode",
            "recommendedAction",
            "allowedActions",
            "similarityKey",
        }
        assert roster_exception["kind"] == "roster"
        assert roster_exception["recommendedAction"] == "choose-roster"
        assert roster_exception["allowedActions"] == ["choose-roster"]
        chosen = before["roster"]["candidateUnitIds"][1]

        status, _headers, body = _post(
            parsed,
            "/api/exception",
            cookie,
            csrf,
            {
                "exceptionId": roster_exception["exceptionId"],
                "action": "choose-roster",
                "rosterUnitId": chosen,
                "applyToSimilar": False,
            },
        )
        assert status == 200
        after = _json(body)
        assert after["roster"]["status"] == "selected"
        assert after["roster"]["rosterUnitId"] == chosen
        assert roster_exception["exceptionId"] not in {
            item["exceptionId"] for item in after["review"]["exceptions"]
        }
        assert after["review"]["coverage"]["unaccountedUnits"] == 0
        assert after["summary"]["proposalDigest"] != before["summary"][
            "proposalDigest"
        ]

        status, _headers, body = _post(parsed, "/api/draft", cookie, csrf, {})
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        assert _run_visible(state, drive)["outcome"] == "draft"
    finally:
        context.__exit__(None, None, None)


def test_roster_candidate_summaries_authorize_trusted_candidate_preview(tmp_path):
    context, state = _ambiguous_state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        client = _json(body)
        csrf = client["csrfToken"]
        assert client["roster"]["candidateSummaries"] == [
            {
                "rosterUnitId": "unit-0001",
                "participantCount": 1,
                "eligible": True,
                "issueCodes": [],
            },
            {
                "rosterUnitId": "unit-0002",
                "participantCount": 1,
                "eligible": True,
                "issueCodes": [],
            },
        ]
        assert PRIVATE_NAME not in repr(client["roster"]["candidateSummaries"])
        candidate_id = client["roster"]["candidateSummaries"][1][
            "rosterUnitId"
        ]
        status, headers, preview = _request(
            parsed,
            "GET",
            f"/api/preview?unitId={candidate_id}",
            cookie=cookie,
        )
        assert status == 200
        assert _header(headers, "Content-Type") == "application/json; charset=utf-8"
        assert _json(preview)["rows"][0][:2] == ["Ho ten", "Ma so nhan vien"]

        status, _headers, body = _post(
            parsed, "/api/draft", cookie, csrf, {}
        )
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        assert _run_visible(state, drive)["outcome"] == "draft"
    finally:
        context.__exit__(None, None, None)


def test_effective_resolution_api_matches_the_consumed_package_snapshot(tmp_path):
    context, state = _effective_resolution_state(tmp_path)
    captured = {}

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200, body
        latest = _json(body)
        csrf = latest["csrfToken"]
        unknown_group = next(
            group
            for group in latest["review"]["organizedGroups"]
            if group["unitKind"] == "pdf-page" and group["role"] == "unknown"
        )
        unit_exception = next(
            item
            for item in latest["review"]["exceptions"]
            if item["kind"] == "unit-cluster"
            and unknown_group["groupId"] in item["groupIds"]
        )
        source_exception = next(
            item
            for item in latest["review"]["exceptions"]
            if item["kind"] == "source"
        )
        status, _headers, body = _post(
            parsed,
            "/api/exception",
            cookie,
            csrf,
            {
                "exceptionId": unit_exception["exceptionId"],
                "action": "assign",
                "role": "acceptance-record",
                "target": {
                    "scope": "individual",
                    "participantHandles": ["participant-0002"],
                },
                "applyToSimilar": False,
            },
        )
        assert status == 200, body
        latest = _json(body)
        status, _headers, body = _post(
            parsed,
            "/api/exception",
            cookie,
            csrf,
            {
                "exceptionId": source_exception["exceptionId"],
                "action": "exclude",
                "reason": "irrelevant",
                "applyToSimilar": False,
            },
        )
        assert status == 200
        latest = _json(body)
        for exception in list(latest["review"]["exceptions"]):
            assert exception["recommendedAction"] == "assign"
            status, _headers, body = _post(
                parsed,
                "/api/exception",
                cookie,
                csrf,
                {
                    "exceptionId": exception["exceptionId"],
                    "action": "accept-recommendation",
                    "applyToSimilar": False,
                },
            )
            assert status == 200, body
            latest = _json(body)
        captured.update(latest)
        assert latest["review"]["exceptions"] == []
        assert latest["review"]["coverage"]["unaccountedUnits"] == 0
        assert latest["summary"]["readyToPrepare"] is True
        resolved_group = next(
            group
            for group in latest["review"]["organizedGroups"]
            if group["memberUnitIds"] == ["unit-0002"]
        )
        assert resolved_group["effectiveResolution"] == {
            "action": "assign",
            "role": "acceptance-record",
            "target": {
                "scope": "individual",
                "participantHandles": ["participant-0002"],
            },
        }
        assert latest["review"]["resolvedExclusions"] == [
            {
                "exceptionId": source_exception["exceptionId"],
                "kind": "source",
                "evidenceId": source_exception["evidenceId"],
                "issueCode": "source-unsupported",
                "reason": "irrelevant",
            }
        ]
        assert PRIVATE_NAME not in repr(latest["review"])
        assert PRIVATE_SECOND_NAME not in repr(latest["review"])

        status, _headers, body = _post(
            parsed,
            "/api/approve",
            cookie,
            csrf,
            {"expectedProposalDigest": latest["summary"]["proposalDigest"]},
        )
        assert status == 200
        return True

    try:
        approved_result = _run_visible(state, drive)
        digest = approved_result["proposalDigest"]
        snapshot = state.consume_approved_package_snapshot(digest)
        unit = next(
            item for item in snapshot.unit_decisions if item.unit_id == "unit-0002"
        )
        source = next(
            item
            for item in snapshot.source_dispositions
            if item.evidence_id
            == captured["review"]["resolvedExclusions"][0]["evidenceId"]
        )
        assert (unit.role, unit.scope, unit.participant_handles) == (
            "acceptance-record",
            "individual",
            ("participant-0002",),
        )
        assert (source.decision, source.reason) == ("excluded", "irrelevant")
    finally:
        context.__exit__(None, None, None)


def test_http_boundary_rejects_malformed_authorization_routes_and_json(tmp_path):
    _source, context, state = _state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(parsed, "GET", "/api/state", cookie=cookie)
        csrf = _json(body)["csrfToken"]

        invalid_requests = [
            ("GET", "/api/state?extra=1", {"Cookie": cookie}, None),
            ("GET", "/private/file", {"Cookie": cookie}, None),
            ("GET", "/upload", {"Cookie": cookie}, None),
            ("GET", "/proxy?url=http%3A%2F%2Fexample.invalid", {"Cookie": cookie}, None),
            ("BREW", "/api/state", {"Cookie": cookie}, None),
            ("PUT", "/api/heartbeat", {"Cookie": cookie}, b"{}"),
            ("POST", "/api/heartbeat", {"Cookie": cookie, "Origin": "http://evil.invalid", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b"{}"),
            ("POST", "/api/heartbeat", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "text/plain", "X-CSRF-Token": csrf}, b"{}"),
            ("POST", "/api/heartbeat", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": "wrong"}, b"{}"),
            ("POST", "/api/heartbeat?extra=1", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b"{}"),
            ("POST", "/api/heartbeat?", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b"{}"),
            ("POST", "/api/heartbeat", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b'{"x":1,"x":2}'),
            ("POST", "/api/heartbeat", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b'{"extra":true}'),
            ("POST", "/api/roster", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b'{"rosterUnitId":true}'),
            ("POST", "/api/unit", {"Cookie": cookie, "Origin": f"http://127.0.0.1:{parsed.port}", "Content-Type": "application/json", "X-CSRF-Token": csrf}, b'{"unitId":"unit-9999","decision":"not-valid"}'),
        ]
        for method, route, headers, request_body in invalid_requests:
            status, response_headers, response_body = _request(
                parsed, method, route, body=request_body, headers=headers
            )
            assert status in {400, 403, 404, 405, 415}
            assert _json(response_body) == {"error": _json(response_body)["error"]}
            _security_headers(response_headers)

        connection = http.client.HTTPConnection(parsed.hostname, parsed.port, timeout=3)
        connection.putrequest("POST", "/api/heartbeat", skip_host=True)
        connection.putheader("Host", f"127.0.0.1:{parsed.port}")
        connection.putheader("Cookie", cookie)
        connection.putheader("Origin", f"http://127.0.0.1:{parsed.port}")
        connection.putheader("Content-Type", "application/json")
        connection.putheader("X-CSRF-Token", csrf)
        connection.putheader("Content-Length", str(1024 * 1024 + 1))
        connection.endheaders()
        response = connection.getresponse()
        assert response.status == 413
        assert _json(response.read()) == {"error": "request-too-large"}
        connection.close()

        status, _headers, body = _post(parsed, "/api/heartbeat", cookie, csrf, {})
        assert status == 200
        assert _json(body) == {"status": "active"}
        status, _headers, body = _post(parsed, "/api/draft", cookie, csrf, {})
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        result = _run_visible(state, drive)
        assert result["outcome"] == "draft"
        assert PRIVATE_NAME not in repr(result)
    finally:
        context.__exit__(None, None, None)


def test_exception_routes_reject_hostile_requests_without_poisoning_session(
    tmp_path,
):
    _source, context, state = _state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        initial = _json(body)
        csrf = initial["csrfToken"]
        initial_digest = initial["summary"]["proposalDigest"]
        cluster = next(
            item
            for item in initial["review"]["exceptions"]
            if item["kind"] == "unit-cluster"
        )
        automatic = next(
            group
            for group in initial["review"]["organizedGroups"]
            if group["state"] == "automatically-organized"
            and group["role"]
        )
        origin = f"http://127.0.0.1:{parsed.port}"
        good_headers = {
            "Cookie": cookie,
            "Origin": origin,
            "Content-Type": "application/json",
            "X-CSRF-Token": csrf,
        }
        private_safe_bodies = []

        def still_usable():
            heartbeat_status, _heartbeat_headers, heartbeat_body = _post(
                parsed, "/api/heartbeat", cookie, csrf, {}
            )
            assert heartbeat_status == 200
            assert _json(heartbeat_body) == {"status": "active"}

        def reject(method, route, *, request_body=b"{}", headers=None, expected=None):
            response_status, response_headers, response_body = _request(
                parsed,
                method,
                route,
                body=request_body,
                headers=good_headers if headers is None else headers,
            )
            assert response_status in (expected or {400, 403, 404, 405, 415})
            parsed_body = _json(response_body)
            assert parsed_body == {"error": parsed_body["error"]}
            assert len(response_body) < 128
            _security_headers(response_headers)
            private_safe_bodies.append(response_body)
            still_usable()

        hostile_json = (
            (
                "/api/exception",
                b'{"exceptionId":"exception-0001","exceptionId":"exception-0001","action":"exclude","reason":"irrelevant","applyToSimilar":false}',
            ),
            ("/api/exception", b'{"action":"exclude","reason":"irrelevant","applyToSimilar":false}'),
            ("/api/exception", b'{"exceptionId":"exception-0001","action":"exclude","reason":"irrelevant","applyToSimilar":false,"extra":0}'),
            ("/api/exception", b'{"exceptionId":true,"action":"exclude","reason":"irrelevant","applyToSimilar":false}'),
            ("/api/exception", b'{"exceptionId":"exception-9999","action":"exclude","reason":"irrelevant","applyToSimilar":false}'),
            ("/api/exception", json.dumps({
                "exceptionId": cluster["exceptionId"],
                "action": "assign",
                "role": "not-a-role",
                "target": {"scope": "case", "participantHandles": []},
                "applyToSimilar": False,
            }, separators=(",", ":")).encode()),
            ("/api/exception", json.dumps({
                "exceptionId": cluster["exceptionId"],
                "action": "assign",
                "role": "service-contract",
                "target": {"scope": "remote", "participantHandles": []},
                "applyToSimilar": False,
            }, separators=(",", ":")).encode()),
            ("/api/exception", json.dumps({
                "exceptionId": cluster["exceptionId"],
                "action": "exclude",
                "reason": "not-a-reason",
                "applyToSimilar": False,
            }, separators=(",", ":")).encode()),
            ("/api/exception", json.dumps({
                "exceptionId": cluster["exceptionId"],
                "action": "split",
                "splitBeforeUnitId": initial["roster"]["rosterUnitId"],
                "applyToSimilar": False,
            }, separators=(",", ":")).encode()),
            (
                "/api/exception/undo",
                b'{"exceptionId":"exception-0001","exceptionId":"exception-0001"}',
            ),
            ("/api/exception/undo", b'{}'),
            ("/api/exception/undo", b'{"exceptionId":false}'),
            ("/api/exception/undo", b'{"exceptionId":"exception-9999"}'),
            (
                "/api/group/reopen",
                b'{"groupId":"group-0001","groupId":"group-0001"}',
            ),
            ("/api/group/reopen", b'{}'),
            ("/api/group/reopen", b'{"groupId":0}'),
            ("/api/group/reopen", b'{"groupId":"group-9999"}'),
        )
        for route, payload in hostile_json:
            reject("POST", route, request_body=payload, expected={400})

        route_bodies = {
            "/api/exception": json.dumps({
                "exceptionId": "exception-9999",
                "action": "exclude",
                "reason": "irrelevant",
                "applyToSimilar": False,
            }, separators=(",", ":")).encode(),
            "/api/exception/undo": b'{"exceptionId":"exception-9999"}',
            "/api/group/reopen": b'{"groupId":"group-9999"}',
        }
        for route, payload in route_bodies.items():
            reject("POST", f"{route}?extra=1", request_body=payload, expected={400})
            reject("GET", route, request_body=None, headers={"Cookie": cookie}, expected={404})
            reject("PUT", route, request_body=payload, expected={405})
            reject(
                "POST",
                route,
                request_body=payload,
                headers={**good_headers, "Host": "evil.invalid"},
                expected={400},
            )
            reject(
                "POST",
                route,
                request_body=payload,
                headers={**good_headers, "Origin": "http://evil.invalid"},
                expected={403},
            )
            reject(
                "POST",
                route,
                request_body=payload,
                headers={**good_headers, "X-CSRF-Token": "wrong"},
                expected={403},
            )
            reject(
                "POST",
                route,
                request_body=payload,
                headers={**good_headers, "Cookie": "bad=token"},
                expected={403},
            )
            reject(
                "POST",
                route,
                request_body=payload,
                headers={**good_headers, "Content-Type": "text/plain"},
                expected={415},
            )
            connection = http.client.HTTPConnection(
                parsed.hostname, parsed.port, timeout=3
            )
            connection.putrequest("POST", route, skip_host=True)
            connection.putheader("Host", f"127.0.0.1:{parsed.port}")
            connection.putheader("Cookie", cookie)
            connection.putheader("Origin", origin)
            connection.putheader("Content-Type", "application/json")
            connection.putheader("X-CSRF-Token", csrf)
            connection.putheader("Content-Length", str(1024 * 1024 + 1))
            connection.endheaders()
            response = connection.getresponse()
            assert response.status == 413
            oversized_body = response.read()
            assert _json(oversized_body) == {"error": "request-too-large"}
            private_safe_bodies.append(oversized_body)
            connection.close()
            still_usable()

        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        assert _json(body)["summary"]["proposalDigest"] == initial_digest

        resolve_payload = {
            "exceptionId": cluster["exceptionId"],
            "action": "exclude",
            "reason": "irrelevant",
            "applyToSimilar": False,
        }
        status, _headers, _body = _post(
            parsed, "/api/exception", cookie, csrf, resolve_payload
        )
        assert status == 200
        reject(
            "POST",
            "/api/exception",
            request_body=json.dumps(resolve_payload, separators=(",", ":")).encode(),
            expected={400},
        )
        status, _headers, _body = _post(
            parsed,
            "/api/exception/undo",
            cookie,
            csrf,
            {"exceptionId": cluster["exceptionId"]},
        )
        assert status == 200
        reject(
            "POST",
            "/api/exception/undo",
            request_body=json.dumps(
                {"exceptionId": cluster["exceptionId"]}, separators=(",", ":")
            ).encode(),
            expected={400},
        )
        status, _headers, _body = _post(
            parsed,
            "/api/group/reopen",
            cookie,
            csrf,
            {"groupId": automatic["groupId"]},
        )
        assert status == 200
        reject(
            "POST",
            "/api/group/reopen",
            request_body=json.dumps(
                {"groupId": automatic["groupId"]}, separators=(",", ":")
            ).encode(),
            expected={400},
        )

        serialized_errors = b"".join(private_safe_bodies)
        assert PRIVATE_NAME.encode() not in serialized_errors
        assert PRIVATE_PATH_PART.encode() not in serialized_errors
        assert csrf.encode() not in serialized_errors

        status, _headers, body = _post(parsed, "/api/draft", cookie, csrf, {})
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        result = _run_visible(state, drive)
        assert result["outcome"] == "draft"
    finally:
        context.__exit__(None, None, None)


def test_huge_ascii_content_length_is_bounded_without_terminating_session(
    tmp_path,
):
    _source, context, state = _state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        csrf = _json(body)["csrfToken"]

        connection = http.client.HTTPConnection(
            parsed.hostname, parsed.port, timeout=3
        )
        connection.putrequest("POST", "/api/heartbeat", skip_host=True)
        connection.putheader("Host", f"127.0.0.1:{parsed.port}")
        connection.putheader("Cookie", cookie)
        connection.putheader("Origin", f"http://127.0.0.1:{parsed.port}")
        connection.putheader("Content-Type", "application/json")
        connection.putheader("X-CSRF-Token", csrf)
        connection.putheader("Content-Length", "9" * 5_000)
        connection.endheaders()
        response = connection.getresponse()
        assert response.status == 413
        assert _json(response.read()) == {"error": "request-too-large"}
        connection.close()

        status, _headers, body = _post(
            parsed, "/api/heartbeat", cookie, csrf, {}
        )
        assert status == 200
        assert _json(body) == {"status": "active"}
        status, _headers, body = _post(parsed, "/api/draft", cookie, csrf, {})
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        result = _run_visible(state, drive)
        assert result["outcome"] == "draft"
    finally:
        context.__exit__(None, None, None)


def test_exception_route_undo_group_reopen_and_group_preview_are_atomic(tmp_path):
    _source, context, state = _state(tmp_path)

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        client_state = _json(body)
        csrf = client_state["csrfToken"]
        review = client_state["review"]
        before_count = len(review["exceptions"])
        before_coverage = review["coverage"]

        roster_unit_id = client_state["roster"]["rosterUnitId"]
        assert any(
            roster_unit_id in group["memberUnitIds"]
            for group in review["organizedGroups"]
        )
        status, headers, body = _request(
            parsed,
            "GET",
            f"/api/preview?unitId={roster_unit_id}",
            cookie=cookie,
        )
        assert status == 200
        assert _header(headers, "Content-Type") == "application/json; charset=utf-8"
        assert _json(body)["rows"][0][:2] == ["Ho ten", "Ma so nhan vien"]

        for route in (
            "/api/preview?unitId=unit-9999",
            f"/api/preview?%75nitId={roster_unit_id}",
            f"/api/preview?unitId={roster_unit_id}&extra=1",
            f"/api/preview?unitId={roster_unit_id}&unitId={roster_unit_id}",
            f"/api/preview?sourcePath=x&unitId={roster_unit_id}",
        ):
            status, _headers, response = _request(
                parsed, "GET", route, cookie=cookie
            )
            assert status in {400, 404}
            assert _json(response)["error"] in {"invalid-request", "preview-not-found"}

        exception = next(
            item for item in review["exceptions"] if item["kind"] == "unit-cluster"
        )
        status, _headers, body = _post(
            parsed,
            "/api/exception",
            cookie,
            csrf,
            {
                "exceptionId": exception["exceptionId"],
                "action": "exclude",
                "reason": "irrelevant",
                "applyToSimilar": False,
            },
        )
        assert status == 200
        resolved = _json(body)
        assert len(resolved["review"]["exceptions"]) == before_count - 1
        assert resolved["review"]["coverage"]["unaccountedUnits"] == 0

        status, _headers, body = _post(
            parsed,
            "/api/exception/undo",
            cookie,
            csrf,
            {"exceptionId": exception["exceptionId"]},
        )
        assert status == 200
        undone = _json(body)
        assert [
            item["exceptionId"] for item in undone["review"]["exceptions"]
        ] == [item["exceptionId"] for item in review["exceptions"]]
        assert undone["review"]["coverage"] == before_coverage

        automatic = next(
            group
            for group in undone["review"]["organizedGroups"]
            if group["state"] == "automatically-organized" and group["role"]
        )
        duplicate = next(
            group
            for group in undone["review"]["organizedGroups"]
            if group["state"] == "automatically-organized" and not group["role"]
        )
        latest = undone
        for group in (automatic, duplicate):
            prior = len(latest["review"]["exceptions"])
            status, _headers, body = _post(
                parsed,
                "/api/group/reopen",
                cookie,
                csrf,
                {"groupId": group["groupId"]},
            )
            assert status == 200
            latest = _json(body)
            assert len(latest["review"]["exceptions"]) == prior + 1
            assert latest["review"]["coverage"]["unaccountedUnits"] == 0

        for removed_route in ("/api/unit", "/api/source"):
            status, _headers, body = _post(
                parsed, removed_route, cookie, csrf, {}
            )
            assert status == 404
            assert _json(body) == {"error": "route-not-found"}

        status, _headers, body = _post(parsed, "/api/draft", cookie, csrf, {})
        assert status == 200
        assert _json(body)["outcome"] == "draft"
        return True

    try:
        result = _run_visible(state, drive)
        assert result["outcome"] == "draft"
        assert PRIVATE_NAME not in repr(result)
    finally:
        context.__exit__(None, None, None)


def test_exception_routes_can_resolve_complete_proposal_then_approve(tmp_path):
    _source, context, state = _state(tmp_path)
    approved_response = {}

    def drive(url):
        parsed, cookie, _headers = _bootstrap(url)
        status, _headers, body = _request(
            parsed, "GET", "/api/state", cookie=cookie
        )
        assert status == 200
        latest = _json(body)
        csrf = latest["csrfToken"]

        while latest["review"]["exceptions"]:
            exception = latest["review"]["exceptions"][0]
            if exception.get("recommendedAction") is not None:
                payload = {
                    "exceptionId": exception["exceptionId"],
                    "action": "accept-recommendation",
                    "applyToSimilar": False,
                }
            else:
                payload = {
                    "exceptionId": exception["exceptionId"],
                    "action": "exclude",
                    "reason": "irrelevant",
                    "applyToSimilar": False,
                }
            status, _headers, body = _post(
                parsed, "/api/exception", cookie, csrf, payload
            )
            assert status == 200
            latest = _json(body)
            assert latest["review"]["coverage"]["unaccountedUnits"] == 0

        status, _headers, body = _post(parsed, "/api/summary", cookie, csrf, {})
        assert status == 200
        summary = _json(body)
        assert summary["readyToPrepare"] is True
        assert PRIVATE_NAME not in repr(summary)

        status, _headers, body = _post(
            parsed,
            "/api/approve",
            cookie,
            csrf,
            {"expectedProposalDigest": "0" * 64},
        )
        assert status == 400
        assert _json(body) == {"error": "invalid-request"}
        status, _headers, body = _post(
            parsed,
            "/api/approve",
            cookie,
            csrf,
            {"expectedProposalDigest": summary["proposalDigest"]},
        )
        assert status == 200
        approved_response.update(_json(body))
        return True

    try:
        result = _run_visible(state, drive)
        assert result == approved_response
        assert result["outcome"] == "approved"
        assert "review" not in result
        assert PRIVATE_NAME not in repr(result)
    finally:
        context.__exit__(None, None, None)


def test_browser_failure_and_timeout_are_fixed_cleanup_errors_without_logs_or_writes(tmp_path, capsys):
    import ctv_proposal_review as review

    source, context, state = _state(tmp_path)
    before = {path.relative_to(source): path.read_bytes() for path in source.rglob("*") if path.is_file()}
    try:
        with pytest.raises(ReviewError, match="^review-browser-open-failed$"):
            run_local_review(state, browser_open=lambda _url: False)
        with pytest.raises(ReviewError, match="^review-timeout$"):
            run_local_review(
                state,
                browser_open=lambda _url: True,
                clock=iter((0.0, 301.0, 301.0)).__next__,
            )
        with pytest.raises(ReviewError, match="^review-timeout$"):
            run_local_review(
                state,
                browser_open=lambda _url: True,
                clock=iter((0.0, 7200.0, 7200.0)).__next__,
            )

        class BindFailure:
            def __init__(self, *_args, **_kwargs):
                raise OSError("PRIVATE-BIND-DETAIL")

        original_server = review.HTTPServer
        review.HTTPServer = BindFailure
        try:
            with pytest.raises(ReviewError, match="^review-server-failed$"):
                run_local_review(state, browser_open=lambda _url: True)
        finally:
            review.HTTPServer = original_server
    finally:
        context.__exit__(None, None, None)
    after = {path.relative_to(source): path.read_bytes() for path in source.rglob("*") if path.is_file()}
    assert after == before
    captured = capsys.readouterr()
    assert captured.out == ""
    assert captured.err == ""
    assert PRIVATE_NAME not in captured.out + captured.err
    assert PRIVATE_PATH_PART not in captured.out + captured.err
